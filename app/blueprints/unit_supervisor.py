from flask import Blueprint, render_template, request, jsonify, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from app.extensions import db
from app.models import *
from datetime import datetime
import json
import jdatetime
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

unit_supervisor_bp = Blueprint('unit_supervisor', __name__, url_prefix='/unit-supervisor')


# ==================== پنل سرپرست واحد - صفحات ====================

@unit_supervisor_bp.route('/dashboard')
@login_required
def dashboard():
    if current_user.role not in ['unit_supervisor', 'admin']:
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('user.dashboard'))
    
    # دریافت واحدهای تحت سرپرستی
    if current_user.role == 'admin':
        supervised_units = Unit.query.limit(1).all()
        if not supervised_units:
            flash('هیچ واحدی در سیستم ثبت نشده است', 'warning')
            return render_template('unit_supervisor/dashboard.html', 
                                  unit_id=None,
                                  unit_name='واحد نمونه (نمایش ادمین)',
                                  department_name='اداره نمونه',
                                  period_title='دوره نمونه',
                                  today_date=jdatetime.datetime.now().strftime('%Y/%m/%d'),
                                  department_color='#667eea')
    else:
        supervised_units = db.session.query(Unit).join(
            UnitSupervisor, UnitSupervisor.unit_id == Unit.id
        ).filter(UnitSupervisor.user_id == current_user.id).all()
        
        if not supervised_units:
            flash('شما به عنوان سرپرست به هیچ واحدی متصل نیستید. لطفاً با ادمین تماس بگیرید.', 'error')
            # ✅ تغییر: به جای ریدایرکت به dashboard، یک صفحه خطا نشان بده
            return render_template('unit_supervisor/dashboard.html', 
                                  unit_id=None,
                                  unit_name='شما به هیچ واحدی متصل نیستید',
                                  department_name='اداره نامشخص',
                                  period_title='تعریف نشده',
                                  today_date=jdatetime.datetime.now().strftime('%Y/%m/%d'),
                                  department_color='#667eea',
                                  error_message='شما به عنوان سرپرست واحد ثبت شده‌اید، اما به هیچ واحدی متصل نیستید. لطفاً با ادمین تماس بگیرید.')
    
    unit = supervised_units[0]
    department = Department.query.get(unit.department_id)
    active_period = WorkPeriod.query.filter_by(is_active=True).first()
    today_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    department_color = department.color if department and department.color else '#667eea'
    
    return render_template('unit_supervisor/dashboard.html', 
                          unit_id=unit.id,
                          unit_name=unit.name,
                          department_name=department.name if department else '-',
                          period_title=active_period.title if active_period else 'تعریف نشده',
                          today_date=today_date,
                          department_color=department_color)

@unit_supervisor_bp.route('/requests')
@login_required
def unit_supervisor_requests_page():
    if current_user.role != 'unit_supervisor':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    return render_template('unit_supervisor/requests_list.html')


# ==================== APIهای سرپرست واحد ====================

@unit_supervisor_bp.route('/api/all-data')
@login_required
def unit_supervisor_api_all_data():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    
    if not supervised_units:
        return jsonify({'error': 'واحد یافت نشد'}), 404
    
    unit = supervised_units[0]
    period_id = request.args.get('period_id', type=int)
    
    dynamic_fields = DynamicField.query.filter_by(is_active=True).order_by(DynamicField.field_order).all()
    
    personnel = Personnel.query.filter_by(unit_id=unit.id, is_deleted=False).all()
    all_periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    
    personnel_data = []
    for p in personnel:
        if period_id and p.period_id != period_id:
            continue
        
        values = {}
        
        if period_id:
            pv_query = PersonnelValue.query.filter_by(personnel_id=p.id, period_id=period_id)
        else:
            pv_query = PersonnelValue.query.filter_by(personnel_id=p.id)
        
        for v in pv_query.all():
            field = DynamicField.query.get(v.field_id)
            if field:
                values[field.title] = v.value_text or v.value_number or v.value_date or ''
        
        period_title = ''
        if p.period_id:
            period = WorkPeriod.query.get(p.period_id)
            if period:
                period_title = f"{period.title} ({period.start_date} - {period.end_date})"
        
        item = {
            'id': p.id,
            'national_code': p.national_code,
            'first_name': p.first_name or '',
            'last_name': p.last_name or '',
            'full_name': p.get_full_name(),
            'phone': p.phone or '',
            'position': p.position or '',
            'period_title': period_title,
            'period_id': p.period_id
        }
        
        for field in dynamic_fields:
            if not field.is_key:
                value = values.get(field.title, '')
                if not value and p.period_id:
                    pv_default = PersonnelValue.query.filter_by(
                        personnel_id=p.id, 
                        field_id=field.id, 
                        period_id=p.period_id
                    ).first()
                    if pv_default:
                        value = pv_default.value_text or pv_default.value_number or pv_default.value_date or ''
                item[field.title] = value
        
        personnel_data.append(item)
    
    visible_fields = [f for f in dynamic_fields if not f.is_key and f.title not in ['نام', 'نام خانوادگی']]
    total_fields = len(visible_fields)
    total_completed_fields = 0
    completed_personnel = 0
    
    for p in personnel_data:
        filled = 0
        for f in visible_fields:
            val = p.get(f.title, '')
            if val and val != '-':
                filled += 1
        total_completed_fields += filled
        if filled == total_fields and total_fields > 0:
            completed_personnel += 1
    
    total_possible = len(personnel_data) * total_fields if total_fields > 0 else 1
    completion_percent = int((total_completed_fields / total_possible) * 100) if total_possible > 0 else 0
    
    return jsonify({
        'personnel': personnel_data,
        'dynamic_fields': [{'id': f.id, 'title': f.title, 'field_type': f.field_type, 'is_key': f.is_key, 'is_required': f.is_required} for f in dynamic_fields],
        'periods': [{'id': p.id, 'title': p.title, 'start_date': p.start_date, 'end_date': p.end_date} for p in all_periods],
        'stats': {
            'total': len(personnel_data),
            'total_fields': total_fields,
            'completed_personnel': completed_personnel,
            'completion_percent': completion_percent,
            'incomplete': len(personnel_data) - completed_personnel
        }
    })


@unit_supervisor_bp.route('/api/periods')
@login_required
def unit_supervisor_api_periods():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    return jsonify([{
        'id': p.id,
        'title': p.title,
        'start_date': p.start_date,
        'end_date': p.end_date,
        'is_active': p.is_active
    } for p in periods])


@unit_supervisor_bp.route('/api/personnel-history/<int:pid>')
@login_required
def unit_supervisor_api_personnel_history(pid):
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    p = Personnel.query.get_or_404(pid)
    fields = DynamicField.query.filter_by(is_active=True).all()
    
    periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    
    history = []
    for period in periods:
        period_data = {
            'period_id': period.id,
            'period_title': period.title,
            'period_start': period.start_date,
            'period_end': period.end_date,
            'values': {}
        }
        
        for field in fields:
            pv = PersonnelValue.query.filter_by(
                personnel_id=pid, 
                field_id=field.id,
                period_id=period.id
            ).first()
            
            if pv:
                value = pv.value_text or pv.value_number or pv.value_date or '-'
            else:
                current_pv = PersonnelValue.query.filter_by(
                    personnel_id=pid, 
                    field_id=field.id,
                    period_id=None
                ).first()
                if current_pv:
                    value = current_pv.value_text or current_pv.value_number or current_pv.value_date or '-'
                else:
                    value = '-'
            
            period_data['values'][field.title] = value
        
        history.append(period_data)
    
    return jsonify({
        'personnel': {
            'id': p.id,
            'national_code': p.national_code,
            'full_name': p.get_full_name()
        },
        'fields': [{'id': f.id, 'title': f.title, 'field_type': f.field_type} for f in fields],
        'history': history
    })


@unit_supervisor_bp.route('/api/pending-requests')
@login_required
def unit_supervisor_api_pending_requests():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    
    if not supervised_units:
        return jsonify({'requests': []})
    
    unit = supervised_units[0]
    
    requests = UnitPersonnelRequest.query.filter_by(
        unit_id=unit.id, 
        request_type='add',
        status='pending'
    ).order_by(UnitPersonnelRequest.created_at.desc()).all()
    
    result = []
    for req in requests:
        data = json.loads(req.data) if req.data else {}
        created_at_jalali = jdatetime.datetime.fromgregorian(datetime=req.created_at).strftime('%Y/%m/%d %H:%M')
        
        result.append({
            'id': req.id,
            'request_type': req.request_type,
            'national_code': data.get('national_code', ''),
            'first_name': data.get('first_name', ''),
            'last_name': data.get('last_name', ''),
            'full_name': f"{data.get('first_name', '')} {data.get('last_name', '')}".strip(),
            'created_at': created_at_jalali
        })
    
    return jsonify({'requests': result})


@unit_supervisor_bp.route('/api/delete-requests')
@login_required
def unit_supervisor_api_delete_requests():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    
    if not supervised_units:
        return jsonify({'requests': []})
    
    unit = supervised_units[0]
    
    requests = UnitPersonnelRequest.query.filter_by(
        unit_id=unit.id, 
        request_type='delete',
        status='pending'
    ).order_by(UnitPersonnelRequest.created_at.desc()).all()
    
    result = []
    for req in requests:
        data = json.loads(req.data) if req.data else {}
        created_at_jalali = jdatetime.datetime.fromgregorian(datetime=req.created_at).strftime('%Y/%m/%d %H:%M')
        
        result.append({
            'id': req.id,
            'request_type': req.request_type,
            'personnel_id': data.get('personnel_id'),
            'national_code': data.get('national_code', ''),
            'first_name': data.get('first_name', ''),
            'last_name': data.get('last_name', ''),
            'full_name': data.get('full_name', ''),
            'delete_reason': data.get('delete_reason', ''),
            'created_at': created_at_jalali
        })
    
    return jsonify({'requests': result})


@unit_supervisor_bp.route('/api/request-add', methods=['POST'])
@login_required
def unit_supervisor_api_request_add():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    
    if not supervised_units:
        return jsonify({'error': 'شما به هیچ واحدی متصل نیستید'}), 400
    
    unit = supervised_units[0]
    data = request.get_json()
    
    period_id = data.get('period_id')
    national_code = data.get('national_code')
    
    if not period_id:
        return jsonify({'error': 'دوره انتخاب نشده است'}), 400
    if not national_code:
        return jsonify({'error': 'کد ملی الزامی است'}), 400
    
    if not (national_code.isdigit() and len(national_code) == 10):
        return jsonify({'error': 'کد ملی باید 10 رقم باشد'}), 400
    
    existing = Personnel.query.filter_by(
        national_code=national_code, 
        period_id=period_id, 
        is_deleted=False
    ).first()
    if existing:
        return jsonify({'error': 'این کد ملی قبلاً در دوره انتخاب شده ثبت شده است'}), 400
    
    request_data = {
        'department_id': unit.department_id,
        'unit_id': unit.id,
        'national_code': national_code,
        'period_id': period_id,
        'requester_note': data.get('note', '')
    }
    
    fields = DynamicField.query.filter_by(is_active=True).all()
    for field in fields:
        value = data.get(field.title)
        if value:
            request_data[field.title] = value
            if field.title == 'نام':
                request_data['first_name'] = value
            if field.title == 'نام خانوادگی':
                request_data['last_name'] = value
            if field.title == 'شماره تماس':
                request_data['phone'] = value
            if field.title == 'سمت':
                request_data['position'] = value
    
    approval = UnitPersonnelRequest(
        unit_id=unit.id,
        requester_id=current_user.id,
        request_type='add',
        data=json.dumps(request_data, ensure_ascii=False),
        status='pending'
    )
    db.session.add(approval)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست شما ثبت شد و منتظر تایید ادمین است'})


@unit_supervisor_bp.route('/api/request-delete', methods=['POST'])
@login_required
def unit_supervisor_api_request_delete():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'داده ارسال نشده است'}), 400
            
        personnel_id = data.get('personnel_id')
        delete_reason = data.get('delete_reason', '')
        
        if not personnel_id:
            return jsonify({'error': 'شناسه پرسنل ارسال نشده است'}), 400
        
        p = Personnel.query.get(personnel_id)
        if not p:
            return jsonify({'error': 'پرسنل یافت نشد'}), 404
        
        supervised_units = db.session.query(Unit).join(
            UnitSupervisor, UnitSupervisor.unit_id == Unit.id
        ).filter(UnitSupervisor.user_id == current_user.id).all()
        
        if not supervised_units:
            return jsonify({'error': 'واحد یافت نشد'}), 404
        
        unit = supervised_units[0]
        
        UnitPersonnelRequest.query.filter_by(
            unit_id=unit.id,
            target_personnel_id=personnel_id,
            request_type='delete',
            status='pending'
        ).delete()
        db.session.commit()
        
        request_data = {
            'personnel_id': p.id,
            'national_code': p.national_code,
            'first_name': p.first_name or '',
            'last_name': p.last_name or '',
            'full_name': p.get_full_name(),
            'phone': p.phone or '',
            'position': p.position or '',
            'delete_reason': delete_reason
        }
        
        approval = UnitPersonnelRequest(
            unit_id=unit.id,
            requester_id=current_user.id,
            request_type='delete',
            target_personnel_id=personnel_id,
            data=json.dumps(request_data, ensure_ascii=False),
            status='pending'
        )
        db.session.add(approval)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'درخواست حذف پرسنل ثبت شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
        return jsonify({'error': f'خطا: {str(e)}'}), 500


@unit_supervisor_bp.route('/api/cancel-request', methods=['POST'])
@login_required
def unit_supervisor_api_cancel_request():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        request_id = data.get('request_id')
        
        if not request_id:
            return jsonify({'error': 'شناسه درخواست ارسال نشده است'}), 400
        
        supervised_units = db.session.query(Unit).join(
            UnitSupervisor, UnitSupervisor.unit_id == Unit.id
        ).filter(UnitSupervisor.user_id == current_user.id).all()
        
        if not supervised_units:
            return jsonify({'error': 'واحد یافت نشد'}), 404
        
        unit = supervised_units[0]
        
        req = UnitPersonnelRequest.query.filter_by(
            id=request_id,
            unit_id=unit.id,
            status='pending'
        ).first()
        
        if not req:
            return jsonify({'error': 'درخواست یافت نشد'}), 404
        
        db.session.delete(req)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'درخواست با موفقیت لغو شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in cancel-request: {e}")
        return jsonify({'error': str(e)}), 500


@unit_supervisor_bp.route('/api/personnel/<int:pid>')
@login_required
def unit_supervisor_api_personnel_get(pid):
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    p = Personnel.query.get_or_404(pid)
    fields = DynamicField.query.filter_by(is_active=True).all()
    
    values = {}
    for v in PersonnelValue.query.filter_by(personnel_id=pid).all():
        field = DynamicField.query.get(v.field_id)
        if field:
            values[field.title] = v.value_text or v.value_number or v.value_date or ''
    
    result = {
        'id': p.id,
        'national_code': p.national_code,
        'first_name': p.first_name or '',
        'last_name': p.last_name or '',
        'phone': p.phone or '',
        'position': p.position or '',
        'period_id': p.period_id
    }
    
    for field in fields:
        result[field.title] = values.get(field.title, '')
    
    return jsonify(result)


@unit_supervisor_bp.route('/api/personnel/<int:pid>/edit', methods=['PUT'])
@login_required
def unit_supervisor_api_personnel_edit(pid):
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        p = Personnel.query.get_or_404(pid)
        data = request.get_json()
        
        p.first_name = data.get('first_name', p.first_name)
        p.last_name = data.get('last_name', p.last_name)
        p.phone = data.get('phone', p.phone)
        p.position = data.get('position', p.position)
        
        period_id = data.get('period_id')
        if period_id and period_id != '' and period_id != 'null':
            p.period_id = int(period_id)
        else:
            p.period_id = None
        
        db.session.commit()
        
        fields = DynamicField.query.filter_by(is_active=True).all()
        for field in fields:
            value = data.get(field.title)
            if value is not None and value != '':
                pv = PersonnelValue.query.filter_by(
                    personnel_id=pid,
                    field_id=field.id,
                    period_id=p.period_id
                ).first()
                
                if pv:
                    if field.field_type == 'text':
                        pv.value_text = value
                    elif field.field_type in ['number', 'decimal']:
                        try:
                            pv.value_number = float(value)
                        except:
                            pv.value_text = value
                    elif field.field_type == 'date':
                        pv.value_date = value
                    pv.updated_at = datetime.now()
                else:
                    pv = PersonnelValue(
                        personnel_id=pid,
                        field_id=field.id,
                        period_id=p.period_id,
                        value_text=value if field.field_type == 'text' else None,
                        value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
                        value_date=value if field.field_type == 'date' else None
                    )
                    db.session.add(pv)
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'اطلاعات ذخیره شد', 'period_id': p.period_id})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ خطا در ویرایش: {e}")
        return jsonify({'error': f'خطا: {str(e)}'}), 500


@unit_supervisor_bp.route('/api/personnel/batch-update', methods=['POST'])
@login_required
def unit_supervisor_api_personnel_batch_update():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = req.get_json()
    field = data.get('field')
    updates = data.get('updates', [])
    
    dynamic_field = DynamicField.query.filter_by(title=field, is_active=True).first()
    
    if not dynamic_field:
        return jsonify({'error': f'فیلد "{field}" یافت نشد'}), 400
    
    count = 0
    for item in updates:
        p = Personnel.query.get(item['id'])
        if p:
            value = item['value']
            
            pv = PersonnelValue.query.filter_by(
                personnel_id=p.id,
                field_id=dynamic_field.id,
                period_id=p.period_id
            ).first()
            
            if pv:
                if dynamic_field.field_type == 'text':
                    pv.value_text = value
                elif dynamic_field.field_type in ['number', 'decimal']:
                    try:
                        pv.value_number = float(value)
                    except:
                        pv.value_text = value
                elif dynamic_field.field_type == 'date':
                    pv.value_date = value
                pv.updated_at = datetime.now()
            else:
                pv = PersonnelValue(
                    personnel_id=p.id,
                    field_id=dynamic_field.id,
                    period_id=p.period_id,
                    value_text=value if dynamic_field.field_type == 'text' else None,
                    value_number=float(value) if dynamic_field.field_type in ['number', 'decimal'] and value else None,
                    value_date=value if dynamic_field.field_type == 'date' else None
                )
                db.session.add(pv)
            count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'count': count, 'message': f'{count} پرسنل با موفقیت به‌روزرسانی شدند'})


@unit_supervisor_bp.route('/export-excel')
@login_required
def unit_supervisor_export_excel():
    if current_user.role != 'unit_supervisor':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('user.dashboard'))
    
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    
    if not supervised_units:
        flash('واحد یافت نشد', 'error')
        return redirect(url_for('user.dashboard'))
    
    unit = supervised_units[0]
    personnel = Personnel.query.filter_by(unit_id=unit.id, is_deleted=False).all()
    fields = DynamicField.query.filter_by(is_active=True).all()
    
    template_obj = ExcelTemplate.query.first()
    if not template_obj:
        template_obj = ExcelTemplate()
    
    def to_persian(num):
        if num is None or num == '':
            return ''
        persian_digits = '۰۱۲۳۴۵۶۷۸۹'
        if isinstance(num, (int, float)):
            return ''.join(persian_digits[int(d)] for d in str(int(num)) if d.isdigit())
        if isinstance(num, str) and num.isdigit():
            return ''.join(persian_digits[int(d)] for d in num)
        if isinstance(num, str):
            digits = ''.join([d for d in num if d.isdigit()])
            if digits:
                return ''.join(persian_digits[int(d)] for d in digits)
        return str(num)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"پرسنل {unit.name}"
    
    header_bg = template_obj.header_bg_color.replace('#', '') if template_obj.header_bg_color else '2c3e50'
    header_text = template_obj.header_text_color.replace('#', '') if template_obj.header_text_color else 'ffffff'
    even_color = template_obj.even_row_color.replace('#', '') if template_obj.even_row_color else 'f8f9fa'
    odd_color = template_obj.odd_row_color.replace('#', '') if template_obj.odd_row_color else 'ffffff'
    font_name = 'Calibri'
    header_font_size = template_obj.header_font_size if template_obj.header_font_size else 12
    data_font_size = template_obj.data_font_size if template_obj.data_font_size else 11
    border_color = template_obj.border_color.replace('#', '') if template_obj.border_color else '000000'
    outer_style = template_obj.outer_border_style if template_obj.outer_border_style else 'thick'
    vertical_style = template_obj.vertical_border_style if template_obj.vertical_border_style else 'thin'
    horizontal_style = template_obj.horizontal_border_style if template_obj.horizontal_border_style else 'thin'
    
    border_map = {
        'thin': Side(border_style='thin', color=border_color),
        'medium': Side(border_style='medium', color=border_color),
        'thick': Side(border_style='thick', color=border_color),
        'double': Side(border_style='double', color=border_color),
        'dashed': Side(border_style='dashed', color=border_color),
        'dotted': Side(border_style='dotted', color=border_color)
    }
    
    outer_side = border_map.get(outer_style, Side(border_style='thin', color=border_color))
    vertical_side = border_map.get(vertical_style, Side(border_style='thin', color=border_color))
    horizontal_side = border_map.get(horizontal_style, Side(border_style='thin', color=border_color))
    
    headers = ['ردیف', 'کد ملی', 'نام', 'نام خانوادگی']
    for f in fields:
        if not f.is_key and f.title not in ['نام', 'نام خانوادگی']:
            headers.append(f.title)
    headers.extend(['شماره تماس', 'سمت'])
    
    total_rows = len(personnel) + 1
    total_cols = len(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name=font_name, size=header_font_size, bold=True, color=header_text)
        cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        top = outer_side
        bottom = outer_side if total_rows == 1 else horizontal_side
        left = outer_side if col == total_cols else vertical_side
        right = outer_side if col == 1 else vertical_side
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)
    
    for idx, p in enumerate(personnel, 1):
        row_num = idx + 1
        bg_color = even_color if (row_num % 2 == 0) else odd_color
        is_last_row = (row_num == total_rows)
        
        cell = ws.cell(row=row_num, column=1, value=to_persian(idx))
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=outer_side, right=vertical_side)
        
        cell = ws.cell(row=row_num, column=2, value=to_persian(p.national_code) if p.national_code else '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        cell = ws.cell(row=row_num, column=3, value=p.first_name or '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        cell = ws.cell(row=row_num, column=4, value=p.last_name or '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        col = 5
        for f in fields:
            if not f.is_key and f.title not in ['نام', 'نام خانوادگی']:
                value_num = ''
                pv = PersonnelValue.query.filter_by(personnel_id=p.id, field_id=f.id, period_id=p.period_id).first()
                if pv:
                    raw = pv.value_text or pv.value_number or pv.value_date or ''
                    if pv.value_number is not None:
                        if pv.value_number == int(pv.value_number):
                            value_num = to_persian(int(pv.value_number))
                        else:
                            value_num = to_persian(str(pv.value_number).replace('.', '/'))
                    else:
                        value_num = to_persian(raw)
                cell = ws.cell(row=row_num, column=col, value=value_num)
                cell.font = Font(name=font_name, size=data_font_size)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
                col += 1
        
        cell = ws.cell(row=row_num, column=col, value=to_persian(p.phone) if p.phone else '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        cell = ws.cell(row=row_num, column=col+1, value=p.position or '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=outer_side)
    
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_idx in range(2, len(personnel) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 30)
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    
    today = jdatetime.datetime.now().strftime('%Y%m%d')
    filename = f"پرسنل_{unit.name}_{today}.xlsx"
    
    return send_file(temp_file.name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)


@unit_supervisor_bp.route('/api/update-personnel-field', methods=['POST'])
@login_required
def unit_supervisor_api_update_personnel_field():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    personnel_id = data.get('personnel_id')
    field_title = data.get('field_title')
    period_id = data.get('period_id')
    value = data.get('value')
    
    p = Personnel.query.get_or_404(personnel_id)
    field = DynamicField.query.filter_by(title=field_title, is_active=True).first()
    
    if not field:
        return jsonify({'error': 'فیلد یافت نشد'}), 404
    
    pv = PersonnelValue.query.filter_by(
        personnel_id=personnel_id,
        field_id=field.id,
        period_id=period_id if period_id else None
    ).first()
    
    if pv:
        if field.field_type == 'text':
            pv.value_text = value
        elif field.field_type in ['number', 'decimal']:
            try:
                pv.value_number = float(value)
            except:
                pv.value_text = value
        elif field.field_type == 'date':
            pv.value_date = value
        pv.updated_at = datetime.now()
    else:
        pv = PersonnelValue(
            personnel_id=personnel_id,
            field_id=field.id,
            period_id=period_id if period_id else None,
            value_text=value if field.field_type == 'text' else None,
            value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
            value_date=value if field.field_type == 'date' else None
        )
        db.session.add(pv)
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'مقدار با موفقیت ذخیره شد'})