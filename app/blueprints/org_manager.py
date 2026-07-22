from flask import Blueprint, render_template, request as req, jsonify, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from app.extensions import db
from app.models import *
from datetime import datetime
import json
import jdatetime
import pandas as pd
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

org_manager_bp = Blueprint('org_manager', __name__, url_prefix='/org-manager')


# ==================== پنل مدیر سازمان - صفحات ====================

@org_manager_bp.route('/dashboard')
@login_required
def dashboard():
    # فقط مدیر سازمان و ادمین دسترسی داشته باشند
    if current_user.role not in ['org_manager', 'admin']:
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('user.dashboard'))
    
    # اگر کاربر ادمین است ولی می‌خواهد پنل مدیر سازمان را ببیند
    # اگر کاربر ادمین است، اجازه بده
    if current_user.role == 'admin':
        organization_name = "سازمان مرکزی (نمایش ادمین)"
    else:
        organization_name = "سازمان مرکزی"
    
    first_dept = Department.query.first()
    if first_dept:
        organization_name = "سازمان"
    
    active_period = WorkPeriod.query.filter_by(is_active=True).first()
    period_title = active_period.title if active_period else 'تعریف نشده'
    today_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    return render_template('org_manager/dashboard.html',
                          organization_name=organization_name,
                          period_title=period_title,
                          today_date=today_date)


# ==================== APIهای مدیر سازمان ====================

@org_manager_bp.route('/api/all-data')
@login_required
def api_all_data():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    departments = Department.query.filter_by(is_active=True).all()
    department_ids = [d.id for d in departments]
    
    units = Unit.query.filter(Unit.department_id.in_(department_ids), Unit.is_active == True).all()
    unit_ids = [u.id for u in units]
    
    personnel = Personnel.query.filter(Personnel.unit_id.in_(unit_ids), Personnel.is_deleted == False).all()
    dynamic_fields = DynamicField.query.filter_by(is_active=True).order_by(DynamicField.field_order).all()
    all_periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    
    personnel_data = []
    for p in personnel:
        unit = Unit.query.get(p.unit_id)
        department = Department.query.get(p.department_id) if p.department_id else None
        values = {}
        
        for v in PersonnelValue.query.filter_by(personnel_id=p.id).all():
            field = DynamicField.query.get(v.field_id)
            if field:
                values[field.title] = v.value_text or v.value_number or v.value_date or ''
        
        period_title = ''
        if p.period_id:
            period = WorkPeriod.query.get(p.period_id)
            if period:
                period_title = f"{period.title} ({period.start_date} - {period.end_date})"
        
        item = {
            'id': p.id,
            'national_code': p.national_code,
            'first_name': p.first_name or '',
            'last_name': p.last_name or '',
            'full_name': p.get_full_name(),
            'phone': p.phone or '',
            'position': p.position or '',
            'department_id': p.department_id,
            'department_name': department.name if department else '-',
            'unit_id': p.unit_id,
            'unit_name': unit.name if unit else '-',
            'period_title': period_title,
            'period_id': p.period_id
        }
        
        for field in dynamic_fields:
            if not field.is_key:
                item[field.title] = values.get(field.title, '')
        
        personnel_data.append(item)
    
    departments_data = [{'id': d.id, 'name': d.name, 'color': d.color} for d in departments]
    units_data = [{'id': u.id, 'name': u.name, 'department_id': u.department_id} for u in units]
    
    return jsonify({
        'personnel': personnel_data,
        'dynamic_fields': [{'id': f.id, 'title': f.title, 'field_type': f.field_type, 'is_key': f.is_key, 'is_required': f.is_required} for f in dynamic_fields],
        'periods': [{'id': p.id, 'title': p.title, 'start_date': p.start_date, 'end_date': p.end_date} for p in all_periods],
        'departments': departments_data,
        'units': units_data
    })


@org_manager_bp.route('/api/periods')
@login_required
def api_periods():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    return jsonify([{
        'id': p.id,
        'title': p.title,
        'start_date': p.start_date,
        'end_date': p.end_date,
        'is_active': p.is_active
    } for p in periods])


@org_manager_bp.route('/api/pending-requests')
@login_required
def api_pending_requests():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    departments = Department.query.filter_by(is_active=True).all()
    department_ids = [d.id for d in departments]
    units = Unit.query.filter(Unit.department_id.in_(department_ids)).all()
    unit_ids = [u.id for u in units]
    
    requests = UnitPersonnelRequest.query.filter(
        UnitPersonnelRequest.unit_id.in_(unit_ids),
        UnitPersonnelRequest.request_type == 'add',
        UnitPersonnelRequest.status == 'pending'
    ).order_by(UnitPersonnelRequest.created_at.desc()).all()
    
    result = []
    for req in requests:
        data = json.loads(req.data) if req.data else {}
        unit = Unit.query.get(req.unit_id)
        created_at_jalali = jdatetime.datetime.fromgregorian(datetime=req.created_at).strftime('%Y/%m/%d %H:%M')
        
        result.append({
            'id': req.id,
            'request_type': req.request_type,
            'national_code': data.get('national_code', ''),
            'first_name': data.get('first_name', ''),
            'last_name': data.get('last_name', ''),
            'full_name': f"{data.get('first_name', '')} {data.get('last_name', '')}".strip(),
            'unit_name': unit.name if unit else '-',
            'created_at': created_at_jalali
        })
    
    return jsonify({'requests': result})


@org_manager_bp.route('/api/delete-requests')
@login_required
def api_delete_requests():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    departments = Department.query.filter_by(is_active=True).all()
    department_ids = [d.id for d in departments]
    units = Unit.query.filter(Unit.department_id.in_(department_ids)).all()
    unit_ids = [u.id for u in units]
    
    requests = UnitPersonnelRequest.query.filter(
        UnitPersonnelRequest.unit_id.in_(unit_ids),
        UnitPersonnelRequest.request_type == 'delete',
        UnitPersonnelRequest.status == 'pending'
    ).order_by(UnitPersonnelRequest.created_at.desc()).all()
    
    result = []
    for req in requests:
        data = json.loads(req.data) if req.data else {}
        unit = Unit.query.get(req.unit_id)
        created_at_jalali = jdatetime.datetime.fromgregorian(datetime=req.created_at).strftime('%Y/%m/%d %H:%M')
        
        result.append({
            'id': req.id,
            'request_type': req.request_type,
            'personnel_id': data.get('personnel_id'),
            'national_code': data.get('national_code', ''),
            'first_name': data.get('first_name', ''),
            'last_name': data.get('last_name', ''),
            'full_name': data.get('full_name', ''),
            'unit_name': unit.name if unit else '-',
            'delete_reason': data.get('delete_reason', ''),
            'created_at': created_at_jalali
        })
    
    return jsonify({'requests': result})


@org_manager_bp.route('/api/cancel-request', methods=['POST'])
@login_required
def api_cancel_request():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = req.get_json()
    request_id = data.get('request_id')
    
    if not request_id:
        return jsonify({'error': 'شناسه درخواست ارسال نشده است'}), 400
    
    req = UnitPersonnelRequest.query.get(request_id)
    if not req:
        return jsonify({'error': 'درخواست یافت نشد'}), 404
    
    db.session.delete(req)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست لغو شد'})


@org_manager_bp.route('/api/request-add', methods=['POST'])
@login_required
def api_request_add():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = req.get_json()
    
    unit_id = data.get('unit_id')
    period_id = data.get('period_id')
    national_code = data.get('national_code')
    department_id = data.get('department_id')
    
    if not department_id:
        return jsonify({'error': 'اداره انتخاب نشده است'}), 400
    if not unit_id:
        return jsonify({'error': 'واحد انتخاب نشده است'}), 400
    if not period_id:
        return jsonify({'error': 'دوره انتخاب نشده است'}), 400
    if not national_code:
        return jsonify({'error': 'کد ملی الزامی است'}), 400
    
    if not (national_code.isdigit() and len(national_code) == 10):
        return jsonify({'error': 'کد ملی باید 10 رقم باشد'}), 400
    
    unit = Unit.query.get(unit_id)
    if not unit:
        return jsonify({'error': 'واحد یافت نشد'}), 404
    
    existing = Personnel.query.filter_by(
        national_code=national_code, 
        period_id=period_id, 
        is_deleted=False
    ).first()
    if existing:
        return jsonify({'error': 'این کد ملی قبلاً در دوره انتخاب شده ثبت شده است'}), 400
    
    request_data = {
        'department_id': department_id,
        'unit_id': unit_id,
        'national_code': national_code,
        'requester_note': data.get('note', ''),
        'period_id': period_id
    }
    
    fields = DynamicField.query.filter_by(is_active=True).all()
    for field in fields:
        value = data.get(field.title)
        if value:
            request_data[field.title] = value
            if field.title == 'نام':
                request_data['first_name'] = value
            if field.title == 'نام خانوادگی':
                request_data['last_name'] = value
            if field.title == 'شماره تماس':
                request_data['phone'] = value
            if field.title == 'سمت':
                request_data['position'] = value
    
    approval = UnitPersonnelRequest(
        unit_id=unit_id,
        requester_id=current_user.id,
        request_type='add',
        data=json.dumps(request_data, ensure_ascii=False),
        status='pending'
    )
    db.session.add(approval)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست شما ثبت شد و منتظر تایید ادمین است'})


@org_manager_bp.route('/api/request-delete', methods=['POST'])
@login_required
def api_request_delete():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = req.get_json()
    personnel_id = data.get('personnel_id')
    delete_reason = data.get('delete_reason', '')
    
    if not personnel_id:
        return jsonify({'error': 'شناسه پرسنل ارسال نشده است'}), 400
    
    p = Personnel.query.get(personnel_id)
    if not p:
        return jsonify({'error': 'پرسنل یافت نشد'}), 404
    
    UnitPersonnelRequest.query.filter_by(
        target_personnel_id=personnel_id,
        request_type='delete',
        status='pending'
    ).delete()
    db.session.commit()
    
    request_data = {
        'personnel_id': p.id,
        'national_code': p.national_code,
        'first_name': p.first_name or '',
        'last_name': p.last_name or '',
        'full_name': p.get_full_name(),
        'phone': p.phone or '',
        'position': p.position or '',
        'delete_reason': delete_reason
    }
    
    approval = UnitPersonnelRequest(
        unit_id=p.unit_id,
        requester_id=current_user.id,
        request_type='delete',
        target_personnel_id=personnel_id,
        data=json.dumps(request_data, ensure_ascii=False),
        status='pending'
    )
    db.session.add(approval)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست حذف پرسنل ثبت شد'})


@org_manager_bp.route('/api/personnel/<int:pid>')
@login_required
def api_personnel_get(pid):
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    p = Personnel.query.get_or_404(pid)
    fields = DynamicField.query.filter_by(is_active=True).all()
    
    values = {}
    for v in PersonnelValue.query.filter_by(personnel_id=pid).all():
        field = DynamicField.query.get(v.field_id)
        if field:
            values[field.title] = v.value_text or v.value_number or v.value_date or ''
    
    result = {
        'id': p.id,
        'national_code': p.national_code,
        'first_name': p.first_name or '',
        'last_name': p.last_name or '',
        'phone': p.phone or '',
        'position': p.position or '',
        'department_id': p.department_id,
        'unit_id': p.unit_id,
        'period_id': p.period_id
    }
    
    for field in fields:
        result[field.title] = values.get(field.title, '')
    
    return jsonify(result)


@org_manager_bp.route('/api/personnel/<int:pid>/edit', methods=['PUT'])
@login_required
def api_personnel_edit(pid):
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        p = Personnel.query.get_or_404(pid)
        data = req.get_json()
        
        p.first_name = data.get('first_name', p.first_name)
        p.last_name = data.get('last_name', p.last_name)
        p.phone = data.get('phone', p.phone)
        p.position = data.get('position', p.position)
        
        period_id = data.get('period_id')
        if period_id and period_id != '' and period_id != 'null':
            p.period_id = int(period_id)
        else:
            p.period_id = None
        
        db.session.commit()
        
        fields = DynamicField.query.filter_by(is_active=True).all()
        for field in fields:
            value = data.get(field.title)
            if value is not None and value != '':
                pv = PersonnelValue.query.filter_by(
                    personnel_id=pid,
                    field_id=field.id,
                    period_id=p.period_id
                ).first()
                
                if pv:
                    if field.field_type == 'text':
                        pv.value_text = value
                    elif field.field_type in ['number', 'decimal']:
                        try:
                            pv.value_number = float(value)
                        except:
                            pv.value_text = value
                    elif field.field_type == 'date':
                        pv.value_date = value
                    pv.updated_at = datetime.now()
                else:
                    pv = PersonnelValue(
                        personnel_id=pid,
                        field_id=field.id,
                        period_id=p.period_id,
                        value_text=value if field.field_type == 'text' else None,
                        value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
                        value_date=value if field.field_type == 'date' else None
                    )
                    db.session.add(pv)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'اطلاعات ذخیره شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
        return jsonify({'error': str(e)}), 500


@org_manager_bp.route('/api/personnel-history/<int:pid>')
@login_required
def api_personnel_history(pid):
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    p = Personnel.query.get_or_404(pid)
    fields = DynamicField.query.filter_by(is_active=True).all()
    periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    
    history = []
    for period in periods:
        period_data = {
            'period_id': period.id,
            'period_title': period.title,
            'period_start': period.start_date,
            'period_end': period.end_date,
            'values': {}
        }
        
        for field in fields:
            pv = PersonnelValue.query.filter_by(
                personnel_id=pid,
                field_id=field.id,
                period_id=period.id
            ).first()
            
            if pv:
                value = pv.value_text or pv.value_number or pv.value_date or '-'
            else:
                pv_default = PersonnelValue.query.filter_by(
                    personnel_id=pid,
                    field_id=field.id,
                    period_id=None
                ).first()
                if pv_default:
                    value = pv_default.value_text or pv_default.value_number or pv_default.value_date or '-'
                else:
                    value = '-'
            
            period_data['values'][field.title] = value
        
        history.append(period_data)
    
    return jsonify({
        'personnel': {
            'id': p.id,
            'national_code': p.national_code,
            'full_name': p.get_full_name()
        },
        'fields': [{'id': f.id, 'title': f.title, 'field_type': f.field_type, 'is_key': f.is_key} for f in fields],
        'history': history
    })


@org_manager_bp.route('/api/personnel/batch-update', methods=['POST'])
@login_required
def api_personnel_batch_update():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = req.get_json()
    field = data.get('field')
    updates = data.get('updates', [])
    
    dynamic_field = DynamicField.query.filter_by(title=field, is_active=True).first()
    if not dynamic_field:
        return jsonify({'error': f'فیلد "{field}" یافت نشد'}), 400
    
    count = 0
    for item in updates:
        p = Personnel.query.get(item['id'])
        if p:
            value = item['value']
            
            pv = PersonnelValue.query.filter_by(
                personnel_id=p.id,
                field_id=dynamic_field.id,
                period_id=p.period_id
            ).first()
            
            if pv:
                if dynamic_field.field_type == 'text':
                    pv.value_text = value
                elif dynamic_field.field_type in ['number', 'decimal']:
                    try:
                        pv.value_number = float(value)
                    except:
                        pv.value_text = value
                elif dynamic_field.field_type == 'date':
                    pv.value_date = value
                pv.updated_at = datetime.now()
            else:
                pv = PersonnelValue(
                    personnel_id=p.id,
                    field_id=dynamic_field.id,
                    period_id=p.period_id,
                    value_text=value if dynamic_field.field_type == 'text' else None,
                    value_number=float(value) if dynamic_field.field_type in ['number', 'decimal'] and value else None,
                    value_date=value if dynamic_field.field_type == 'date' else None
                )
                db.session.add(pv)
            count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'count': count})


@org_manager_bp.route('/api/import-excel', methods=['POST'])
@login_required
def api_import_excel():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    if 'excel_file' not in req.files:
        return jsonify({'error': 'فایلی ارسال نشده'}), 400
    
    file = req.files['excel_file']
    if file.filename == '':
        return jsonify({'error': 'فایلی انتخاب نشده'}), 400
    
    try:
        df = pd.read_excel(file, dtype=str).fillna('')
        success, error = 0, 0
        
        for _, row in df.iterrows():
            code = str(row.get('کد ملی', '')).strip().replace('.0', '')
            if not code.isdigit() or len(code) != 10:
                error += 1
                continue
            
            if Personnel.query.filter_by(national_code=code).first():
                error += 1
                continue
            
            name = str(row.get('نام', '')).strip()
            family = str(row.get('نام خانوادگی', '')).strip()
            if not name or not family:
                error += 1
                continue
            
            dept_name = str(row.get('نام اداره', '')).strip()
            unit_name = str(row.get('نام واحد', '')).strip()
            
            dept = Department.query.filter_by(name=dept_name).first()
            if not dept:
                error += 1
                continue
            
            unit = Unit.query.filter_by(name=unit_name, department_id=dept.id).first()
            if not unit:
                error += 1
                continue
            
            p = Personnel(
                national_code=code,
                first_name=name,
                last_name=family,
                phone=str(row.get('شماره تماس', '')).strip(),
                position=str(row.get('سمت', '')).strip(),
                hire_date=str(row.get('تاریخ استخدام', '')).strip(),
                department_id=dept.id,
                unit_id=unit.id
            )
            db.session.add(p)
            success += 1
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'✅ {success} پرسنل اضافه شد. ❌ {error} خطا'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@org_manager_bp.route('/api/departments')
@login_required
def api_departments():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    departments = Department.query.filter_by(is_active=True).all()
    result = []
    for dept in departments:
        managers = db.session.query(User).join(
            DepartmentManager, DepartmentManager.user_id == User.id
        ).filter(DepartmentManager.department_id == dept.id).all()
        managers_names = [m.get_full_name() for m in managers]
        
        supervisors_count = db.session.query(UnitSupervisor).join(
            Unit, UnitSupervisor.unit_id == Unit.id
        ).filter(Unit.department_id == dept.id).count()
        
        result.append({
            'id': dept.id,
            'name': dept.name,
            'color': dept.color,
            'managers': managers_names,
            'supervisors_count': supervisors_count
        })
    
    return jsonify(result)


@org_manager_bp.route('/api/units')
@login_required
def api_units():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    units = Unit.query.filter_by(is_active=True).all()
    result = []
    for unit in units:
        supervisors = db.session.query(User).join(
            UnitSupervisor, UnitSupervisor.user_id == User.id
        ).filter(UnitSupervisor.unit_id == unit.id).all()
        supervisors_names = [s.get_full_name() for s in supervisors]
        
        result.append({
            'id': unit.id,
            'name': unit.name,
            'department_id': unit.department_id,
            'supervisors': supervisors_names
        })
    
    return jsonify(result)


@org_manager_bp.route('/export-excel')
@login_required
def export_excel():
    if current_user.role not in ['org_manager', 'admin']:
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('user.dashboard'))
    
    try:
        departments = Department.query.filter_by(is_active=True).all()
        department_ids = [d.id for d in departments]
        
        units = Unit.query.filter(Unit.department_id.in_(department_ids), Unit.is_active == True).all()
        unit_ids = [u.id for u in units]
        
        personnel = Personnel.query.filter(Personnel.unit_id.in_(unit_ids), Personnel.is_deleted == False).all()
        fields = DynamicField.query.filter_by(is_active=True).order_by(DynamicField.field_order).all()
        
        template = ExcelTemplate.query.first()
        if not template:
            template = ExcelTemplate()
        
        def to_persian(num):
            if num is None or num == '':
                return ''
            persian_digits = '۰۱۲۳۴۵۶۷۸۹'
            if isinstance(num, (int, float)):
                return ''.join(persian_digits[int(d)] for d in str(int(num)) if d.isdigit())
            if isinstance(num, str) and num.isdigit():
                return ''.join(persian_digits[int(d)] for d in num)
            if isinstance(num, str):
                digits = ''.join([d for d in num if d.isdigit()])
                if digits:
                    return ''.join(persian_digits[int(d)] for d in digits)
            return str(num)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "لیست پرسنل"
        
        header_bg = template.header_bg_color.replace('#', '') if template.header_bg_color else '2c3e50'
        header_text = template.header_text_color.replace('#', '') if template.header_text_color else 'ffffff'
        even_color = template.even_row_color.replace('#', '') if template.even_row_color else 'f8f9fa'
        odd_color = template.odd_row_color.replace('#', '') if template.odd_row_color else 'ffffff'
        font_name = 'Calibri'
        header_font_size = template.header_font_size if template.header_font_size else 12
        data_font_size = template.data_font_size if template.data_font_size else 11
        border_color = template.border_color.replace('#', '') if template.border_color else '000000'
        outer_style = template.outer_border_style if template.outer_border_style else 'thick'
        vertical_style = template.vertical_border_style if template.vertical_border_style else 'thin'
        horizontal_style = template.horizontal_border_style if template.horizontal_border_style else 'thin'
        
        border_map = {
            'thin': Side(border_style='thin', color=border_color),
            'medium': Side(border_style='medium', color=border_color),
            'thick': Side(border_style='thick', color=border_color),
            'double': Side(border_style='double', color=border_color),
            'dashed': Side(border_style='dashed', color=border_color),
            'dotted': Side(border_style='dotted', color=border_color)
        }
        
        outer_side = border_map.get(outer_style, Side(border_style='thin', color=border_color))
        vertical_side = border_map.get(vertical_style, Side(border_style='thin', color=border_color))
        horizontal_side = border_map.get(horizontal_style, Side(border_style='thin', color=border_color))
        
        headers = ['ردیف', 'کد ملی', 'نام', 'نام خانوادگی', 'اداره', 'واحد', 'سمت', 'شماره تماس', 'دوره']
        
        for f in fields:
            if hasattr(f, 'title') and not getattr(f, 'is_key', False) and f.title not in ['نام', 'نام خانوادگی']:
                headers.append(f.title)
        
        personnel_data = []
        for p in personnel:
            unit = Unit.query.get(p.unit_id)
            department = Department.query.get(p.department_id)
            
            period_title = ''
            if p.period_id:
                period = WorkPeriod.query.get(p.period_id)
                if period:
                    period_title = f"{period.title}"
            
            item = {
                'id': p.id,
                'national_code': p.national_code,
                'first_name': p.first_name or '',
                'last_name': p.last_name or '',
                'full_name': p.get_full_name(),
                'phone': p.phone or '',
                'position': p.position or '',
                'department_name': department.name if department else '-',
                'unit_name': unit.name if unit else '-',
                'period_title': period_title
            }
            
            for v in PersonnelValue.query.filter_by(personnel_id=p.id).all():
                field = DynamicField.query.get(v.field_id)
                if field and not field.is_key and field.title not in ['نام', 'نام خانوادگی']:
                    item[field.title] = v.value_text or v.value_number or v.value_date or ''
            
            personnel_data.append(item)
        
        total_rows = len(personnel_data) + 1
        total_cols = len(headers)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(name=font_name, size=header_font_size, bold=True, color=header_text)
            cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            top = outer_side
            bottom = outer_side if total_rows == 1 else horizontal_side
            left = outer_side if col == total_cols else vertical_side
            right = outer_side if col == 1 else vertical_side
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)
        
        for idx, p in enumerate(personnel_data, 1):
            row = idx + 1
            bg_color = even_color if (row % 2 == 0) else odd_color
            is_last_row = (row == total_rows)
            
            cell = ws.cell(row=row, column=1, value=to_persian(idx))
            cell.font = Font(name=font_name, size=data_font_size)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(top=outer_side if row == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=outer_side, right=vertical_side)
            
            col = 2
            for key in ['national_code', 'first_name', 'last_name', 'department_name', 'unit_name', 'position']:
                value = p.get(key, '')
                if key == 'national_code':
                    value = to_persian(value)
                elif key in ['phone']:
                    value = to_persian(value)
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = Font(name=font_name, size=data_font_size)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='center' if key in ['national_code', 'phone'] else 'right', vertical='center')
                cell.border = Border(top=outer_side if row == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
                col += 1
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        
        today = jdatetime.datetime.now().strftime('%Y%m%d')
        filename = f"لیست_پرسنل_سازمان_{today}.xlsx"
        
        return send_file(
            temp_file.name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error in export_excel: {e}")
        import traceback
        traceback.print_exc()
        flash('خطا در تولید فایل اکسل', 'error')
        return redirect(url_for('dashboard'))