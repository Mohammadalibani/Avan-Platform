from flask import Flask, render_template, redirect, url_for, request, flash, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from sqlalchemy.exc import IntegrityError
import jdatetime
import json
import os
import shutil
import glob

app = Flask(__name__)
app.config['SECRET_KEY'] = 'avan-samaneye-pishrafteh-1403'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///avan_system.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'لطفا برای دسترسی به این صفحه وارد شوید'

# ========== بهینه‌سازی SQLite ==========
with app.app_context():
    from sqlalchemy import text
    try:
        # فعال کردن حالت WAL (Write-Ahead Logging) - خیلی مهم برای سرعت
        db.session.execute(text("PRAGMA journal_mode=WAL;"))
        
        # کاهش همگام‌سازی برای سرعت بیشتر
        db.session.execute(text("PRAGMA synchronous=NORMAL;"))
        
        # افزایش کش دیتابیس به 20MB
        db.session.execute(text("PRAGMA cache_size=-20000;"))
        
        # ذخیره جدول‌های موقت در حافظه
        db.session.execute(text("PRAGMA temp_store=MEMORY;"))
        
        # افزایش حافظه موقت
        db.session.execute(text("PRAGMA page_size=4096;"))
        
        db.session.commit()
        print("✅ SQLite بهینه شد (WAL, cache=20MB, synchronous=NORMAL)")
    except Exception as e:
        print(f"⚠️ خطا در بهینه‌سازی SQLite: {e}")

# ========== کش ساده درون حافظه (جایگزین Memcached) ==========
from functools import wraps
import hashlib
import json
import time
from collections import OrderedDict

class SimpleCache:
    def __init__(self, maxsize=200, default_ttl=300):
        self.cache = OrderedDict()
        self.maxsize = maxsize
        self.default_ttl = default_ttl
    
    def get(self, key):
        if key in self.cache:
            value, expiry = self.cache[key]
            if time.time() < expiry:
                self.cache.move_to_end(key)
                return value
            del self.cache[key]
        return None
    
    def set(self, key, value, ttl=None):
        ttl = ttl or self.default_ttl
        if key in self.cache:
            del self.cache[key]
        elif len(self.cache) >= self.maxsize:
            self.cache.popitem(last=False)
        self.cache[key] = (value, time.time() + ttl)
    
    def flush_all(self):
        self.cache.clear()


# ========== راه‌اندازی کش ==========
cache = SimpleCache(maxsize=500, default_ttl=300)
CACHE_AVAILABLE = True
print("✅ کش درون حافظه فعال شد")


# ========== تابع کش کردن APIها ==========
def cached(ttl=300):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if not CACHE_AVAILABLE:
                return func(*args, **kwargs)
            
            from flask import request
            from flask_login import current_user
            
            user_id = current_user.id if current_user.is_authenticated else 0
            cache_key = hashlib.md5(f"{request.path}:{user_id}:{request.args}".encode()).hexdigest()
            
            cached_data = cache.get(cache_key)
            if cached_data:
                return json.loads(cached_data)
            
            result = func(*args, **kwargs)
            
            try:
                if not isinstance(result, tuple):
                    cache.set(cache_key, json.dumps(result), ttl)
            except:
                pass
            
            return result
        return wrapper
    return decorator


# ========== تابع پاک کردن کش ==========
def invalidate_cache():
    cache.flush_all()
    print("✅ کش پاک شد")
    
    

# ==================== مدل‌ها ====================
class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    national_code = db.Column(db.String(10), unique=True, nullable=False)
    username = db.Column(db.String(10), unique=True, nullable=False)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    phone = db.Column(db.String(20), nullable=True)  # ← این خط را اضافه کن
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(30), nullable=False, default='subordinate')
    personnel_code = db.Column(db.String(20), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    is_approved = db.Column(db.Boolean, default=False)
    profile_picture = db.Column(db.String(200), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    last_login = db.Column(db.DateTime, nullable=True)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    def get_full_name(self):
        return f"{self.first_name} {self.last_name}"
    def get_jalali_created_date(self):
        return jdatetime.datetime.fromgregorian(datetime=self.created_at).strftime('%Y/%m/%d')
    def get_role_persian(self):
        roles = {
            'admin': 'مدیر کل سیستم',
            'org_manager': 'مدیر سازمان',
            'dept_manager': 'مدیر اداره',
            'hr_manager': 'مدیر منابع انسانی',
            'unit_supervisor': 'سرپرست واحد',
            'subordinate': 'کاربر عادی'
        }
        return roles.get(self.role, 'نامشخص')


class UnitPersonnelRequest(db.Model):
    __tablename__ = 'unit_personnel_requests'
    
    id = db.Column(db.Integer, primary_key=True)
    unit_id = db.Column(db.Integer, db.ForeignKey('units.id'), nullable=False)
    requester_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    request_type = db.Column(db.String(20), nullable=False)
    target_personnel_id = db.Column(db.Integer, nullable=True)  # این خط مهم است
    data = db.Column(db.Text, nullable=True)
    status = db.Column(db.String(20), default='pending')
    admin_note = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    reviewed_at = db.Column(db.DateTime, nullable=True)

class Department(db.Model):
    __tablename__ = 'departments'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    color = db.Column(db.String(7), default='#3498db')
    description = db.Column(db.String(500), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    is_active = db.Column(db.Boolean, default=True)
    
    def get_jalali_created_date(self):
        import jdatetime
        return jdatetime.datetime.fromgregorian(datetime=self.created_at).strftime('%Y/%m/%d')
        
class DepartmentManager(db.Model):
    __tablename__ = 'department_managers'
    id = db.Column(db.Integer, primary_key=True)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)


class WorkPeriod(db.Model):
    __tablename__ = 'work_periods'
    
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100), nullable=False)
    start_date = db.Column(db.String(20), nullable=False)
    end_date = db.Column(db.String(20), nullable=False)
    deadline = db.Column(db.String(20), nullable=True)  # ← این خط را اضافه کنید
    display_order = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
@app.route('/admin/add-order-column')
@login_required
def add_order_column():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    import sqlite3
    db_path = os.path.join(os.path.dirname(__file__), 'instance', 'avan_system.db')
    
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("PRAGMA table_info(work_periods)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'display_order' not in columns:
            cursor.execute("ALTER TABLE work_periods ADD COLUMN display_order INTEGER DEFAULT 0")
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'ستون order اضافه شد'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
class Unit(db.Model):
    __tablename__ = 'units'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=False)
    description = db.Column(db.String(500), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    is_active = db.Column(db.Boolean, default=True)
    needs_approval = db.Column(db.Boolean, default=True)
    department = db.relationship('Department', backref='units')
    
    def get_jalali_created_date(self):
        import jdatetime
        return jdatetime.datetime.fromgregorian(datetime=self.created_at).strftime('%Y/%m/%d')
        
class PersonnelApprovalRequest(db.Model):
    __tablename__ = 'personnel_approval_requests'
    
    id = db.Column(db.Integer, primary_key=True)
    request_type = db.Column(db.String(20), nullable=False)  # add, delete
    personnel_data = db.Column(db.Text, nullable=True)  # JSON داده‌های پرسنل (برای add)
    personnel_id = db.Column(db.Integer, db.ForeignKey('personnel.id'), nullable=True)  # برای delete
    requester_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    unit_id = db.Column(db.Integer, db.ForeignKey('units.id'), nullable=False)
    status = db.Column(db.String(20), default='pending')  # pending, approved, rejected
    admin_note = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    reviewed_at = db.Column(db.DateTime, nullable=True)
    
    requester = db.relationship('User', foreign_keys=[requester_id])
    unit = db.relationship('Unit')

class UnitSupervisor(db.Model):
    __tablename__ = 'unit_supervisors'
    id = db.Column(db.Integer, primary_key=True)
    unit_id = db.Column(db.Integer, db.ForeignKey('units.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)

class Personnel(db.Model):
    __tablename__ = 'personnel'
    
    id = db.Column(db.Integer, primary_key=True)
    national_code = db.Column(db.String(10), unique=True, nullable=False)
    first_name = db.Column(db.String(50), nullable=True)
    last_name = db.Column(db.String(50), nullable=True)
    phone = db.Column(db.String(20), nullable=True)
    position = db.Column(db.String(100), nullable=True)
    hire_date = db.Column(db.String(20), nullable=True)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=True)
    unit_id = db.Column(db.Integer, db.ForeignKey('units.id'), nullable=False)
    period_id = db.Column(db.Integer, nullable=True)  # ← این خط را اضافه کن
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    is_deleted = db.Column(db.Boolean, default=False)
    
    department = db.relationship('Department', backref='personnel_list')
    unit = db.relationship('Unit', backref='personnel_list')
    
    def get_full_name(self):
        return f"{self.first_name or ''} {self.last_name or ''}".strip()

class ActivityLog(db.Model):
    __tablename__ = 'activity_logs'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    user_name = db.Column(db.String(100), nullable=True)
    message = db.Column(db.Text, nullable=False)
    badge = db.Column(db.String(50), nullable=True)
    log_type = db.Column(db.String(30), default='info')
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    user = db.relationship('User', foreign_keys=[user_id])
    
    def get_jalali_date(self):
        return jdatetime.datetime.fromgregorian(datetime=self.created_at).strftime('%Y/%m/%d %H:%M:%S')

class PersonnelValue(db.Model):
    __tablename__ = 'personnel_values'
    
    id = db.Column(db.Integer, primary_key=True)
    personnel_id = db.Column(db.Integer, db.ForeignKey('personnel.id'), nullable=False)
    field_id = db.Column(db.Integer, db.ForeignKey('dynamic_fields.id'), nullable=False)
    period_id = db.Column(db.Integer, nullable=True)
    value_text = db.Column(db.Text, nullable=True)
    value_number = db.Column(db.Float, nullable=True)
    value_date = db.Column(db.String(20), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    
    personnel = db.relationship('Personnel', backref='values')
    field = db.relationship('DynamicField', backref='values')


# ==================== مدل DynamicField (در جای صحیح) ====================
class DynamicField(db.Model):
    __tablename__ = 'dynamic_fields'
    
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100), nullable=False)
    field_type = db.Column(db.String(20), nullable=False)  # text, number, date, decimal
    is_required = db.Column(db.Boolean, default=False)
    is_locked = db.Column(db.Boolean, default=False)
    is_monitoring = db.Column(db.Boolean, default=False)
    is_key = db.Column(db.Boolean, default=False)
    field_order = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.now)
    is_active = db.Column(db.Boolean, default=True)

class ExcelTemplate(db.Model):
    __tablename__ = 'excel_templates'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, default='قالب پیش‌فرض')
    header_bg_color = db.Column(db.String(7), default='#2c3e50')
    header_text_color = db.Column(db.String(7), default='#ffffff')
    even_row_color = db.Column(db.String(7), default='#f8f9fa')
    odd_row_color = db.Column(db.String(7), default='#ffffff')
    outer_border_style = db.Column(db.String(20), default='thick')
    vertical_border_style = db.Column(db.String(20), default='thin')
    horizontal_border_style = db.Column(db.String(20), default='dotted')
    border_color = db.Column(db.String(7), default='#000000')
    font_name = db.Column(db.String(50), default='B Nazanin')
    header_font_size = db.Column(db.Integer, default=12)
    data_font_size = db.Column(db.Integer, default=11)

class ApprovalRequest(db.Model):
    __tablename__ = 'approval_requests'
    id = db.Column(db.Integer, primary_key=True)
    request_type = db.Column(db.String(20), nullable=False)
    requester_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    reviewer_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    target_personnel_id = db.Column(db.Integer, db.ForeignKey('personnel.id'), nullable=True)
    data = db.Column(db.Text, nullable=True)
    requester_note = db.Column(db.Text, nullable=True)
    admin_note = db.Column(db.Text, nullable=True)
    status = db.Column(db.String(20), default='pending')
    created_at = db.Column(db.DateTime, default=datetime.now)
    reviewed_at = db.Column(db.DateTime, nullable=True)

class Setting(db.Model):
    __tablename__ = 'settings'
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.Text, nullable=True)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    @staticmethod
    def get(key, default=None):
        setting = Setting.query.filter_by(key=key).first()
        return setting.value if setting else default
    @staticmethod
    def set(key, value):
        setting = Setting.query.filter_by(key=key).first()
        if setting:
            setting.value = value
        else:
            setting = Setting(key=key, value=value)
            db.session.add(setting)
        db.session.commit()

# ==================== مدل تیکت (سیستم پیام رسانی) ====================
class Ticket(db.Model):
    __tablename__ = 'tickets'
    
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text, nullable=False)
    sender_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    receiver_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    status = db.Column(db.String(20), default='open')
    priority = db.Column(db.String(20), default='normal')
    message_type = db.Column(db.String(20), default='ticket')  # <-- اضافه کن: 'ticket' یا 'message'
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)

class TicketReply(db.Model):
    __tablename__ = 'ticket_replies'
    
    id = db.Column(db.Integer, primary_key=True)
    ticket_id = db.Column(db.Integer, db.ForeignKey('tickets.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    message = db.Column(db.Text, nullable=False)
    is_admin_reply = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    ticket = db.relationship('Ticket', backref='replies')
    user = db.relationship('User', backref='ticket_replies')

# ==================== مدل مدارک کاربران ====================
class UserDocument(db.Model):
    __tablename__ = 'user_documents'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    doc_type = db.Column(db.String(50), nullable=False)
    doc_title = db.Column(db.String(100), nullable=True)
    doc_filename = db.Column(db.String(200), nullable=False)
    doc_original_name = db.Column(db.String(200), nullable=True)
    doc_size = db.Column(db.Integer, default=0)
    status = db.Column(db.String(20), default='pending')
    admin_note = db.Column(db.Text, nullable=True)
    reviewed_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    reviewed_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    
    # 🔧 اصلاح: اضافه کردن foreign_keys برای مشخص کردن کدام کلید خارجی استفاده شود
    user = db.relationship('User', foreign_keys=[user_id], backref='documents')
    reviewer = db.relationship('User', foreign_keys=[reviewed_by], backref='reviewed_documents')
# ==================== مدل درخواست‌های پرسنل (سیستم جدید) ====================

class Request(db.Model):
    """مدل پایه برای همه درخواست‌های پرسنل"""
    __tablename__ = 'requests'
    
    id = db.Column(db.Integer, primary_key=True)
    
    # نوع درخواست
    request_type = db.Column(db.String(50), nullable=False)
    # مقادیر: overtime, deficiency, daily_mission, official_mission, arbaeen, annual_leave, hourly_leave
    
    # اطلاعات فرستنده و گیرنده
    requester_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    unit_supervisor_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    
    # وضعیت‌ها: pending_unit (منتظر سرپرست), approved (تأیید نهایی), rejected (رد), revision (نیاز به اصلاح)
    status = db.Column(db.String(30), default='pending_unit')
    
    # تاریخ‌ها
    request_date = db.Column(db.DateTime, default=datetime.now)
    reviewed_at = db.Column(db.DateTime, nullable=True)
    
    # اطلاعات اصلاح و رد
    revision_note = db.Column(db.Text, nullable=True)
    reject_reason = db.Column(db.Text, nullable=True)
    
    # داده‌های اضافی (JSON)
    extra_data = db.Column(db.Text, nullable=True)
    
    # پیوست فایل
    attachment = db.Column(db.String(500), nullable=True)
    attachment_filename = db.Column(db.String(200), nullable=True)
    
    # روابط
    requester = db.relationship('User', foreign_keys=[requester_id], backref='sent_requests')
    unit_supervisor = db.relationship('User', foreign_keys=[unit_supervisor_id], backref='received_requests')
    
    def get_extra_data(self):
        """دریافت داده‌های اضافی به صورت دیکشنری"""
        if self.extra_data:
            return json.loads(self.extra_data)
        return {}
    
    def set_extra_data(self, data):
        """ذخیره داده‌های اضافی"""
        self.extra_data = json.dumps(data, ensure_ascii=False)
    
    def get_status_persian(self):
        """دریافت وضعیت به فارسی"""
        status_map = {
            'pending_unit': '⏳ در انتظار تایید سرپرست',
            'approved': '✅ تایید شده',
            'rejected': '❌ رد شده',
            'revision': '🔄 نیاز به اصلاح دارد'
        }
        return status_map.get(self.status, self.status)
    
    def get_request_type_persian(self):
        """دریافت نوع درخواست به فارسی"""
        type_map = {
            'overtime': 'اضافه کار ساعتی',
            'deficiency': 'ثبت نواقص',
            'daily_mission': 'ماموریت روزانه',
            'official_mission': 'ماموریت اداری',
            'arbaeen': 'سفر اربعین',
            'annual_leave': 'مرخصی روزانه',
            'hourly_leave': 'مرخصی ساعتی'
        }
        return type_map.get(self.request_type, self.request_type)
    
    def to_dict(self):
        """تبدیل به دیکشنری برای JSON"""
        extra = self.get_extra_data()
        return {
            'id': self.id,
            'request_type': self.request_type,
            'request_type_persian': self.get_request_type_persian(),
            'requester_id': self.requester_id,
            'requester_name': self.requester.get_full_name() if self.requester else '-',
            'status': self.status,
            'status_persian': self.get_status_persian(),
            'request_date': self.request_date.strftime('%Y/%m/%d %H:%M'),
            'reviewed_at': self.reviewed_at.strftime('%Y/%m/%d %H:%M') if self.reviewed_at else None,
            'revision_note': self.revision_note,
            'reject_reason': self.reject_reason,
            'extra_data': extra,
            'has_attachment': bool(self.attachment)
        }

# ==================== مدل انتصاب پرسنل به واحدها (با تاریخچه) ====================

class PersonnelAssignment(db.Model):
    """مدل انتصاب پرسنل به واحدها (با قابلیت انتقال و تاریخچه)"""
    __tablename__ = 'personnel_assignments'
    
    id = db.Column(db.Integer, primary_key=True)
    personnel_id = db.Column(db.Integer, db.ForeignKey('personnel.id'), nullable=False)
    unit_id = db.Column(db.Integer, db.ForeignKey('units.id'), nullable=False)
    period_id = db.Column(db.Integer, db.ForeignKey('work_periods.id'), nullable=True)
    
    # نوع انتصاب: 'initial' (اولیه), 'transfer' (انتقال), 'promotion' (ارتقا)
    assignment_type = db.Column(db.String(30), default='initial')
    
    # تاریخ شروع انتصاب
    start_date = db.Column(db.String(20), nullable=False)  # تاریخ شمسی
    
    # تاریخ پایان انتصاب (برای انتصابات قبلی)
    end_date = db.Column(db.String(20), nullable=True)
    
    # آیا انتصاب فعلی فعال است؟
    is_active = db.Column(db.Boolean, default=True)
    
    # توضیحات (دلیل انتقال، ارتقا، ...)
    description = db.Column(db.Text, nullable=True)
    
    # ثبت‌کننده
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    # روابط
    personnel = db.relationship('Personnel', backref='assignments')
    unit = db.relationship('Unit')
    period = db.relationship('WorkPeriod')
    creator = db.relationship('User', foreign_keys=[created_by])
    
    def to_dict(self):
        return {
            'id': self.id,
            'personnel_id': self.personnel_id,
            'personnel_name': self.personnel.get_full_name() if self.personnel else '-',
            'unit_id': self.unit_id,
            'unit_name': self.unit.name if self.unit else '-',
            'department_id': self.unit.department_id if self.unit else None,
            'department_name': self.unit.department.name if self.unit and self.unit.department else '-',
            'period_id': self.period_id,
            'period_title': self.period.title if self.period else '-',
            'assignment_type': self.assignment_type,
            'assignment_type_persian': {
                'initial': 'انتصاب اولیه',
                'transfer': 'انتقال',
                'promotion': 'ارتقا'
            }.get(self.assignment_type, self.assignment_type),
            'start_date': self.start_date,
            'end_date': self.end_date,
            'is_active': self.is_active,
            'description': self.description,
            'created_by_name': self.creator.get_full_name() if self.creator else '-',
            'created_at': self.created_at.strftime('%Y/%m/%d %H:%M')
        }


# اطمینان از وجود پوشه آپلود درخواست‌ها
UPLOAD_FOLDER = os.path.join(app.root_path, 'static', 'uploads', 'requests')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def init_default_settings():
    with app.app_context():
        if not Setting.query.filter_by(key='base_url').first():
            Setting.set('base_url', '10.86.109.219')
        if not Setting.query.filter_by(key='port').first():
            Setting.set('port', '5000')
        if not Setting.query.filter_by(key='backup_hour').first():
            Setting.set('backup_hour', '23')
        if not Setting.query.filter_by(key='session_timeout').first():
            Setting.set('session_timeout', '30')
        if not Setting.query.filter_by(key='install_date').first():
            Setting.set('install_date', datetime.now().strftime('%Y/%m/%d'))

def init_default_admin():
    with app.app_context():
        admin = User.query.filter_by(role='admin').first()
        if not admin:
            admin = User(national_code='1234567890', username='1234567890', first_name='مدیر', last_name='سیستم', role='admin', personnel_code='ADMIN001', is_active=True, is_approved=True)
            admin.set_password('1234')
            db.session.add(admin)
            db.session.commit()
            print("✅ ادمین پیش‌فرض ایجاد شد")

@app.route('/api/work/messages/<int:msg_id>/reply', methods=['POST'])
@login_required
def api_work_message_reply(msg_id):
    """پاسخ به پیام اصلاح کارکرد"""
    original_msg = WorkRevisionMessage.query.get_or_404(msg_id)
    
    # بررسی دسترسی: فقط گیرنده اصلی می‌تواند پاسخ دهد
    if original_msg.to_user_id != current_user.id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    reply_message = data.get('message', '').strip()
    
    if not reply_message:
        return jsonify({'error': 'متن پاسخ الزامی است'}), 400
    
    # ایجاد پیام جدید (برعکس: فرستنده و گیرنده جابجا می‌شوند)
    new_msg = WorkRevisionMessage(
        work_status_id=original_msg.work_status_id,
        personnel_id=original_msg.personnel_id,
        period_id=original_msg.period_id,
        from_role=current_user.role,
        from_user_id=current_user.id,
        to_role=original_msg.from_role,
        to_user_id=original_msg.from_user_id,
        message=reply_message,
        is_read=False
    )
    db.session.add(new_msg)
    db.session.commit()
    
    # ارسال اعلان
    send_workflow_notification(
        to_user_id=original_msg.from_user_id,
        title=f"پاسخ به پیام اصلاح کارکرد",
        message=f"{current_user.get_full_name()} به پیام شما پاسخ داد: {reply_message[:100]}",
        link=None
    )
    
    return jsonify({'success': True})

    
@app.route('/api/work/messages/<int:msg_id>/delete', methods=['DELETE'])
@login_required
def api_work_message_delete(msg_id):
    """حذف پیام (فقط برای گیرنده)"""
    msg = WorkRevisionMessage.query.get_or_404(msg_id)
    
    if msg.to_user_id != current_user.id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    db.session.delete(msg)
    db.session.commit()
    
    return jsonify({'success': True})

# ==================== توابع کمکی وضعیت کارکرد ====================

def get_personnel_work_status(personnel_id, period_id):
    """دریافت وضعیت کارکرد یک پرسنل در یک دوره"""
    return PersonnelWorkStatus.query.filter_by(
        personnel_id=personnel_id, 
        period_id=period_id
    ).first()


def create_or_update_work_status(personnel_id, period_id, status, **kwargs):
    """ایجاد یا بروزرسانی وضعیت کارکرد"""
    work_status = get_personnel_work_status(personnel_id, period_id)
    
    if not work_status:
        work_status = PersonnelWorkStatus(
            personnel_id=personnel_id,
            period_id=period_id,
            status=status
        )
        db.session.add(work_status)
    
    work_status.status = status
    
    for key, value in kwargs.items():
        if hasattr(work_status, key):
            setattr(work_status, key, value)
    
    work_status.updated_at = datetime.now()
    db.session.commit()
    
    return work_status


def can_edit_personnel(user, personnel, period_id):
    """بررسی دسترسی ویرایش پرسنل بر اساس سلسله مراتب و وضعیت"""
    work_status = get_personnel_work_status(personnel.id, period_id)
    
    if not work_status or work_status.status == 'draft':
        return True
    
    if user.role == 'org_manager':
        return True
    
    if user.role == 'dept_manager':
        # اضافه کردن 'unit_pending'
        return work_status.status in ['draft', 'unit_pending', 'unit_approved', 'revision']
    
    if user.role == 'unit_supervisor':
        return work_status.status in ['draft', 'revision']
    
    return False


def can_delete_personnel(user, personnel, period_id):
    """بررسی دسترسی حذف پرسنل بر اساس سلسله مراتب و وضعیت"""
    work_status = get_personnel_work_status(personnel.id, period_id)
    
    if user.role == 'org_manager':
        return True
    
    if user.role == 'dept_manager':
        # مدیر اداره می‌تواند در حالت‌های پیش‌نویس، در انتظار تایید خودش، و اصلاح حذف کند
        return work_status.status in ['draft', 'unit_pending', 'unit_approved', 'revision']
    
    if user.role == 'unit_supervisor':
        # سرپرست واحد فقط در حالت پیش‌نویس و اصلاح می‌تواند حذف کند
        return work_status.status in ['draft', 'revision']
    
    return False


def send_workflow_notification(to_user_id, title, message, notification_type='workflow', link=None):
    """ارسال اعلان گردش کار"""
    if not to_user_id:
        return
    
    notif = Notification(
        user_id=to_user_id,
        notification_type=notification_type,
        title=title,
        message=message,
        link=link,
        is_read=False
    )
    db.session.add(notif)
    db.session.commit()
    print(f"✅ اعلان ارسال شد به کاربر {to_user_id}: {title}")


def create_work_message(work_status_id, personnel_id, period_id, from_user, to_user_id, to_role, message, message_type='general'):
    """ایجاد پیام در صندوق پیام"""
    if not to_user_id:
        return None
    
    # دریافت اطلاعات پرسنل و دوره
    personnel = Personnel.query.get(personnel_id)
    period = WorkPeriod.query.get(period_id)
    
    # ساخت عنوان و متن پیام بر اساس نوع
    if message_type == 'approve':
        prefix = "✅ تایید کارکرد"
        body = f"مدیر اداره {from_user.get_full_name()} کارکرد پرسنل {personnel.get_full_name()} را برای دوره {period.title} تایید کرد."
        if message:
            body += f"\n\n📝 توضیحات:\n{message}"
    elif message_type == 'direct_approve':
        prefix = "✅ تایید مستقیم کارکرد"
        body = f"مدیر اداره {from_user.get_full_name()} کارکرد پرسنل {personnel.get_full_name()} را برای دوره {period.title} مستقیماً تایید کرد."
        if message:
            body += f"\n\n📝 توضیحات:\n{message}"
    elif message_type == 'revision':
        prefix = "🔄 اصلاح کارکرد"
        body = f"مدیر اداره {from_user.get_full_name()} درخواست اصلاح کارکرد پرسنل {personnel.get_full_name()} را برای دوره {period.title} صادر کرد."
        if message:
            body += f"\n\n📝 توضیحات اصلاح:\n{message}"
        else:
            body += f"\n\nلطفاً اطلاعات را ویرایش و مجدداً ارسال کنید."
    else:
        prefix = "📝 پیام جدید"
        body = message
    
    final_message = f"{prefix}\n\n{body}"
    
    msg = WorkRevisionMessage(
        work_status_id=work_status_id,
        personnel_id=personnel_id,
        period_id=period_id,
        from_role=from_user.role,
        from_user_id=from_user.id,
        to_role=to_role,
        to_user_id=to_user_id,
        message=final_message,
        is_read=False,
        message_type=message_type
    )
    db.session.add(msg)
    db.session.commit()
    print(f"✅ پیام ایجاد شد به کاربر {to_user_id}: {prefix}")
    return msg

def send_revision_message(work_status_id, from_user, to_role, to_user_id, message_text):
    """ارسال پیام اصلاح به صندوق پیام"""
    work_status = PersonnelWorkStatus.query.get(work_status_id)
    if not work_status:
        print(f"⚠️ work_status with id {work_status_id} not found")
        return
    
    msg = WorkRevisionMessage(
        work_status_id=work_status_id,
        personnel_id=work_status.personnel_id,
        period_id=work_status.period_id,
        from_role=from_user.role,
        from_user_id=from_user.id,
        to_role=to_role,
        to_user_id=to_user_id,
        message=message_text,
        is_read=False
    )
    db.session.add(msg)
    db.session.commit()
    
    # همچنین اعلان ارسال کن
    send_workflow_notification(
        to_user_id=to_user_id,
        title=f"پیام اصلاح کارکرد از {from_user.get_full_name()}",
        message=message_text[:200] if message_text else "درخواست اصلاح کارکرد",
        link=None
    )
    
@app.route('/api/work/bulk-submit', methods=['POST'])
@login_required
def api_work_bulk_submit():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    personnel_ids = data.get('personnel_ids', [])
    period_id = data.get('period_id')
    
    if not period_id:
        return jsonify({'error': 'دوره انتخاب نشده است'}), 400
    
    if not personnel_ids:
        return jsonify({'error': 'هیچ پرسنلی انتخاب نشده است'}), 400
    
    success_count = 0
    errors = []
    
    for pid in personnel_ids:
        try:
            # بازسازی درخواست تکی برای هر پرسنل
            personnel = Personnel.query.get(pid)
            if not personnel:
                continue
            
            # بررسی فیلدهای اجباری
            dynamic_fields = DynamicField.query.filter_by(is_active=True, is_required=True).all()
            missing_fields = []
            for field in dynamic_fields:
                if field.is_key:
                    continue
                pv = PersonnelValue.query.filter_by(
                    personnel_id=pid,
                    field_id=field.id,
                    period_id=period_id
                ).first()
                if not pv or not (pv.value_text or pv.value_number or pv.value_date):
                    missing_fields.append(field.title)
            
            if missing_fields:
                errors.append(f"{personnel.get_full_name()}: {', '.join(missing_fields)}")
                continue
            
            # ایجاد یا بروزرسانی وضعیت
            work_status = create_or_update_work_status(
                personnel_id=pid,
                period_id=period_id,
                status='unit_pending'
            )
            work_status.unit_approved_at = datetime.now()
            work_status.unit_approver_id = current_user.id
            db.session.commit()
            
            # ارسال اعلان به مدیر اداره
            dept_manager = DepartmentManager.query.filter_by(department_id=personnel.department_id).first()
            if dept_manager:
                send_workflow_notification(
                    to_user_id=dept_manager.user_id,
                    title=f"درخواست تایید کارکرد دوره",
                    message=f"سرپرست واحد {current_user.get_full_name()} کارکرد پرسنل {personnel.get_full_name()} را تایید کرده است.",
                    link=None
                )
            
            success_count += 1
            
        except Exception as e:
            errors.append(str(e))
    
    return jsonify({
        'success': True,
        'count': success_count,
        'errors': errors if errors else None
    })
    
# ==================== API برای مدیریت پرسنل (فیلدهای داینامیک) ====================
@app.route('/admin/api/fields')
@login_required
def admin_api_fields():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    fields = DynamicField.query.filter_by(is_active=True).order_by(DynamicField.field_order).all()
    return jsonify([{
        'id': f.id,
        'title': f.title,
        'field_type': f.field_type,
        'is_required': f.is_required,
        'is_key': f.is_key
    } for f in fields])


@app.route('/admin/api/departments')
@login_required
@cached(ttl=300)
def admin_api_departments():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    depts = Department.query.filter_by(is_active=True).all()
    return jsonify([{'id': d.id, 'name': d.name} for d in depts])


@app.route('/admin/api/units')
@login_required
def admin_api_units():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    dept_id = request.args.get('department_id')
    if not dept_id:
        return jsonify([])
    
    units = Unit.query.filter_by(department_id=dept_id, is_active=True).all()
    return jsonify([{'id': u.id, 'name': u.name} for u in units])


@app.route('/admin/api/personnel')
@login_required
def admin_api_personnel():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    personnel = Personnel.query.filter_by(is_deleted=False).all()
    result = []
    for p in personnel:
        dept = Department.query.get(p.department_id)
        unit = Unit.query.get(p.unit_id)
        values = PersonnelValue.query.filter_by(personnel_id=p.id).all()
        result.append({
            'id': p.id,
            'national_code': p.national_code,
            'first_name': p.first_name,
            'last_name': p.last_name,
            'phone': p.phone,
            'position': p.position,
            'hire_date': p.hire_date,
            'department_name': dept.name if dept else '-',
            'unit_name': unit.name if unit else '-',
            'values': [{'field_id': v.field_id, 'value_text': v.value_text, 'value_number': v.value_number, 'value_date': v.value_date} for v in values]
        })
    return jsonify(result)


@app.route('/admin/api/personnel/create', methods=['POST'])
@login_required
def admin_api_personnel_create():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    
    # بررسی یکتایی کد ملی
    national_code = data.get('national_code')
    if not national_code:
        return jsonify({'error': 'کد ملی الزامی است'}), 400
    if Personnel.query.filter_by(national_code=national_code).first():
        return jsonify({'error': 'کد ملی تکراری است'}), 400
    
    # ایجاد پرسنل جدید
    personnel = Personnel(
        national_code=national_code,
        first_name=data.get('first_name', ''),
        last_name=data.get('last_name', ''),
        phone=data.get('phone', ''),
        position=data.get('position', ''),
        hire_date=data.get('hire_date', ''),
        department_id=data.get('department_id'),
        unit_id=data.get('unit_id')
    )
    db.session.add(personnel)
    db.session.commit()
    
    # ذخیره مقادیر فیلدهای داینامیک
    fields = DynamicField.query.filter_by(is_active=True).all()
    for field in fields:
        value = data.get(str(field.id))
        if value and value != '':
            pv = PersonnelValue(
                personnel_id=personnel.id,
                field_id=field.id,
                value_text=value if field.field_type == 'text' else None,
                value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
                value_date=value if field.field_type == 'date' else None
            )
            db.session.add(pv)
    
    db.session.commit()
    return jsonify({'success': True, 'message': 'پرسنل با موفقیت اضافه شد'})


@app.route('/admin/api/personnel/<int:pid>/delete', methods=['DELETE'])
@login_required
def admin_api_personnel_delete(pid):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    p = Personnel.query.get_or_404(pid)
    p.is_deleted = True
    db.session.commit()
    return jsonify({'success': True, 'message': 'پرسنل حذف شد'})

@app.route('/admin/api/personnel/<int:pid>')
@login_required
def admin_api_personnel_get(pid):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    p = Personnel.query.get_or_404(pid)
    dept = Department.query.get(p.department_id)
    unit = Unit.query.get(p.unit_id)
    values = PersonnelValue.query.filter_by(personnel_id=p.id).all()
    
    return jsonify({
        'id': p.id,
        'national_code': p.national_code,
        'first_name': p.first_name,
        'last_name': p.last_name,
        'phone': p.phone,
        'position': p.position,
        'hire_date': p.hire_date,
        'department_id': p.department_id,
        'unit_id': p.unit_id,
        'department_name': dept.name if dept else '-',
        'unit_name': unit.name if unit else '-',
        'values': [{'field_id': v.field_id, 'value_text': v.value_text, 'value_number': v.value_number, 'value_date': v.value_date} for v in values]
    })


@app.route('/admin/api/personnel/<int:pid>/edit', methods=['PUT'])
@login_required
def admin_api_personnel_edit(pid):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    p = Personnel.query.get_or_404(pid)
    data = request.get_json()
    
    # بروزرسانی اطلاعات پایه
    p.first_name = data.get('first_name', p.first_name)
    p.last_name = data.get('last_name', p.last_name)
    p.phone = data.get('phone', p.phone)
    p.position = data.get('position', p.position)
    p.hire_date = data.get('hire_date', p.hire_date)
    p.department_id = data.get('department_id', p.department_id)
    p.unit_id = data.get('unit_id', p.unit_id)
    
    # بروزرسانی کد ملی اگر تغییر کرده
    national_code = data.get('national_code')
    if national_code and national_code != p.national_code:
        if Personnel.query.filter_by(national_code=national_code).first():
            return jsonify({'error': 'کد ملی تکراری است'}), 400
        p.national_code = national_code
    
    db.session.commit()
    
    # بروزرسانی مقادیر فیلدهای داینامیک
    fields = DynamicField.query.filter_by(is_active=True).all()
    for field in fields:
        value = data.get(str(field.id))
        if value:
            pv = PersonnelValue.query.filter_by(personnel_id=pid, field_id=field.id).first()
            if pv:
                pv.value_text = value if field.field_type == 'text' else None
                pv.value_number = float(value) if field.field_type in ['number', 'decimal'] and value else None
                pv.value_date = value if field.field_type == 'date' else None
                pv.updated_at = datetime.now()
            else:
                pv = PersonnelValue(
                    personnel_id=pid,
                    field_id=field.id,
                    value_text=value if field.field_type == 'text' else None,
                    value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
                    value_date=value if field.field_type == 'date' else None
                )
                db.session.add(pv)
    
    db.session.commit()
    return jsonify({'success': True})

# ==================== مسیرهای عمومی ====================
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password) and user.is_active and user.is_approved:
            # ========== این خط مهم است ==========
            user.last_login = datetime.now()
            db.session.commit()
            # ===================================
            login_user(user)
            flash(f'خوش آمدید {user.get_full_name()}', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('کد ملی یا رمز عبور اشتباه است', 'error')
    
    return render_template('login.html')

@app.route('/test-update-login')
@login_required
def test_update_login():
    current_user.last_login = datetime.now()
    db.session.commit()
    return f"✅ last_login برای {current_user.get_full_name()} به {current_user.last_login} بروزرسانی شد"

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('شما از سامانه خارج شدید', 'info')
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.role == 'admin':
        return redirect(url_for('admin_dashboard'))
    elif current_user.role == 'org_manager':
        return redirect(url_for('org_manager_dashboard'))
    elif current_user.role == 'dept_manager':
        return redirect(url_for('dept_manager_dashboard'))
    elif current_user.role == 'hr_manager':
        return redirect(url_for('hr_manager_dashboard'))
    elif current_user.role == 'unit_supervisor':
        return redirect(url_for('unit_supervisor_dashboard'))
    else:
        return redirect(url_for('subordinate_dashboard'))
        
        
@app.route('/admin/api/users-passwords')
@login_required
def admin_api_users_passwords():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    users = User.query.all()
    result = {}
    for u in users:
        # رمز پیش‌فرض: 4 رقم آخر کد ملی
        result[u.id] = u.national_code[-4:] if u.national_code else '1234'
    
    return jsonify(result)
    
    
# ==================== سیستم تیکت (فقط یک بار) ====================
@app.route('/tickets')
@login_required
def tickets():
    return render_template('tickets.html')


@app.route('/api/tickets')
@login_required
def api_tickets():
    admin_user = User.query.filter_by(role='admin').first()
    admin_id = admin_user.id if admin_user else 1
    
    if current_user.role == 'admin':
        tickets = Ticket.query.filter(
            db.or_(
                Ticket.receiver_id == current_user.id,
                Ticket.sender_id == current_user.id
            )
        ).order_by(Ticket.created_at.desc()).all()
    else:
        tickets = Ticket.query.filter(
            db.or_(
                Ticket.sender_id == current_user.id,
                db.and_(
                    Ticket.receiver_id == current_user.id,
                    Ticket.message_type == 'message'
                )
            )
        ).order_by(Ticket.created_at.desc()).all()
    
    result = []
    for t in tickets:
        sender = User.query.get(t.sender_id)
        receiver = User.query.get(t.receiver_id)
        msg_type = getattr(t, 'message_type', 'ticket')
        
        # ✅ اصلاح شده: فقط status='open' هایلایت بشه
        is_unread = (t.receiver_id == current_user.id and t.status == 'open')
        
        result.append({
            'id': t.id,
            'title': t.title,
            'message': t.message,
            'sender_name': sender.get_full_name() if sender else '-',
            'receiver_name': receiver.get_full_name() if receiver else '-',
            'status': t.status,
            'priority': t.priority,
            'message_type': msg_type,
            'created_at': t.created_at.strftime('%Y/%m/%d %H:%M'),
            'reply_count': len(t.replies),
            'is_unread': is_unread
        })
    
    # محاسبه تعداد خوانده نشده برای هدر (فقط open)
    unread_count = Ticket.query.filter(
        Ticket.receiver_id == current_user.id,
        Ticket.status == 'open'
    ).count()
    
    return jsonify({
        'messages': result,
        'unread_count': unread_count
    })

@app.route('/api/tickets/create', methods=['POST'])
@login_required
def api_ticket_create():
    data = request.get_json()
    title = data.get('title', '').strip()
    message = data.get('message', '').strip()
    priority = data.get('priority', 'normal')
    msg_type = data.get('message_type', 'ticket')  # 'ticket' یا 'message'
    
    if not title or not message:
        return jsonify({'error': 'عنوان و متن پیام الزامی است'}), 400
    
    if current_user.role == 'admin':
        receiver_id = data.get('receiver_id')
        if not receiver_id:
            return jsonify({'error': 'لطفاً کاربر گیرنده را انتخاب کنید'}), 400
        receiver_id = int(receiver_id)
    else:
        admin_user = User.query.filter_by(role='admin').first()
        if not admin_user:
            return jsonify({'error': 'ادمین در سیستم ثبت نشده است'}), 400
        receiver_id = admin_user.id
    
    receiver = User.query.get(receiver_id)
    if not receiver:
        return jsonify({'error': 'کاربر گیرنده یافت نشد'}), 400
    
    ticket = Ticket(
        title=title,
        message=message,
        sender_id=current_user.id,
        receiver_id=receiver_id,
        priority=priority,
        message_type=msg_type
    )
    db.session.add(ticket)
    db.session.commit()
    
    return jsonify({'success': True})


@app.route('/api/tickets/<int:ticket_id>')
@login_required
def api_ticket_get(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    
    if current_user.role != 'admin' and ticket.sender_id != current_user.id and ticket.receiver_id != current_user.id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # ========== اضافه کن: اگر کاربر گیرنده است و تیکت رو می‌بینه ==========
    if ticket.receiver_id == current_user.id and ticket.status == 'open':
        ticket.status = 'in_progress'
        db.session.commit()
    
    sender = User.query.get(ticket.sender_id)
    receiver = User.query.get(ticket.receiver_id)
    
    replies = []
    for r in ticket.replies:
        reply_user = User.query.get(r.user_id)
        replies.append({
            'id': r.id,
            'message': r.message,
            'is_admin_reply': r.is_admin_reply,
            'user_name': reply_user.get_full_name() if reply_user else '-',
            'created_at': r.created_at.strftime('%Y/%m/%d %H:%M')
        })
    
    return jsonify({
        'id': ticket.id,
        'title': ticket.title,
        'message': ticket.message,
        'sender_name': sender.get_full_name() if sender else '-',
        'receiver_name': receiver.get_full_name() if receiver else '-',
        'status': ticket.status,
        'priority': ticket.priority,
        'message_type': getattr(ticket, 'message_type', 'ticket'),
        'created_at': ticket.created_at.strftime('%Y/%m/%d %H:%M'),
        'replies': replies
    })


@app.route('/api/tickets/<int:ticket_id>/reply', methods=['POST'])
@login_required
def api_ticket_reply(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    
    if current_user.id != ticket.sender_id and current_user.id != ticket.receiver_id and current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    message = data.get('message', '').strip()
    
    if not message:
        return jsonify({'error': 'متن پاسخ الزامی است'}), 400
    
    is_admin_reply = (current_user.role == 'admin')
    
    reply = TicketReply(
        ticket_id=ticket_id,
        user_id=current_user.id,
        message=message,
        is_admin_reply=is_admin_reply
    )
    db.session.add(reply)
    
    # تغییر وضعیت تیکت
    if is_admin_reply:
        new_status = data.get('status', 'in_progress')
        ticket.status = new_status
    else:
        ticket.status = 'in_progress'
    
    ticket.updated_at = datetime.now()
    db.session.commit()
    
    # محاسبه تعداد خوانده نشده جدید برای گیرنده (ادمین)
    unread_count = Ticket.query.filter(
        Ticket.receiver_id == current_user.id,
        Ticket.status.in_(['open', 'in_progress'])
    ).count()
    
    # برگردوندن وضعیت جدید برای این تیکت
    is_now_unread = (ticket.receiver_id == current_user.id and ticket.status in ['open', 'in_progress'])
    
    return jsonify({
        'success': True,
        'unread_count': unread_count,
        'is_unread': is_now_unread
    })



# ==================== APIهای کامل برای پنل مدیر سازمان ====================

@app.route('/org-manager/api/all-data')
@login_required
@cached(ttl=120)  # کش 2 دقیقه
def org_manager_api_all_data():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # دریافت همه ادارات
    departments = Department.query.filter_by(is_active=True).all()
    department_ids = [d.id for d in departments]
    
    # دریافت همه واحدها
    units = Unit.query.filter(Unit.department_id.in_(department_ids), Unit.is_active == True).all()
    unit_ids = [u.id for u in units]
    
    # دریافت پرسنل همه واحدها
    personnel = Personnel.query.filter(Personnel.unit_id.in_(unit_ids), Personnel.is_deleted == False).all()
    
    # دریافت فیلدهای داینامیک
    dynamic_fields = DynamicField.query.filter_by(is_active=True).order_by(DynamicField.field_order).all()
    
    # دریافت همه دوره‌ها
    all_periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    
    personnel_data = []
    for p in personnel:
        unit = Unit.query.get(p.unit_id)
        department = Department.query.get(p.department_id) if p.department_id else None
        values = {}
        
        # دریافت مقادیر فیلدها
        for v in PersonnelValue.query.filter_by(personnel_id=p.id).all():
            field = DynamicField.query.get(v.field_id)
            if field:
                values[field.title] = v.value_text or v.value_number or v.value_date or ''
        
        # دریافت عنوان دوره
        period_title = ''
        if p.period_id:
            period = WorkPeriod.query.get(p.period_id)
            if period:
                period_title = f"{period.title} ({period.start_date} - {period.end_date})"
        
        item = {
            'id': p.id,
            'national_code': p.national_code,
            'first_name': p.first_name or '',
            'last_name': p.last_name or '',
            'full_name': p.get_full_name(),
            'phone': p.phone or '',
            'position': p.position or '',
            'department_id': p.department_id,
            'department_name': department.name if department else '-',
            'unit_id': p.unit_id,
            'unit_name': unit.name if unit else '-',
            'period_title': period_title,
            'period_id': p.period_id
        }
        
        # اضافه کردن فیلدهای داینامیک
        for field in dynamic_fields:
            if not field.is_key:
                item[field.title] = values.get(field.title, '')
        
        personnel_data.append(item)
    
    # داده‌های ادارات
    departments_data = [{'id': d.id, 'name': d.name, 'color': d.color} for d in departments]
    
    # داده‌های واحدها با department_id
    units_data = [{'id': u.id, 'name': u.name, 'department_id': u.department_id} for u in units]
    
    return jsonify({
        'personnel': personnel_data,
        'dynamic_fields': [{'id': f.id, 'title': f.title, 'field_type': f.field_type, 'is_key': f.is_key, 'is_required': f.is_required} for f in dynamic_fields],
        'periods': [{'id': p.id, 'title': p.title, 'start_date': p.start_date, 'end_date': p.end_date} for p in all_periods],
        'departments': departments_data,
        'units': units_data
    })


@app.route('/org-manager/api/periods')
@login_required
def org_manager_api_periods():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    return jsonify([{
        'id': p.id,
        'title': p.title,
        'start_date': p.start_date,
        'end_date': p.end_date,
        'is_active': p.is_active
    } for p in periods])


@app.route('/org-manager/api/pending-requests')
@login_required
def org_manager_api_pending_requests():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # دریافت همه واحدهای سازمان
    departments = Department.query.filter_by(is_active=True).all()
    department_ids = [d.id for d in departments]
    units = Unit.query.filter(Unit.department_id.in_(department_ids)).all()
    unit_ids = [u.id for u in units]
    
    requests = UnitPersonnelRequest.query.filter(
        UnitPersonnelRequest.unit_id.in_(unit_ids),
        UnitPersonnelRequest.request_type == 'add',
        UnitPersonnelRequest.status == 'pending'
    ).order_by(UnitPersonnelRequest.created_at.desc()).all()
    
    result = []
    for req in requests:
        data = json.loads(req.data) if req.data else {}
        unit = Unit.query.get(req.unit_id)
        created_at_jalali = jdatetime.datetime.fromgregorian(datetime=req.created_at).strftime('%Y/%m/%d %H:%M')
        
        result.append({
            'id': req.id,
            'request_type': req.request_type,
            'national_code': data.get('national_code', ''),
            'first_name': data.get('first_name', ''),
            'last_name': data.get('last_name', ''),
            'full_name': f"{data.get('first_name', '')} {data.get('last_name', '')}".strip(),
            'unit_name': unit.name if unit else '-',
            'created_at': created_at_jalali
        })
    
    return jsonify({'requests': result})


@app.route('/org-manager/api/delete-requests')
@login_required
def org_manager_api_delete_requests():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # دریافت همه واحدهای سازمان
    departments = Department.query.filter_by(is_active=True).all()
    department_ids = [d.id for d in departments]
    units = Unit.query.filter(Unit.department_id.in_(department_ids)).all()
    unit_ids = [u.id for u in units]
    
    requests = UnitPersonnelRequest.query.filter(
        UnitPersonnelRequest.unit_id.in_(unit_ids),
        UnitPersonnelRequest.request_type == 'delete',
        UnitPersonnelRequest.status == 'pending'
    ).order_by(UnitPersonnelRequest.created_at.desc()).all()
    
    result = []
    for req in requests:
        data = json.loads(req.data) if req.data else {}
        unit = Unit.query.get(req.unit_id)
        created_at_jalali = jdatetime.datetime.fromgregorian(datetime=req.created_at).strftime('%Y/%m/%d %H:%M')
        
        result.append({
            'id': req.id,
            'request_type': req.request_type,
            'personnel_id': data.get('personnel_id'),
            'national_code': data.get('national_code', ''),
            'first_name': data.get('first_name', ''),
            'last_name': data.get('last_name', ''),
            'full_name': data.get('full_name', ''),
            'unit_name': unit.name if unit else '-',
            'delete_reason': data.get('delete_reason', ''),
            'created_at': created_at_jalali
        })
    
    return jsonify({'requests': result})


@app.route('/org-manager/api/cancel-request', methods=['POST'])
@login_required
def org_manager_api_cancel_request():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    request_id = data.get('request_id')
    
    if not request_id:
        return jsonify({'error': 'شناسه درخواست ارسال نشده است'}), 400
    
    req = UnitPersonnelRequest.query.get(request_id)
    if not req:
        return jsonify({'error': 'درخواست یافت نشد'}), 404
    
    db.session.delete(req)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست لغو شد'})


@app.route('/org-manager/api/request-add', methods=['POST'])
@login_required
def org_manager_api_request_add():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    print("📥 داده دریافتی در org_manager:", data)  # برای دیباگ
    
    unit_id = data.get('unit_id')
    period_id = data.get('period_id')
    national_code = data.get('national_code')  # ← این خط مهم است
    department_id = data.get('department_id')
    
    # اعتبارسنجی
    if not department_id:
        return jsonify({'error': 'اداره انتخاب نشده است'}), 400
    if not unit_id:
        return jsonify({'error': 'واحد انتخاب نشده است'}), 400
    if not period_id:
        return jsonify({'error': 'دوره انتخاب نشده است'}), 400
    if not national_code:
        return jsonify({'error': 'کد ملی الزامی است'}), 400
    
    # اعتبارسنجی کد ملی (10 رقم)
    if not (national_code.isdigit() and len(national_code) == 10):
        return jsonify({'error': 'کد ملی باید 10 رقم باشد'}), 400
    
    unit = Unit.query.get(unit_id)
    if not unit:
        return jsonify({'error': 'واحد یافت نشد'}), 404
    
    # بررسی تکراری نبودن کد ملی در دوره
    existing = Personnel.query.filter_by(
        national_code=national_code, 
        period_id=period_id, 
        is_deleted=False
    ).first()
    if existing:
        return jsonify({'error': 'این کد ملی قبلاً در دوره انتخاب شده ثبت شده است'}), 400
    
    # جمع‌آوری داده‌های پرسنل
    request_data = {
        'department_id': department_id,
        'unit_id': unit_id,
        'national_code': national_code,
        'requester_note': data.get('note', ''),
        'period_id': period_id
    }
    
    fields = DynamicField.query.filter_by(is_active=True).all()
    for field in fields:
        value = data.get(field.title)
        if value:
            request_data[field.title] = value
            if field.title == 'نام':
                request_data['first_name'] = value
            if field.title == 'نام خانوادگی':
                request_data['last_name'] = value
            if field.title == 'شماره تماس':
                request_data['phone'] = value
            if field.title == 'سمت':
                request_data['position'] = value
    
    # ذخیره درخواست
    approval = UnitPersonnelRequest(
        unit_id=unit_id,
        requester_id=current_user.id,
        request_type='add',
        data=json.dumps(request_data, ensure_ascii=False),
        status='pending'
    )
    db.session.add(approval)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست شما ثبت شد و منتظر تایید ادمین است'})

@app.route('/org-manager/api/request-delete', methods=['POST'])
@login_required
def org_manager_api_request_delete():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    personnel_id = data.get('personnel_id')
    delete_reason = data.get('delete_reason', '')
    
    if not personnel_id:
        return jsonify({'error': 'شناسه پرسنل ارسال نشده است'}), 400
    
    p = Personnel.query.get(personnel_id)
    if not p:
        return jsonify({'error': 'پرسنل یافت نشد'}), 404
    
    # حذف درخواست قبلی
    UnitPersonnelRequest.query.filter_by(
        target_personnel_id=personnel_id,
        request_type='delete',
        status='pending'
    ).delete()
    db.session.commit()
    
    request_data = {
        'personnel_id': p.id,
        'national_code': p.national_code,
        'first_name': p.first_name or '',
        'last_name': p.last_name or '',
        'full_name': p.get_full_name(),
        'phone': p.phone or '',
        'position': p.position or '',
        'delete_reason': delete_reason
    }
    
    approval = UnitPersonnelRequest(
        unit_id=p.unit_id,
        requester_id=current_user.id,
        request_type='delete',
        target_personnel_id=personnel_id,
        data=json.dumps(request_data, ensure_ascii=False),
        status='pending'
    )
    db.session.add(approval)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست حذف پرسنل ثبت شد'})


@app.route('/org-manager/api/personnel/<int:pid>')
@login_required
def org_manager_api_personnel_get(pid):
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    p = Personnel.query.get_or_404(pid)
    fields = DynamicField.query.filter_by(is_active=True).all()
    
    values = {}
    for v in PersonnelValue.query.filter_by(personnel_id=pid).all():
        field = DynamicField.query.get(v.field_id)
        if field:
            values[field.title] = v.value_text or v.value_number or v.value_date or ''
    
    result = {
        'id': p.id,
        'national_code': p.national_code,
        'first_name': p.first_name or '',
        'last_name': p.last_name or '',
        'phone': p.phone or '',
        'position': p.position or '',
        'department_id': p.department_id,
        'unit_id': p.unit_id,
        'period_id': p.period_id
    }
    
    for field in fields:
        result[field.title] = values.get(field.title, '')
    
    return jsonify(result)


@app.route('/org-manager/api/personnel/<int:pid>/edit', methods=['PUT'])
@login_required
def org_manager_api_personnel_edit(pid):
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        p = Personnel.query.get_or_404(pid)
        data = request.get_json()
        
        # به‌روزرسانی فیلدهای پایه
        p.first_name = data.get('first_name', p.first_name)
        p.last_name = data.get('last_name', p.last_name)
        p.phone = data.get('phone', p.phone)
        p.position = data.get('position', p.position)
        
        # به‌روزرسانی period_id
        period_id = data.get('period_id')
        if period_id and period_id != '' and period_id != 'null':
            p.period_id = int(period_id)
        else:
            p.period_id = None
        
        db.session.commit()
        
        # به‌روزرسانی فیلدهای داینامیک
        fields = DynamicField.query.filter_by(is_active=True).all()
        for field in fields:
            value = data.get(field.title)
            if value is not None and value != '':
                pv = PersonnelValue.query.filter_by(
                    personnel_id=pid,
                    field_id=field.id,
                    period_id=p.period_id
                ).first()
                
                if pv:
                    if field.field_type == 'text':
                        pv.value_text = value
                    elif field.field_type in ['number', 'decimal']:
                        try:
                            pv.value_number = float(value)
                        except:
                            pv.value_text = value
                    elif field.field_type == 'date':
                        pv.value_date = value
                    pv.updated_at = datetime.now()
                else:
                    pv = PersonnelValue(
                        personnel_id=pid,
                        field_id=field.id,
                        period_id=p.period_id,
                        value_text=value if field.field_type == 'text' else None,
                        value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
                        value_date=value if field.field_type == 'date' else None
                    )
                    db.session.add(pv)
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'اطلاعات ذخیره شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/org-manager/api/personnel-history/<int:pid>')
@login_required
def org_manager_api_personnel_history(pid):
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    p = Personnel.query.get_or_404(pid)
    fields = DynamicField.query.filter_by(is_active=True).all()
    periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    
    history = []
    for period in periods:
        period_data = {
            'period_id': period.id,
            'period_title': period.title,
            'period_start': period.start_date,
            'period_end': period.end_date,
            'values': {}
        }
        
        for field in fields:
            pv = PersonnelValue.query.filter_by(
                personnel_id=pid,
                field_id=field.id,
                period_id=period.id
            ).first()
            
            if pv:
                value = pv.value_text or pv.value_number or pv.value_date or '-'
            else:
                pv_default = PersonnelValue.query.filter_by(
                    personnel_id=pid,
                    field_id=field.id,
                    period_id=None
                ).first()
                if pv_default:
                    value = pv_default.value_text or pv_default.value_number or pv_default.value_date or '-'
                else:
                    value = '-'
            
            period_data['values'][field.title] = value
        
        history.append(period_data)
    
    return jsonify({
        'personnel': {
            'id': p.id,
            'national_code': p.national_code,
            'full_name': p.get_full_name()
        },
        'fields': [{'id': f.id, 'title': f.title, 'field_type': f.field_type, 'is_key': f.is_key} for f in fields],
        'history': history
    })


@app.route('/org-manager/api/personnel/batch-update', methods=['POST'])
@login_required
def org_manager_api_personnel_batch_update():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    field = data.get('field')
    updates = data.get('updates', [])
    
    dynamic_field = DynamicField.query.filter_by(title=field, is_active=True).first()
    if not dynamic_field:
        return jsonify({'error': f'فیلد "{field}" یافت نشد'}), 400
    
    count = 0
    for item in updates:
        p = Personnel.query.get(item['id'])
        if p:
            value = item['value']
            
            pv = PersonnelValue.query.filter_by(
                personnel_id=p.id,
                field_id=dynamic_field.id,
                period_id=p.period_id
            ).first()
            
            if pv:
                if dynamic_field.field_type == 'text':
                    pv.value_text = value
                elif dynamic_field.field_type in ['number', 'decimal']:
                    try:
                        pv.value_number = float(value)
                    except:
                        pv.value_text = value
                elif dynamic_field.field_type == 'date':
                    pv.value_date = value
                pv.updated_at = datetime.now()
            else:
                pv = PersonnelValue(
                    personnel_id=p.id,
                    field_id=dynamic_field.id,
                    period_id=p.period_id,
                    value_text=value if dynamic_field.field_type == 'text' else None,
                    value_number=float(value) if dynamic_field.field_type in ['number', 'decimal'] and value else None,
                    value_date=value if dynamic_field.field_type == 'date' else None
                )
                db.session.add(pv)
            count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'count': count})


@app.route('/org-manager/api/import-excel', methods=['POST'])
@login_required
def org_manager_api_import_excel():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    if 'excel_file' not in request.files:
        return jsonify({'error': 'فایلی ارسال نشده'}), 400
    
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'error': 'فایلی انتخاب نشده'}), 400
    
    try:
        import pandas as pd
        df = pd.read_excel(file, dtype=str).fillna('')
        success, error = 0, 0
        
        for _, row in df.iterrows():
            code = str(row.get('کد ملی', '')).strip().replace('.0', '')
            if not code.isdigit() or len(code) != 10:
                error += 1
                continue
            
            if Personnel.query.filter_by(national_code=code).first():
                error += 1
                continue
            
            name = str(row.get('نام', '')).strip()
            family = str(row.get('نام خانوادگی', '')).strip()
            if not name or not family:
                error += 1
                continue
            
            dept_name = str(row.get('نام اداره', '')).strip()
            unit_name = str(row.get('نام واحد', '')).strip()
            
            dept = Department.query.filter_by(name=dept_name).first()
            if not dept:
                error += 1
                continue
            
            unit = Unit.query.filter_by(name=unit_name, department_id=dept.id).first()
            if not unit:
                error += 1
                continue
            
            p = Personnel(
                national_code=code,
                first_name=name,
                last_name=family,
                phone=str(row.get('شماره تماس', '')).strip(),
                position=str(row.get('سمت', '')).strip(),
                hire_date=str(row.get('تاریخ استخدام', '')).strip(),
                department_id=dept.id,
                unit_id=unit.id
            )
            db.session.add(p)
            success += 1
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'✅ {success} پرسنل اضافه شد. ❌ {error} خطا'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 400
# ==================== پنل مدیر سازمان - مسیر اصلی ====================
@app.route('/org-manager/dashboard')
@login_required
def org_manager_dashboard():
    # اجازه دسترسی به ادمین و مدیر سازمان
    if current_user.role not in ['org_manager', 'admin']:
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    # دریافت نام سازمان
    organization_name = "سازمان مرکزی"
    first_dept = Department.query.first()
    if first_dept:
        organization_name = "سازمان"
    
    # دریافت دوره فعال
    active_period = WorkPeriod.query.filter_by(is_active=True).first()
    period_title = active_period.title if active_period else 'تعریف نشده'
    today_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    return render_template('org_manager/dashboard.html',
                          organization_name=organization_name,
                          period_title=period_title,
                          today_date=today_date)
                          
# ==================== APIهای اضافه شده برای پنل مدیر سازمان ====================

@app.route('/org-manager/api/departments')
@login_required
def org_manager_api_departments():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    departments = Department.query.filter_by(is_active=True).all()
    result = []
    for dept in departments:
        managers = db.session.query(User).join(
            DepartmentManager, DepartmentManager.user_id == User.id
        ).filter(DepartmentManager.department_id == dept.id).all()
        managers_names = [m.get_full_name() for m in managers]
        
        supervisors_count = db.session.query(UnitSupervisor).join(
            Unit, UnitSupervisor.unit_id == Unit.id
        ).filter(Unit.department_id == dept.id).count()
        
        result.append({
            'id': dept.id,
            'name': dept.name,
            'color': dept.color,
            'managers': managers_names,
            'supervisors_count': supervisors_count
        })
    
    return jsonify(result)


@app.route('/org-manager/api/units')
@login_required
def org_manager_api_units():
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    units = Unit.query.filter_by(is_active=True).all()
    result = []
    for unit in units:
        supervisors = db.session.query(User).join(
            UnitSupervisor, UnitSupervisor.user_id == User.id
        ).filter(UnitSupervisor.unit_id == unit.id).all()
        supervisors_names = [s.get_full_name() for s in supervisors]
        
        result.append({
            'id': unit.id,
            'name': unit.name,
            'department_id': unit.department_id,
            'supervisors': supervisors_names
        })
    
    return jsonify(result)
    
def export_personnel_list_to_excel(personnel_data, fields, template):
    """تولید فایل اکسل لیست پرسنل (برای مدیر سازمان) با قالب کامل و فونت Calibri"""
    import tempfile
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "لیست پرسنل"
    
    # تبدیل عدد به فارسی
    def to_persian(num):
        if num is None or num == '':
            return ''
        persian_digits = '۰۱۲۳۴۵۶۷۸۹'
        if isinstance(num, (int, float)):
            return ''.join(persian_digits[int(d)] for d in str(int(num)) if d.isdigit())
        if isinstance(num, str) and num.isdigit():
            return ''.join(persian_digits[int(d)] for d in num)
        # برای شماره تلفن که ممکن است شامل کاراکترهای غیرعددی باشد
        if isinstance(num, str):
            digits = ''.join([d for d in num if d.isdigit()])
            if digits:
                return ''.join(persian_digits[int(d)] for d in digits)
        return str(num)
    
    # رنگ‌ها از قالب
    header_bg = template.header_bg_color.replace('#', '') if template and template.header_bg_color else '2c3e50'
    header_text = template.header_text_color.replace('#', '') if template and template.header_text_color else 'ffffff'
    even_color = template.even_row_color.replace('#', '') if template and template.even_row_color else 'f8f9fa'
    odd_color = template.odd_row_color.replace('#', '') if template and template.odd_row_color else 'ffffff'
    font_name = 'Calibri'  # تغییر به Calibri
    header_font_size = template.header_font_size if template and template.header_font_size else 12
    data_font_size = template.data_font_size if template and template.data_font_size else 11
    border_color = template.border_color.replace('#', '') if template and template.border_color else '000000'
    outer_style = template.outer_border_style if template and template.outer_border_style else 'thick'
    vertical_style = template.vertical_border_style if template and template.vertical_border_style else 'thin'
    horizontal_style = template.horizontal_border_style if template and template.horizontal_border_style else 'thin'
    
    # نگاشت استایل border
    border_map = {
        'thin': Side(border_style='thin', color=border_color),
        'medium': Side(border_style='medium', color=border_color),
        'thick': Side(border_style='thick', color=border_color),
        'double': Side(border_style='double', color=border_color),
        'dashed': Side(border_style='dashed', color=border_color),
        'dotted': Side(border_style='dotted', color=border_color)
    }
    
    outer_side = border_map.get(outer_style, Side(border_style='thin', color=border_color))
    vertical_side = border_map.get(vertical_style, Side(border_style='thin', color=border_color))
    horizontal_side = border_map.get(horizontal_style, Side(border_style='thin', color=border_color))
    
    # ساخت هدرها
    headers = ['ردیف', 'کد ملی', 'نام', 'نام خانوادگی', 'اداره', 'واحد', 'سمت', 'شماره تماس', 'دوره']
    
    # اضافه کردن فیلدهای داینامیک
    for f in fields:
        if hasattr(f, 'title') and not getattr(f, 'is_key', False) and f.title not in ['نام', 'نام خانوادگی']:
            headers.append(f.title)
    
    total_rows = len(personnel_data) + 1
    total_cols = len(headers)
    
    # تنظیم هدر با border کامل
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name=font_name, size=header_font_size, bold=True, color=header_text)
        cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # تنظیم border برای هدر
        top = outer_side
        bottom = outer_side if total_rows == 1 else horizontal_side
        left = outer_side if col == total_cols else vertical_side
        right = outer_side if col == 1 else vertical_side
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)
    
    # نوشتن داده‌ها با border کامل
    for idx, p in enumerate(personnel_data, 1):
        row = idx + 1
        bg_color = even_color if (row % 2 == 0) else odd_color
        is_last_row = (row == total_rows)
        
        # ستون ردیف (عدد فارسی)
        cell = ws.cell(row=row, column=1, value=to_persian(idx))
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            top=outer_side if row == 2 else horizontal_side,
            bottom=outer_side if is_last_row else horizontal_side,
            left=outer_side,
            right=vertical_side
        )
        
        # کد ملی (عدد فارسی)
        cell = ws.cell(row=row, column=2, value=to_persian(p.get('national_code', '')))
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            top=outer_side if row == 2 else horizontal_side,
            bottom=outer_side if is_last_row else horizontal_side,
            left=vertical_side,
            right=vertical_side
        )
        
        # نام
        cell = ws.cell(row=row, column=3, value=p.get('first_name', ''))
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(
            top=outer_side if row == 2 else horizontal_side,
            bottom=outer_side if is_last_row else horizontal_side,
            left=vertical_side,
            right=vertical_side
        )
        
        # نام خانوادگی
        cell = ws.cell(row=row, column=4, value=p.get('last_name', ''))
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(
            top=outer_side if row == 2 else horizontal_side,
            bottom=outer_side if is_last_row else horizontal_side,
            left=vertical_side,
            right=vertical_side
        )
        
        # اداره
        cell = ws.cell(row=row, column=5, value=p.get('department_name', ''))
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(
            top=outer_side if row == 2 else horizontal_side,
            bottom=outer_side if is_last_row else horizontal_side,
            left=vertical_side,
            right=vertical_side
        )
        
        # واحد
        cell = ws.cell(row=row, column=6, value=p.get('unit_name', ''))
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(
            top=outer_side if row == 2 else horizontal_side,
            bottom=outer_side if is_last_row else horizontal_side,
            left=vertical_side,
            right=vertical_side
        )
        
        # سمت
        cell = ws.cell(row=row, column=7, value=p.get('position', ''))
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(
            top=outer_side if row == 2 else horizontal_side,
            bottom=outer_side if is_last_row else horizontal_side,
            left=vertical_side,
            right=vertical_side
        )
        
        # شماره تماس (عدد فارسی)
        phone_val = p.get('phone', '')
        cell = ws.cell(row=row, column=8, value=to_persian(phone_val))
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            top=outer_side if row == 2 else horizontal_side,
            bottom=outer_side if is_last_row else horizontal_side,
            left=vertical_side,
            right=vertical_side
        )
        
        # دوره
        cell = ws.cell(row=row, column=9, value=p.get('period_title', ''))
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            top=outer_side if row == 2 else horizontal_side,
            bottom=outer_side if is_last_row else horizontal_side,
            left=vertical_side,
            right=vertical_side
        )
        
        col = 10
        for f in fields:
            if hasattr(f, 'title') and not getattr(f, 'is_key', False) and f.title not in ['نام', 'نام خانوادگی']:
                value = p.get(f.title, '')
                # اگر فیلد از نوع عددی است، به فارسی تبدیل کن
                if f.field_type in ['number', 'decimal'] and value:
                    try:
                        value = to_persian(float(value))
                    except:
                        pass
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = Font(name=font_name, size=data_font_size)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                is_last_col = (col == total_cols)
                cell.border = Border(
                    top=outer_side if row == 2 else horizontal_side,
                    bottom=outer_side if is_last_row else horizontal_side,
                    left=vertical_side,
                    right=outer_side if is_last_col else vertical_side
                )
                col += 1
    
    # تنظیم عرض ستون‌ها
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    return temp_file.name

@app.route('/org-manager/export-excel')
@login_required
def org_manager_export_excel():
    """خروجی اکسل لیست پرسنل برای مدیر سازمان"""
    if current_user.role not in ['org_manager', 'admin']:
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # دریافت همه ادارات
        departments = Department.query.filter_by(is_active=True).all()
        department_ids = [d.id for d in departments]
        
        # دریافت همه واحدها
        units = Unit.query.filter(Unit.department_id.in_(department_ids), Unit.is_active == True).all()
        unit_ids = [u.id for u in units]
        
        # دریافت پرسنل
        personnel = Personnel.query.filter(Personnel.unit_id.in_(unit_ids), Personnel.is_deleted == False).all()
        
        # دریافت فیلدهای داینامیک
        fields = DynamicField.query.filter_by(is_active=True).order_by(DynamicField.field_order).all()
        
        # دریافت قالب اکسل
        template = ExcelTemplate.query.first()
        if not template:
            template = ExcelTemplate()
        
        personnel_data = []
        for p in personnel:
            unit = Unit.query.get(p.unit_id)
            department = Department.query.get(p.department_id)
            
            # دریافت عنوان دوره
            period_title = ''
            if p.period_id:
                period = WorkPeriod.query.get(p.period_id)
                if period:
                    period_title = f"{period.title}"
            
            item = {
                'id': p.id,
                'national_code': p.national_code,
                'first_name': p.first_name or '',
                'last_name': p.last_name or '',
                'full_name': p.get_full_name(),
                'phone': p.phone or '',
                'position': p.position or '',
                'department_name': department.name if department else '-',
                'unit_name': unit.name if unit else '-',
                'period_title': period_title
            }
            
            # اضافه کردن مقادیر فیلدهای داینامیک
            for v in PersonnelValue.query.filter_by(personnel_id=p.id).all():
                field = DynamicField.query.get(v.field_id)
                if field and not field.is_key and field.title not in ['نام', 'نام خانوادگی']:
                    item[field.title] = v.value_text or v.value_number or v.value_date or ''
            
            personnel_data.append(item)
        
        excel_file = export_personnel_list_to_excel(personnel_data, fields, template)
        today = jdatetime.datetime.now().strftime('%Y%m%d')
        filename = f"لیست_پرسنل_سازمان_{today}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error in org_manager_export_excel: {e}")
        import traceback
        traceback.print_exc()
        flash('خطا در تولید فایل اکسل', 'error')
        return redirect(url_for('org_manager_dashboard'))


# ==================== پنل ادمین ====================
@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    stats = {
        'total_users': User.query.count(),
        'total_departments': Department.query.count(),
        'total_units': Unit.query.count(),
        'total_personnel': Personnel.query.filter_by(is_deleted=False).count(),
        'pending_users': User.query.filter_by(is_approved=False).count()
    }
    return render_template('admin/dashboard.html', stats=stats)


@app.route('/admin/users')
@login_required
def admin_users():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    query = User.query
    
    total_count = query.count()
    active_count = query.filter_by(is_active=True, is_approved=True).count()
    pending_count = query.filter_by(is_approved=False, is_active=True).count()
    inactive_count = query.filter_by(is_active=False).count()
    
    pagination = query.order_by(User.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    # اضافه کردن now برای محاسبه آنلاین
    from datetime import datetime
    now = datetime.now()
    
    return render_template('admin/users.html',
                          users=pagination.items,
                          page=page,
                          per_page=per_page,
                          total_pages=pagination.pages,
                          total_count=total_count,
                          active_count=active_count,
                          pending_count=pending_count,
                          inactive_count=inactive_count,
                          now=now)  # ← این خط مهم است
                          
@app.route('/admin/users/<int:user_id>')
@login_required
def admin_user_get(user_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    user = User.query.get_or_404(user_id)
    return jsonify({
        'id': user.id,
        'national_code': user.national_code,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'phone': user.phone or '',  # ← اضافه کن
        'role': user.role,
        'personnel_code': user.personnel_code
    })
    
    
    
@app.route('/admin/users/create', methods=['POST'])
@login_required
def admin_user_create():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        
        national_code = data.get('national_code', '').strip()
        first_name = data.get('first_name', '').strip()
        last_name = data.get('last_name', '').strip()
        phone = data.get('phone', '').strip()
        role = data.get('role', 'subordinate')
        personnel_code = data.get('personnel_code', '').strip()
        password = data.get('password', '').strip()
        
        if not national_code or not first_name or not last_name:
            return jsonify({'error': 'فیلدهای الزامی را پر کنید'}), 400
        
        if len(national_code) != 10 or not national_code.isdigit():
            return jsonify({'error': 'کد ملی باید 10 رقم باشد'}), 400
        
        if User.query.filter_by(national_code=national_code).first():
            return jsonify({'error': 'این کد ملی قبلاً ثبت شده است'}), 400
        
        if not password:
            password = national_code[-4:]
        
        user = User(
            national_code=national_code,
            username=national_code,
            first_name=first_name,
            last_name=last_name,
            phone=phone if phone else None,
            role=role,
            personnel_code=personnel_code if personnel_code else None,
            is_active=True,
            is_approved=True
        )
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        # ========== ذخیره انتصاب برای کاربر جدید ==========
        unit_id = data.get('assign_unit_id')
        start_date = data.get('assign_start_date')
        
        if unit_id and start_date:
            unit_obj = Unit.query.get(unit_id)
            department_id = unit_obj.department_id if unit_obj else None
            
            # ایجاد پرسنل
            personnel = Personnel(
                national_code=national_code,
                first_name=first_name,
                last_name=last_name,
                phone=phone or '',
                department_id=department_id,
                unit_id=unit_id,
                period_id=None
            )
            db.session.add(personnel)
            db.session.commit()
            
            # ایجاد انتصاب
            new_assignment = PersonnelAssignment(
                personnel_id=personnel.id,
                unit_id=unit_id,
                period_id=None,
                start_date=start_date,
                assignment_type='initial',
                description='انتصاب از طریق مدیریت کاربران',
                created_by=current_user.id,
                is_active=True
            )
            db.session.add(new_assignment)
            db.session.commit()
        
        return jsonify({'success': True, 'message': f'کاربر {first_name} {last_name} اضافه شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
        return jsonify({'error': f'خطا در سرور: {str(e)}'}), 500
        

    
    
@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])  # ← به POST تغییر دهید
@login_required
def admin_user_delete(user_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    user = User.query.get_or_404(user_id)
    
    if user.id == current_user.id:
        return jsonify({'error': 'نمی‌توانید خود را حذف کنید'}), 400
    
    try:
        # حذف روابط مرتبط
        DepartmentManager.query.filter_by(user_id=user_id).delete()
        UnitSupervisor.query.filter_by(user_id=user_id).delete()
        Ticket.query.filter_by(sender_id=user_id).delete()
        Ticket.query.filter_by(receiver_id=user_id).delete()
        ApprovalRequest.query.filter_by(requester_id=user_id).delete()
        ApprovalRequest.query.filter_by(reviewer_id=user_id).delete()
        
        db.session.delete(user)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'کاربر حذف شد'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
def admin_user_reset_password(user_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    user = User.query.get_or_404(user_id)
    new_password = user.national_code[-4:]
    user.set_password(new_password)
    db.session.commit()
    print(f"✅ رمز کاربر {user.get_full_name()} به {new_password} تغییر کرد")
    return jsonify({'success': True, 'message': f'رمز با موفقیت به {new_password} تغییر یافت'})
    

@app.route('/admin/api/check-password/<int:user_id>')
@login_required
def admin_api_check_password(user_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    user = User.query.get_or_404(user_id)
    # تست با رمز پیش‌فرض
    default_pass = user.national_code[-4:]
    check_default = user.check_password(default_pass)
    
    return jsonify({
        'user_id': user.id,
        'name': user.get_full_name(),
        'default_password': default_pass,
        'default_password_works': check_default,
        'password_hash_preview': user.password_hash[:50] + '...'
    })
    
    
    
@app.route('/admin/test-reset-password/<int:user_id>/<new_pass>')
@login_required
def admin_test_reset_password(user_id, new_pass):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    user = User.query.get_or_404(user_id)
    old_hash = user.password_hash
    user.set_password(new_pass)
    db.session.commit()
    new_hash = user.password_hash
    
    return jsonify({
        'success': True,
        'user': user.get_full_name(),
        'new_password': new_pass,
        'old_hash': old_hash[:50] + '...',
        'new_hash': new_hash[:50] + '...',
        'password_changed': old_hash != new_hash
    })
    
@app.route('/admin/users/bulk-delete', methods=['POST'])
@login_required
def admin_users_bulk_delete():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    deleted_count = 0
    errors = []
    
    for uid in data.get('user_ids', []):
        user = User.query.get(uid)
        if user and user.id != current_user.id:
            try:
                # حذف روابط مرتبط
                DepartmentManager.query.filter_by(user_id=uid).delete()
                UnitSupervisor.query.filter_by(user_id=uid).delete()
                Ticket.query.filter_by(sender_id=uid).delete()
                Ticket.query.filter_by(receiver_id=uid).delete()
                ApprovalRequest.query.filter_by(requester_id=uid).delete()
                ApprovalRequest.query.filter_by(reviewer_id=uid).delete()
                
                db.session.delete(user)
                deleted_count += 1
            except Exception as e:
                errors.append(str(e))
    
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'خطا در حذف گروهی: {str(e)}'}), 500
    
    message = f'{deleted_count} کاربر با موفقیت حذف شد'
    if errors:
        message += f' - خطاها: {", ".join(errors[:3])}'
    
    return jsonify({'success': True, 'message': message})

@app.route('/admin/users/import-excel', methods=['POST'])
@login_required
def admin_users_import_excel():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    if 'excel_file' not in request.files:
        return jsonify({'error': 'فایلی ارسال نشده'}), 400
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'error': 'فایلی انتخاب نشده'}), 400
    try:
        import pandas as pd
        df = pd.read_excel(file, dtype=str).fillna('')
        success, error = 0, 0
        for _, row in df.iterrows():
            code = str(row.get('کد ملی', '')).strip().replace('.0', '')
            if len(code) < 10:
                code = code.zfill(10)
            if not code.isdigit() or len(code) != 10:
                error += 1
                continue
            if User.query.filter_by(national_code=code).first():
                error += 1
                continue
            name = str(row.get('نام', '')).strip()
            family = str(row.get('نام خانوادگی', '')).strip()
            if not name or not family:
                error += 1
                continue
            role_text = str(row.get('نقش', 'subordinate')).strip()
            if role_text in ['مدیر کل سیستم', 'admin']:
                role = 'admin'
            elif role_text in ['مدیر سازمان', 'org_manager']:
                role = 'org_manager'
            elif role_text in ['مدیر اداره', 'dept_manager']:
                role = 'dept_manager'
            elif role_text in ['سرپرست واحد', 'unit_supervisor']:
                role = 'unit_supervisor'
            else:
                role = 'subordinate'
            user = User(national_code=code, username=code, first_name=name, last_name=family, role=role, is_active=True, is_approved=True)
            user.set_password(code[-4:])
            db.session.add(user)
            success += 1
        db.session.commit()
        return jsonify({'success': True, 'message': f'✅ {success} کاربر اضافه شد. ❌ {error} خطا'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# ==================== مدیریت ادارات ====================
@app.route('/admin/departments')
@login_required
def admin_departments():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    departments = Department.query.order_by(Department.created_at.desc()).all()
    users = User.query.filter(User.role.in_(['org_manager', 'dept_manager'])).all()
    
    # دریافت واحدها
    units = Unit.query.all()
    
    # دریافت سرپرستان واحدها (به صورت دیکشنری)
    unit_supervisors = {}
    for us in UnitSupervisor.query.all():
        if us.unit_id not in unit_supervisors:
            unit_supervisors[us.unit_id] = []
        unit_supervisors[us.unit_id].append(us.user_id)
    
    # دریافت مدیران ادارات
    dept_managers = {}
    for dm in DepartmentManager.query.all():
        if dm.department_id not in dept_managers:
            dept_managers[dm.department_id] = []
        dept_managers[dm.department_id].append(dm.user_id)
    
    return render_template('admin/departments.html', 
                          departments=departments, 
                          users=users, 
                          dept_managers=dept_managers,
                          units=units,
                          unit_supervisors=unit_supervisors)
                          
@app.route('/admin/departments/create', methods=['POST'])
@login_required
def admin_department_create():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        print("Received data:", data)  # برای دیباگ
        
        if Department.query.filter_by(name=data['name']).first():
            return jsonify({'error': 'این اداره قبلاً ثبت شده است'}), 400
        
        dept = Department(
            name=data['name'],
            color=data.get('color', '#3498db'),
            description=data.get('description', '')
        )
        db.session.add(dept)
        db.session.commit()
        
        for user_id in data.get('manager_ids', []):
            dm = DepartmentManager(department_id=dept.id, user_id=user_id)
            db.session.add(dm)
        db.session.commit()
        invalidate_cache()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
        invalidate_cache()
        return jsonify({'error': str(e)}), 500

@app.route('/admin/departments/<int:dept_id>/edit', methods=['POST'])
@login_required
def admin_department_edit(dept_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    dept = Department.query.get_or_404(dept_id)
    data = request.get_json()
    dept.name = data.get('name', dept.name)
    dept.color = data.get('color', dept.color)
    dept.description = data.get('description', dept.description)
    DepartmentManager.query.filter_by(department_id=dept_id).delete()
    for user_id in data.get('manager_ids', []):
        dm = DepartmentManager(department_id=dept_id, user_id=user_id)
        db.session.add(dm)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/departments/<int:dept_id>/delete', methods=['POST'])
@login_required
def admin_department_delete(dept_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        dept = Department.query.get_or_404(dept_id)
        
        # 1. پیدا کردن همه واحدهای این اداره
        units = Unit.query.filter_by(department_id=dept_id).all()
        unit_ids = [u.id for u in units]
        
        # 2. حذف سرپرستان واحدها
        UnitSupervisor.query.filter(UnitSupervisor.unit_id.in_(unit_ids)).delete(synchronize_session=False)
        
        # 3. حذف پرسنل و مقادیر فیلدهای داینامیک
        for unit_id in unit_ids:
            personnel_list = Personnel.query.filter_by(unit_id=unit_id).all()
            for p in personnel_list:
                PersonnelValue.query.filter_by(personnel_id=p.id).delete()
                PersonnelWorkStatus.query.filter_by(personnel_id=p.id).delete()
            Personnel.query.filter_by(unit_id=unit_id).delete()
        
        # 4. حذف واحدها
        Unit.query.filter_by(department_id=dept_id).delete()
        
        # 5. حذف مدیران اداره
        DepartmentManager.query.filter_by(department_id=dept_id).delete()
        
        # 6. حذف درخواست‌های مرتبط با این اداره
        UnitPersonnelRequest.query.filter(
            UnitPersonnelRequest.unit_id.in_(unit_ids)
        ).delete(synchronize_session=False)
        
        # 7. حالا اداره رو حذف کن
        db.session.delete(dept)
        db.session.commit()
        
        invalidate_cache()
        return jsonify({'success': True, 'message': 'اداره و همه وابستگی‌ها حذف شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ خطا در حذف اداره: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/units')
@login_required
def admin_units():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    departments = Department.query.all()
    units = Unit.query.order_by(Unit.created_at.desc()).all()
    users = User.query.filter_by(role='unit_supervisor').all()
    
    unit_supervisors = {}
    for us in UnitSupervisor.query.all():
        if us.unit_id not in unit_supervisors:
            unit_supervisors[us.unit_id] = []
        unit_supervisors[us.unit_id].append(us.user_id)
    
    for dept in departments:
        dept.supervisors_count = UnitSupervisor.query.join(Unit).filter(Unit.department_id == dept.id).count()
        dept.units = Unit.query.filter_by(department_id=dept.id).all()
    
    # اضافه کردن زمان حال برای محاسبه آنلاین (اگه نیاز باشه)
    from datetime import datetime
    now = datetime.now()
    
    return render_template('admin/units.html', 
                          departments=departments,
                          units=units,
                          users=users,
                          unit_supervisors=unit_supervisors,
                          now=now)

@app.route('/admin/units/create', methods=['POST'])
@login_required
def admin_unit_create():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        
        unit = Unit(
            name=data['name'],
            department_id=data['department_id'],
            description=data.get('description', ''),
            needs_approval=data.get('needs_approval', True)
        )
        db.session.add(unit)
        db.session.commit()
        
        for user_id in data.get('supervisor_ids', []):
            us = UnitSupervisor(unit_id=unit.id, user_id=user_id)
            db.session.add(us)
        db.session.commit()
        
        # دریافت نام سرپرستان
        supervisors = []
        for us in UnitSupervisor.query.filter_by(unit_id=unit.id).all():
            user = User.query.get(us.user_id)
            if user:
                supervisors.append(user.get_full_name())
        
        return jsonify({
            'success': True,
            'unit': {
                'id': unit.id,
                'name': unit.name,
                'description': unit.description,
                'needs_approval': unit.needs_approval,
                'created_at': unit.created_at.strftime('%Y/%m/%d'),
                'supervisors': supervisors
            }
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in admin_unit_create: {e}")
        invalidate_cache()
        return jsonify({'error': str(e)}), 500


@app.route('/admin/units/<int:unit_id>/edit', methods=['POST'])
@login_required
def admin_unit_edit(unit_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    unit = Unit.query.get_or_404(unit_id)
    data = request.get_json()
    unit.name = data.get('name', unit.name)
    unit.department_id = data.get('department_id', unit.department_id)
    unit.description = data.get('description', unit.description)
    unit.needs_approval = data.get('needs_approval', unit.needs_approval)
    UnitSupervisor.query.filter_by(unit_id=unit_id).delete()
    for user_id in data.get('supervisor_ids', []):
        us = UnitSupervisor(unit_id=unit_id, user_id=user_id)
        db.session.add(us)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/units/<int:unit_id>/delete', methods=['POST'])
@login_required
def admin_unit_delete(unit_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        unit = Unit.query.get_or_404(unit_id)
        
        # حذف روابط سرپرستان
        UnitSupervisor.query.filter_by(unit_id=unit_id).delete()
        
        # حذف روابط پرسنل (PersonnelValue)
        personnel_list = Personnel.query.filter_by(unit_id=unit_id).all()
        for p in personnel_list:
            PersonnelValue.query.filter_by(personnel_id=p.id).delete()
        
        # حذف پرسنل
        Personnel.query.filter_by(unit_id=unit_id).delete()
        
        # حذف واحد
        db.session.delete(unit)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'واحد با موفقیت حذف شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in delete unit: {e}")
        return jsonify({'error': f'خطا در حذف واحد: {str(e)}'}), 500

@app.route('/admin/personnel')
@login_required
def admin_personnel():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    personnel = Personnel.query.filter_by(is_deleted=False).all()
    departments = Department.query.all()
    
    # آمار
    stats = {
        'total_departments': Department.query.count(),
        'total_units': Unit.query.count(),
        'total_personnel': len(personnel),
        'total_managers': DepartmentManager.query.count(),
        'total_supervisors': UnitSupervisor.query.count()
    }
    
    # آمار هر اداره
    for dept in departments:
        dept.units_count = Unit.query.filter_by(department_id=dept.id).count()
        dept.personnel_count = Personnel.query.filter_by(department_id=dept.id, is_deleted=False).count()
        dept.managers_count = DepartmentManager.query.filter_by(department_id=dept.id).count()
        dept.units = Unit.query.filter_by(department_id=dept.id).all()
    
    # فیلدهای داینامیک - فیلد کلید را جداگانه نگه می‌داریم
    all_fields = DynamicField.query.filter_by(is_active=True).order_by(DynamicField.field_order).all()
    dynamic_fields = [f for f in all_fields if not f.is_key]  # فیلدهای غیر کلید
    key_field = [f for f in all_fields if f.is_key]  # فیلد کلید (کد ملی)
    
    # داده‌های پرسنل با مقادیر داینامیک
    personnel_data = []
    for p in personnel:
        values = {}
        for v in PersonnelValue.query.filter_by(personnel_id=p.id).all():
            values[v.field_id] = v.value_text or v.value_number or v.value_date or '-'
        personnel_data.append({
            'id': p.id,
            'national_code': p.national_code,
            'department_id': p.department_id,
            'unit_id': p.unit_id,
            'unit_name': p.unit.name if p.unit else '-',
            'dynamic_values': values
        })
    
    all_units = Unit.query.all()
    
    return render_template('admin/personnel.html', 
                          personnel=personnel_data,
                          departments=departments,
                          all_units=all_units,
                          stats=stats,
                          dynamic_fields=dynamic_fields,
                          key_field=key_field)
                          

    
@app.route('/admin/personnel/<int:pid>')
@login_required
def admin_personnel_get(pid):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    p = Personnel.query.get_or_404(pid)
    
    # دریافت فیلدهای داینامیک
    dynamic_fields = DynamicField.query.filter_by(is_active=True).all()
    
    # دریافت مقادیر فیلدهای داینامیک
    dynamic_values = {}
    for v in PersonnelValue.query.filter_by(personnel_id=pid).all():
        field = DynamicField.query.get(v.field_id)
        if field:
            dynamic_values[field.id] = v.value_text or v.value_number or v.value_date or ''
    
    # دریافت period_id از خود پرسنل (نه از PersonnelValue)
    period_id = p.period_id
    
    return jsonify({
        'id': p.id,
        'national_code': p.national_code,
        'department_id': p.department_id,
        'unit_id': p.unit_id,
        'period_id': period_id,  # این مهم است
        'dynamic_values': dynamic_values
    })

@app.route('/admin/personnel/<int:pid>/edit', methods=['POST'])
@login_required
def admin_personnel_edit(pid):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        p = Personnel.query.get_or_404(pid)
        data = request.get_json()
        
        # ========== دریافت و ذخیره first_name و last_name ==========
        first_name = data.get('first_name', '').strip()
        last_name = data.get('last_name', '').strip()
        
        if not first_name:
            name_field = DynamicField.query.filter_by(title='نام', is_active=True).first()
            if name_field:
                first_name = data.get(str(name_field.id), '').strip()
        
        if not last_name:
            family_field = DynamicField.query.filter_by(title='نام خانوادگی', is_active=True).first()
            if family_field:
                last_name = data.get(str(family_field.id), '').strip()
        
        p.first_name = first_name
        p.last_name = last_name
        
        # بروزرسانی سایر فیلدها
        phone = data.get('phone', '').strip()
        if not phone:
            phone_field = DynamicField.query.filter_by(title='شماره تماس', is_active=True).first()
            if phone_field:
                phone = data.get(str(phone_field.id), '').strip()
        p.phone = phone
        
        position = data.get('position', '').strip()
        if not position:
            position_field = DynamicField.query.filter_by(title='سمت', is_active=True).first()
            if position_field:
                position = data.get(str(position_field.id), '').strip()
        p.position = position
        
        # بروزرسانی کد ملی اگر تغییر کرده
        new_national_code = data.get('national_code')
        if new_national_code and new_national_code != p.national_code:
            if not new_national_code.isdigit() or len(new_national_code) != 10:
                return jsonify({'error': 'کد ملی باید 10 رقم باشد'}), 400
            existing = Personnel.query.filter_by(
                national_code=new_national_code, 
                period_id=p.period_id
            ).first()
            if existing and existing.id != pid:
                return jsonify({'error': 'کد ملی تکراری است'}), 400
            p.national_code = new_national_code
        
        p.department_id = data.get('department_id', p.department_id)
        p.unit_id = data.get('unit_id', p.unit_id)
        
        period_id = data.get('period_id')
        if period_id and period_id != '' and period_id != 'null':
            p.period_id = int(period_id)
        else:
            p.period_id = None
        
        db.session.commit()
        
        # حذف مقادیر قبلی فیلدهای داینامیک برای این دوره
        PersonnelValue.query.filter_by(personnel_id=pid, period_id=p.period_id).delete()
        
        # ذخیره مقادیر جدید فیلدهای داینامیک
        fields = DynamicField.query.filter_by(is_active=True).all()
        for field in fields:
            value = data.get(str(field.id))
            if not value and field.title:
                value = data.get(field.title)
            
            if value and str(value).strip():
                pv = PersonnelValue(
                    personnel_id=pid,
                    field_id=field.id,
                    period_id=p.period_id,
                    value_text=value if field.field_type == 'text' else None,
                    value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
                    value_date=value if field.field_type == 'date' else None
                )
        db.session.add(pv)
        db.session.commit()
        invalidate_cache()
        return jsonify({'success': True, 'message': 'پرسنل با موفقیت ویرایش شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ خطا در ویرایش: {str(e)}")
        invalidate_cache()
        return jsonify({'error': str(e)}), 500

# تابع admin_personnel_delete - جایگزین کن
@app.route('/admin/personnel/<int:pid>/delete', methods=['POST'])
@login_required
def admin_personnel_delete(pid):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        p = Personnel.query.get_or_404(pid)
        
        # ⚠️ این دو خط اضافه شود - حذف وضعیت‌های کارکرد مرتبط
        PersonnelWorkStatus.query.filter_by(personnel_id=pid).delete()
        
        # حذف مقادیر فیلدهای داینامیک
        PersonnelValue.query.filter_by(personnel_id=pid).delete()
        
        # حذف کامل پرسنل
        db.session.delete(p)
        db.session.commit()
        invalidate_cache()
        return jsonify({'success': True, 'message': 'پرسنل با موفقیت حذف شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in personnel delete: {e}")
        invalidate_cache()
        return jsonify({'error': str(e)}), 500
        
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in personnel delete: {e}")
        invalidate_cache()
        return jsonify({'error': str(e)}), 500

# تابع admin_personnel_bulk_delete - جایگزین کن
@app.route('/admin/personnel/bulk-delete', methods=['POST'])
@login_required
def admin_personnel_bulk_delete():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    data = request.get_json()
    count = 0
    for pid in data.get('ids', []):
        p = Personnel.query.get(pid)
        if p:
            # ⚠️ این خط اضافه شود
            PersonnelWorkStatus.query.filter_by(personnel_id=pid).delete()
            PersonnelValue.query.filter_by(personnel_id=pid).delete()
            db.session.delete(p)
            count += 1
    db.session.commit()
    return jsonify({'success': True, 'message': f'{count} پرسنل حذف شدند'})

@app.route('/admin/personnel/import-excel', methods=['POST'])
@login_required
def admin_personnel_import_excel():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    if 'excel_file' not in request.files:
        return jsonify({'error': 'فایلی ارسال نشده'}), 400
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'error': 'فایلی انتخاب نشده'}), 400
    
    try:
        import pandas as pd
        df = pd.read_excel(file, dtype=str).fillna('')
        
        dynamic_fields = DynamicField.query.filter_by(is_active=True).all()
        field_map = {f.title: f for f in dynamic_fields}
        
        success, error = 0, 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                code = str(row.get('کد ملی', '')).strip().replace('.0', '')
                if not code.isdigit() or len(code) != 10:
                    error += 1
                    continue
                
                if Personnel.query.filter_by(national_code=code).first():
                    error += 1
                    continue
                
                name = str(row.get('نام', '')).strip()
                family = str(row.get('نام خانوادگی', '')).strip()
                if not name or not family:
                    error += 1
                    continue
                
                dept_name = str(row.get('نام اداره', '')).strip()
                unit_name = str(row.get('نام واحد', '')).strip()
                period_title = str(row.get('دوره', '')).strip()
                phone = str(row.get('شماره تماس', '')).strip()
                position = str(row.get('سمت', '')).strip()
                
                dept = Department.query.filter_by(name=dept_name).first()
                if not dept:
                    error += 1
                    continue
                
                unit = Unit.query.filter_by(name=unit_name, department_id=dept.id).first()
                if not unit:
                    error += 1
                    continue
                
                period_id = None
                if period_title:
                    period = WorkPeriod.query.filter_by(title=period_title).first()
                    if period:
                        period_id = period.id
                
                if not period_id:
                    active_period = WorkPeriod.query.filter_by(is_active=True).first()
                    if active_period:
                        period_id = active_period.id
                
                p = Personnel(
                    national_code=code,
                    first_name=name,
                    last_name=family,
                    phone=phone if phone else None,
                    position=position if position else None,
                    department_id=dept.id,
                    unit_id=unit.id,
                    period_id=period_id
                )
                db.session.add(p)
                db.session.flush()
                
                for field_name, field in field_map.items():
                    if field.is_key or field.title in ['نام', 'نام خانوادگی', 'شماره تماس', 'سمت']:
                        continue
                    value = str(row.get(field.title, '')).strip()
                    if value and value != 'nan':
                        pv = PersonnelValue(
                            personnel_id=p.id,
                            field_id=field.id,
                            period_id=period_id,
                            value_text=value if field.field_type == 'text' else None,
                            value_number=float(value) if field.field_type in ['number', 'decimal'] else None,
                            value_date=value if field.field_type == 'date' else None
                        )
                        db.session.add(pv)
                
                success += 1
            except Exception as e:
                error += 1
                errors.append(str(e))
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'✅ {success} پرسنل اضافه شد. ❌ {error} خطا'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'خطا در پردازش فایل: {str(e)}'}), 400
        
@app.route('/admin/api/dynamic-fields')
@login_required
def admin_api_dynamic_fields():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    fields = DynamicField.query.filter_by(is_active=True).order_by(DynamicField.field_order).all()
    return jsonify([{
        'id': f.id,
        'title': f.title,
        'field_type': f.field_type,
        'is_required': f.is_required,
        'is_key': f.is_key,
        'placeholder': ''
    } for f in fields])
    
# ==================== درخواست‌های افزودن/حذف پرسنل ====================
@app.route('/api/personnel/request-add', methods=['POST'])
@login_required
def api_personnel_request_add():
    if current_user.role not in ['dept_manager', 'unit_supervisor', 'org_manager']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    unit_id = data.get('unit_id')
    
    unit = Unit.query.get(unit_id)
    if not unit:
        return jsonify({'error': 'واحد یافت نشد'}), 404
    
    # جمع‌آوری داده‌های پرسنل
    personnel_data = {
        'first_name': data.get('first_name', ''),
        'last_name': data.get('last_name', ''),
        'phone': data.get('phone', ''),
        'position': data.get('position', ''),
        'hire_date': data.get('hire_date', ''),
        'department_id': data.get('department_id'),
        'unit_id': unit_id
    }
    
    # فیلدهای داینامیک
    fields = DynamicField.query.filter_by(is_active=True).all()
    for field in fields:
        value = data.get(str(field.id))
        if value:
            personnel_data[f'field_{field.id}'] = value
            if field.is_key:
                personnel_data['national_code'] = value
    
    if unit.needs_approval:
        # ایجاد درخواست تایید
        approval = ApprovalRequest(
            request_type='add',
            requester_id=current_user.id,
            data=json.dumps(personnel_data),
            requester_note=data.get('requester_note', ''),
            status='pending'
        )
        db.session.add(approval)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'needs_approval': True, 
            'message': '✅ درخواست افزودن پرسنل ثبت شد. منتظر تایید ادمین باشید.'
        })
    else:
        # ثبت مستقیم
        return redirect(url_for('admin_api_personnel_create'), code=307, data=personnel_data)


@app.route('/api/personnel/request-delete', methods=['POST'])
@login_required
def api_personnel_request_delete():
    if current_user.role not in ['dept_manager', 'unit_supervisor', 'org_manager']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    personnel_id = data.get('personnel_id')
    unit_id = data.get('unit_id')
    
    p = Personnel.query.get(personnel_id)
    if not p:
        return jsonify({'error': 'پرسنل یافت نشد'}), 404
    
    unit = Unit.query.get(unit_id)
    if not unit:
        return jsonify({'error': 'واحد یافت نشد'}), 404
    
    if unit.needs_approval:
        approval = ApprovalRequest(
            request_type='delete',
            requester_id=current_user.id,
            target_personnel_id=personnel_id,
            data=json.dumps({'full_name': p.get_full_name(), 'national_code': p.national_code}),
            requester_note=data.get('requester_note', ''),
            status='pending'
        )
        db.session.add(approval)
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'needs_approval': True, 
            'message': '✅ درخواست حذف پرسنل ثبت شد. منتظر تایید ادمین باشید.'
        })
    else:
        # حذف مستقیم
        p.is_deleted = True
        db.session.commit()
        return jsonify({'success': True, 'message': 'پرسنل با موفقیت حذف شد'})

# ==================== مدیریت دوره‌های کارکرد ====================
@app.route('/admin/periods')
@login_required
def admin_periods():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    periods = WorkPeriod.query.order_by(WorkPeriod.created_at.desc()).all()
    return render_template('admin/periods.html', periods=periods)

@app.route('/admin/periods/create', methods=['POST'])
@login_required
def admin_period_create():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    data = request.get_json()
    period = WorkPeriod(
        title=data['title'], 
        start_date=data['start_date'], 
        end_date=data['end_date'],
        deadline=data.get('deadline', '')  # ← این خط را اضافه کنید
    )
    db.session.add(period)
    db.session.commit()
    invalidate_cache()
    return jsonify({'success': True})

@app.route('/admin/add-period-columns')
@login_required
def add_period_columns():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    import sqlite3
    db_path = os.path.join(os.path.dirname(__file__), 'instance', 'avan_system.db')
    
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("PRAGMA table_info(work_periods)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'deadline' not in columns:
            cursor.execute("ALTER TABLE work_periods ADD COLUMN deadline VARCHAR(20)")
        if 'display_order' not in columns:
            cursor.execute("ALTER TABLE work_periods ADD COLUMN display_order INTEGER DEFAULT 0")
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'ستون‌ها اضافه شدند'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
@app.route('/admin/periods/<int:period_id>/edit', methods=['PUT'])
@login_required
def admin_period_edit(period_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    period = WorkPeriod.query.get_or_404(period_id)
    data = request.get_json()
    period.title = data.get('title', period.title)
    period.start_date = data.get('start_date', period.start_date)
    period.end_date = data.get('end_date', period.end_date)
    period.deadline = data.get('deadline', period.deadline)  # ← این خط را اضافه کنید
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/periods/<int:period_id>/delete', methods=['DELETE'])
@login_required
def admin_period_delete(period_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    period = WorkPeriod.query.get_or_404(period_id)
    db.session.delete(period)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/periods/<int:period_id>/set-active', methods=['POST'])
@login_required
def admin_period_set_active(period_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    WorkPeriod.query.update({WorkPeriod.is_active: False})
    period = WorkPeriod.query.get_or_404(period_id)
    period.is_active = True
    db.session.commit()
    return jsonify({'success': True, 'message': f'دوره {period.title} فعال شد'})

@app.route('/admin/test-periods')
@login_required
def test_periods():
    try:
        periods = WorkPeriod.query.all()
        return jsonify({
            'success': True,
            'count': len(periods),
            'periods': [{'id': p.id, 'title': p.title} for p in periods]
        })
    except Exception as e:
        import traceback
        return jsonify({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500
# ==================== قالب گزارش اکسل ====================
@app.route('/admin/excel-template')
@login_required
def admin_excel_template():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    template = ExcelTemplate.query.first()
    if not template:
        template = ExcelTemplate(name='قالب پیش‌فرض')
        db.session.add(template)
        db.session.commit()
    return render_template('admin/excel_template.html', template=template)

@app.route('/admin/excel-template/save', methods=['POST'])
@login_required
def admin_excel_template_save():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    data = request.get_json()
    template = ExcelTemplate.query.first()
    if not template:
        template = ExcelTemplate()
        db.session.add(template)
    template.name = data.get('name', 'قالب پیش‌فرض')
    template.header_bg_color = data.get('header_bg_color', '#2c3e50')
    template.header_text_color = data.get('header_text_color', '#ffffff')
    template.even_row_color = data.get('even_row_color', '#f8f9fa')
    template.odd_row_color = data.get('odd_row_color', '#ffffff')
    template.outer_border_style = data.get('outer_border_style', 'thick')
    template.vertical_border_style = data.get('vertical_border_style', 'thin')
    template.horizontal_border_style = data.get('horizontal_border_style', 'dotted')
    template.border_color = data.get('border_color', '#000000')
    template.font_name = data.get('font_name', 'B Nazanin')
    template.header_font_size = data.get('header_font_size', 12)
    template.data_font_size = data.get('data_font_size', 11)
    db.session.commit()
    return jsonify({'success': True})

# ==================== درخواست‌های تایید ====================
def _get_data_summary(request_type, data):
    if request_type == 'add':
        return f"افزودن پرسنل: {data.get('first_name', '')} {data.get('last_name', '')}"
    elif request_type == 'edit':
        return f"ویرایش پرسنل: {data.get('first_name', '')} {data.get('last_name', '')}"
    else:
        return f"حذف پرسنل"
        
@app.route('/admin/personnel/create', methods=['POST'])
@login_required
def admin_personnel_create():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        print(f"📥 دریافت داده برای ایجاد پرسنل: {data}")
        
        if not data:
            return jsonify({'error': 'داده ارسال نشده است'}), 400
        
        # دریافت کد ملی
        national_code = data.get('national_code', '').strip()
        if not national_code:
            key_field = DynamicField.query.filter_by(is_key=True, is_active=True).first()
            if key_field:
                national_code = data.get(str(key_field.id), '').strip()
        
        if not national_code:
            return jsonify({'error': 'کد ملی الزامی است'}), 400
        
        if not national_code.isdigit() or len(national_code) != 10:
            return jsonify({'error': 'کد ملی باید 10 رقم باشد'}), 400
        
        # دریافت period_id
        period_id = data.get('period_id')
        if period_id and period_id != '' and period_id != 'null':
            period_id = int(period_id)
        else:
            period_id = None
        
        # ========== دریافت first_name و last_name از دیتای ارسالی ==========
        first_name = data.get('first_name', '').strip()
        last_name = data.get('last_name', '').strip()
        
        # اگر در دیتا نبود، از فیلدهای داینامیک بگیر
        if not first_name:
            name_field = DynamicField.query.filter_by(title='نام', is_active=True).first()
            if name_field:
                first_name = data.get(str(name_field.id), '').strip()
        
        if not last_name:
            family_field = DynamicField.query.filter_by(title='نام خانوادگی', is_active=True).first()
            if family_field:
                last_name = data.get(str(family_field.id), '').strip()
        
        # دریافت سایر فیلدها
        phone = data.get('phone', '').strip()
        position = data.get('position', '').strip()
        
        if not phone:
            phone_field = DynamicField.query.filter_by(title='شماره تماس', is_active=True).first()
            if phone_field:
                phone = data.get(str(phone_field.id), '').strip()
        
        if not position:
            position_field = DynamicField.query.filter_by(title='سمت', is_active=True).first()
            if position_field:
                position = data.get(str(position_field.id), '').strip()
        
        # بررسی تکراری بودن در دوره انتخابی
        if period_id:
            existing = Personnel.query.filter_by(
                national_code=national_code, 
                period_id=period_id, 
                is_deleted=False
            ).first()
            if existing:
                return jsonify({'error': f'کد ملی {national_code} قبلاً در این دوره ثبت شده است'}), 400
        
        # ایجاد پرسنل جدید با first_name و last_name
        personnel = Personnel(
            national_code=national_code,
            first_name=first_name,
            last_name=last_name,
            phone=phone,
            position=position,
            department_id=data.get('department_id'),
            unit_id=data.get('unit_id'),
            period_id=period_id
        )
        db.session.add(personnel)
        db.session.flush()
        
        # ذخیره مقادیر فیلدهای داینامیک
        fields = DynamicField.query.filter_by(is_active=True).all()
        for field in fields:
            value = data.get(str(field.id))
            if not value and field.title:
                value = data.get(field.title)
            
            if value and str(value).strip():
                pv = PersonnelValue(
                    personnel_id=personnel.id,
                    field_id=field.id,
                    period_id=period_id,
                    value_text=value if field.field_type == 'text' else None,
                    value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
                    value_date=value if field.field_type == 'date' else None
                )
                db.session.add(pv)
        
        db.session.commit()
        print(f"✅ پرسنل ایجاد شد: {first_name} {last_name} - کد ملی: {national_code}")
        invalidate_cache()
        return jsonify({'success': True, 'message': 'پرسنل با موفقیت اضافه شد', 'id': personnel.id})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ خطا در ایجاد پرسنل: {str(e)}")
        import traceback
        traceback.print_exc()
        invalidate_cache()
        return jsonify({'error': f'خطا در سرور: {str(e)}'}), 500


# ==================== درخواست‌های تایید ادمین ====================
# اضافه کن به بخش ادمین

@app.route('/admin/api/approvals-data')
@login_required
def admin_api_approvals_data():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    pending = []
    approved = []
    rejected = []
    
    for req in UnitPersonnelRequest.query.order_by(UnitPersonnelRequest.created_at.desc()).all():
        requester = User.query.get(req.requester_id)
        data = json.loads(req.data) if req.data else {}
        unit = Unit.query.get(req.unit_id)
        
        item = {
            'id': req.id,
            'request_type': req.request_type,
            'requester_name': requester.get_full_name() if requester else '-',
            'unit_name': unit.name if unit else '-',
            'national_code': data.get('national_code', ''),
            'full_name': f"{data.get('first_name', '')} {data.get('last_name', '')}".strip(),
            'data': data,
            'created_at': req.created_at.strftime('%Y/%m/%d %H:%M'),
            'admin_note': req.admin_note,
            'requester_note': data.get('requester_note', '')
        }
        
        if req.status == 'pending':
            pending.append(item)
        elif req.status == 'approved':
            approved.append(item)
        else:
            rejected.append(item)
    
    return jsonify({'pending': pending, 'approved': approved, 'rejected': rejected})


@app.route('/admin/approvals/bulk-delete', methods=['POST'])
@login_required
def admin_approvals_bulk_delete():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    ids = data.get('ids', [])
    req_type = data.get('type', 'approved')
    
    count = 0
    for rid in ids:
        req = UnitPersonnelRequest.query.get(rid)
        if req and req.status == req_type:
            db.session.delete(req)
            count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'message': f'{count} درخواست حذف شد'})

@app.route('/admin/approvals')
@login_required
def admin_approvals():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    # دریافت درخواست‌های پرسنل واحدها
    pending_requests = []
    approved_requests = []
    rejected_requests = []
    
    # دریافت فیلدهای داینامیک برای نمایش صحیح
    dynamic_fields = DynamicField.query.filter_by(is_active=True).all()
    key_field = next((f for f in dynamic_fields if f.is_key), None)
    name_field = next((f for f in dynamic_fields if f.title == 'نام'), None)
    family_field = next((f for f in dynamic_fields if f.title == 'نام خانوادگی'), None)
    
    for req in UnitPersonnelRequest.query.order_by(UnitPersonnelRequest.created_at.desc()).all():
        requester = User.query.get(req.requester_id)
        data = json.loads(req.data) if req.data else {}
        unit = Unit.query.get(req.unit_id)
        
        # دریافت نام و نام خانوادگی از دیتا
        first_name = data.get('first_name', '')
        if not first_name and name_field:
            first_name = data.get(name_field.title, '')
        
        last_name = data.get('last_name', '')
        if not last_name and family_field:
            last_name = data.get(family_field.title, '')
        
        national_code = data.get('national_code', '')
        if not national_code and key_field:
            national_code = data.get(key_field.title, '')
        
        item = {
            'id': req.id,
            'request_type': req.request_type,
            'requester_name': requester.get_full_name() if requester else '-',
            'unit_name': unit.name if unit else '-',
            'national_code': national_code,
            'first_name': first_name,
            'last_name': last_name,
            'full_name': f"{first_name} {last_name}".strip(),
            'data': data,
            'created_at': req.created_at.strftime('%Y/%m/%d %H:%M'),
            'requester_note': data.get('requester_note', ''),
            'status': req.status
        }
        
        if req.status == 'pending':
            pending_requests.append(item)
        elif req.status == 'approved':
            approved_requests.append(item)
        else:
            rejected_requests.append(item)
    
    return render_template('admin/approvals.html', 
                          pending_requests=pending_requests,
                          approved_requests=approved_requests,
                          rejected_requests=rejected_requests)


@app.route('/admin/approvals/<int:req_id>/approve', methods=['POST'])
@login_required
def admin_approval_approve(req_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    req = UnitPersonnelRequest.query.get_or_404(req_id)
    data = request.get_json()
    admin_note = data.get('admin_note', '')
    
    if req.request_type == 'add':
        # اضافه کردن پرسنل جدید
        request_data = json.loads(req.data) if req.data else {}
        
        national_code = request_data.get('national_code', '')
        if not national_code:
            return jsonify({'error': 'کد ملی یافت نشد'}), 400
        
        # بررسی تکراری نبودن
        existing = Personnel.query.filter_by(national_code=national_code, is_deleted=False).first()
        if existing:
            return jsonify({'error': 'این کد ملی قبلاً ثبت شده است'}), 400
        
        # ایجاد پرسنل جدید
        personnel = Personnel(
            national_code=national_code,
            department_id=request_data.get('department_id'),
            unit_id=request_data.get('unit_id'),
            first_name=request_data.get('first_name', ''),
            last_name=request_data.get('last_name', ''),
            phone=request_data.get('phone', ''),
            position=request_data.get('position', '')
        )
        db.session.add(personnel)
        db.session.commit()
        
        # ذخیره فیلدهای داینامیک
        fields = DynamicField.query.filter_by(is_active=True).all()
        for field in fields:
            value = request_data.get(field.title)
            if value:
                pv = PersonnelValue(
                    personnel_id=personnel.id,
                    field_id=field.id,
                    value_text=value if field.field_type == 'text' else None,
                    value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
                    value_date=value if field.field_type == 'date' else None
                )
                db.session.add(pv)
        db.session.commit()
        
    elif req.request_type == 'delete':
        # حذف پرسنل
        target_id = req.target_personnel_id
        if target_id:
            p = Personnel.query.get(target_id)
            if p:
                p.is_deleted = True
                db.session.commit()
    
    req.status = 'approved'
    req.admin_note = admin_note
    req.reviewed_at = datetime.now()
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست با موفقیت تایید شد'})


@app.route('/admin/approvals/<int:req_id>/reject', methods=['POST'])
@login_required
def admin_approval_reject(req_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    req = UnitPersonnelRequest.query.get_or_404(req_id)
    data = request.get_json()
    admin_note = data.get('admin_note', '')
    
    req.status = 'rejected'
    req.admin_note = admin_note
    req.reviewed_at = datetime.now()
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست رد شد'})


@app.route('/admin/approvals/<int:req_id>/modify', methods=['POST'])
@login_required
def admin_approval_modify(req_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        req = UnitPersonnelRequest.query.get_or_404(req_id)
        data = request.get_json()
        admin_note = data.get('admin_note', '')
        
        req.admin_note = admin_note
        db.session.commit()
        
        # دریافت اطلاعات پرسنل از درخواست
        request_data = json.loads(req.data) if req.data else {}
        national_code = request_data.get('national_code', '')
        first_name = request_data.get('first_name', '')
        last_name = request_data.get('last_name', '')
        full_name = f"{first_name} {last_name}".strip()
        
        request_type_text = "افزودن پرسنل" if req.request_type == 'add' else "حذف پرسنل"
        
        # ایجاد اعلان برای سرپرست (درخواست‌دهنده)
        notification = Notification(
            user_id=req.requester_id,
            request_id=req.id,
            request_type=req.request_type,
            personnel_national_code=national_code,
            personnel_full_name=full_name,
            title=f"درخواست {request_type_text} نیاز به اصلاح دارد",
            message=f"درخواست {request_type_text} برای کد ملی {national_code or 'نامشخص'} توسط ادمین برگشت داده شد.",
            admin_note=admin_note,
            is_read=False
        )
        db.session.add(notification)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'یادداشت اضافه شد و اعلان ارسال شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in modify: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/notifications')
@login_required
def api_notifications():
    notifications = Notification.query.filter_by(
        user_id=current_user.id, 
        is_read=False
    ).order_by(Notification.created_at.desc()).all()
    
    return jsonify([{
        'id': n.id,
        'request_id': n.request_id,
        'request_type': n.request_type,
        'request_type_text': 'افزودن پرسنل' if n.request_type == 'add' else 'حذف پرسنل',
        'personnel_national_code': n.personnel_national_code,
        'personnel_full_name': n.personnel_full_name,
        'title': n.title,
        'message': n.message,
        'admin_note': n.admin_note,
        'created_at': n.created_at.strftime('%Y/%m/%d %H:%M')
    } for n in notifications])

@app.route('/api/notifications/<int:notif_id>/read', methods=['POST'])
@login_required
def api_notification_read(notif_id):
    notif = Notification.query.get_or_404(notif_id)
    if notif.user_id != current_user.id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    notif.is_read = True
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/notifications/read-all', methods=['POST'])
@login_required
def api_notifications_read_all():
    Notification.query.filter_by(user_id=current_user.id, is_read=False).update({'is_read': True})
    db.session.commit()
    return jsonify({'success': True})

# ==================== مدل وضعیت کارکرد پرسنل (سلسله مراتبی) ====================
class PersonnelWorkStatus(db.Model):
    __tablename__ = 'personnel_work_status'
    
    id = db.Column(db.Integer, primary_key=True)
    personnel_id = db.Column(db.Integer, db.ForeignKey('personnel.id'), nullable=False)
    period_id = db.Column(db.Integer, db.ForeignKey('work_periods.id'), nullable=False)
    
    # وضعیت‌ها: draft, unit_pending, unit_approved, dept_pending, dept_approved, org_pending, org_approved, revision
    status = db.Column(db.String(30), default='draft')
    
    # زمان‌های تایید
    unit_approved_at = db.Column(db.DateTime, nullable=True)
    dept_approved_at = db.Column(db.DateTime, nullable=True)
    org_approved_at = db.Column(db.DateTime, nullable=True)
    
    # تاییدکنندگان
    unit_approver_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    dept_approver_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    org_approver_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    
    # اطلاعات اصلاح
    revision_note = db.Column(db.Text, nullable=True)
    revision_from_role = db.Column(db.String(30), nullable=True)  # dept_manager, org_manager
    revision_from_user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    
    # روابط
    personnel = db.relationship('Personnel', backref='work_statuses')
    period = db.relationship('WorkPeriod', backref='work_statuses')
    unit_approver = db.relationship('User', foreign_keys=[unit_approver_id])
    dept_approver = db.relationship('User', foreign_keys=[dept_approver_id])
    org_approver = db.relationship('User', foreign_keys=[org_approver_id])
    
    __table_args__ = (
        db.UniqueConstraint('personnel_id', 'period_id', name='unique_personnel_period_work'),
    )


# ==================== مدل پیام‌های اصلاح کارکرد (صندوق پیام رفت و برگشتی) ====================
class WorkRevisionMessage(db.Model):
    __tablename__ = 'work_revision_messages'
    
    id = db.Column(db.Integer, primary_key=True)
    work_status_id = db.Column(db.Integer, db.ForeignKey('personnel_work_status.id'), nullable=False)
    personnel_id = db.Column(db.Integer, db.ForeignKey('personnel.id'), nullable=False)
    period_id = db.Column(db.Integer, db.ForeignKey('work_periods.id'), nullable=False)
    
    from_role = db.Column(db.String(30), nullable=False)
    from_user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    to_role = db.Column(db.String(30), nullable=False)
    to_user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    
    message = db.Column(db.Text, nullable=False)
    message_type = db.Column(db.String(30), default='general')  # approve, direct_approve, revision, general
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    # روابط
    work_status = db.relationship('PersonnelWorkStatus', backref='revision_messages')
    personnel = db.relationship('Personnel')
    period = db.relationship('WorkPeriod')
    from_user = db.relationship('User', foreign_keys=[from_user_id])
    to_user = db.relationship('User', foreign_keys=[to_user_id])

# ==================== مدل صندوق پیام (جدا از تیکت) ====================

class Message(db.Model):
    __tablename__ = 'messages'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text, nullable=False)
    sender_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    receiver_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    sender = db.relationship('User', foreign_keys=[sender_id])
    receiver = db.relationship('User', foreign_keys=[receiver_id])


class MessageReply(db.Model):
    __tablename__ = 'message_replies'
    id = db.Column(db.Integer, primary_key=True)
    message_id = db.Column(db.Integer, db.ForeignKey('messages.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    message = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    message_rel = db.relationship('Message', backref='replies')
    user = db.relationship('User')

class Notification(db.Model):
    __tablename__ = 'notifications'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    notification_type = db.Column(db.String(30), default='workflow')
    title = db.Column(db.String(200), nullable=False)
    message = db.Column(db.Text, nullable=False)
    link = db.Column(db.String(500), nullable=True)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    
    # این 5 خط را اضافه کن
    request_id = db.Column(db.Integer, nullable=True)
    request_type = db.Column(db.String(20), nullable=True)
    personnel_national_code = db.Column(db.String(10), nullable=True)
    personnel_full_name = db.Column(db.String(100), nullable=True)
    admin_note = db.Column(db.Text, nullable=True)
    
    user = db.relationship('User', backref='notifications')
    
def apply_approval_request(req):
    """اعمال درخواست تایید شده"""
    data = json.loads(req.data) if req.data else {}
    
    if req.request_type == 'add':
        # ایجاد پرسنل جدید
        national_code = data.get('national_code')
        if national_code:
            # بررسی تکراری نبودن
            existing = Personnel.query.filter_by(national_code=national_code, is_deleted=False).first()
            if existing:
                return
        
        personnel = Personnel(
            national_code=national_code,
            first_name=data.get('first_name', ''),
            last_name=data.get('last_name', ''),
            phone=data.get('phone', ''),
            position=data.get('position', ''),
            hire_date=data.get('hire_date', ''),
            department_id=data.get('department_id'),
            unit_id=data.get('unit_id')
        )
        db.session.add(personnel)
        db.session.commit()
        
        # ذخیره فیلدهای داینامیک
        fields = DynamicField.query.filter_by(is_active=True).all()
        for field in fields:
            value = data.get(f'field_{field.id}')
            if value:
                pv = PersonnelValue(
                    personnel_id=personnel.id,
                    field_id=field.id,
                    value_text=value if field.field_type == 'text' else None,
                    value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
                    value_date=value if field.field_type == 'date' else None
                )
                db.session.add(pv)
        db.session.commit()
        
    elif req.request_type == 'delete' and req.target_personnel_id:
        p = Personnel.query.get(req.target_personnel_id)
        if p:
            p.is_deleted = True
            db.session.commit()

@app.route('/work-messages')
@login_required
def work_messages():
    return render_template('work_messages.html')

@app.route('/api/work/dept-revision', methods=['POST'])
@login_required
def api_work_dept_revision():
    """مدیر اداره: اصلاح کارکرد و ارسال پیام به سرپرست"""
    if current_user.role not in ['dept_manager', 'org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    personnel_id = data.get('personnel_id')
    period_id = data.get('period_id')
    revision_note = data.get('revision_note', '')
    
    # بررسی وجود personnel
    personnel = Personnel.query.get_or_404(personnel_id)
    
    # ایجاد یا دریافت work_status
    work_status = get_personnel_work_status(personnel_id, period_id)
    if not work_status:
        # اگر وجود نداشت، ایجاد کن
        work_status = PersonnelWorkStatus(
            personnel_id=personnel_id,
            period_id=period_id,
            status='revision'
        )
        db.session.add(work_status)
        db.session.commit()
    
    work_status.status = 'revision'
    work_status.revision_note = revision_note
    work_status.revision_from_role = 'dept_manager'
    work_status.revision_from_user_id = current_user.id
    db.session.commit()
    
    # پیدا کردن سرپرست واحد
    unit_supervisor = UnitSupervisor.query.filter_by(unit_id=personnel.unit_id).first()
    
    if unit_supervisor:
        send_revision_message(
            work_status_id=work_status.id,
            from_user=current_user,
            to_role='unit_supervisor',
            to_user_id=unit_supervisor.user_id,
            message_text=revision_note or f"کارکرد دوره {period_id} نیاز به اصلاح دارد"
        )
    
    return jsonify({'success': True, 'message': 'کارکرد برای اصلاح به سرپرست برگشت داده شد'})

@app.route('/api/work/statuses')
@login_required
def api_work_statuses():
    period_id = request.args.get('period_id', type=int)
    if not period_id:
        return jsonify([])
    
    if current_user.role == 'unit_supervisor':
        supervised_units = db.session.query(Unit).join(
            UnitSupervisor, UnitSupervisor.unit_id == Unit.id
        ).filter(UnitSupervisor.user_id == current_user.id).all()
        unit_ids = [u.id for u in supervised_units]
        # فقط پرسنلی که period_id برابر با دوره انتخاب شده دارند
        personnel = Personnel.query.filter(
            Personnel.unit_id.in_(unit_ids), 
            Personnel.is_deleted == False,
            Personnel.period_id == period_id  # ← این خط اضافه شده
        ).all()
        personnel_ids = [p.id for p in personnel]
    else:
        personnel = Personnel.query.filter_by(period_id=period_id, is_deleted=False).all()
        personnel_ids = [p.id for p in personnel]
    
    statuses = PersonnelWorkStatus.query.filter(
        PersonnelWorkStatus.personnel_id.in_(personnel_ids),
        PersonnelWorkStatus.period_id == period_id
    ).all()
    
    return jsonify([{
        'personnel_id': s.personnel_id,
        'status': s.status
    } for s in statuses])
# ==================== تنظیمات سامانه ====================
@app.route('/admin/settings')
@login_required
def admin_settings():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    settings = {
        'base_url': Setting.get('base_url', '10.86.109.219'),
        'port': Setting.get('port', '5000'),
        'backup_hour': int(Setting.get('backup_hour', '23')),
        'session_timeout': int(Setting.get('session_timeout', '30')),
        'install_date': Setting.get('install_date', datetime.now().strftime('%Y/%m/%d'))
    }
    
    # دریافت دوره‌ها برای حذف
    periods = WorkPeriod.query.order_by(WorkPeriod.created_at.desc()).all()
    
    backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    backups = []
    for f in sorted(glob.glob(os.path.join(backup_dir, 'avan_backup_*.db')), reverse=True)[:10]:
        stat = os.stat(f)
        size = round(stat.st_size / 1024, 1)
        name = os.path.basename(f)
        date = datetime.fromtimestamp(stat.st_mtime).strftime('%Y/%m/%d %H:%M')
        backups.append({'name': name, 'size': size, 'date': date})
    
    return render_template('admin/settings.html', settings=settings, backups=backups, periods=periods)


# ==================== تنظیمات شبکه (دامنه/IP) ====================

@app.route('/admin/settings/network', methods=['POST'])
@login_required
def admin_settings_network():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    base_url = data.get('base_url', '').strip()
    port = data.get('port', 5000)
    
    if not base_url:
        return jsonify({'error': 'آدرس دامنه یا IP نمی‌تواند خالی باشد'}), 400
    
    # ذخیره در دیتابیس
    Setting.set('base_url', base_url)
    Setting.set('port', str(port))
    
    # به‌روزرسانی فایل کانفیگ
    config_path = os.path.join(os.path.dirname(__file__), 'avan_config.json')
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
        else:
            config = {}
        
        config['base_url'] = base_url
        config['port'] = port
        
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"خطا در ذخیره کانفیگ: {e}")
    
    return jsonify({'success': True, 'message': f'آدرس به {base_url}:{port} تغییر یافت'})
@app.route('/admin/settings/backup-schedule', methods=['POST'])
@login_required
def admin_settings_backup_schedule():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    data = request.get_json()
    Setting.set('backup_hour', str(data.get('backup_hour', 23)))
    return jsonify({'success': True})

@app.route('/admin/settings/security', methods=['POST'])
@login_required
def admin_settings_security():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    data = request.get_json()
    Setting.set('session_timeout', str(data.get('session_timeout', 30)))
    return jsonify({'success': True})

@app.route('/admin/settings/create-backup', methods=['POST'])
@login_required
def admin_settings_create_backup():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
    os.makedirs(backup_dir, exist_ok=True)
    db_path = os.path.join(os.path.dirname(__file__), 'instance', 'avan_system.db')
    backup_name = f"avan_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    backup_path = os.path.join(backup_dir, backup_name)
    try:
        shutil.copy2(db_path, backup_path)
        return jsonify({'success': True, 'message': f'بکاپ با نام {backup_name} ایجاد شد'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/settings/restore-backup', methods=['POST'])
@login_required
def admin_settings_restore_backup():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    data = request.get_json()
    backup_name = data.get('backup_name')
    backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
    backup_path = os.path.join(backup_dir, backup_name)
    if not os.path.exists(backup_path):
        return jsonify({'error': 'فایل بکاپ یافت نشد'}), 404
    db_path = os.path.join(os.path.dirname(__file__), 'instance', 'avan_system.db')
    try:
        shutil.copy2(backup_path, db_path)
        return jsonify({'success': True, 'message': 'بکاپ با موفقیت بازیابی شد'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



@app.route('/admin/settings/clear-logs', methods=['POST'])
@login_required
def admin_settings_clear_logs():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    logs_dir = os.path.join(os.path.dirname(__file__), 'logs')
    os.makedirs(logs_dir, exist_ok=True)
    log_file = os.path.join(logs_dir, 'app.log')
    if os.path.exists(log_file):
        open(log_file, 'w').close()
    return jsonify({'success': True, 'message': 'لاگ‌های سیستم پاک شدند'})

# ==================== صفحات کاربری ====================
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'change_password':
            old_password = request.form.get('old_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            if not current_user.check_password(old_password):
                flash('رمز عبور فعلی اشتباه است', 'error')
            elif new_password != confirm_password:
                flash('رمز عبور جدید و تکرار آن مطابقت ندارند', 'error')
            elif len(new_password) < 4:
                flash('رمز عبور جدید باید حداقل 4 کاراکتر باشد', 'error')
            else:
                current_user.set_password(new_password)
                db.session.commit()
                flash('رمز عبور با موفقیت تغییر یافت', 'success')
        
        elif action == 'update_profile':
            first_name = request.form.get('first_name')
            last_name = request.form.get('last_name')
            personnel_code = request.form.get('personnel_code')
            
            if first_name:
                current_user.first_name = first_name
            if last_name:
                current_user.last_name = last_name
            if personnel_code:
                current_user.personnel_code = personnel_code
            db.session.commit()
            flash('اطلاعات با موفقیت به‌روزرسانی شد', 'success')
        
        return redirect(url_for('profile'))
    
    is_admin = (current_user.role == 'admin')
    return render_template('profile.html', user=current_user, is_admin=is_admin)

# ==================== پنل مدیر اداره (DEpt Manager) ====================
@app.route('/dept-manager/dashboard')
@login_required
def dept_manager_dashboard():
    if current_user.role != 'dept_manager':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    # دریافت اداره مدیر
    dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
    if not dept_manager:
        flash('شما به هیچ اداره‌ای متصل نیستید', 'error')
        return redirect(url_for('dashboard'))
    
    department = Department.query.get(dept_manager.department_id)
    active_period = WorkPeriod.query.filter_by(is_active=True).first()
    today_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    return render_template('dept_manager/dashboard.html', 
                          department_name=department.name if department else '-',
                          period_title=active_period.title if active_period else 'تعریف نشده',
                          today_date=today_date)


@app.route('/dept-manager/api/all-data')
@login_required
@cached(ttl=120)
def dept_manager_api_all_data():
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # دریافت اداره مدیر
    dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
    if not dept_manager:
        return jsonify({'error': 'اداره یافت نشد'}), 404
    
    department_id = dept_manager.department_id
    
    # دریافت همه واحدهای اداره
    units = Unit.query.filter_by(department_id=department_id, is_active=True).all()
    unit_ids = [u.id for u in units]
    
    # دریافت پرسنل همه واحدهای اداره
    personnel = Personnel.query.filter(Personnel.unit_id.in_(unit_ids), Personnel.is_deleted == False).all()
    
    # دریافت فیلدهای داینامیک
    dynamic_fields = DynamicField.query.filter_by(is_active=True).order_by(DynamicField.field_order).all()
    
    # دریافت همه دوره‌ها
    all_periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    
    personnel_data = []
    for p in personnel:
        unit = Unit.query.get(p.unit_id)
        values = {}
        
        # دریافت مقادیر فیلدها
        for v in PersonnelValue.query.filter_by(personnel_id=p.id).all():
            field = DynamicField.query.get(v.field_id)
            if field:
                values[field.title] = v.value_text or v.value_number or v.value_date or ''
        
        # دریافت عنوان دوره
        period_title = ''
        if p.period_id:
            period = WorkPeriod.query.get(p.period_id)
            if period:
                period_title = f"{period.title} ({period.start_date} - {period.end_date})"
        
        item = {
            'id': p.id,
            'national_code': p.national_code,
            'first_name': p.first_name or '',
            'last_name': p.last_name or '',
            'full_name': p.get_full_name(),
            'phone': p.phone or '',
            'position': p.position or '',
            'unit_id': p.unit_id,
            'unit_name': unit.name if unit else '-',
            'period_title': period_title,
            'period_id': p.period_id
        }
        
        # اضافه کردن فیلدهای داینامیک
        for field in dynamic_fields:
            if not field.is_key:
                item[field.title] = values.get(field.title, '')
        
        personnel_data.append(item)
    
    # محاسبه آمار
    visible_fields = [f for f in dynamic_fields if not f.is_key and f.title not in ['نام', 'نام خانوادگی']]
    total_fields = len(visible_fields)
    total_completed_fields = 0
    completed_personnel = 0
    
    for p in personnel_data:
        filled = 0
        for f in visible_fields:
            val = p.get(f.title, '')
            if val and val != '-':
                filled += 1
        total_completed_fields += filled
        if filled == total_fields and total_fields > 0:
            completed_personnel += 1
    
    total_possible = len(personnel_data) * total_fields if total_fields > 0 else 1
    completion_percent = int((total_completed_fields / total_possible) * 100) if total_possible > 0 else 0
    
    return jsonify({
        'personnel': personnel_data,
        'dynamic_fields': [{'id': f.id, 'title': f.title, 'field_type': f.field_type, 'is_key': f.is_key, 'is_required': f.is_required} for f in dynamic_fields],
        'periods': [{'id': p.id, 'title': p.title, 'start_date': p.start_date, 'end_date': p.end_date} for p in all_periods],
        'units': [{'id': u.id, 'name': u.name} for u in units],
        'stats': {
            'total': len(personnel_data),
            'total_fields': total_fields,
            'completed_personnel': completed_personnel,
            'completion_percent': completion_percent,
            'incomplete': len(personnel_data) - completed_personnel
        }
    })


@app.route('/dept-manager/api/pending-requests')
@login_required
def dept_manager_api_pending_requests():
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # دریافت اداره مدیر
    dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
    if not dept_manager:
        return jsonify({'requests': []})
    
    # دریافت همه واحدهای اداره
    units = Unit.query.filter_by(department_id=dept_manager.department_id, is_active=True).all()
    unit_ids = [u.id for u in units]
    
    # دریافت درخواست‌های افزودن
    requests = UnitPersonnelRequest.query.filter(
        UnitPersonnelRequest.unit_id.in_(unit_ids),
        UnitPersonnelRequest.request_type == 'add',
        UnitPersonnelRequest.status == 'pending'
    ).order_by(UnitPersonnelRequest.created_at.desc()).all()
    
    result = []
    for req in requests:
        data = json.loads(req.data) if req.data else {}
        unit = Unit.query.get(req.unit_id)
        created_at_jalali = jdatetime.datetime.fromgregorian(datetime=req.created_at).strftime('%Y/%m/%d %H:%M')
        
        result.append({
            'id': req.id,
            'request_type': req.request_type,
            'national_code': data.get('national_code', ''),
            'first_name': data.get('first_name', ''),
            'last_name': data.get('last_name', ''),
            'full_name': f"{data.get('first_name', '')} {data.get('last_name', '')}".strip(),
            'unit_name': unit.name if unit else '-',
            'created_at': created_at_jalali
        })
    
    return jsonify({'requests': result})


@app.route('/dept-manager/api/delete-requests')
@login_required
def dept_manager_api_delete_requests():
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # دریافت اداره مدیر
    dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
    if not dept_manager:
        return jsonify({'requests': []})
    
    # دریافت همه واحدهای اداره
    units = Unit.query.filter_by(department_id=dept_manager.department_id, is_active=True).all()
    unit_ids = [u.id for u in units]
    
    # دریافت درخواست‌های حذف
    requests = UnitPersonnelRequest.query.filter(
        UnitPersonnelRequest.unit_id.in_(unit_ids),
        UnitPersonnelRequest.request_type == 'delete',
        UnitPersonnelRequest.status == 'pending'
    ).order_by(UnitPersonnelRequest.created_at.desc()).all()
    
    result = []
    for req in requests:
        data = json.loads(req.data) if req.data else {}
        unit = Unit.query.get(req.unit_id)
        created_at_jalali = jdatetime.datetime.fromgregorian(datetime=req.created_at).strftime('%Y/%m/%d %H:%M')
        
        result.append({
            'id': req.id,
            'request_type': req.request_type,
            'personnel_id': data.get('personnel_id'),
            'national_code': data.get('national_code', ''),
            'first_name': data.get('first_name', ''),
            'last_name': data.get('last_name', ''),
            'full_name': data.get('full_name', ''),
            'unit_name': unit.name if unit else '-',
            'delete_reason': data.get('delete_reason', ''),
            'created_at': created_at_jalali
        })
    
    return jsonify({'requests': result})


@app.route('/dept-manager/api/cancel-request', methods=['POST'])
@login_required
def dept_manager_api_cancel_request():
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    request_id = data.get('request_id')
    
    if not request_id:
        return jsonify({'error': 'شناسه درخواست ارسال نشده است'}), 400
    
    req = UnitPersonnelRequest.query.get(request_id)
    if not req:
        return jsonify({'error': 'درخواست یافت نشد'}), 404
    
    db.session.delete(req)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست لغو شد'})


@app.route('/dept-manager/api/request-add', methods=['POST'])
@login_required
def dept_manager_api_request_add():
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    print("📥 داده دریافتی:", data)  # برای دیباگ
    
    unit_id = data.get('unit_id')
    period_id = data.get('period_id')
    national_code = data.get('national_code')  # ← این خط مهم است
    
    # اعتبارسنجی
    if not unit_id:
        return jsonify({'error': 'واحد انتخاب نشده است'}), 400
    if not period_id:
        return jsonify({'error': 'دوره انتخاب نشده است'}), 400
    if not national_code:
        return jsonify({'error': 'کد ملی الزامی است'}), 400
    
    # اعتبارسنجی کد ملی (10 رقم)
    if not (national_code.isdigit() and len(national_code) == 10):
        return jsonify({'error': 'کد ملی باید 10 رقم باشد'}), 400
    
    unit = Unit.query.get(unit_id)
    if not unit:
        return jsonify({'error': 'واحد یافت نشد'}), 404
    
    # بررسی تکراری نبودن کد ملی در دوره
    existing = Personnel.query.filter_by(
        national_code=national_code, 
        period_id=period_id, 
        is_deleted=False
    ).first()
    if existing:
        return jsonify({'error': 'این کد ملی قبلاً در دوره انتخاب شده ثبت شده است'}), 400
    
    # جمع‌آوری داده‌های پرسنل
    request_data = {
        'department_id': unit.department_id,
        'unit_id': unit_id,
        'national_code': national_code,  # ← حتماً این خط را اضافه کنید
        'requester_note': data.get('note', ''),
        'period_id': period_id
    }
    
    fields = DynamicField.query.filter_by(is_active=True).all()
    for field in fields:
        value = data.get(field.title)
        if value:
            request_data[field.title] = value
            if field.title == 'نام':
                request_data['first_name'] = value
            if field.title == 'نام خانوادگی':
                request_data['last_name'] = value
            if field.title == 'شماره تماس':
                request_data['phone'] = value
            if field.title == 'سمت':
                request_data['position'] = value
    
    # ذخیره درخواست
    approval = UnitPersonnelRequest(
        unit_id=unit_id,
        requester_id=current_user.id,
        request_type='add',
        data=json.dumps(request_data, ensure_ascii=False),
        status='pending'
    )
    db.session.add(approval)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست شما ثبت شد و منتظر تایید ادمین است'})

@app.route('/dept-manager/api/request-delete', methods=['POST'])
@login_required
def dept_manager_api_request_delete():
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    personnel_id = data.get('personnel_id')
    delete_reason = data.get('delete_reason', '')
    
    if not personnel_id:
        return jsonify({'error': 'شناسه پرسنل ارسال نشده است'}), 400
    
    p = Personnel.query.get(personnel_id)
    if not p:
        return jsonify({'error': 'پرسنل یافت نشد'}), 404
    
    # حذف درخواست قبلی
    UnitPersonnelRequest.query.filter_by(
        target_personnel_id=personnel_id,
        request_type='delete',
        status='pending'
    ).delete()
    db.session.commit()
    
    request_data = {
        'personnel_id': p.id,
        'national_code': p.national_code,
        'first_name': p.first_name or '',
        'last_name': p.last_name or '',
        'full_name': p.get_full_name(),
        'phone': p.phone or '',
        'position': p.position or '',
        'delete_reason': delete_reason
    }
    
    approval = UnitPersonnelRequest(
        unit_id=p.unit_id,
        requester_id=current_user.id,
        request_type='delete',
        target_personnel_id=personnel_id,
        data=json.dumps(request_data, ensure_ascii=False),
        status='pending'
    )
    db.session.add(approval)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست حذف پرسنل ثبت شد'})


@app.route('/dept-manager/api/personnel/batch-update', methods=['POST'])
@login_required
def dept_manager_api_personnel_batch_update():
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    field = data.get('field')
    updates = data.get('updates', [])
    
    dynamic_field = DynamicField.query.filter_by(title=field, is_active=True).first()
    
    if not dynamic_field:
        return jsonify({'error': f'فیلد "{field}" یافت نشد'}), 400
    
    count = 0
    for item in updates:
        p = Personnel.query.get(item['id'])
        if p:
            value = item['value']
            
            pv = PersonnelValue.query.filter_by(
                personnel_id=p.id,
                field_id=dynamic_field.id,
                period_id=p.period_id
            ).first()
            
            if pv:
                if dynamic_field.field_type == 'text':
                    pv.value_text = value
                elif dynamic_field.field_type in ['number', 'decimal']:
                    try:
                        pv.value_number = float(value)
                    except:
                        pv.value_text = value
                elif dynamic_field.field_type == 'date':
                    pv.value_date = value
                pv.updated_at = datetime.now()
            else:
                pv = PersonnelValue(
                    personnel_id=p.id,
                    field_id=dynamic_field.id,
                    period_id=p.period_id,
                    value_text=value if dynamic_field.field_type == 'text' else None,
                    value_number=float(value) if dynamic_field.field_type in ['number', 'decimal'] and value else None,
                    value_date=value if dynamic_field.field_type == 'date' else None
                )
                db.session.add(pv)
            count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'count': count, 'message': f'{count} پرسنل با موفقیت به‌روزرسانی شدند'})

@app.route('/dept-manager/export-excel')
@login_required
def dept_manager_export_excel():
    if current_user.role != 'dept_manager':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
    if not dept_manager:
        flash('شما به هیچ اداره‌ای متصل نیستید', 'error')
        return redirect(url_for('dept_manager_dashboard'))
    
    # دریافت همه واحدهای اداره
    units = Unit.query.filter_by(department_id=dept_manager.department_id, is_active=True).all()
    unit_ids = [u.id for u in units]
    
    personnel = Personnel.query.filter(Personnel.unit_id.in_(unit_ids), Personnel.is_deleted == False).all()
    fields = DynamicField.query.filter_by(is_active=True).all()
    
    # دریافت قالب از دیتابیس
    template_obj = ExcelTemplate.query.first()
    if not template_obj:
        template_obj = ExcelTemplate()
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile
    
    # تبدیل عدد به فارسی
    def to_persian(num):
        if num is None or num == '':
            return ''
        persian_digits = '۰۱۲۳۴۵۶۷۸۹'
        if isinstance(num, (int, float)):
            return ''.join(persian_digits[int(d)] for d in str(int(num)) if d.isdigit())
        if isinstance(num, str) and num.isdigit():
            return ''.join(persian_digits[int(d)] for d in num)
        if isinstance(num, str):
            digits = ''.join([d for d in num if d.isdigit()])
            if digits:
                return ''.join(persian_digits[int(d)] for d in digits)
        return str(num)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"پرسنل اداره"
    
    header_bg = template_obj.header_bg_color.replace('#', '') if template_obj.header_bg_color else '2c3e50'
    header_text = template_obj.header_text_color.replace('#', '') if template_obj.header_text_color else 'ffffff'
    even_color = template_obj.even_row_color.replace('#', '') if template_obj.even_row_color else 'f8f9fa'
    odd_color = template_obj.odd_row_color.replace('#', '') if template_obj.odd_row_color else 'ffffff'
    font_name = 'Calibri'  # فونت ثابت Calibri
    header_font_size = template_obj.header_font_size if template_obj.header_font_size else 12
    data_font_size = template_obj.data_font_size if template_obj.data_font_size else 11
    border_color = template_obj.border_color.replace('#', '') if template_obj.border_color else '000000'
    outer_style = template_obj.outer_border_style if template_obj.outer_border_style else 'thick'
    vertical_style = template_obj.vertical_border_style if template_obj.vertical_border_style else 'thin'
    horizontal_style = template_obj.horizontal_border_style if template_obj.horizontal_border_style else 'thin'
    
    border_map = {
        'thin': Side(border_style='thin', color=border_color),
        'medium': Side(border_style='medium', color=border_color),
        'thick': Side(border_style='thick', color=border_color),
        'double': Side(border_style='double', color=border_color),
        'dashed': Side(border_style='dashed', color=border_color),
        'dotted': Side(border_style='dotted', color=border_color)
    }
    
    outer_side = border_map.get(outer_style, Side(border_style='thin', color=border_color))
    vertical_side = border_map.get(vertical_style, Side(border_style='thin', color=border_color))
    horizontal_side = border_map.get(horizontal_style, Side(border_style='thin', color=border_color))
    
    headers = ['ردیف', 'کد ملی', 'نام', 'نام خانوادگی', 'واحد']
    for f in fields:
        if not f.is_key and f.title not in ['نام', 'نام خانوادگی']:
            headers.append(f.title)
    headers.extend(['شماره تماس', 'سمت'])
    
    total_rows = len(personnel) + 1
    total_cols = len(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name=font_name, size=header_font_size, bold=True, color=header_text)
        cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        top = outer_side
        bottom = outer_side if total_rows == 1 else horizontal_side
        left = outer_side if col == total_cols else vertical_side
        right = outer_side if col == 1 else vertical_side
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)
    
    for idx, p in enumerate(personnel, 1):
        row_num = idx + 1
        bg_color = even_color if (row_num % 2 == 0) else odd_color
        is_last_row = (row_num == total_rows)
        
        unit = Unit.query.get(p.unit_id)
        
        # ردیف
        cell = ws.cell(row=row_num, column=1, value=to_persian(idx))
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=outer_side, right=vertical_side)
        
        # کد ملی
        cell = ws.cell(row=row_num, column=2, value=to_persian(p.national_code) if p.national_code else '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        # نام
        cell = ws.cell(row=row_num, column=3, value=p.first_name or '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        # نام خانوادگی
        cell = ws.cell(row=row_num, column=4, value=p.last_name or '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        # واحد
        cell = ws.cell(row=row_num, column=5, value=unit.name if unit else '-')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        col = 6
        for f in fields:
            if not f.is_key and f.title not in ['نام', 'نام خانوادگی']:
                value_num = ''
                pv = PersonnelValue.query.filter_by(personnel_id=p.id, field_id=f.id, period_id=p.period_id).first()
                if pv:
                    raw = pv.value_text or pv.value_number or pv.value_date or ''
                    if pv.value_number is not None:
                        if pv.value_number == int(pv.value_number):
                            value_num = to_persian(int(pv.value_number))
                        else:
                            value_num = to_persian(str(pv.value_number).replace('.', '/'))
                    else:
                        value_num = to_persian(raw)
                cell = ws.cell(row=row_num, column=col, value=value_num)
                cell.font = Font(name=font_name, size=data_font_size)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
                col += 1
        
        # شماره تماس
        cell = ws.cell(row=row_num, column=col, value=to_persian(p.phone) if p.phone else '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        # سمت
        cell = ws.cell(row=row_num, column=col+1, value=p.position or '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=outer_side)
    
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_idx in range(2, len(personnel) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 30)
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    
    today = jdatetime.datetime.now().strftime('%Y%m%d')
    filename = f"پرسنل_اداره_{today}.xlsx"
    
    return send_file(temp_file.name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
    
# ==================== اضافه کردن این دو API به بخش پنل مدیر اداره ====================

@app.route('/dept-manager/api/personnel/<int:pid>')
@login_required
def dept_manager_api_personnel_get(pid):
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
    if not dept_manager:
        return jsonify({'error': 'اداره یافت نشد'}), 404
    
    p = Personnel.query.get_or_404(pid)
    unit = Unit.query.get(p.unit_id)
    if not unit or unit.department_id != dept_manager.department_id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    fields = DynamicField.query.filter_by(is_active=True).all()
    values = {}
    for v in PersonnelValue.query.filter_by(personnel_id=pid).all():
        field = DynamicField.query.get(v.field_id)
        if field:
            values[field.title] = v.value_text or v.value_number or v.value_date or ''
    
    result = {'id': p.id, 'national_code': p.national_code, 'first_name': p.first_name or '', 'last_name': p.last_name or '', 'phone': p.phone or '', 'position': p.position or '', 'period_id': p.period_id}
    for field in fields:
        result[field.title] = values.get(field.title, '')
    return jsonify(result)


@app.route('/dept-manager/api/personnel-history/<int:pid>')
@login_required
def dept_manager_api_personnel_history(pid):
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
    if not dept_manager:
        return jsonify({'error': 'اداره یافت نشد'}), 404
    
    p = Personnel.query.get_or_404(pid)
    unit = Unit.query.get(p.unit_id)
    if not unit or unit.department_id != dept_manager.department_id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    fields = DynamicField.query.filter_by(is_active=True).all()
    periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    
    history = []
    for period in periods:
        period_data = {'period_id': period.id, 'period_title': period.title, 'period_start': period.start_date, 'period_end': period.end_date, 'values': {}}
        for field in fields:
            pv = PersonnelValue.query.filter_by(personnel_id=pid, field_id=field.id, period_id=period.id).first()
            if pv:
                value = pv.value_text or pv.value_number or pv.value_date or '-'
            else:
                pv_default = PersonnelValue.query.filter_by(personnel_id=pid, field_id=field.id, period_id=None).first()
                value = (pv_default.value_text or pv_default.value_number or pv_default.value_date or '-') if pv_default else '-'
            period_data['values'][field.title] = value
        history.append(period_data)
    
    return jsonify({'personnel': {'id': p.id, 'national_code': p.national_code, 'full_name': p.get_full_name()}, 'fields': [{'id': f.id, 'title': f.title, 'field_type': f.field_type, 'is_key': f.is_key} for f in fields], 'history': history})


@app.route('/dept-manager/api/personnel/<int:pid>/edit', methods=['PUT'])
@login_required
def dept_manager_api_personnel_edit(pid):
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
    if not dept_manager:
        return jsonify({'error': 'اداره یافت نشد'}), 404
    
    p = Personnel.query.get_or_404(pid)
    unit = Unit.query.get(p.unit_id)
    if not unit or unit.department_id != dept_manager.department_id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        p.first_name = data.get('first_name', p.first_name)
        p.last_name = data.get('last_name', p.last_name)
        p.phone = data.get('phone', p.phone)
        p.position = data.get('position', p.position)
        period_id = data.get('period_id')
        p.period_id = int(period_id) if period_id and period_id != '' and period_id != 'null' else None
        db.session.commit()
        
        fields = DynamicField.query.filter_by(is_active=True).all()
        for field in fields:
            value = data.get(field.title)
            if value is not None and value != '':
                pv = PersonnelValue.query.filter_by(personnel_id=pid, field_id=field.id, period_id=p.period_id).first()
                if pv:
                    if field.field_type == 'text':
                        pv.value_text = value
                    elif field.field_type in ['number', 'decimal']:
                        try:
                            pv.value_number = float(value)
                        except:
                            pv.value_text = value
                    elif field.field_type == 'date':
                        pv.value_date = value
                    pv.updated_at = datetime.now()
                else:
                    pv = PersonnelValue(
                        personnel_id=pid,
                        field_id=field.id,
                        period_id=p.period_id,
                        value_text=value if field.field_type == 'text' else None,
                        value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
                        value_date=value if field.field_type == 'date' else None
                    )
                    db.session.add(pv)
        db.session.commit()
        return jsonify({'success': True, 'message': 'اطلاعات با موفقیت ذخیره شد'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/dept-manager/api/department-color')
@login_required
def dept_manager_api_department_color():
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
    if not dept_manager:
        return jsonify({'color': '#667eea'})
    department = Department.query.get(dept_manager.department_id)
    return jsonify({'color': department.color if department else '#667eea'})


@app.route('/dept-manager/api/periods')
@login_required
def dept_manager_api_periods():
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    return jsonify([{'id': p.id, 'title': p.title, 'start_date': p.start_date, 'end_date': p.end_date} for p in periods])
    
# ==================== پنل کاربر عادی ====================
@app.route('/subordinate/dashboard')
@login_required
def subordinate_dashboard():
    if current_user.role != 'subordinate':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    today_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    total_forms = 0
    today_forms = 0
    rank = 0
    
    return render_template('subordinate/dashboard.html', 
                          total_forms=total_forms,
                          today_forms=today_forms,
                          rank=rank,
                          today_date=today_date)

@app.route('/subordinate/profile', methods=['GET', 'POST'])
@login_required
def subordinate_profile():
    if current_user.role != 'subordinate':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'change_password':
            old_password = request.form.get('old_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            if not current_user.check_password(old_password):
                flash('رمز عبور فعلی اشتباه است', 'error')
            elif new_password != confirm_password:
                flash('رمز عبور جدید و تکرار آن مطابقت ندارند', 'error')
            elif len(new_password) < 4:
                flash('رمز عبور جدید باید حداقل 4 کاراکتر باشد', 'error')
            else:
                current_user.set_password(new_password)
                db.session.commit()
                flash('رمز عبور با موفقیت تغییر یافت', 'success')
        elif action == 'update_profile':
            first_name = request.form.get('first_name')
            last_name = request.form.get('last_name')
            personnel_code = request.form.get('personnel_code')
            
            if first_name:
                current_user.first_name = first_name
            if last_name:
                current_user.last_name = last_name
            if personnel_code:
                current_user.personnel_code = personnel_code
            db.session.commit()
            flash('اطلاعات با موفقیت به‌روزرسانی شد', 'success')
        
        return redirect(url_for('subordinate_profile'))
    
    return render_template('subordinate/profile.html', user=current_user)


@app.route('/subordinate/upload-avatar', methods=['POST'])
@login_required
def subordinate_upload_avatar():
    if current_user.role != 'subordinate':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    if 'avatar' not in request.files:
        return jsonify({'error': 'فایلی ارسال نشده'}), 400
    
    file = request.files['avatar']
    if file.filename == '':
        return jsonify({'error': 'فایلی انتخاب نشده'}), 400
    
    upload_dir = os.path.join(app.root_path, 'static', 'uploads')
    os.makedirs(upload_dir, exist_ok=True)
    
    filename = f"user_{current_user.id}_{int(datetime.now().timestamp())}.jpg"
    file.save(os.path.join(upload_dir, filename))
    
    if current_user.profile_picture:
        old_file = os.path.join(upload_dir, current_user.profile_picture)
        if os.path.exists(old_file):
            os.remove(old_file)
    
    current_user.profile_picture = filename
    db.session.commit()
    
    return jsonify({'success': True})



@app.route('/unit-supervisor/api/update-personnel-field', methods=['POST'])
@login_required
def unit_supervisor_api_update_personnel_field():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    personnel_id = data.get('personnel_id')
    field_title = data.get('field_title')
    period_id = data.get('period_id')
    value = data.get('value')
    
    p = Personnel.query.get_or_404(personnel_id)
    field = DynamicField.query.filter_by(title=field_title, is_active=True).first()
    
    if not field:
        return jsonify({'error': 'فیلد یافت نشد'}), 404
    
    # ذخیره مقدار در دوره مشخص
    pv = PersonnelValue.query.filter_by(
        personnel_id=personnel_id,
        field_id=field.id,
        period_id=period_id if period_id else None
    ).first()
    
    if pv:
        if field.field_type == 'text':
            pv.value_text = value
        elif field.field_type in ['number', 'decimal']:
            try:
                pv.value_number = float(value)
            except:
                pv.value_text = value
        elif field.field_type == 'date':
            pv.value_date = value
        pv.updated_at = datetime.now()
    else:
        pv = PersonnelValue(
            personnel_id=personnel_id,
            field_id=field.id,
            period_id=period_id if period_id else None,
            value_text=value if field.field_type == 'text' else None,
            value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
            value_date=value if field.field_type == 'date' else None
        )
        db.session.add(pv)
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'مقدار با موفقیت ذخیره شد'})
    
    
# ==================== پنل سرپرست واحد ====================
@app.route('/unit-supervisor/dashboard')
@login_required
def unit_supervisor_dashboard():
    # فقط سرپرست واحد و ادمین دسترسی داشته باشند
    if current_user.role not in ['unit_supervisor', 'admin']:
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    
    if not supervised_units and current_user.role != 'admin':
        flash('شما به عنوان سرپرست به هیچ واحدی متصل نیستید', 'warning')
        return redirect(url_for('dashboard'))
    
    # اگر ادمین است، اولین واحد را بگیر
    if current_user.role == 'admin' and not supervised_units:
        unit = Unit.query.first()
        department = Department.query.get(unit.department_id) if unit else None
        unit_name = unit.name if unit else 'واحد نمونه'
        department_name = department.name if department else 'اداره نمونه'
    else:
        unit = supervised_units[0]
        department = Department.query.get(unit.department_id)
        unit_name = unit.name
        department_name = department.name if department else '-'
    
    active_period = WorkPeriod.query.filter_by(is_active=True).first()
    today_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    return render_template('unit_supervisor/dashboard.html', 
                          unit_id=unit.id,
                          unit_name=unit_name,
                          department_name=department_name,
                          period_title=active_period.title if active_period else 'تعریف نشده',
                          today_date=today_date)

@app.route('/unit-supervisor/api/all-data')
@login_required
@cached(ttl=120)
def unit_supervisor_api_all_data():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    
    if not supervised_units:
        return jsonify({'error': 'واحد یافت نشد'}), 404
    
    unit = supervised_units[0]
    period_id = request.args.get('period_id', type=int)
    
    dynamic_fields = DynamicField.query.filter_by(is_active=True).order_by(DynamicField.field_order).all()
    
    # دریافت پرسنل واحد
    personnel = Personnel.query.filter_by(unit_id=unit.id, is_deleted=False).all()
    
    # دریافت همه دوره‌ها برای نمایش در فیلتر
    all_periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    
    personnel_data = []
    for p in personnel:
        # فیلتر دوره: اگر دوره خاصی انتخاب شده و پرسنل آن دوره را ندارد، رد کن
        if period_id and p.period_id != period_id:
            continue
        
        values = {}
        
        # دریافت مقادیر فیلدها برای دوره انتخاب شده
        if period_id:
            pv_query = PersonnelValue.query.filter_by(personnel_id=p.id, period_id=period_id)
        else:
            pv_query = PersonnelValue.query.filter_by(personnel_id=p.id)
        
        for v in pv_query.all():
            field = DynamicField.query.get(v.field_id)
            if field:
                values[field.title] = v.value_text or v.value_number or v.value_date or ''
        
        # دریافت عنوان دوره برای نمایش
        period_title = ''
        if p.period_id:
            period = WorkPeriod.query.get(p.period_id)
            if period:
                period_title = f"{period.title} ({period.start_date} - {period.end_date})"
        
        item = {
            'id': p.id,
            'national_code': p.national_code,
            'first_name': p.first_name or '',
            'last_name': p.last_name or '',
            'full_name': p.get_full_name(),
            'phone': p.phone or '',
            'position': p.position or '',
            'period_title': period_title,
            'period_id': p.period_id
        }
        
        # اضافه کردن فیلدهای داینامیک
        for field in dynamic_fields:
            if not field.is_key:
                value = values.get(field.title, '')
                if not value and p.period_id:
                    pv_default = PersonnelValue.query.filter_by(
                        personnel_id=p.id, 
                        field_id=field.id, 
                        period_id=p.period_id
                    ).first()
                    if pv_default:
                        value = pv_default.value_text or pv_default.value_number or pv_default.value_date or ''
                item[field.title] = value
        
        personnel_data.append(item)
    
    # محاسبه آمار
    visible_fields = [f for f in dynamic_fields if not f.is_key and f.title not in ['نام', 'نام خانوادگی']]
    total_fields = len(visible_fields)
    total_completed_fields = 0
    completed_personnel = 0
    
    for p in personnel_data:
        filled = 0
        for f in visible_fields:
            val = p.get(f.title, '')
            if val and val != '-':
                filled += 1
        total_completed_fields += filled
        if filled == total_fields and total_fields > 0:
            completed_personnel += 1
    
    total_possible = len(personnel_data) * total_fields if total_fields > 0 else 1
    completion_percent = int((total_completed_fields / total_possible) * 100) if total_possible > 0 else 0
    
    return jsonify({
        'personnel': personnel_data,
        'dynamic_fields': [{'id': f.id, 'title': f.title, 'field_type': f.field_type, 'is_key': f.is_key, 'is_required': f.is_required} for f in dynamic_fields],
        'periods': [{'id': p.id, 'title': p.title, 'start_date': p.start_date, 'end_date': p.end_date} for p in all_periods],
        'stats': {
            'total': len(personnel_data),
            'total_fields': total_fields,
            'completed_personnel': completed_personnel,
            'completion_percent': completion_percent,
            'incomplete': len(personnel_data) - completed_personnel
        }
    })


@app.route('/unit-supervisor/api/periods')
@login_required
def unit_supervisor_api_periods():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    return jsonify([{
        'id': p.id,
        'title': p.title,
        'start_date': p.start_date,
        'end_date': p.end_date,
        'is_active': p.is_active
    } for p in periods])


@app.route('/unit-supervisor/api/personnel-history/<int:pid>')
@login_required
def unit_supervisor_api_personnel_history(pid):
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    p = Personnel.query.get_or_404(pid)
    fields = DynamicField.query.filter_by(is_active=True).all()
    
    periods = WorkPeriod.query.order_by(WorkPeriod.start_date.desc()).all()
    
    history = []
    for period in periods:
        period_data = {
            'period_id': period.id,
            'period_title': period.title,
            'period_start': period.start_date,
            'period_end': period.end_date,
            'values': {}
        }
        
        for field in fields:
            pv = PersonnelValue.query.filter_by(
                personnel_id=pid, 
                field_id=field.id,
                period_id=period.id
            ).first()
            
            if pv:
                value = pv.value_text or pv.value_number or pv.value_date or '-'
            else:
                current_pv = PersonnelValue.query.filter_by(
                    personnel_id=pid, 
                    field_id=field.id,
                    period_id=None
                ).first()
                if current_pv:
                    value = current_pv.value_text or current_pv.value_number or current_pv.value_date or '-'
                else:
                    value = '-'
            
            period_data['values'][field.title] = value
        
        history.append(period_data)
    
    return jsonify({
        'personnel': {
            'id': p.id,
            'national_code': p.national_code,
            'full_name': p.get_full_name()
        },
        'fields': [{'id': f.id, 'title': f.title, 'field_type': f.field_type} for f in fields],
        'history': history
    })


@app.route('/unit-supervisor/api/pending-requests')
@login_required
def unit_supervisor_api_pending_requests():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    
    if not supervised_units:
        return jsonify({'requests': []})
    
    unit = supervised_units[0]
    
    requests = UnitPersonnelRequest.query.filter_by(
        unit_id=unit.id, 
        request_type='add',
        status='pending'
    ).order_by(UnitPersonnelRequest.created_at.desc()).all()
    
    result = []
    for req in requests:
        data = json.loads(req.data) if req.data else {}
        created_at_jalali = jdatetime.datetime.fromgregorian(datetime=req.created_at).strftime('%Y/%m/%d %H:%M')
        
        result.append({
            'id': req.id,
            'request_type': req.request_type,
            'national_code': data.get('national_code', ''),
            'first_name': data.get('first_name', ''),
            'last_name': data.get('last_name', ''),
            'full_name': f"{data.get('first_name', '')} {data.get('last_name', '')}".strip(),
            'created_at': created_at_jalali
        })
    
    return jsonify({'requests': result})


@app.route('/unit-supervisor/api/delete-requests')
@login_required
def unit_supervisor_api_delete_requests():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    
    if not supervised_units:
        return jsonify({'requests': []})
    
    unit = supervised_units[0]
    
    requests = UnitPersonnelRequest.query.filter_by(
        unit_id=unit.id, 
        request_type='delete',
        status='pending'
    ).order_by(UnitPersonnelRequest.created_at.desc()).all()
    
    result = []
    for req in requests:
        data = json.loads(req.data) if req.data else {}
        created_at_jalali = jdatetime.datetime.fromgregorian(datetime=req.created_at).strftime('%Y/%m/%d %H:%M')
        
        result.append({
            'id': req.id,
            'request_type': req.request_type,
            'personnel_id': data.get('personnel_id'),
            'national_code': data.get('national_code', ''),
            'first_name': data.get('first_name', ''),
            'last_name': data.get('last_name', ''),
            'full_name': data.get('full_name', ''),
            'delete_reason': data.get('delete_reason', ''),
            'created_at': created_at_jalali
        })
    
    return jsonify({'requests': result})


# ==================== اصلاح 1: تابع unit_supervisor_api_request_add ====================
# این تابع را در بخش پنل سرپرست واحد پیدا کنید و جایگزین کنید

@app.route('/unit-supervisor/api/request-add', methods=['POST'])
@login_required
def unit_supervisor_api_request_add():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    
    if not supervised_units:
        return jsonify({'error': 'شما به هیچ واحدی متصل نیستید'}), 400
    
    unit = supervised_units[0]
    data = request.get_json()
    
    # دریافت داده‌ها - IMPORTANT: period_id must be captured
    period_id = data.get('period_id')
    national_code = data.get('national_code')
    
    # اعتبارسنجی
    if not period_id:
        return jsonify({'error': 'دوره انتخاب نشده است'}), 400
    if not national_code:
        return jsonify({'error': 'کد ملی الزامی است'}), 400
    
    # اعتبارسنجی کد ملی (10 رقم)
    if not (national_code.isdigit() and len(national_code) == 10):
        return jsonify({'error': 'کد ملی باید 10 رقم باشد'}), 400
    
    # بررسی تکراری نبودن کد ملی در دوره
    existing = Personnel.query.filter_by(
        national_code=national_code, 
        period_id=period_id, 
        is_deleted=False
    ).first()
    if existing:
        return jsonify({'error': 'این کد ملی قبلاً در دوره انتخاب شده ثبت شده است'}), 400
    
    # جمع‌آوری داده‌های پرسنل - حتما period_id را ذخیره کن
    request_data = {
        'department_id': unit.department_id,
        'unit_id': unit.id,
        'national_code': national_code,
        'period_id': period_id,  # ← این خیلی مهم است
        'requester_note': data.get('note', '')
    }
    
    # اضافه کردن first_name و last_name از فیلدهای داینامیک
    fields = DynamicField.query.filter_by(is_active=True).all()
    for field in fields:
        value = data.get(field.title)
        if value:
            request_data[field.title] = value
            if field.title == 'نام':
                request_data['first_name'] = value
            if field.title == 'نام خانوادگی':
                request_data['last_name'] = value
            if field.title == 'شماره تماس':
                request_data['phone'] = value
            if field.title == 'سمت':
                request_data['position'] = value
    
    # ذخیره درخواست
    approval = UnitPersonnelRequest(
        unit_id=unit.id,
        requester_id=current_user.id,
        request_type='add',
        data=json.dumps(request_data, ensure_ascii=False),
        status='pending'
    )
    db.session.add(approval)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'درخواست شما ثبت شد و منتظر تایید ادمین است'})


@app.route('/unit-supervisor/api/request-delete', methods=['POST'])
@login_required
def unit_supervisor_api_request_delete():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'داده ارسال نشده است'}), 400
            
        personnel_id = data.get('personnel_id')
        delete_reason = data.get('delete_reason', '')
        
        if not personnel_id:
            return jsonify({'error': 'شناسه پرسنل ارسال نشده است'}), 400
        
        p = Personnel.query.get(personnel_id)
        if not p:
            return jsonify({'error': 'پرسنل یافت نشد'}), 404
        
        supervised_units = db.session.query(Unit).join(
            UnitSupervisor, UnitSupervisor.unit_id == Unit.id
        ).filter(UnitSupervisor.user_id == current_user.id).all()
        
        if not supervised_units:
            return jsonify({'error': 'واحد یافت نشد'}), 404
        
        unit = supervised_units[0]
        
        # حذف درخواست قبلی
        UnitPersonnelRequest.query.filter_by(
            unit_id=unit.id,
            target_personnel_id=personnel_id,
            request_type='delete',
            status='pending'
        ).delete()
        db.session.commit()
        
        request_data = {
            'personnel_id': p.id,
            'national_code': p.national_code,
            'first_name': p.first_name or '',
            'last_name': p.last_name or '',
            'full_name': p.get_full_name(),
            'phone': p.phone or '',
            'position': p.position or '',
            'delete_reason': delete_reason
        }
        
        approval = UnitPersonnelRequest(
            unit_id=unit.id,
            requester_id=current_user.id,
            request_type='delete',
            target_personnel_id=personnel_id,
            data=json.dumps(request_data, ensure_ascii=False),
            status='pending'
        )
        db.session.add(approval)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'درخواست حذف پرسنل ثبت شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
        return jsonify({'error': f'خطا: {str(e)}'}), 500


@app.route('/unit-supervisor/api/cancel-request', methods=['POST'])
@login_required
def unit_supervisor_api_cancel_request():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        request_id = data.get('request_id')
        
        if not request_id:
            return jsonify({'error': 'شناسه درخواست ارسال نشده است'}), 400
        
        supervised_units = db.session.query(Unit).join(
            UnitSupervisor, UnitSupervisor.unit_id == Unit.id
        ).filter(UnitSupervisor.user_id == current_user.id).all()
        
        if not supervised_units:
            return jsonify({'error': 'واحد یافت نشد'}), 404
        
        unit = supervised_units[0]
        
        req = UnitPersonnelRequest.query.filter_by(
            id=request_id,
            unit_id=unit.id,
            status='pending'
        ).first()
        
        if not req:
            return jsonify({'error': 'درخواست یافت نشد'}), 404
        
        db.session.delete(req)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'درخواست با موفقیت لغو شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in cancel-request: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/unit-supervisor/api/personnel/<int:pid>')
@login_required
def unit_supervisor_api_personnel_get(pid):
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    p = Personnel.query.get_or_404(pid)
    fields = DynamicField.query.filter_by(is_active=True).all()
    
    values = {}
    for v in PersonnelValue.query.filter_by(personnel_id=pid).all():
        field = DynamicField.query.get(v.field_id)
        if field:
            values[field.title] = v.value_text or v.value_number or v.value_date or ''
    
    result = {
        'id': p.id,
        'national_code': p.national_code,
        'first_name': p.first_name or '',
        'last_name': p.last_name or '',
        'phone': p.phone or '',
        'position': p.position or '',
        'period_id': p.period_id  # ← این خط را اضافه کن
    }
    
    for field in fields:
        result[field.title] = values.get(field.title, '')
    
    return jsonify(result)

@app.route('/unit-supervisor/api/personnel/<int:pid>/edit', methods=['PUT'])
@login_required
def unit_supervisor_api_personnel_edit(pid):
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        p = Personnel.query.get_or_404(pid)
        data = request.get_json()
        
        print(f"📝 دریافت داده برای ویرایش پرسنل {pid}: {data}")
        
        # به‌روزرسانی فیلدهای پایه
        p.first_name = data.get('first_name', p.first_name)
        p.last_name = data.get('last_name', p.last_name)
        p.phone = data.get('phone', p.phone)
        p.position = data.get('position', p.position)
        
        # به‌روزرسانی period_id
        period_id = data.get('period_id')
        if period_id and period_id != '' and period_id != 'null':
            p.period_id = int(period_id)
        else:
            p.period_id = None
        
        db.session.commit()
        
        # ❌ این خط را حذف کنید - قبلاً همه مقادیر را پاک می‌کرد!
        # PersonnelValue.query.filter_by(personnel_id=pid, period_id=p.period_id).delete()
        
        # ✅ به جای آن، فقط مقادیر موجود را به‌روزرسانی یا اضافه کنید
        fields = DynamicField.query.filter_by(is_active=True).all()
        for field in fields:
            value = data.get(field.title)
            if value is not None and value != '':
                # بررسی وجود مقدار قبلی
                pv = PersonnelValue.query.filter_by(
                    personnel_id=pid,
                    field_id=field.id,
                    period_id=p.period_id
                ).first()
                
                if pv:
                    # به‌روزرسانی مقدار موجود
                    if field.field_type == 'text':
                        pv.value_text = value
                    elif field.field_type in ['number', 'decimal']:
                        try:
                            pv.value_number = float(value)
                        except:
                            pv.value_text = value
                    elif field.field_type == 'date':
                        pv.value_date = value
                    pv.updated_at = datetime.now()
                else:
                    # ایجاد مقدار جدید
                    pv = PersonnelValue(
                        personnel_id=pid,
                        field_id=field.id,
                        period_id=p.period_id,
                        value_text=value if field.field_type == 'text' else None,
                        value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
                        value_date=value if field.field_type == 'date' else None
                    )
                    db.session.add(pv)
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'اطلاعات ذخیره شد', 'period_id': p.period_id})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ خطا در ویرایش: {e}")
        return jsonify({'error': f'خطا: {str(e)}'}), 500
        
@app.route('/unit-supervisor/api/personnel/batch-update', methods=['POST'])
@login_required
def unit_supervisor_api_personnel_batch_update():
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    field = data.get('field')  # نام فیلد داینامیک
    updates = data.get('updates', [])
    
    # پیدا کردن فیلد داینامیک
    dynamic_field = DynamicField.query.filter_by(title=field, is_active=True).first()
    
    if not dynamic_field:
        return jsonify({'error': f'فیلد "{field}" یافت نشد'}), 400
    
    count = 0
    for item in updates:
        p = Personnel.query.get(item['id'])
        if p:
            value = item['value']
            
            # ذخیره مقدار در PersonnelValue
            pv = PersonnelValue.query.filter_by(
                personnel_id=p.id,
                field_id=dynamic_field.id,
                period_id=p.period_id
            ).first()
            
            if pv:
                # به‌روزرسانی مقدار موجود
                if dynamic_field.field_type == 'text':
                    pv.value_text = value
                elif dynamic_field.field_type in ['number', 'decimal']:
                    try:
                        pv.value_number = float(value)
                    except:
                        pv.value_text = value
                elif dynamic_field.field_type == 'date':
                    pv.value_date = value
                pv.updated_at = datetime.now()
            else:
                # ایجاد مقدار جدید
                pv = PersonnelValue(
                    personnel_id=p.id,
                    field_id=dynamic_field.id,
                    period_id=p.period_id,
                    value_text=value if dynamic_field.field_type == 'text' else None,
                    value_number=float(value) if dynamic_field.field_type in ['number', 'decimal'] and value else None,
                    value_date=value if dynamic_field.field_type == 'date' else None
                )
                db.session.add(pv)
            count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'count': count, 'message': f'{count} پرسنل با موفقیت به‌روزرسانی شدند'})


@app.route('/unit-supervisor/export-excel')
@login_required
def unit_supervisor_export_excel():
    if current_user.role != 'unit_supervisor':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    
    if not supervised_units:
        flash('واحد یافت نشد', 'error')
        return redirect(url_for('dashboard'))
    
    unit = supervised_units[0]
    personnel = Personnel.query.filter_by(unit_id=unit.id, is_deleted=False).all()
    fields = DynamicField.query.filter_by(is_active=True).all()
    
    # دریافت قالب از دیتابیس
    template_obj = ExcelTemplate.query.first()
    if not template_obj:
        template_obj = ExcelTemplate()
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile
    
    # تبدیل عدد به فارسی
    def to_persian(num):
        if num is None or num == '':
            return ''
        persian_digits = '۰۱۲۳۴۵۶۷۸۹'
        if isinstance(num, (int, float)):
            return ''.join(persian_digits[int(d)] for d in str(int(num)) if d.isdigit())
        if isinstance(num, str) and num.isdigit():
            return ''.join(persian_digits[int(d)] for d in num)
        if isinstance(num, str):
            digits = ''.join([d for d in num if d.isdigit()])
            if digits:
                return ''.join(persian_digits[int(d)] for d in digits)
        return str(num)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"پرسنل {unit.name}"
    
    header_bg = template_obj.header_bg_color.replace('#', '') if template_obj.header_bg_color else '2c3e50'
    header_text = template_obj.header_text_color.replace('#', '') if template_obj.header_text_color else 'ffffff'
    even_color = template_obj.even_row_color.replace('#', '') if template_obj.even_row_color else 'f8f9fa'
    odd_color = template_obj.odd_row_color.replace('#', '') if template_obj.odd_row_color else 'ffffff'
    font_name = 'Calibri'  # فونت ثابت Calibri
    header_font_size = template_obj.header_font_size if template_obj.header_font_size else 12
    data_font_size = template_obj.data_font_size if template_obj.data_font_size else 11
    border_color = template_obj.border_color.replace('#', '') if template_obj.border_color else '000000'
    outer_style = template_obj.outer_border_style if template_obj.outer_border_style else 'thick'
    vertical_style = template_obj.vertical_border_style if template_obj.vertical_border_style else 'thin'
    horizontal_style = template_obj.horizontal_border_style if template_obj.horizontal_border_style else 'thin'
    
    border_map = {
        'thin': Side(border_style='thin', color=border_color),
        'medium': Side(border_style='medium', color=border_color),
        'thick': Side(border_style='thick', color=border_color),
        'double': Side(border_style='double', color=border_color),
        'dashed': Side(border_style='dashed', color=border_color),
        'dotted': Side(border_style='dotted', color=border_color)
    }
    
    outer_side = border_map.get(outer_style, Side(border_style='thin', color=border_color))
    vertical_side = border_map.get(vertical_style, Side(border_style='thin', color=border_color))
    horizontal_side = border_map.get(horizontal_style, Side(border_style='thin', color=border_color))
    
    headers = ['ردیف', 'کد ملی', 'نام', 'نام خانوادگی']
    for f in fields:
        if not f.is_key and f.title not in ['نام', 'نام خانوادگی']:
            headers.append(f.title)
    headers.extend(['شماره تماس', 'سمت'])
    
    total_rows = len(personnel) + 1
    total_cols = len(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name=font_name, size=header_font_size, bold=True, color=header_text)
        cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        top = outer_side
        bottom = outer_side if total_rows == 1 else horizontal_side
        left = outer_side if col == total_cols else vertical_side
        right = outer_side if col == 1 else vertical_side
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)
    
    for idx, p in enumerate(personnel, 1):
        row_num = idx + 1
        bg_color = even_color if (row_num % 2 == 0) else odd_color
        is_last_row = (row_num == total_rows)
        
        # ردیف
        cell = ws.cell(row=row_num, column=1, value=to_persian(idx))
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=outer_side, right=vertical_side)
        
        # کد ملی
        cell = ws.cell(row=row_num, column=2, value=to_persian(p.national_code) if p.national_code else '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        # نام
        cell = ws.cell(row=row_num, column=3, value=p.first_name or '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        # نام خانوادگی
        cell = ws.cell(row=row_num, column=4, value=p.last_name or '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        col = 5
        for f in fields:
            if not f.is_key and f.title not in ['نام', 'نام خانوادگی']:
                value_num = ''
                pv = PersonnelValue.query.filter_by(personnel_id=p.id, field_id=f.id, period_id=p.period_id).first()
                if pv:
                    raw = pv.value_text or pv.value_number or pv.value_date or ''
                    if pv.value_number is not None:
                        if pv.value_number == int(pv.value_number):
                            value_num = to_persian(int(pv.value_number))
                        else:
                            value_num = to_persian(str(pv.value_number).replace('.', '/'))
                    else:
                        value_num = to_persian(raw)
                cell = ws.cell(row=row_num, column=col, value=value_num)
                cell.font = Font(name=font_name, size=data_font_size)
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
                col += 1
        
        # شماره تماس
        cell = ws.cell(row=row_num, column=col, value=to_persian(p.phone) if p.phone else '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=vertical_side)
        
        # سمت
        cell = ws.cell(row=row_num, column=col+1, value=p.position or '')
        cell.font = Font(name=font_name, size=data_font_size)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = Border(top=outer_side if row_num == 2 else horizontal_side, bottom=outer_side if is_last_row else horizontal_side, left=vertical_side, right=outer_side)
    
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row_idx in range(2, len(personnel) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 3, 30)
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    
    today = jdatetime.datetime.now().strftime('%Y%m%d')
    filename = f"پرسنل_{unit.name}_{today}.xlsx"
    
    return send_file(temp_file.name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)
# ==================== اضافه کردن دوره کارکرد به فرم پرسنل ====================
@app.route('/admin/api/periods')
@login_required
# @cached(ttl=300)  # ← این خط را کامنت کنید
def admin_api_periods():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        periods = WorkPeriod.query.order_by(WorkPeriod.display_order).all()
        
        result = []
        for p in periods:
            created_at_jalali = ''
            if p.created_at:
                created_at_jalali = jdatetime.datetime.fromgregorian(datetime=p.created_at).strftime('%Y/%m/%d')
            
            result.append({
                'id': p.id,
                'title': p.title,
                'start_date': p.start_date,
                'end_date': p.end_date,
                'deadline': p.deadline or '',
                'is_active': p.is_active,
                'display_order': p.display_order,
                'created_at': p.created_at.isoformat() if p.created_at else None,
                'created_at_jalali': created_at_jalali
            })
        return jsonify(result)
    except Exception as e:
        print(f"❌ خطا: {e}")
        return jsonify({'error': str(e)}), 500

# ==================== پیش‌نمایش پنل‌ها برای ادمین ====================
@app.route('/admin/org-manager-preview')
@login_required
def admin_org_manager_preview():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    departments = Department.query.all()
    departments_data = []
    for dept in departments:
        managers = db.session.query(User).join(
            DepartmentManager, DepartmentManager.user_id == User.id
        ).filter(DepartmentManager.department_id == dept.id).all()
        personnel_list = Personnel.query.filter_by(department_id=dept.id, is_deleted=False).all()
        personnel_count = len(personnel_list)
        completed = sum(1 for p in personnel_list if p.first_name and p.last_name and p.phone)
        completion_percent = int((completed / personnel_count) * 100) if personnel_count > 0 else 0
        
        if completion_percent >= 80:
            status = 'عالی'
            status_color = '#10b981'
        elif completion_percent >= 50:
            status = 'متوسط'
            status_color = '#f59e0b'
        else:
            status = 'ضعیف'
            status_color = '#ef4444'
        
        departments_data.append({
            'id': dept.id,
            'name': dept.name,
            'color': dept.color,
            'managers': [m.get_full_name() for m in managers],
            'personnel_count': personnel_count,
            'completion_percent': completion_percent,
            'status': status,
            'status_color': status_color,
            'incomplete_count': personnel_count - completed
        })
    
    active_period = WorkPeriod.query.filter_by(is_active=True).first()
    period_title = active_period.title if active_period else 'تعریف نشده'
    today_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    return render_template('admin/org_manager_preview.html', 
                          departments=departments_data,
                          period_title=period_title,
                          today_date=today_date)

@app.route('/admin/dept-manager-preview')
@login_required
def admin_dept_manager_preview():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    departments = Department.query.all()
    dept_id = request.args.get('dept_id', type=int)
    if dept_id:
        selected_dept = Department.query.get(dept_id)
    else:
        selected_dept = departments[0] if departments else None
    
    units_data = []
    if selected_dept:
        units = Unit.query.filter_by(department_id=selected_dept.id, is_active=True).all()
        for unit in units:
            personnel_list = Personnel.query.filter_by(unit_id=unit.id, is_deleted=False).all()
            total = len(personnel_list)
            completed = sum(1 for p in personnel_list if p.first_name and p.last_name and p.phone)
            completion_percent = int((completed / total) * 100) if total > 0 else 0
            
            units_data.append({
                'id': unit.id,
                'name': unit.name,
                'total_personnel': total,
                'completion_percent': completion_percent,
                'incomplete_count': total - completed
            })
    
    active_period = WorkPeriod.query.filter_by(is_active=True).first()
    period_title = active_period.title if active_period else 'تعریف نشده'
    today_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    return render_template('admin/dept_manager_preview.html', 
                          departments=departments,
                          selected_dept=selected_dept,
                          units=units_data,
                          period_title=period_title,
                          today_date=today_date)

@app.route('/admin/unit-supervisor-preview')
@login_required
def admin_unit_supervisor_preview():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    units = Unit.query.filter_by(is_active=True).all()
    units_data = []
    for unit in units:
        dept = Department.query.get(unit.department_id)
        personnel_list = Personnel.query.filter_by(unit_id=unit.id, is_deleted=False).all()
        total = len(personnel_list)
        completed = sum(1 for p in personnel_list if p.first_name and p.last_name and p.phone)
        completion_percent = int((completed / total) * 100) if total > 0 else 0
        
        units_data.append({
            'id': unit.id,
            'name': unit.name,
            'department_name': dept.name if dept else '-',
            'total_personnel': total,
            'completion_percent': completion_percent,
            'incomplete_count': total - completed
        })
    
    active_period = WorkPeriod.query.filter_by(is_active=True).first()
    period_title = active_period.title if active_period else 'تعریف نشده'
    today_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    return render_template('admin/unit_supervisor_preview.html', 
                          units=units_data,
                          period_title=period_title,
                          today_date=today_date)

# ==================== مدیریت فیلدهای داینامیک ====================
@app.route('/admin/fields')
@login_required
def admin_fields():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    fields = DynamicField.query.order_by(DynamicField.field_order).all()
    return render_template('admin/fields.html', fields=fields)


@app.route('/admin/fields/create', methods=['POST'])
@login_required
def admin_field_create():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    max_order = db.session.query(db.func.max(DynamicField.field_order)).scalar() or 0
    
    field = DynamicField(
        title=data['title'],
        field_type=data['field_type'],
        is_required=data.get('is_required', False),
        is_locked=data.get('is_locked', False),
        is_monitoring=data.get('is_monitoring', False),
        is_key=data.get('is_key', False),
        is_active=data.get('is_active', True),  # ← این خط را اضافه کنید
        field_order=max_order + 1
    )
    db.session.add(field)
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/admin/fields/<int:field_id>')
@login_required
def admin_field_get(field_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    field = DynamicField.query.get_or_404(field_id)
    return jsonify({
        'id': field.id,
        'title': field.title,
        'field_type': field.field_type,
        'is_required': field.is_required,
        'is_locked': field.is_locked,
        'is_monitoring': field.is_monitoring,
        'is_key': field.is_key,
        'is_active': field.is_active  # ← این خط را اضافه کنید
    })

@app.route('/admin/fields/<int:field_id>/edit', methods=['PUT'])
@login_required
def admin_field_edit(field_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    field = DynamicField.query.get_or_404(field_id)
    data = request.get_json()
    
    field.title = data.get('title', field.title)
    field.field_type = data.get('field_type', field.field_type)
    field.is_required = data.get('is_required', field.is_required)
    field.is_locked = data.get('is_locked', field.is_locked)
    field.is_monitoring = data.get('is_monitoring', field.is_monitoring)
    field.is_key = data.get('is_key', field.is_key)
    field.is_active = data.get('is_active', field.is_active)  # ← این خط را اضافه کنید
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/fields/<int:field_id>/delete', methods=['DELETE'])
@login_required
def admin_field_delete(field_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        field = DynamicField.query.get_or_404(field_id)
        
        # ابتدا مقادیر مرتبط را حذف کنید
        PersonnelValue.query.filter_by(field_id=field.id).delete()
        
        # سپس خود فیلد را حذف کنید
        db.session.delete(field)
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting field: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/admin/fields/<int:field_id>/toggle-status', methods=['POST'])
@login_required
def admin_field_toggle_status(field_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        field = DynamicField.query.get_or_404(field_id)
        data = request.get_json()
        is_active = data.get('is_active', True)
        
        # غیرفعال کردن فیلد کلید (کد ملی) مجاز نیست
        if field.is_key and not is_active:
            return jsonify({'error': 'فیلد کلید (کد ملی) نمی‌تواند غیرفعال شود'}), 400
        
        field.is_active = is_active
        db.session.commit()
        
        return jsonify({'success': True, 'is_active': field.is_active})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error toggling field status: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/admin/fields/reorder', methods=['POST'])
@login_required
def admin_fields_reorder():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        orders = data.get('orders', [])
        
        for item in orders:
            field = DynamicField.query.get(item['id'])
            if field:
                field.field_order = item['order']
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
    
@app.route('/admin/tickets')
@login_required
def admin_tickets():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    return redirect(url_for('tickets'))
    
@app.route('/api/tickets/<int:ticket_id>/delete', methods=['DELETE'])
@login_required
def api_ticket_delete(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    
    if current_user.role != 'admin' and ticket.sender_id != current_user.id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # حذف پاسخ‌ها
    for reply in ticket.replies:
        db.session.delete(reply)
    
    # حذف تیکت
    db.session.delete(ticket)
    db.session.commit()
    
    # محاسبه تعداد خوانده نشده جدید
    unread_count = Ticket.query.filter(
        Ticket.receiver_id == current_user.id,
        Ticket.status.in_(['open', 'in_progress'])
    ).count()
    
    return jsonify({
        'success': True,
        'message': 'با موفقیت حذف شد',
        'unread_count': unread_count
    })


@app.route('/debug/tickets/<int:ticket_id>')
@login_required
def debug_ticket(ticket_id):
    ticket = Ticket.query.get_or_404(ticket_id)
    return jsonify({
        'id': ticket.id,
        'status': ticket.status,
        'receiver_id': ticket.receiver_id,
        'current_user_id': current_user.id,
        'is_unread': (ticket.receiver_id == current_user.id and ticket.status in ['open', 'in_progress'])
    })

@app.route('/dept-manager/api/units/<int:unit_id>/personnel')
@login_required
def dept_manager_api_unit_personnel(unit_id):
    if current_user.role != 'dept_manager':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # بررسی دسترسی به واحد (مدیر اداره باید به این واحد دسترسی داشته باشد)
    dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
    if not dept_manager:
        return jsonify({'error': 'شما به هیچ اداره‌ای متصل نیستید'}), 403
    
    unit = Unit.query.get_or_404(unit_id)
    if unit.department_id != dept_manager.department_id:
        return jsonify({'error': 'شما به این واحد دسترسی ندارید'}), 403
    
    personnel = Personnel.query.filter_by(unit_id=unit_id, is_deleted=False).all()
    
    result = []
    for p in personnel:
        result.append({
            'id': p.id,
            'national_code': p.national_code,
            'full_name': p.get_full_name(),
            'first_name': p.first_name,
            'last_name': p.last_name,
            'phone': p.phone,
            'position': p.position,
            'hire_date': p.hire_date,
            'is_completed': bool(p.first_name and p.last_name and p.phone)
        })
    
    return jsonify(result)
    
# ==================== APIهای گردش کار (سلسله مراتبی) ====================

@app.route('/api/work/submit', methods=['POST'])
@login_required
def api_work_submit():
    """سرپرست واحد: تایید موقت کارکرد پرسنل در یک دوره"""
    if current_user.role not in ['unit_supervisor', 'dept_manager', 'org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    personnel_id = data.get('personnel_id')
    period_id = data.get('period_id')
    
    if not personnel_id or not period_id:
        return jsonify({'error': 'شناسه پرسنل و دوره الزامی است'}), 400
    
    personnel = Personnel.query.get_or_404(personnel_id)
    period = WorkPeriod.query.get_or_404(period_id)
    
    # بررسی اینکه کاربر دسترسی به این پرسنل دارد
    if current_user.role == 'unit_supervisor':
        # بررسی سرپرست بودن واحد این پرسنل
        supervised_units = db.session.query(Unit).join(
            UnitSupervisor, UnitSupervisor.unit_id == Unit.id
        ).filter(UnitSupervisor.user_id == current_user.id).all()
        unit_ids = [u.id for u in supervised_units]
        if personnel.unit_id not in unit_ids:
            return jsonify({'error': 'شما به این پرسنل دسترسی ندارید'}), 403
    
    # بررسی تکمیل بودن همه فیلدهای الزامی
    dynamic_fields = DynamicField.query.filter_by(is_active=True, is_required=True).all()
    missing_fields = []
    
    for field in dynamic_fields:
        if field.is_key:
            continue
        # بررسی مقدار پرسنل در این دوره
        pv = PersonnelValue.query.filter_by(
            personnel_id=personnel_id,
            field_id=field.id,
            period_id=period_id
        ).first()
        
        if not pv or not (pv.value_text or pv.value_number or pv.value_date):
            missing_fields.append(field.title)
    
    if missing_fields:
        return jsonify({
            'error': f'فیلدهای زیر الزامی هستند و تکمیل نشده‌اند: {", ".join(missing_fields)}',
            'missing_fields': missing_fields
        }), 400
    
    # ایجاد یا بروزرسانی وضعیت
    work_status = create_or_update_work_status(
        personnel_id=personnel_id,
        period_id=period_id,
        status='unit_pending'  # در انتظار تایید مدیر اداره
    )
    
    # بروزرسانی اطلاعات تاییدکننده
    work_status.unit_approved_at = datetime.now()
    work_status.unit_approver_id = current_user.id
    db.session.commit()
    
    # پیدا کردن مدیر اداره
    dept_manager = DepartmentManager.query.filter_by(department_id=personnel.department_id).first()
    
    if dept_manager:
        dept_manager_user = User.query.get(dept_manager.user_id)
        if dept_manager_user:
            # ارسال اعلان به مدیر اداره
            send_workflow_notification(
                to_user_id=dept_manager_user.id,
                title=f"درخواست تایید کارکرد دوره {period.title}",
                message=f"سرپرست واحد {current_user.get_full_name()} کارکرد پرسنل {personnel.get_full_name()} را تایید کرده است. لطفاً بررسی کنید.",
                link=f"/dept-manager/dashboard?period={period_id}&personnel={personnel_id}"
            )
    
    return jsonify({
        'success': True,
        'message': f'کارکرد پرسنل {personnel.get_full_name()} برای تایید مدیر اداره ارسال شد',
        'status': 'unit_pending'
    })


@app.route('/api/work/dept-approve', methods=['POST'])
@login_required
def api_work_dept_approve():
    """مدیر اداره: تایید یا اصلاح کارکرد"""
    if current_user.role not in ['dept_manager', 'org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    personnel_id = data.get('personnel_id')
    period_id = data.get('period_id')
    action = data.get('action')
    note = data.get('note', '') or data.get('revision_note', '')
    recipient = data.get('recipient', 'org_manager')  # org_manager, unit_supervisor, both
    
    if not personnel_id or not period_id:
        return jsonify({'error': 'شناسه پرسنل و دوره الزامی است'}), 400
    
    personnel = Personnel.query.get_or_404(personnel_id)
    period = WorkPeriod.query.get_or_404(period_id)
    
    # بررسی دسترسی
    if current_user.role == 'dept_manager':
        dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
        if not dept_manager or dept_manager.department_id != personnel.department_id:
            return jsonify({'error': 'شما به این پرسنل دسترسی ندارید'}), 403
    
    # دریافت یا ایجاد work_status
    work_status = get_personnel_work_status(personnel_id, period_id)
    if not work_status:
        work_status = PersonnelWorkStatus(
            personnel_id=personnel_id,
            period_id=period_id,
            status='draft'
        )
        db.session.add(work_status)
        db.session.commit()
    
    # پیدا کردن گیرندگان
    org_manager = User.query.filter_by(role='org_manager').first()
    if not org_manager:
        org_manager = User.query.filter_by(role='admin').first()
    
    unit_supervisor = UnitSupervisor.query.filter_by(unit_id=personnel.unit_id).first()
    supervisor_user = User.query.get(unit_supervisor.user_id) if unit_supervisor else None
    
    final_note = note
    
    # ========== تایید معمولی ==========
    if action == 'approve':
        if work_status.status != 'unit_pending':
            return jsonify({'error': f'وضعیت فعلی {work_status.status} اجازه تایید را نمی‌دهد. ابتدا سرپرست واحد باید تایید کند.'}), 400
        
        work_status.status = 'dept_pending'
        work_status.dept_approved_at = datetime.now()
        work_status.dept_approver_id = current_user.id
        db.session.commit()
        
        # 1. اعلان به مدیر سازمان (همیشه)
        if org_manager:
            send_workflow_notification(
                to_user_id=org_manager.id,
                title=f"✅ درخواست تایید کارکرد - {personnel.get_full_name()}",
                message=f"مدیر اداره {current_user.get_full_name()} کارکرد پرسنل {personnel.get_full_name()} را برای دوره {period.title} تایید کرد. منتظر تایید نهایی شماست.",
                link=None
            )
        
        # 2. اعلان به سرپرست واحد (همیشه)
        if supervisor_user:
            send_workflow_notification(
                to_user_id=supervisor_user.id,
                title=f"✅ تایید کارکرد - {personnel.get_full_name()}",
                message=f"مدیر اداره {current_user.get_full_name()} کارکرد پرسنل {personnel.get_full_name()} را برای دوره {period.title} تایید کرد.",
                link=None
            )
        
        # ساخت متن پیام تایید
        if final_note:
            approve_message = f"✅ تایید کارکرد توسط مدیر اداره.\n\n📝 توضیحات: {final_note}"
        else:
            approve_message = "✅ تایید کارکرد توسط مدیر اداره."
        
        # 3. ارسال پیام به صندوق بر اساس انتخاب
        if recipient == 'org_manager' and org_manager:
            create_work_message(
                work_status_id=work_status.id,
                personnel_id=personnel_id,
                period_id=period_id,
                from_user=current_user,
                to_user_id=org_manager.id,
                to_role='org_manager',
                message=approve_message,
                message_type='approve'
            )
        elif recipient == 'unit_supervisor' and supervisor_user:
            create_work_message(
                work_status_id=work_status.id,
                personnel_id=personnel_id,
                period_id=period_id,
                from_user=current_user,
                to_user_id=supervisor_user.id,
                to_role='unit_supervisor',
                message=approve_message,
                message_type='approve'
            )
        elif recipient == 'both':
            if org_manager:
                create_work_message(
                    work_status_id=work_status.id,
                    personnel_id=personnel_id,
                    period_id=period_id,
                    from_user=current_user,
                    to_user_id=org_manager.id,
                    to_role='org_manager',
                    message=approve_message,
                    message_type='approve'
                )
            if supervisor_user:
                create_work_message(
                    work_status_id=work_status.id,
                    personnel_id=personnel_id,
                    period_id=period_id,
                    from_user=current_user,
                    to_user_id=supervisor_user.id,
                    to_role='unit_supervisor',
                    message=approve_message,
                    message_type='approve'
                )
        
        return jsonify({
            'success': True,
            'message': f'کارکرد پرسنل {personnel.get_full_name()} برای تایید مدیر سازمان ارسال شد',
            'status': 'dept_pending'
        })
    
    # ========== تایید مستقیم ==========
    elif action == 'direct_approve':
        work_status.status = 'dept_pending'
        work_status.dept_approved_at = datetime.now()
        work_status.dept_approver_id = current_user.id
        db.session.commit()
        
        # 1. اعلان به مدیر سازمان (همیشه)
        if org_manager:
            send_workflow_notification(
                to_user_id=org_manager.id,
                title=f"✅ تایید مستقیم کارکرد - {personnel.get_full_name()}",
                message=f"مدیر اداره {current_user.get_full_name()} کارکرد پرسنل {personnel.get_full_name()} را برای دوره {period.title} مستقیماً تایید کرد. منتظر تایید نهایی شماست.",
                link=None
            )
        
        # 2. اعلان به سرپرست واحد (همیشه)
        if supervisor_user:
            send_workflow_notification(
                to_user_id=supervisor_user.id,
                title=f"✅ تایید مستقیم کارکرد - {personnel.get_full_name()}",
                message=f"مدیر اداره {current_user.get_full_name()} کارکرد پرسنل {personnel.get_full_name()} را برای دوره {period.title} مستقیماً تایید کرد.",
                link=None
            )
        
        # ساخت متن پیام تایید مستقیم
        if final_note:
            direct_approve_message = f"✅ تایید مستقیم کارکرد توسط مدیر اداره.\n\n📝 توضیحات: {final_note}"
        else:
            direct_approve_message = "✅ تایید مستقیم کارکرد توسط مدیر اداره."
        
        # 3. ارسال پیام به صندوق بر اساس انتخاب
        if recipient == 'org_manager' and org_manager:
            create_work_message(
                work_status_id=work_status.id,
                personnel_id=personnel_id,
                period_id=period_id,
                from_user=current_user,
                to_user_id=org_manager.id,
                to_role='org_manager',
                message=direct_approve_message,
                message_type='direct_approve'
            )
        elif recipient == 'unit_supervisor' and supervisor_user:
            create_work_message(
                work_status_id=work_status.id,
                personnel_id=personnel_id,
                period_id=period_id,
                from_user=current_user,
                to_user_id=supervisor_user.id,
                to_role='unit_supervisor',
                message=direct_approve_message,
                message_type='direct_approve'
            )
        elif recipient == 'both':
            if org_manager:
                create_work_message(
                    work_status_id=work_status.id,
                    personnel_id=personnel_id,
                    period_id=period_id,
                    from_user=current_user,
                    to_user_id=org_manager.id,
                    to_role='org_manager',
                    message=direct_approve_message,
                    message_type='direct_approve'
                )
            if supervisor_user:
                create_work_message(
                    work_status_id=work_status.id,
                    personnel_id=personnel_id,
                    period_id=period_id,
                    from_user=current_user,
                    to_user_id=supervisor_user.id,
                    to_role='unit_supervisor',
                    message=direct_approve_message,
                    message_type='direct_approve'
                )
        
        return jsonify({
            'success': True,
            'message': f'کارکرد پرسنل {personnel.get_full_name()} با موفقیت تایید شد',
            'status': 'dept_pending'
        })
    
    # ========== اصلاح کارکرد ==========
    elif action == 'revision':
        if work_status.status != 'unit_pending':
            return jsonify({'error': f'وضعیت فعلی {work_status.status} اجازه اصلاح را نمی‌دهد'}), 400
        
        work_status.status = 'revision'
        work_status.revision_note = final_note
        work_status.revision_from_role = 'dept_manager'
        work_status.revision_from_user_id = current_user.id
        db.session.commit()
        
        # 1. اعلان به سرپرست واحد (همیشه)
        if supervisor_user:
            send_workflow_notification(
                to_user_id=supervisor_user.id,
                title=f"🔄 درخواست اصلاح کارکرد - {personnel.get_full_name()}",
                message=f"مدیر اداره {current_user.get_full_name()} درخواست اصلاح کارکرد پرسنل {personnel.get_full_name()} را برای دوره {period.title} صادر کرد.",
                link=None
            )
        
        # 2. اعلان به مدیر سازمان (همیشه - برای اطلاع)
        if org_manager:
            send_workflow_notification(
                to_user_id=org_manager.id,
                title=f"🔄 اطلاع از اصلاح کارکرد - {personnel.get_full_name()}",
                message=f"مدیر اداره {current_user.get_full_name()} کارکرد پرسنل {personnel.get_full_name()} را به سرپرست واحد برگشت داد.",
                link=None
            )
        
        # ساخت متن پیام اصلاح
        if final_note:
            revision_message = f"🔄 درخواست اصلاح کارکرد توسط مدیر اداره.\n\n📝 توضیحات اصلاح: {final_note}"
        else:
            revision_message = "🔄 درخواست اصلاح کارکرد توسط مدیر اداره."
        
        # 3. پیام به صندوق سرپرست واحد (اجباری)
        if supervisor_user:
            create_work_message(
                work_status_id=work_status.id,
                personnel_id=personnel_id,
                period_id=period_id,
                from_user=current_user,
                to_user_id=supervisor_user.id,
                to_role='unit_supervisor',
                message=revision_message,
                message_type='revision'
            )
        
        # 4. اگر انتخاب شده توضیحات به مدیر سازمان هم برود
        if (recipient == 'org_manager' or recipient == 'both') and org_manager:
            info_message = f"🔄 اطلاع از اصلاح کارکرد.\n\nکارکرد پرسنل {personnel.get_full_name()} به سرپرست واحد برگشت داده شد.\n\n📝 دلیل: {final_note}" if final_note else "🔄 اطلاع از اصلاح کارکرد.\n\nکارکرد پرسنل به سرپرست واحد برگشت داده شد."
            create_work_message(
                work_status_id=work_status.id,
                personnel_id=personnel_id,
                period_id=period_id,
                from_user=current_user,
                to_user_id=org_manager.id,
                to_role='org_manager',
                message=info_message,
                message_type='revision'
            )
        
        return jsonify({
            'success': True,
            'message': f'کارکرد پرسنل {personnel.get_full_name()} برای اصلاح به سرپرست واحد برگشت داده شد',
            'status': 'revision'
        })
    
    return jsonify({'error': 'عملیات نامعتبر'}), 400

@app.route('/api/work/conversations')
@login_required
def api_work_conversations():
    """دریافت لیست مکالمات برای کاربر جاری"""
    try:
        # دریافت همه پیام‌هایی که کاربر در آنها فرستنده یا گیرنده است
        messages = WorkRevisionMessage.query.filter(
            (WorkRevisionMessage.from_user_id == current_user.id) | 
            (WorkRevisionMessage.to_user_id == current_user.id)
        ).order_by(WorkRevisionMessage.created_at.desc()).all()
        
        # گروه‌بندی بر اساس personnel_id + period_id
        conversations = {}
        for msg in messages:
            key = f"{msg.personnel_id}_{msg.period_id}"
            if key not in conversations:
                personnel = Personnel.query.get(msg.personnel_id)
                period = WorkPeriod.query.get(msg.period_id)
                work_status = PersonnelWorkStatus.query.filter_by(
                    personnel_id=msg.personnel_id,
                    period_id=msg.period_id
                ).first()
                
                from_user = User.query.get(msg.from_user_id)
                
                # تبدیل تاریخ به شمسی
                created_at_jalali = jdatetime.datetime.fromgregorian(datetime=msg.created_at).strftime('%Y/%m/%d %H:%M')
                
                # تعیین وضعیت نمایشی
                status_display = 'پیش‌نویس'
                if work_status:
                    if work_status.status == 'unit_pending':
                        status_display = 'در انتظار تایید مدیر اداره'
                    elif work_status.status == 'dept_pending':
                        status_display = 'در انتظار تایید مدیر سازمان'
                    elif work_status.status == 'org_approved':
                        status_display = 'تایید نهایی شده'
                    elif work_status.status == 'revision':
                        status_display = 'نیاز به اصلاح دارد'
                
                conversations[key] = {
                    'personnel_id': msg.personnel_id,
                    'period_id': msg.period_id,
                    'personnel_name': personnel.get_full_name() if personnel else f'پرسنل {msg.personnel_id}',
                    'period_title': period.title if period else f'دوره {msg.period_id}',
                    'status': work_status.status if work_status else 'draft',
                    'status_display': status_display,
                    'last_message_preview': msg.message[:150],
                    'last_message_date': created_at_jalali,
                    'last_from_role': msg.from_role,
                    'last_from_name': from_user.get_full_name() if from_user else '-',
                    'reply_count': 1
                }
            else:
                conversations[key]['reply_count'] += 1
                # به‌روزرسانی آخرین پیام اگر جدیدتر باشد
                if msg.created_at > datetime.strptime(conversations[key]['last_message_date'], '%Y/%m/%d %H:%M'):
                    from_user = User.query.get(msg.from_user_id)
                    created_at_jalali = jdatetime.datetime.fromgregorian(datetime=msg.created_at).strftime('%Y/%m/%d %H:%M')
                    conversations[key]['last_message_preview'] = msg.message[:150]
                    conversations[key]['last_message_date'] = created_at_jalali
                    conversations[key]['last_from_role'] = msg.from_role
                    conversations[key]['last_from_name'] = from_user.get_full_name() if from_user else '-'
        
        result = list(conversations.values())
        result.sort(key=lambda x: x['last_message_date'], reverse=True)
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error in api_work_conversations: {e}")
        import traceback
        traceback.print_exc()
        return jsonify([]), 200
        
@app.route('/api/work/org-approve', methods=['POST'])
@login_required
def api_work_org_approve():
    """مدیر سازمان: تایید نهایی یا اصلاح"""
    if current_user.role not in ['org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    personnel_id = data.get('personnel_id')
    period_id = data.get('period_id')
    action = data.get('action')
    note = data.get('note', '')
    revision_note = data.get('revision_note', '')
    recipient = data.get('recipient', 'dept_manager')
    
    if not personnel_id or not period_id:
        return jsonify({'error': 'شناسه پرسنل و دوره الزامی است'}), 400
    
    personnel = Personnel.query.get_or_404(personnel_id)
    period = WorkPeriod.query.get_or_404(period_id)
    
    work_status = get_personnel_work_status(personnel_id, period_id)
    
    if not work_status:
        return jsonify({'error': 'وضعیت کارکردی برای این پرسنل در این دوره یافت نشد'}), 404
    
    if work_status.status != 'dept_pending':
        return jsonify({'error': f'وضعیت فعلی {work_status.status} اجازه این عملیات را نمی‌دهد'}), 400
    
    if action == 'approve':
        # تایید نهایی مدیر سازمان
        work_status.status = 'org_approved'
        work_status.org_approved_at = datetime.now()
        work_status.org_approver_id = current_user.id
        db.session.commit()
        
        # بروزرسانی پرسنل (اختیاری: ذخیره work_status_id در خود پرسنل)
        personnel.work_status_id = work_status.id
        db.session.commit()
        
        # اعلان به مدیر اداره و سرپرست واحد
        dept_manager = DepartmentManager.query.filter_by(department_id=personnel.department_id).first()
        unit_supervisor = UnitSupervisor.query.filter_by(unit_id=personnel.unit_id).first()
        
        if dept_manager:
            send_workflow_notification(
                to_user_id=dept_manager.user_id,
                title=f"✅ تایید نهایی کارکرد دوره {period.title}",
                message=f"مدیر سازمان کارکرد پرسنل {personnel.get_full_name()} را تأیید نهایی کرد.",
                link=f"/dept-manager/dashboard"
            )
        
        if unit_supervisor:
            send_workflow_notification(
                to_user_id=unit_supervisor.user_id,
                title=f"✅ تایید نهایی کارکرد دوره {period.title}",
                message=f"مدیر سازمان کارکرد پرسنل {personnel.get_full_name()} را تأیید نهایی کرد.",
                link=f"/unit-supervisor/dashboard"
            )
        
        return jsonify({
            'success': True,
            'message': f'کارکرد پرسنل {personnel.get_full_name()} با موفقیت تأیید نهایی شد',
            'status': 'org_approved'
        })
    
    elif action == 'revision':
        # اصلاح - برگشت به مدیر اداره
        work_status.status = 'revision'
        work_status.revision_note = revision_note
        work_status.revision_from_role = 'org_manager'
        work_status.revision_from_user_id = current_user.id
        db.session.commit()
        
        # پیدا کردن مدیر اداره
        dept_manager = DepartmentManager.query.filter_by(department_id=personnel.department_id).first()
        
        if dept_manager:
            # ارسال پیام اصلاح
            send_revision_message(
                work_status_id=work_status.id,
                from_user=current_user,
                to_role='dept_manager',
                to_user_id=dept_manager.user_id,
                message_text=revision_note or f"کارکرد دوره {period.title} نیاز به اصلاح دارد. لطفاً بررسی و اصلاح کنید."
            )
            
            send_workflow_notification(
                to_user_id=dept_manager.user_id,
                title=f"🔄 درخواست اصلاح کارکرد دوره {period.title}",
                message=f"مدیر سازمان {current_user.get_full_name()} کارکرد پرسنل {personnel.get_full_name()} را برگشت داده است. دلیل: {revision_note or 'بدون توضیح'}",
                link=f"/dept-manager/dashboard?period={period_id}"
            )
        
        return jsonify({
            'success': True,
            'message': f'کارکرد پرسنل {personnel.get_full_name()} برای اصلاح به مدیر اداره برگشت داده شد',
            'status': 'revision'
        })
    
    return jsonify({'error': 'عملیات نامعتبر'}), 400

@app.route('/api/user-info')
@login_required
def api_user_info():
    """دریافت اطلاعات کاربر جاری"""
    return jsonify({
        'id': current_user.id,
        'national_code': current_user.national_code,
        'first_name': current_user.first_name,
        'last_name': current_user.last_name,
        'full_name': current_user.get_full_name(),
        'role': current_user.role,
        'username': current_user.username,
        'profile_picture': current_user.profile_picture  # ← اضافه کن
    })

@app.route('/api/work/status/<int:personnel_id>/<int:period_id>')
@login_required
def api_work_status(personnel_id, period_id):
    """دریافت وضعیت کارکرد یک پرسنل در یک دوره"""
    work_status = get_personnel_work_status(personnel_id, period_id)
    
    if not work_status:
        return jsonify({
            'status': 'draft',
            'can_edit': can_edit_personnel(current_user, Personnel.query.get(personnel_id), period_id),
            'can_delete': can_delete_personnel(current_user, Personnel.query.get(personnel_id), period_id)
        })
    
    # بررسی دسترسی برای نمایش
    personnel = Personnel.query.get(personnel_id)
    is_accessible = False
    
    if current_user.role == 'admin':
        is_accessible = True
    elif current_user.role == 'org_manager':
        is_accessible = True
    elif current_user.role == 'dept_manager':
        dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
        if dept_manager and dept_manager.department_id == personnel.department_id:
            is_accessible = True
    elif current_user.role == 'unit_supervisor':
        supervised_units = db.session.query(Unit).join(
            UnitSupervisor, UnitSupervisor.unit_id == Unit.id
        ).filter(UnitSupervisor.user_id == current_user.id).all()
        unit_ids = [u.id for u in supervised_units]
        if personnel.unit_id in unit_ids:
            is_accessible = True
    
    if not is_accessible:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    return jsonify({
        'id': work_status.id,
        'status': work_status.status,
        'status_persian': {
            'draft': 'پیش‌نویس',
            'unit_pending': 'در انتظار تایید مدیر اداره',
            'dept_pending': 'در انتظار تایید مدیر سازمان',
            'org_approved': 'تایید نهایی شده',
            'revision': 'نیاز به اصلاح دارد'
        }.get(work_status.status, work_status.status),
        'unit_approved_at': work_status.unit_approved_at.strftime('%Y/%m/%d %H:%M') if work_status.unit_approved_at else None,
        'dept_approved_at': work_status.dept_approved_at.strftime('%Y/%m/%d %H:%M') if work_status.dept_approved_at else None,
        'org_approved_at': work_status.org_approved_at.strftime('%Y/%m/%d %H:%M') if work_status.org_approved_at else None,
        'revision_note': work_status.revision_note,
        'can_edit': can_edit_personnel(current_user, personnel, period_id),
        'can_delete': can_delete_personnel(current_user, personnel, period_id),
        'can_approve': (
            (current_user.role == 'unit_supervisor' and work_status.status == 'draft') or
            (current_user.role == 'dept_manager' and work_status.status == 'unit_pending') or
            (current_user.role == 'org_manager' and work_status.status == 'dept_pending')
        ),
        'can_revision': (
            (current_user.role == 'dept_manager' and work_status.status == 'unit_pending') or
            (current_user.role == 'org_manager' and work_status.status == 'dept_pending')
        )
    })


@app.route('/api/work/messages')
@login_required
def api_work_messages():
    """دریافت پیام‌های صندوق پیام برای کاربر جاری"""
    messages = WorkRevisionMessage.query.filter_by(to_user_id=current_user.id).order_by(
        WorkRevisionMessage.created_at.desc()
    ).all()
    
    # شمارش پیام‌های خوانده نشده
    unread_count = sum(1 for msg in messages if not msg.is_read)
    
    result = []
    for msg in messages:
        personnel = Personnel.query.get(msg.personnel_id)
        period = WorkPeriod.query.get(msg.period_id)
        from_user = User.query.get(msg.from_user_id)
        
        created_at_jalali = jdatetime.datetime.fromgregorian(datetime=msg.created_at).strftime('%Y/%m/%d %H:%M')
        
        result.append({
            'id': msg.id,
            'personnel_id': msg.personnel_id,
            'period_id': msg.period_id,
            'personnel_name': personnel.get_full_name() if personnel else '-',
            'period_title': period.title if period else '-',
            'from_role': msg.from_role,
            'from_user_id': msg.from_user_id,
            'from_user_name': from_user.get_full_name() if from_user else '-',
            'message': msg.message,
            'is_read': msg.is_read,
            'created_at': created_at_jalali,
            'created_at_raw': msg.created_at.isoformat()
        })
    
    return jsonify({
        'unread_count': unread_count,
        'messages': result
    })

@app.route('/api/work/conversation/<int:personnel_id>/<int:period_id>')
@login_required
def api_work_conversation_detail(personnel_id, period_id):
    """دریافت همه پیام‌های یک مکالمه"""
    try:
        messages = WorkRevisionMessage.query.filter_by(
            personnel_id=personnel_id,
            period_id=period_id
        ).order_by(WorkRevisionMessage.created_at.asc()).all()
        
        if not messages:
            return jsonify({'error': 'مکالمه یافت نشد'}), 404
        
        first_msg = messages[0]
        first_from_user = User.query.get(first_msg.from_user_id)
        work_status = PersonnelWorkStatus.query.filter_by(
            personnel_id=personnel_id,
            period_id=period_id
        ).first()
        
        # تبدیل تاریخ اول به شمسی
        first_date_jalali = jdatetime.datetime.fromgregorian(datetime=first_msg.created_at).strftime('%Y/%m/%d %H:%M')
        
        replies = []
        for msg in messages[1:]:
            from_user = User.query.get(msg.from_user_id)
            # تبدیل تاریخ به شمسی
            created_at_jalali = jdatetime.datetime.fromgregorian(datetime=msg.created_at).strftime('%Y/%m/%d %H:%M')
            
            replies.append({
                'id': msg.id,
                'from_role': msg.from_role,
                'from_user_id': msg.from_user_id,
                'from_user_name': from_user.get_full_name() if from_user else '-',
                'message': msg.message,
                'created_at': created_at_jalali,  # تاریخ شمسی
                'is_read': msg.is_read,
                'to_user_id': msg.to_user_id
            })
        
        # علامت‌گذاری پیام‌های دریافتی به عنوان خوانده شده
        for msg in messages:
            if msg.to_user_id == current_user.id and not msg.is_read:
                msg.is_read = True
        db.session.commit()
        
        return jsonify({
            'first_message': first_msg.message,
            'first_message_date': first_date_jalali,  # تاریخ شمسی
            'first_from_name': first_from_user.get_full_name() if first_from_user else '-',
            'first_from_role': first_msg.from_role,
            'first_status': work_status.status if work_status else 'draft',
            'replies': replies
        })
        
    except Exception as e:
        print(f"Error in api_work_conversation_detail: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/work/send-message', methods=['POST'])
@login_required
def api_work_send_message():
    """ارسال پیام جدید در یک مکالمه"""
    try:
        data = request.get_json()
        personnel_id = data.get('personnel_id')
        period_id = data.get('period_id')
        message = data.get('message', '').strip()
        
        if not personnel_id or not period_id:
            return jsonify({'error': 'اطلاعات ناقص است'}), 400
        
        if not message:
            return jsonify({'error': 'متن پیام نمی‌تواند خالی باشد'}), 400
        
        personnel = Personnel.query.get(personnel_id)
        if not personnel:
            return jsonify({'error': 'پرسنل یافت نشد'}), 404
        
        # پیدا کردن آخرین پیام برای تعیین گیرنده
        last_msg = WorkRevisionMessage.query.filter_by(
            personnel_id=personnel_id,
            period_id=period_id
        ).order_by(WorkRevisionMessage.created_at.desc()).first()
        
        # تعیین گیرنده
        if last_msg:
            # اگر پیام قبلی وجود دارد، گیرنده = فرستنده پیام قبلی
            to_user_id = last_msg.from_user_id
            to_role = last_msg.from_role
        else:
            # اولین پیام: باید گیرنده مشخص شود (مدیر اداره یا مدیر سازمان)
            # بر اساس نقش کاربر فعلی
            if current_user.role == 'dept_manager':
                # مدیر اداره به مدیر سازمان پیام می‌دهد
                org_manager = User.query.filter_by(role='org_manager').first()
                if not org_manager:
                    org_manager = User.query.filter_by(role='admin').first()
                to_user_id = org_manager.id if org_manager else 1
                to_role = 'org_manager'
            else:
                # سرپرست واحد به مدیر اداره پیام می‌دهد
                dept_manager = DepartmentManager.query.filter_by(department_id=personnel.department_id).first()
                to_user_id = dept_manager.user_id if dept_manager else 1
                to_role = 'dept_manager'
        
        # دریافت work_status_id
        work_status = get_personnel_work_status(personnel_id, period_id)
        work_status_id = work_status.id if work_status else None
        
        # ایجاد پیام جدید
        new_msg = WorkRevisionMessage(
            work_status_id=work_status_id,
            personnel_id=personnel_id,
            period_id=period_id,
            from_role=current_user.role,
            from_user_id=current_user.id,
            to_role=to_role,
            to_user_id=to_user_id,
            message=message,
            is_read=False
        )
        db.session.add(new_msg)
        db.session.commit()
        
        created_at_jalali = jdatetime.datetime.fromgregorian(datetime=new_msg.created_at).strftime('%Y/%m/%d %H:%M')
        
        # ارسال اعلان به گیرنده
        send_workflow_notification(
            to_user_id=to_user_id,
            title=f"پیام جدید در مورد {personnel.get_full_name()}",
            message=message[:100],
            link=None
        )
        
        return jsonify({
            'success': True,
            'message': 'پیام با موفقیت ارسال شد',
            'new_message': {
                'id': new_msg.id,
                'from_role': new_msg.from_role,
                'from_user_id': new_msg.from_user_id,
                'from_user_name': current_user.get_full_name(),
                'message': new_msg.message,
                'created_at': created_at_jalali,
                'is_sent_by_me': True
            }
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/work/messages/<int:msg_id>/read', methods=['POST'])
@login_required
def api_work_message_read(msg_id):
    """علامت‌گذاری پیام به عنوان خوانده شده"""
    msg = WorkRevisionMessage.query.get_or_404(msg_id)
    
    if msg.to_user_id != current_user.id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    msg.is_read = True
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/api/work/dept-direct-approve', methods=['POST'])
@login_required
def api_work_dept_direct_approve():
    """مدیر اداره: تایید مستقیم کارکرد (بدون نیاز به تایید سرپرست واحد)"""
    if current_user.role not in ['dept_manager', 'org_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    personnel_id = data.get('personnel_id')
    period_id = data.get('period_id')
    note = data.get('note', '')
    recipient = data.get('recipient', 'org_manager')
    
    if not personnel_id or not period_id:
        return jsonify({'error': 'شناسه پرسنل و دوره الزامی است'}), 400
    
    personnel = Personnel.query.get_or_404(personnel_id)
    period = WorkPeriod.query.get_or_404(period_id)
    
    # بررسی دسترسی مدیر اداره
    if current_user.role == 'dept_manager':
        dept_manager = DepartmentManager.query.filter_by(user_id=current_user.id).first()
        if not dept_manager or dept_manager.department_id != personnel.department_id:
            return jsonify({'error': 'شما به این پرسنل دسترسی ندارید'}), 403
    
    # ایجاد یا بروزرسانی وضعیت - تایید مستقیم به dept_pending می‌رود
    work_status = get_personnel_work_status(personnel_id, period_id)
    if not work_status:
        work_status = PersonnelWorkStatus(
            personnel_id=personnel_id,
            period_id=period_id,
            status='dept_pending'
        )
        db.session.add(work_status)
    else:
        work_status.status = 'dept_pending'
    
    work_status.dept_approved_at = datetime.now()
    work_status.dept_approver_id = current_user.id
    db.session.commit()
    
    # ارسال اعلان به مدیر سازمان
    org_manager = User.query.filter_by(role='org_manager').first()
    if not org_manager:
        org_manager = User.query.filter_by(role='admin').first()
    
    if org_manager:
        send_workflow_notification(
            to_user_id=org_manager.id,
            title=f"درخواست تایید نهایی کارکرد دوره {period.title} (تایید مستقیم)",
            message=f"مدیر اداره {current_user.get_full_name()} کارکرد پرسنل {personnel.get_full_name()} را مستقیماً تایید کرده است.",
            link=f"/org-manager/dashboard?period={period_id}&personnel={personnel_id}"
        )
    
    # اگر توضیح نوشته شده، به صندوق پیام مدیر سازمان ارسال کن
    if note and recipient == 'org_manager':
        send_revision_message(
            work_status_id=work_status.id,
            from_user=current_user,
            to_role='org_manager',
            to_user_id=org_manager.id,
            message_text=note
        )
    
    return jsonify({
        'success': True,
        'message': f'کارکرد پرسنل {personnel.get_full_name()} با موفقیت تایید شد',
        'status': 'dept_pending'
    })

@app.route('/api/work/mark-conversation-read', methods=['POST'])
@login_required
def api_work_mark_conversation_read():
    """علامت‌گذاری همه پیام‌های یک پرسنل به عنوان خوانده شده"""
    try:
        data = request.get_json()
        if not data:
            data = {}
        
        personnel_id = data.get('personnel_id') or request.args.get('personnel_id')
        
        if not personnel_id:
            return jsonify({'error': 'شناسه پرسنل الزامی است'}), 400
        
        updated = WorkRevisionMessage.query.filter_by(
            personnel_id=personnel_id,
            to_user_id=current_user.id,
            is_read=False
        ).update({'is_read': True})
        
        db.session.commit()
        
        return jsonify({'success': True, 'updated': updated})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/work/send-reply', methods=['POST'])
@login_required
def api_work_send_reply():
    """ارسال پاسخ در یک مکالمه"""
    try:
        data = request.get_json()
        personnel_id = data.get('personnel_id')
        period_id = data.get('period_id')
        message = data.get('message', '').strip()
        
        if not personnel_id or not period_id:
            return jsonify({'error': 'اطلاعات ناقص است'}), 400
        
        if not message:
            return jsonify({'error': 'متن پیام نمی‌تواند خالی باشد'}), 400
        
        personnel = Personnel.query.get(personnel_id)
        if not personnel:
            return jsonify({'error': 'پرسنل یافت نشد'}), 404
        
        # پیدا کردن آخرین پیام برای تعیین گیرنده
        last_msg = WorkRevisionMessage.query.filter_by(
            personnel_id=personnel_id,
            period_id=period_id
        ).order_by(WorkRevisionMessage.created_at.desc()).first()
        
        if last_msg:
            to_user_id = last_msg.from_user_id
            to_role = last_msg.from_role
        else:
            # اولین پیام: گیرنده را بر اساس نقش تعیین کن
            if current_user.role == 'unit_supervisor':
                dept_manager = DepartmentManager.query.filter_by(department_id=personnel.department_id).first()
                to_user_id = dept_manager.user_id if dept_manager else 1
                to_role = 'dept_manager'
            elif current_user.role == 'dept_manager':
                org_manager = User.query.filter_by(role='org_manager').first()
                to_user_id = org_manager.id if org_manager else 1
                to_role = 'org_manager'
            else:
                to_user_id = 1
                to_role = 'admin'
        
        work_status = get_personnel_work_status(personnel_id, period_id)
        work_status_id = work_status.id if work_status else None
        
        new_msg = WorkRevisionMessage(
            work_status_id=work_status_id,
            personnel_id=personnel_id,
            period_id=period_id,
            from_role=current_user.role,
            from_user_id=current_user.id,
            to_role=to_role,
            to_user_id=to_user_id,
            message=message,
            is_read=False
        )
        db.session.add(new_msg)
        db.session.commit()
        
        # ارسال اعلان به گیرنده
        send_workflow_notification(
            to_user_id=to_user_id,
            title=f"پیام جدید در مکالمه {personnel.get_full_name()}",
            message=message[:100],
            link=None
        )
        
        return jsonify({'success': True, 'message': 'پیام با موفقیت ارسال شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in api_work_send_reply: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/work/delete-conversation', methods=['DELETE'])
@login_required
def api_work_delete_conversation():
    """حذف یک مکالمه کامل"""
    try:
        personnel_id = request.args.get('personnel_id')
        period_id = request.args.get('period_id')
        
        if not personnel_id or not period_id:
            return jsonify({'error': 'شناسه پرسنل و دوره الزامی است'}), 400
        
        deleted = WorkRevisionMessage.query.filter(
            WorkRevisionMessage.personnel_id == personnel_id,
            WorkRevisionMessage.period_id == period_id,
            db.or_(
                WorkRevisionMessage.from_user_id == current_user.id,
                WorkRevisionMessage.to_user_id == current_user.id
            )
        ).delete()
        
        db.session.commit()
        
        return jsonify({'success': True, 'deleted': deleted})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error in api_work_delete_conversation: {e}")
        return jsonify({'error': str(e)}), 500
        
# ==================== APIهای جدید داشبورد ادمین ====================

@app.route('/admin/api/dashboard-data')
@login_required
@cached(ttl=60)  # کش 60 ثانیه
def admin_api_dashboard_data():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        # ادارات
        departments = Department.query.all()
        departments_data = []
        for d in departments:
            managers = db.session.query(User).join(DepartmentManager).filter(DepartmentManager.department_id == d.id).all()
            units_count = Unit.query.filter_by(department_id=d.id).count()
            personnel_count = Personnel.query.filter_by(department_id=d.id, is_deleted=False).count()
            departments_data.append({
                'id': d.id,
                'name': d.name or '',
                'color': d.color or '#3498db',
                'managers': ', '.join([m.get_full_name() for m in managers]) if managers else '-',
                'units_count': units_count,
                'personnel_count': personnel_count
            })
        
        # واحدها
        units = Unit.query.all()
        units_data = []
        all_units_list = []
        for u in units:
            dept = Department.query.get(u.department_id)
            supervisors = db.session.query(User).join(UnitSupervisor).filter(UnitSupervisor.unit_id == u.id).all()
            personnel_count = Personnel.query.filter_by(unit_id=u.id, is_deleted=False).count()
            units_data.append({
                'id': u.id,
                'name': u.name or '',
                'department_name': dept.name if dept else '-',
                'supervisors': ', '.join([s.get_full_name() for s in supervisors]) if supervisors else '-',
                'personnel_count': personnel_count
            })
            all_units_list.append({
                'id': u.id,
                'name': u.name or '',
                'department_id': u.department_id
            })
        
        # پرسنل - اضافه شدن فیلدهای جدید برای نمایش در جدول
        personnel = Personnel.query.filter_by(is_deleted=False).all()
        personnel_data = []
        dynamic_fields = DynamicField.query.filter_by(is_active=True).all()
        
        # دریافت همه دوره‌ها برای تبدیل شناسه به عنوان
        all_periods = {p.id: p.title for p in WorkPeriod.query.all()}
        
        for p in personnel:
            dept = Department.query.get(p.department_id)
            unit = Unit.query.get(p.unit_id)
            
            # عنوان دوره
            period_title = all_periods.get(p.period_id, '-') if p.period_id else '-'
            
            # تاریخ ایجاد به شمسی
            created_at_jalali = jdatetime.datetime.fromgregorian(datetime=p.created_at).strftime('%Y/%m/%d') if p.created_at else '-'
            
            # دریافت مقادیر فیلدهای داینامیک برای این پرسنل (برای نمایش در جدول)
            dynamic_values = {}
            for v in PersonnelValue.query.filter_by(personnel_id=p.id).all():
                field = DynamicField.query.get(v.field_id)
                if field:
                    # مقدار را بر اساس نوع فیلد انتخاب کن
                    value = v.value_text or v.value_number or v.value_date or ''
                    dynamic_values[field.id] = value
            
            # بررسی تکمیل بودن اطلاعات
            filled = 0
            total_required = 0
            for f in dynamic_fields:
                if f.is_key or f.title in ['نام', 'نام خانوادگی']:
                    continue
                total_required += 1
                pv = PersonnelValue.query.filter_by(
                    personnel_id=p.id, 
                    field_id=f.id, 
                    period_id=p.period_id
                ).first()
                if pv and (pv.value_text or pv.value_number or pv.value_date):
                    filled += 1
            is_complete = total_required > 0 and filled == total_required if total_required > 0 else True
            
            personnel_data.append({
                'id': p.id,
                'national_code': p.national_code or '',
                'full_name': p.get_full_name() or '',
                'first_name': p.first_name or '',
                'last_name': p.last_name or '',
                'phone': p.phone or '',
                'position': p.position or '',
                'department_id': p.department_id,
                'department_name': dept.name if dept else '-',
                'unit_id': p.unit_id,
                'unit_name': unit.name if unit else '-',
                'period_id': p.period_id,
                'period_title': period_title,
                'created_at_jalali': created_at_jalali,
                'dynamic_values': dynamic_values,
                'is_complete': is_complete
            })
        
        # کاربران
        users = User.query.all()
        users_data = [{
    'id': u.id,
    'full_name': u.get_full_name(),
    'national_code': u.national_code,
    'phone': u.phone or '',  # ← این خط را اضافه کنید
    'role_persian': u.get_role_persian()
} for u in users]
        
        # پیشرفت دوره‌ها
        periods = WorkPeriod.query.all()
        periods_progress = []
        for period in periods:
            total_personnel = Personnel.query.filter_by(period_id=period.id, is_deleted=False).count()
            approved_count = PersonnelWorkStatus.query.filter_by(period_id=period.id, status='org_approved').count()
            completion_percent = int((approved_count / total_personnel) * 100) if total_personnel > 0 else 0
            periods_progress.append({
                'id': period.id,
                'title': period.title or '',
                'total_count': total_personnel,
                'approved_count': approved_count,
                'completion_percent': completion_percent
            })
        
        # پرسنل ناقص (با جزئیات بیشتر برای نمایش در داشبورد)
        incomplete_personnel = []
        for p in personnel_data:
            if not p['is_complete']:
                incomplete_personnel.append({
                    'id': p['id'],
                    'national_code': p['national_code'],
                    'full_name': p['full_name'],
                    'department_name': p['department_name'],
                    'unit_name': p['unit_name']
                })
        
        # وضعیت‌های کارکرد
        work_statuses = PersonnelWorkStatus.query.all()
        work_statuses_data = [{'personnel_id': ws.personnel_id, 'status': ws.status} for ws in work_statuses]
        
        # تعداد کاربران آنلاین (15 دقیقه اخیر)
        from datetime import timedelta
        fifteen_min_ago = datetime.now() - timedelta(minutes=15)
        online_count = User.query.filter(User.last_login > fifteen_min_ago).count()
        
        # لاگ فعالیت‌ها
        activity_logs = []
        recent_statuses = PersonnelWorkStatus.query.order_by(PersonnelWorkStatus.updated_at.desc()).limit(30).all()
        for ws in recent_statuses:
            personnel = Personnel.query.get(ws.personnel_id)
            period = WorkPeriod.query.get(ws.period_id)
            if personnel and period:
                status_text = ''
                badge = ''
                type_log = ''
                if ws.status == 'unit_pending':
                    status_text = f'توسط سرپرست واحد تایید شد'
                    badge = 'در انتظار مدیر اداره'
                    type_log = 'pending'
                elif ws.status == 'dept_pending':
                    status_text = f'توسط مدیر اداره تایید شد'
                    badge = 'در انتظار مدیر سازمان'
                    type_log = 'pending'
                elif ws.status == 'org_approved':
                    status_text = f'تایید نهایی شد'
                    badge = 'تایید نهایی'
                    type_log = 'approve'
                elif ws.status == 'revision':
                    status_text = f'نیاز به اصلاح دارد'
                    badge = 'اصلاح'
                    type_log = 'pending'
                else:
                    continue
                
                activity_logs.append({
                    'time': ws.updated_at.strftime('%H:%M %Y/%m/%d') if ws.updated_at else '',
                    'message': f'📅 دوره {period.title} - 👤 {personnel.get_full_name()} - {status_text}',
                    'badge': badge,
                    'type': type_log
                })
        
        activity_logs = sorted(activity_logs, key=lambda x: x['time'], reverse=True)[:20]
        
        result = {
            'departments': departments_data,
            'units': units_data,
            'all_units': all_units_list,
            'personnel': personnel_data,
            'users': users_data,
            'periods_progress': periods_progress,
            'incomplete_personnel': incomplete_personnel,
            'work_statuses': work_statuses_data,
            'online_count': online_count,
            'activity_logs': activity_logs
        }
        
        return jsonify(result)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/admin/api/online-users')
@login_required
def admin_api_online_users():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    from datetime import timedelta
    # تغییر از 15 دقیقه به 60 دقیقه
    one_hour_ago = datetime.now() - timedelta(minutes=60)
    online_users = User.query.filter(User.last_login > one_hour_ago).all()
    
    return jsonify([{
        'id': u.id, 'full_name': u.get_full_name(),
        'national_code': u.national_code, 'role_persian': u.get_role_persian()
    } for u in online_users])


@app.route('/admin/api/users-list')
@login_required
def admin_api_users_list():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    users = User.query.all()
    return jsonify([{
        'id': u.id, 'full_name': u.get_full_name(),
        'role_persian': u.get_role_persian()
    } for u in users])


@app.route('/admin/api/send-message', methods=['POST'])
@login_required
def admin_api_send_message():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    to_user_id = data.get('user_id')
    message_text = data.get('message', '').strip()
    
    if not to_user_id or not message_text:
        return jsonify({'error': 'اطلاعات ناقص است'}), 400
    
    to_user = User.query.get(to_user_id)
    if not to_user:
        return jsonify({'error': 'کاربر یافت نشد'}), 404
    
    # درست کردن متغیر message_type که تعریف نشده بود
    message_type = data.get('message_type', 'message')  # اضافه شد
    
    ticket = Ticket(
        title=f"پیام از مدیریت",
        message=message_text,
        sender_id=current_user.id,
        receiver_id=to_user_id,
        priority='normal',
        message_type=message_type  # حالا تعریف شده
    )
    db.session.add(ticket)
    db.session.commit()
    
    send_workflow_notification(
        to_user_id=to_user_id,
        title=f"پیام جدید از مدیریت",
        message=message_text[:100],
        link="/tickets"
    )
    
    return jsonify({'success': True, 'message': 'پیام با موفقیت ارسال شد'})

    
@app.route('/admin/inbox')
@login_required
def admin_inbox():
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    tickets = Ticket.query.filter(
        Ticket.receiver_id == current_user.id
    ).order_by(Ticket.created_at.desc()).all()
    
    return render_template('admin/inbox.html', tickets=tickets)


@app.route('/admin/api/online-users-debug')
@login_required
def admin_api_online_users_debug():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    from datetime import timedelta
    fifteen_min_ago = datetime.now() - timedelta(minutes=15)
    
    all_users = User.query.all()
    online_users = User.query.filter(User.last_login > fifteen_min_ago).all()
    
    result = {
        'now': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'fifteen_min_ago': fifteen_min_ago.strftime('%Y-%m-%d %H:%M:%S'),
        'all_users': [{'id': u.id, 'name': u.get_full_name(), 'last_login': u.last_login.strftime('%Y-%m-%d %H:%M:%S') if u.last_login else 'هرگز'} for u in all_users],
        'online_users': [{'id': u.id, 'name': u.get_full_name(), 'role': u.role} for u in online_users],
        'online_count': len(online_users)
    }
    
    return jsonify(result)
    
    
@app.route('/test-online')
@login_required
def test_online():
    from datetime import timedelta
    fifteen_min_ago = datetime.now() - timedelta(minutes=15)
    
    online_users = User.query.filter(User.last_login > fifteen_min_ago).all()
    
    result = "<html dir='rtl'><body style='font-family:Tahoma; padding:20px;'>"
    result += f"<h2>زمان حال: {datetime.now().strftime('%H:%M:%S')}</h2>"
    result += f"<h2>۱۵ دقیقه قبل: {fifteen_min_ago.strftime('%H:%M:%S')}</h2>"
    result += "<h3>کاربران آنلاین:</h3><ul>"
    for u in online_users:
        result += f"<li>🟢 {u.get_full_name()} ({u.role}) - آخرین لاگین: {u.last_login}</li>"
    if not online_users:
        result += "<li>❌ هیچ کاربر آنلاینی نیست!</li>"
    result += "</ul></body></html>"
    
    return result
    
# فقط این بخش را در انتهای فایل قرار بده (بقیه کدهای تکراری را حذف کن)

# ==================== APIهای صندوق پیام ادمین (فقط یک بار) ====================

@app.route('/admin/api/inbox')
@login_required
def admin_api_inbox():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    tickets = Ticket.query.filter(
        Ticket.receiver_id == current_user.id
    ).order_by(Ticket.created_at.desc()).all()
    
    result = []
    for t in tickets:
        sender = User.query.get(t.sender_id)
        result.append({
            'id': t.id,
            'title': t.title,
            'message': t.message,
            'sender_name': sender.get_full_name() if sender else '-',
            'status': t.status,
            'priority': t.priority,
            'is_read': t.status != 'open',
            'created_at': t.created_at.strftime('%Y/%m/%d %H:%M')
        })
    
    return jsonify(result)


@app.route('/admin/api/inbox-count')
@login_required
def admin_api_inbox_count():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    unread_count = Ticket.query.filter(
        Ticket.receiver_id == current_user.id,
        Ticket.status == 'open'
    ).count()
    
    return jsonify({'unread_count': unread_count})


@app.route('/admin/api/ticket/<int:ticket_id>')
@login_required
def admin_api_ticket_get(ticket_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    ticket = Ticket.query.get_or_404(ticket_id)
    sender = User.query.get(ticket.sender_id)
    
    replies = []
    for r in ticket.replies:
        user = User.query.get(r.user_id)
        replies.append({
            'id': r.id,
            'user_name': user.get_full_name() if user else '-',
            'message': r.message,
            'created_at': r.created_at.strftime('%Y/%m/%d %H:%M')
        })
    
    return jsonify({
        'id': ticket.id,
        'title': ticket.title,
        'message': ticket.message,
        'sender_name': sender.get_full_name() if sender else '-',
        'status': ticket.status,
        'priority': ticket.priority,
        'created_at': ticket.created_at.strftime('%Y/%m/%d %H:%M'),
        'replies': replies
    })


@app.route('/admin/api/ticket/<int:ticket_id>/read', methods=['POST'])
@login_required
def admin_api_ticket_read(ticket_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    ticket = Ticket.query.get_or_404(ticket_id)
    if ticket.status == 'open':
        ticket.status = 'in_progress'
        db.session.commit()
    
    return jsonify({'success': True})


@app.route('/admin/api/ticket/<int:ticket_id>/reply', methods=['POST'])
@login_required
def admin_api_ticket_reply(ticket_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    ticket = Ticket.query.get_or_404(ticket_id)
    data = request.get_json()
    message = data.get('message', '').strip()
    
    if not message:
        return jsonify({'error': 'متن پاسخ الزامی است'}), 400
    
    reply = TicketReply(
        ticket_id=ticket_id,
        user_id=current_user.id,
        message=message,
        is_admin_reply=True
    )
    db.session.add(reply)
    ticket.status = 'in_progress'
    db.session.commit()
    
    send_workflow_notification(
        to_user_id=ticket.sender_id,
        title=f"پاسخ به تیکت: {ticket.title}",
        message=f"ادمین به تیکت شما پاسخ داد: {message[:100]}",
        link="/tickets"
    )
    
    return jsonify({'success': True})

@app.route('/admin/api/department-detail/<int:dept_id>')
@login_required
def admin_api_department_detail(dept_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    dept = Department.query.get_or_404(dept_id)
    manager_ids = [dm.user_id for dm in DepartmentManager.query.filter_by(department_id=dept_id).all()]
    
    return jsonify({
        'id': dept.id,
        'name': dept.name,
        'color': dept.color,
        'description': dept.description,
        'manager_ids': manager_ids
    })
    
    
@app.route('/admin/api/unit-detail/<int:unit_id>')
@login_required
def admin_api_unit_detail(unit_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    unit = Unit.query.get_or_404(unit_id)
    supervisor_ids = [us.user_id for us in UnitSupervisor.query.filter_by(unit_id=unit_id).all()]
    
    return jsonify({
        'id': unit.id,
        'name': unit.name,
        'department_id': unit.department_id,
        'description': unit.description,
        'needs_approval': unit.needs_approval,
        'supervisor_ids': supervisor_ids
    })
    
@app.route('/admin/units/test-delete/<int:unit_id>')
@login_required
def admin_unit_test_delete(unit_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        unit = Unit.query.get(unit_id)
        if not unit:
            return jsonify({'error': 'واحد یافت نشد'}), 404
        
        return jsonify({
            'id': unit.id,
            'name': unit.name,
            'department_id': unit.department_id,
            'has_personnel': Personnel.query.filter_by(unit_id=unit_id).count() > 0
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
        
@app.route('/admin/personnel/download-template')
@login_required
def admin_personnel_download_template():
    """دانلود قالب اکسل خالی برای آپلود پرسنل"""
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import tempfile
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قالب پرسنل"
    
    # هدرهای پایه
    headers = ['کد ملی', 'نام', 'نام خانوادگی', 'نام اداره', 'نام واحد', 'دوره', 'شماره تماس', 'سمت']
    
    # اضافه کردن فیلدهای داینامیک فعال
    dynamic_fields = DynamicField.query.filter_by(is_active=True).all()
    for field in dynamic_fields:
        if not field.is_key and field.title not in ['نام', 'نام خانوادگی', 'شماره تماس', 'سمت']:
            headers.append(field.title)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color='ffffff')
        cell.fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    
    return send_file(temp_file.name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='قالب_پرسنل_آوان.xlsx')

# اضافه کن به بخش APIهای ادمین

@app.route('/admin/api/units-all')
@login_required
def admin_api_units_all():
    """دریافت همه واحدها برای فیلترها"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    units = Unit.query.filter_by(is_active=True).all()
    return jsonify([{
        'id': u.id,
        'name': u.name,
        'department_id': u.department_id
    } for u in units])


@app.route('/admin/personnel/import-excel-smart', methods=['POST'])
@login_required
def admin_personnel_import_excel_smart():
    """آپلود هوشمند اکسل پرسنل - با تشخیص خودکار ستون‌ها"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    if 'personnel_data' not in request.form:
        return jsonify({'error': 'داده‌ای ارسال نشده است'}), 400
    
    try:
        personnel_list = json.loads(request.form['personnel_data'])
        dynamic_fields = DynamicField.query.filter_by(is_active=True).all()
        field_dict = {f.id: f for f in dynamic_fields}
        
        success_count = 0
        error_count = 0
        errors = []
        
        for item in personnel_list:
            try:
                national_code = item.get('national_code')
                existing = Personnel.query.filter_by(national_code=national_code, is_deleted=False).first()
                if existing:
                    error_count += 1
                    errors.append(f"کد ملی {national_code} تکراری است")
                    continue
                
                personnel = Personnel(
                    national_code=national_code,
                    first_name=item.get('first_name'),
                    last_name=item.get('last_name'),
                    phone=item.get('phone') or None,
                    position=item.get('position') or None,
                    department_id=item.get('department_id'),
                    unit_id=item.get('unit_id'),
                    period_id=item.get('period_id')
                )
                db.session.add(personnel)
                db.session.flush()
                
                for field_id, value in item.get('dynamic_values', {}).items():
                    field_id = int(field_id)
                    if value and value != '':
                        field = field_dict.get(field_id)
                        if field:
                            pv = PersonnelValue(
                                personnel_id=personnel.id,
                                field_id=field_id,
                                period_id=item.get('period_id'),
                                value_text=value if field.field_type == 'text' else None,
                                value_number=float(value) if field.field_type in ['number', 'decimal'] and value else None,
                                value_date=value if field.field_type == 'date' else None
                            )
                            db.session.add(pv)
                
                success_count += 1
            except Exception as e:
                error_count += 1
                errors.append(str(e))
        
        db.session.commit()
        
        message = f"✅ {success_count} پرسنل با موفقیت اضافه شد."
        if error_count > 0:
            message += f" ❌ {error_count} خطا"
        
        return jsonify({'success': True, 'message': message})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'خطا در پردازش: {str(e)}'}), 400



# ==================== آپلود تصاویر برای تنظیمات ظاهر ====================


@app.route('/admin/settings/upload-login-logo', methods=['POST'])
@login_required
def admin_upload_login_logo():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': 'فایلی ارسال نشده'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'فایلی انتخاب نشده'}), 400
    
    old_file = Setting.get('login_logo', '')
    if old_file:
        old_path = os.path.join(app.root_path, 'static', 'uploads', old_file)
        if os.path.exists(old_path):
            os.remove(old_path)
    
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'png'
    filename = f"login_logo_{int(datetime.now().timestamp())}.{ext}"
    file.save(os.path.join(app.root_path, 'static', 'uploads', filename))
    
    Setting.set('login_logo', f'/static/uploads/{filename}')
    
    return jsonify({'success': True, 'url': f'/static/uploads/{filename}'})


@app.route('/admin/settings/upload-login-bg', methods=['POST'])
@login_required
def admin_upload_login_bg():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': 'فایلی ارسال نشده'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'فایلی انتخاب نشده'}), 400
    
    # حذف فایل قبلی
    old_file = Setting.get('login_bg', '')
    if old_file:
        old_path = os.path.join(app.root_path, 'static', 'uploads', old_file.replace('/static/uploads/', ''))
        if os.path.exists(old_path):
            os.remove(old_path)
    
    # ذخیره فایل جدید
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
    filename = f"login_bg_{int(datetime.now().timestamp())}.{ext}"
    file.save(os.path.join(app.root_path, 'static', 'uploads', filename))
    
    Setting.set('login_bg', f'/static/uploads/{filename}')
    
    return jsonify({'success': True, 'message': 'تصویر پس‌زمینه با موفقیت آپلود شد', 'url': f'/static/uploads/{filename}'})

@app.route('/admin/settings/upload-favicon', methods=['POST'])
@login_required
def admin_upload_favicon():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': 'فایلی ارسال نشده'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'فایلی انتخاب نشده'}), 400
    
    old_file = Setting.get('favicon', '')
    if old_file:
        old_path = os.path.join(app.root_path, 'static', 'uploads', old_file)
        if os.path.exists(old_path):
            os.remove(old_path)
    
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'ico'
    filename = f"favicon_{int(datetime.now().timestamp())}.{ext}"
    file.save(os.path.join(app.root_path, 'static', 'uploads', filename))
    
    Setting.set('favicon', f'/static/uploads/{filename}')
    
    return jsonify({'success': True, 'url': f'/static/uploads/{filename}'})





@app.route('/admin/settings/delete-backup', methods=['POST'])
@login_required
def admin_settings_delete_backup():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    backup_name = data.get('backup_name')
    backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
    backup_path = os.path.join(backup_dir, backup_name)
    
    if not os.path.exists(backup_path):
        return jsonify({'error': 'فایل یافت نشد'}), 404
    
    try:
        os.remove(backup_path)
        return jsonify({'success': True, 'message': 'فایل بکاپ حذف شد'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/settings/download-backup/<backup_name>')
@login_required
def admin_settings_download_backup(backup_name):
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    backup_dir = os.path.join(os.path.dirname(__file__), 'backups')
    backup_path = os.path.join(backup_dir, backup_name)
    
    if not os.path.exists(backup_path):
        flash('فایل بکاپ یافت نشد', 'error')
        return redirect(url_for('admin_settings'))
    
    return send_file(backup_path, as_attachment=True, download_name=backup_name)


@app.route('/admin/settings/upload-header-logo-nav', methods=['POST'])
@login_required
def admin_upload_header_logo_nav():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': 'فایلی ارسال نشده'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'فایلی انتخاب نشده'}), 400
    
    old_file = Setting.get('header_logo', '')
    if old_file:
        old_path = os.path.join(app.root_path, 'static', 'uploads', old_file)
        if os.path.exists(old_path):
            os.remove(old_path)
    
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'png'
    filename = f"header_logo_nav_{int(datetime.now().timestamp())}.{ext}"
    file.save(os.path.join(app.root_path, 'static', 'uploads', filename))
    
    Setting.set('header_logo', f'/static/uploads/{filename}')
    
    return jsonify({'success': True, 'url': f'/static/uploads/{filename}'})


@app.route('/admin/settings/upload-index-bg', methods=['POST'])
@login_required
def admin_upload_index_bg():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': 'فایلی ارسال نشده'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'فایلی انتخاب نشده'}), 400
    
    old_file = Setting.get('index_bg', '')
    if old_file:
        old_path = os.path.join(app.root_path, 'static', 'uploads', old_file)
        if os.path.exists(old_path):
            os.remove(old_path)
    
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
    filename = f"index_bg_{int(datetime.now().timestamp())}.{ext}"
    file.save(os.path.join(app.root_path, 'static', 'uploads', filename))
    
    Setting.set('index_bg', f'/static/uploads/{filename}')
    
    return jsonify({'success': True, 'url': f'/static/uploads/{filename}'})
    
    
# ==================== تنظیمات ظاهر سامانه ====================

@app.route('/api/settings/current', methods=['GET'])
def api_settings_current():
    """API عمومی برای دریافت تنظیمات ظاهر سامانه (بدون نیاز به لاگین)"""
    return jsonify({
        'site_title': Setting.get('site_title', 'سامانه کارکرد آوان'),
        'site_subtitle': Setting.get('site_subtitle', 'آغازگر ورود اطلاعات نوین'),
        'site_logo_url': Setting.get('site_logo', ''),
        'footer_text': Setting.get('footer_text', 'تمامی حقوق سامانه آوان متعلق به اداره کل نظارت و پشتیبانی امور مشتریان می باشد. 1403 - 1404'),
        'login_logo_url': Setting.get('login_logo', ''),
        'login_bg_url': Setting.get('login_bg', ''),
        'favicon_url': Setting.get('favicon', ''),
        'login_title': Setting.get('login_title', 'ورود به سامانه'),
        'login_subtitle': Setting.get('login_subtitle', 'مدیریت یکپارچه کارکرد و پیشرفته ورود اطلاعات'),
        'index_bg_gradient': Setting.get('index_bg_gradient', 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)'),
        'index_bg_url': Setting.get('index_bg', ''),
        'login_bg_gradient': Setting.get('login_bg_gradient', ''),
        'header_title': Setting.get('header_title', 'سامانه آوان'),
        'header_bg_color': Setting.get('header_bg_color', '#1e293b'),
        'header_text_color': Setting.get('header_text_color', '#ffffff'),
        'header_logo_url': Setting.get('header_logo', '')
    })


@app.route('/admin/settings/current', methods=['GET'])
@login_required
def admin_settings_current():
    """API ادمین برای دریافت تنظیمات (با لاگین)"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    return jsonify({
        'site_title': Setting.get('site_title', 'سامانه کارکرد آوان'),
        'site_subtitle': Setting.get('site_subtitle', 'آغازگر ورود اطلاعات نوین'),
        'site_logo_url': Setting.get('site_logo', ''),
        'footer_text': Setting.get('footer_text', 'تمامی حقوق سامانه آوان متعلق به اداره کل نظارت و پشتیبانی امور مشتریان می باشد. 1403 - 1404'),
        'login_logo_url': Setting.get('login_logo', ''),
        'login_bg_url': Setting.get('login_bg', ''),
        'favicon_url': Setting.get('favicon', ''),
        'base_url': Setting.get('base_url', '10.86.109.219'),
        'port': Setting.get('port', '5000'),
        'login_title': Setting.get('login_title', 'ورود به سامانه'),
        'login_subtitle': Setting.get('login_subtitle', 'مدیریت یکپارچه کارکرد و پیشرفته ورود اطلاعات'),
        'header_title': Setting.get('header_title', 'سامانه آوان'),
        'header_bg_color': Setting.get('header_bg_color', '#1e293b'),
        'header_text_color': Setting.get('header_text_color', '#ffffff'),
        'header_logo_url': Setting.get('header_logo', ''),
        'index_bg_gradient': Setting.get('index_bg_gradient', 'linear-gradient(135deg, #667eea 0%, #764ba2 100%)'),
        'index_bg_url': Setting.get('index_bg', ''),
        'login_bg_gradient': Setting.get('login_bg_gradient', '')
    })


@app.route('/admin/settings/appearance', methods=['POST'])
@login_required
def admin_settings_appearance():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    print("=" * 50)
    print("📥 دریافت تنظیمات ظاهر:")
    print(f"  site_title: {data.get('site_title')}")
    print(f"  site_subtitle: {data.get('site_subtitle')}")
    print(f"  login_title: {data.get('login_title')}")
    print(f"  login_subtitle: {data.get('login_subtitle')}")
    print(f"  footer_text: {data.get('footer_text')}")
    print(f"  header_title: {data.get('header_title')}")
    print(f"  header_bg_color: {data.get('header_bg_color')}")
    print(f"  header_text_color: {data.get('header_text_color')}")
    print(f"  index_bg_gradient: {data.get('index_bg_gradient')}")
    print(f"  login_bg_gradient: {data.get('login_bg_gradient')}")
    print("=" * 50)
    
    Setting.set('site_title', data.get('site_title', 'سامانه کارکرد آوان'))
    Setting.set('site_subtitle', data.get('site_subtitle', 'آغازگر ورود اطلاعات نوین'))
    Setting.set('login_title', data.get('login_title', 'ورود به سامانه'))
    Setting.set('login_subtitle', data.get('login_subtitle', 'مدیریت یکپارچه کارکرد و پیشرفته ورود اطلاعات'))
    Setting.set('footer_text', data.get('footer_text', ''))
    Setting.set('header_title', data.get('header_title', 'سامانه آوان'))
    Setting.set('header_bg_color', data.get('header_bg_color', '#1e293b'))
    Setting.set('header_text_color', data.get('header_text_color', '#ffffff'))
    Setting.set('index_bg_gradient', data.get('index_bg_gradient', ''))
    Setting.set('login_bg_gradient', data.get('login_bg_gradient', ''))
    
    return jsonify({'success': True, 'message': 'تنظیمات ظاهر ذخیره شد'})


# ==================== آپلود تصاویر ====================

@app.route('/admin/settings/upload-header-logo', methods=['POST'])
@login_required
def admin_upload_header_logo():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': 'فایلی ارسال نشده'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'فایلی انتخاب نشده'}), 400
    
    # حذف فایل قبلی
    old_file = Setting.get('site_logo', '')
    if old_file:
        old_path = os.path.join(app.root_path, 'static', 'uploads', old_file.replace('/static/uploads/', ''))
        if os.path.exists(old_path):
            os.remove(old_path)
    
    # ذخیره فایل جدید
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'png'
    filename = f"header_logo_{int(datetime.now().timestamp())}.{ext}"
    file.save(os.path.join(app.root_path, 'static', 'uploads', filename))
    
    Setting.set('site_logo', f'/static/uploads/{filename}')
    
    return jsonify({'success': True, 'message': 'لوگو با موفقیت آپلود شد', 'url': f'/static/uploads/{filename}'})


@app.route('/admin/settings/reset-image', methods=['POST'])
@login_required
def admin_settings_reset_image():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    image_type = data.get('image_type')
    
    if image_type == 'headerLogo':
        old_file = Setting.get('site_logo', '')
        Setting.set('site_logo', '')
        if old_file:
            old_path = os.path.join(app.root_path, 'static', 'uploads', old_file.replace('/static/uploads/', ''))
            if os.path.exists(old_path):
                os.remove(old_path)
    elif image_type == 'loginLogo':
        old_file = Setting.get('login_logo', '')
        Setting.set('login_logo', '')
        if old_file:
            old_path = os.path.join(app.root_path, 'static', 'uploads', old_file.replace('/static/uploads/', ''))
            if os.path.exists(old_path):
                os.remove(old_path)
    elif image_type == 'loginBg':
        old_file = Setting.get('login_bg', '')
        Setting.set('login_bg', '')
        if old_file:
            old_path = os.path.join(app.root_path, 'static', 'uploads', old_file.replace('/static/uploads/', ''))
            if os.path.exists(old_path):
                os.remove(old_path)
    elif image_type == 'favicon':
        old_file = Setting.get('favicon', '')
        Setting.set('favicon', '')
        if old_file:
            old_path = os.path.join(app.root_path, 'static', 'uploads', old_file.replace('/static/uploads/', ''))
            if os.path.exists(old_path):
                os.remove(old_path)
    else:
        return jsonify({'error': 'نوع نامعتبر'}), 400
    
    return jsonify({'success': True, 'message': 'تصویر با موفقیت حذف شد'})


# ==================== مدیریت داده (حذف‌های گروهی) ====================
@app.route('/admin/settings/delete-personnel-by-period', methods=['POST'])
@login_required
def admin_settings_delete_personnel_by_period():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    period_id = data.get('period_id')
    
    if not period_id:
        return jsonify({'error': 'دوره انتخاب نشده است'}), 400
    
    period = WorkPeriod.query.get(period_id)
    if not period:
        return jsonify({'error': 'دوره یافت نشد'}), 404
    
    personnel_list = Personnel.query.filter_by(period_id=period_id).all()
    count = 0
    for p in personnel_list:
        PersonnelValue.query.filter_by(personnel_id=p.id).delete()
        PersonnelWorkStatus.query.filter_by(personnel_id=p.id).delete()
        db.session.delete(p)
        count += 1
    
    db.session.commit()
    return jsonify({'success': True, 'message': f'{count} پرسنل دوره {period.title} حذف شدند'})


@app.route('/admin/settings/clear-all-personnel', methods=['POST'])
@login_required
def admin_settings_clear_all_personnel():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    PersonnelValue.query.delete()
    PersonnelWorkStatus.query.delete()
    count = Personnel.query.delete()
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'{count} پرسنل حذف شدند'})


@app.route('/admin/settings/clear-all-users', methods=['POST'])
@login_required
def admin_settings_clear_all_users():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    users = User.query.filter(User.id != current_user.id).all()
    count = len(users)
    
    for user in users:
        DepartmentManager.query.filter_by(user_id=user.id).delete()
        UnitSupervisor.query.filter_by(user_id=user.id).delete()
        db.session.delete(user)
    
    db.session.commit()
    return jsonify({'success': True, 'message': f'{count} کاربر (غیر از ادمین) حذف شدند'})


@app.route('/admin/settings/reset-all-data', methods=['POST'])
@login_required
def admin_settings_reset_all_data():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        PersonnelValue.query.delete()
        PersonnelWorkStatus.query.delete()
        Personnel.query.delete()
        UnitSupervisor.query.delete()
        Unit.query.delete()
        DepartmentManager.query.delete()
        Department.query.delete()
        WorkPeriod.query.delete()
        DynamicField.query.delete()
        ApprovalRequest.query.delete()
        UnitPersonnelRequest.query.delete()
        
        users = User.query.filter(User.id != current_user.id).all()
        for user in users:
            db.session.delete(user)
        
        db.session.commit()
        
        init_default_admin()
        
        return jsonify({'success': True, 'message': 'همه داده‌ها با موفقیت پاک شدند. سامانه به حالت اولیه بازگشت.'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
        
        
# ==================== APIهای صندوق پیام (جدا از تیکت) ====================

@app.route('/api/messages/create', methods=['POST'])
@login_required
def api_message_create():
    """ارسال پیام جدید به صندوق پیام (بدون تبدیل به تیکت)"""
    data = request.get_json()
    title = data.get('title', '').strip()
    message = data.get('message', '').strip()
    receiver_id = data.get('receiver_id')
    
    if not title or not message:
        return jsonify({'error': 'عنوان و متن پیام الزامی است'}), 400
    
    if not receiver_id:
        return jsonify({'error': 'گیرنده مشخص نشده است'}), 400
    
    receiver = User.query.get(receiver_id)
    if not receiver:
        return jsonify({'error': 'کاربر گیرنده یافت نشد'}), 400
    
    msg = Message(
        title=title,
        message=message,
        sender_id=current_user.id,
        receiver_id=receiver_id
    )
    db.session.add(msg)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'پیام با موفقیت ارسال شد'})


@app.route('/admin/api/messages')
@login_required
def admin_api_messages():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # اگر جدول messages وجود ندارد، لیست خالی برگردان
    try:
        messages = Message.query.filter_by(receiver_id=current_user.id).order_by(Message.created_at.desc()).all()
    except:
        return jsonify([])
    
    result = []
    for m in messages:
        sender = User.query.get(m.sender_id)
        result.append({
            'id': m.id,
            'title': m.title,
            'message': m.message,
            'sender_name': sender.get_full_name() if sender else '-',
            'is_read': m.is_read,
            'created_at': m.created_at.strftime('%Y/%m/%d %H:%M')
        })
    
    return jsonify(result)

@app.route('/admin/api/message/<int:msg_id>')
@login_required
def admin_api_message_get(msg_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        msg = Message.query.get_or_404(msg_id)
    except:
        return jsonify({'error': 'پیام یافت نشد'}), 404
    
    sender = User.query.get(msg.sender_id)
    
    replies = []
    for r in msg.replies:
        user = User.query.get(r.user_id)
        replies.append({
            'id': r.id,
            'user_name': user.get_full_name() if user else '-',
            'message': r.message,
            'created_at': r.created_at.strftime('%Y/%m/%d %H:%M')
        })
    
    return jsonify({
        'id': msg.id,
        'title': msg.title,
        'message': msg.message,
        'sender_name': sender.get_full_name() if sender else '-',
        'created_at': msg.created_at.strftime('%Y/%m/%d %H:%M'),
        'replies': replies
    })
    
    

@app.route('/admin/api/message/<int:msg_id>/read', methods=['POST'])
@login_required
def admin_api_message_read(msg_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        msg = Message.query.get_or_404(msg_id)
        msg.is_read = True
        db.session.commit()
    except:
        return jsonify({'error': 'خطا'}), 404
    
    return jsonify({'success': True})
    
@app.route('/admin/api/message/<int:msg_id>/reply', methods=['POST'])
@login_required
def admin_api_message_reply(msg_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        msg = Message.query.get_or_404(msg_id)
    except:
        return jsonify({'error': 'پیام یافت نشد'}), 404
    
    data = request.get_json()
    message = data.get('message', '').strip()
    
    if not message:
        return jsonify({'error': 'متن پاسخ الزامی است'}), 400
    
    try:
        reply = MessageReply(
            message_id=msg_id,
            user_id=current_user.id,
            message=message
        )
        db.session.add(reply)
        db.session.commit()
    except:
        return jsonify({'error': 'خطا در ذخیره پاسخ'}), 500
    
    return jsonify({'success': True})
    
# ==================== APIهای صندوق پیام ====================

@app.route('/api/messages/send', methods=['POST'])
@login_required
def api_message_send():
    if current_user.role != 'admin':
        return jsonify({'error': 'فقط ادمین می‌تواند پیام ارسال کند'}), 403
    
    data = request.get_json()
    title = data.get('title', '').strip()
    message = data.get('message', '').strip()
    receiver_id = data.get('receiver_id')
    
    if not title or not message:
        return jsonify({'error': 'عنوان و متن پیام الزامی است'}), 400
    
    if not receiver_id:
        return jsonify({'error': 'گیرنده مشخص نشده است'}), 400
    
    receiver = User.query.get(receiver_id)
    if not receiver:
        return jsonify({'error': 'کاربر گیرنده یافت نشد'}), 404
    
    msg = Message(
        title=title,
        message=message,
        sender_id=current_user.id,
        receiver_id=receiver_id
    )
    db.session.add(msg)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'پیام با موفقیت ارسال شد'})


@app.route('/api/messages')
@login_required
def api_messages():
    messages = Message.query.filter_by(receiver_id=current_user.id).order_by(Message.created_at.desc()).all()
    
    result = []
    for m in messages:
        sender = User.query.get(m.sender_id)
        result.append({
            'id': m.id,
            'title': m.title,
            'message': m.message,
            'sender_name': sender.get_full_name() if sender else '-',
            'is_read': m.is_read,
            'created_at': m.created_at.strftime('%Y/%m/%d %H:%M')
        })
    
    return jsonify(result)


@app.route('/api/messages/<int:msg_id>')
@login_required
def api_message_get(msg_id):
    msg = Message.query.get_or_404(msg_id)
    
    if msg.receiver_id != current_user.id and msg.sender_id != current_user.id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    sender = User.query.get(msg.sender_id)
    
    replies = []
    for r in msg.replies:
        user = User.query.get(r.user_id)
        replies.append({
            'id': r.id,
            'user_name': user.get_full_name() if user else '-',
            'message': r.message,
            'created_at': r.created_at.strftime('%Y/%m/%d %H:%M')
        })
    
    if msg.receiver_id == current_user.id and not msg.is_read:
        msg.is_read = True
        db.session.commit()
    
    return jsonify({
        'id': msg.id,
        'title': msg.title,
        'message': msg.message,
        'sender_name': sender.get_full_name() if sender else '-',
        'created_at': msg.created_at.strftime('%Y/%m/%d %H:%M'),
        'replies': replies
    })


@app.route('/api/messages/<int:msg_id>/reply', methods=['POST'])
@login_required
def api_message_reply(msg_id):
    msg = Message.query.get_or_404(msg_id)
    
    if msg.receiver_id != current_user.id and msg.sender_id != current_user.id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    message = data.get('message', '').strip()
    
    if not message:
        return jsonify({'error': 'متن پاسخ الزامی است'}), 400
    
    reply = MessageReply(
        message_id=msg_id,
        user_id=current_user.id,
        message=message
    )
    db.session.add(reply)
    db.session.commit()
    
    return jsonify({'success': True})
    
@app.route('/admin/api/clear-logs', methods=['POST'])
@login_required
def admin_api_clear_logs():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # حذف همه رکوردهای PersonnelWorkStatus (لاگ‌های فعالیت)
    try:
        deleted_count = PersonnelWorkStatus.query.delete()
        db.session.commit()
        return jsonify({'success': True, 'message': f'{deleted_count} لاگ با موفقیت حذف شد'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/admin/personnel/copy-to-period', methods=['POST'])
@login_required
def admin_personnel_copy_to_period():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        source_period_id = data.get('source_period_id')
        target_period_id = data.get('target_period_id')
        duplicate_action = data.get('duplicate_action', 'skip')
        personnel_ids = data.get('personnel_ids', [])
        
        if not source_period_id or not target_period_id:
            return jsonify({'error': 'دوره مبدأ و مقصد الزامی است'}), 400
        
        # دریافت پرسنل
        if personnel_ids and len(personnel_ids) > 0:
            source_personnel = Personnel.query.filter(
                Personnel.id.in_(personnel_ids),
                Personnel.is_deleted == False
            ).all()
        else:
            source_personnel = Personnel.query.filter_by(
                period_id=source_period_id,
                is_deleted=False
            ).all()
        
        success_count = 0
        skip_count = 0
        
        for source_p in source_personnel:
            # بررسی وجود در دوره مقصد
            existing = Personnel.query.filter_by(
                national_code=source_p.national_code,
                period_id=target_period_id,
                is_deleted=False
            ).first()
            
            if existing:
                if duplicate_action == 'skip':
                    skip_count += 1
                    continue
                else:
                    # حذف قبلی
                    PersonnelValue.query.filter_by(personnel_id=existing.id).delete()
                    PersonnelWorkStatus.query.filter_by(personnel_id=existing.id).delete()
                    db.session.delete(existing)
                    db.session.flush()
            
            # ایجاد پرسنل جدید
            new_personnel = Personnel(
                national_code=source_p.national_code,
                first_name=source_p.first_name or '',
                last_name=source_p.last_name or '',
                phone=source_p.phone or '',
                position=source_p.position or '',
                department_id=source_p.department_id,
                unit_id=source_p.unit_id,
                period_id=target_period_id
            )
            db.session.add(new_personnel)
            db.session.flush()
            
            # کپی مقادیر فیلدهای داینامیک
            for val in PersonnelValue.query.filter_by(
                personnel_id=source_p.id,
                period_id=source_period_id
            ).all():
                new_val = PersonnelValue(
                    personnel_id=new_personnel.id,
                    field_id=val.field_id,
                    period_id=target_period_id,
                    value_text=val.value_text,
                    value_number=val.value_number,
                    value_date=val.value_date
                )
                db.session.add(new_val)
            
            success_count += 1
        
        db.session.commit()
        
        message = f"✅ {success_count} پرسنل با موفقیت کپی شدند."
        if skip_count > 0:
            message += f" ⏭️ {skip_count} پرسنل تکراری نادیده گرفته شدند."
        
        return jsonify({'success': True, 'message': message})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
        
        
@app.route('/admin/fix-unique-constraint', methods=['GET'])
@login_required
def fix_unique_constraint():
    """حذف محدودیت unique از فیلد national_code در جدول personnel"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        # حذف ایندکس unique
        with db.engine.connect() as conn:
            # پیدا کردن و حذف ایندکس unique
            result = conn.execute("SELECT name FROM sqlite_master WHERE type='index' AND tbl_name='personnel'")
            indexes = result.fetchall()
            for idx in indexes:
                if 'national_code' in idx[0] or 'unique' in idx[0].lower():
                    conn.execute(f"DROP INDEX IF EXISTS {idx[0]}")
                    print(f" Dropped index: {idx[0]}")
            
            # بازسازی جدول بدون unique constraint
            conn.execute("PRAGMA foreign_keys=OFF")
            conn.execute("CREATE TABLE personnel_temp AS SELECT * FROM personnel")
            conn.execute("DROP TABLE personnel")
            conn.execute("""
                CREATE TABLE personnel (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    national_code VARCHAR(10) NOT NULL,
                    first_name VARCHAR(50),
                    last_name VARCHAR(50),
                    phone VARCHAR(20),
                    position VARCHAR(100),
                    hire_date VARCHAR(20),
                    department_id INTEGER NOT NULL REFERENCES departments(id),
                    unit_id INTEGER NOT NULL REFERENCES units(id),
                    period_id INTEGER,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    is_deleted BOOLEAN DEFAULT 0
                )
            """)
            conn.execute("INSERT INTO personnel SELECT * FROM personnel_temp")
            conn.execute("DROP TABLE personnel_temp")
            conn.execute("PRAGMA foreign_keys=ON")
            
        return jsonify({
            'success': True,
            'message': 'محدودیت یکتایی کد ملی با موفقیت حذف شد. اکنون می‌توانید پرسنل را بین دوره‌ها کپی کنید.'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
        
        
@app.route('/admin/personnel/simple-copy', methods=['POST'])
@login_required
def admin_personnel_simple_copy():
    """کپی ساده پرسنل - با حذف و ایجاد مجدد"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        source_period_id = data.get('source_period_id')
        target_period_id = data.get('target_period_id')
        
        source_personnel = Personnel.query.filter_by(period_id=source_period_id, is_deleted=False).all()
        
        success = 0
        for p in source_personnel:
            try:
                # حذف پرسنل تکراری در دوره مقصد (اگر وجود داشته باشد)
                existing = Personnel.query.filter_by(
                    national_code=p.national_code,
                    period_id=target_period_id
                ).first()
                if existing:
                    # حذف وابستگی‌ها
                    PersonnelValue.query.filter_by(personnel_id=existing.id, period_id=target_period_id).delete()
                    PersonnelWorkStatus.query.filter_by(personnel_id=existing.id, period_id=target_period_id).delete()
                    db.session.delete(existing)
                    db.session.commit()
                
                # ایجاد پرسنل جدید
                new_p = Personnel(
                    national_code=p.national_code,
                    first_name=p.first_name,
                    last_name=p.last_name,
                    phone=p.phone,
                    position=p.position,
                    department_id=p.department_id,
                    unit_id=p.unit_id,
                    period_id=target_period_id
                )
                db.session.add(new_p)
                db.session.flush()
                
                # کپی مقادیر فیلدها
                for val in PersonnelValue.query.filter_by(personnel_id=p.id, period_id=source_period_id).all():
                    new_val = PersonnelValue(
                        personnel_id=new_p.id,
                        field_id=val.field_id,
                        period_id=target_period_id,
                        value_text=val.value_text,
                        value_number=val.value_number,
                        value_date=val.value_date
                    )
                    db.session.add(new_val)
                
                success += 1
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{success} پرسنل با موفقیت کپی شدند',
            'total': len(source_personnel),
            'copied': success
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
@app.route('/admin/fix-names', methods=['GET'])
@login_required
def fix_names():
    """انتقال نام و نام خانوادگی از PersonnelValue به Personnel"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        updated = 0
        # پیدا کردن فیلدهای داینامیک نام و نام خانوادگی
        name_field = DynamicField.query.filter_by(title='نام', is_active=True).first()
        family_field = DynamicField.query.filter_by(title='نام خانوادگی', is_active=True).first()
        
        if not name_field or not family_field:
            return jsonify({'error': 'فیلدهای نام یا نام خانوادگی یافت نشد'}), 404
        
        # دریافت همه پرسنل
        all_personnel = Personnel.query.filter_by(is_deleted=False).all()
        
        for p in all_personnel:
            # دریافت مقدار نام از PersonnelValue
            name_pv = PersonnelValue.query.filter_by(
                personnel_id=p.id,
                field_id=name_field.id,
                period_id=p.period_id
            ).first()
            
            # دریافت مقدار نام خانوادگی از PersonnelValue
            family_pv = PersonnelValue.query.filter_by(
                personnel_id=p.id,
                field_id=family_field.id,
                period_id=p.period_id
            ).first()
            
            updated_flag = False
            
            if name_pv and name_pv.value_text and (not p.first_name or p.first_name == ''):
                p.first_name = name_pv.value_text
                updated_flag = True
                
            if family_pv and family_pv.value_text and (not p.last_name or p.last_name == ''):
                p.last_name = family_pv.value_text
                updated_flag = True
            
            if updated_flag:
                updated += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{updated} پرسنل با موفقیت به‌روزرسانی شدند.',
            'updated': updated
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# اضافه کنید به بخش مدیریت دوره‌ها در app.py
@app.route('/admin/periods/update-order', methods=['POST'])
@login_required
def admin_periods_update_order():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        data = request.get_json()
        orders = data.get('orders', [])
        
        for item in orders:
            period = WorkPeriod.query.get(item['id'])
            if period:
                period.display_order = item['display_order']
        
        db.session.commit()
        return jsonify({'success': True, 'message': 'ترتیب با موفقیت ذخیره شد'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
        
# ==================== APIهای مدیریت لاگ ====================

@app.route('/admin/api/add-log', methods=['POST'])
@login_required
def admin_api_add_log():
    """ذخیره لاگ جدید در دیتابیس"""
    if current_user.role not in ['admin', 'org_manager', 'dept_manager']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    message = data.get('message', '')
    badge = data.get('badge', '')
    log_type = data.get('type', 'info')
    
    log = ActivityLog(
        user_id=current_user.id,
        user_name=current_user.get_full_name(),
        message=message,
        badge=badge,
        log_type=log_type
    )
    db.session.add(log)
    db.session.commit()
    
    return jsonify({'success': True})


@app.route('/admin/api/get-logs', methods=['GET'])
@login_required
def admin_api_get_logs():
    """دریافت لاگ‌ها از دیتابیس"""
    if current_user.role not in ['admin', 'org_manager', 'dept_manager']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    limit = request.args.get('limit', 200, type=int)
    logs = ActivityLog.query.order_by(ActivityLog.created_at.desc()).limit(limit).all()
    
    result = []
    for log in logs:
        result.append({
            'id': log.id,
            'user_name': log.user_name or '-',
            'message': log.message,
            'badge': log.badge or 'فعالیت',
            'type': log.log_type,
            'time': log.get_jalali_date()
        })
    
    return jsonify({'logs': result})


@app.route('/admin/api/delete-logs', methods=['POST'])
@login_required
def admin_api_delete_logs():
    """حذف لاگ‌های انتخاب شده (فقط ادمین)"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    log_ids = data.get('ids', [])
    
    if not log_ids:
        return jsonify({'error': 'هیچ لاگی انتخاب نشده است'}), 400
    
    deleted_count = ActivityLog.query.filter(ActivityLog.id.in_(log_ids)).delete(synchronize_session=False)
    db.session.commit()
    
    return jsonify({'success': True, 'deleted_count': deleted_count})


@app.route('/admin/api/clear-all-logs', methods=['POST'])
@login_required
def admin_api_clear_all_logs():
    """پاک کردن همه لاگ‌ها (فقط ادمین)"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    deleted_count = ActivityLog.query.delete()
    db.session.commit()
    
    return jsonify({'success': True, 'deleted_count': deleted_count})
    
    
@app.route('/admin/create-logs-table')
@login_required
def create_logs_table():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        db.create_all()  # این کار جدول جدید را ایجاد می‌کند
        return jsonify({'success': True, 'message': 'جدول لاگ‌ها ایجاد شد'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
@app.route('/admin/add-deadline-column')
@login_required
def add_deadline_column():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    import sqlite3
    db_path = os.path.join(os.path.dirname(__file__), 'instance', 'avan_system.db')
    
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("PRAGMA table_info(work_periods)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'deadline' not in columns:
            cursor.execute("ALTER TABLE work_periods ADD COLUMN deadline VARCHAR(20)")
            conn.commit()
            return jsonify({'success': True, 'message': 'ستون ددلاین اضافه شد'})
        else:
            return jsonify({'success': True, 'message': 'ستون ددلاین قبلاً وجود دارد'})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
# ==================== APIهای مدیریت مدارک ====================

# آپلود مدرک
@app.route('/api/documents/upload', methods=['POST'])
@login_required
def api_upload_document():
    try:
        doc_type = request.form.get('doc_type')
        if doc_type not in ['شناسنامه', 'مدرک تحصیلی', 'کارت ملی', 'کارت پایان خدمت', 'سوء پیشینه', 'سایر']:
            return jsonify({'error': 'نوع مدرک نامعتبر است'}), 400
        
        if 'file' not in request.files:
            return jsonify({'error': 'فایلی ارسال نشده'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'فایلی انتخاب نشده'}), 400
        
        # بررسی پسوند
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if ext not in ['jpg', 'jpeg', 'png']:
            return jsonify({'error': 'فقط فایل‌های JPG و PNG مجاز هستند'}), 400
        
        # بررسی حجم (حداکثر 5MB)
        file.seek(0, 2)
        file_size = file.tell() // 1024  # کیلوبایت
        file.seek(0)
        if file_size > 5120:
            return jsonify({'error': 'حجم فایل نباید بیشتر از 5 مگابایت باشد'}), 400
        
        # ایجاد نام یکتا
        timestamp = int(datetime.now().timestamp())
        filename = f"doc_{current_user.id}_{doc_type}_{timestamp}.{ext}"
        
        # ذخیره فایل
        upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'documents')
        os.makedirs(upload_dir, exist_ok=True)
        file.save(os.path.join(upload_dir, filename))
        
        # حذف مدرک قبلی از همین نوع (در صورت وجود)
        old_doc = UserDocument.query.filter_by(user_id=current_user.id, doc_type=doc_type).first()
        if old_doc:
            old_path = os.path.join(upload_dir, old_doc.doc_filename)
            if os.path.exists(old_path):
                os.remove(old_path)
            db.session.delete(old_doc)
        
        # ذخیره در دیتابیس
        doc = UserDocument(
            user_id=current_user.id,
            doc_type=doc_type,
            doc_filename=filename,
            doc_original_name=file.filename,
            doc_size=file_size,
            status='pending'
        )
        db.session.add(doc)
        db.session.commit()
        
        # اعلان به مدیر منابع انسانی
        hr_managers = User.query.filter_by(role='hr_manager').all()
        for hr in hr_managers:
            send_workflow_notification(
                to_user_id=hr.id,
                title=f"📎 مدرک جدید از {current_user.get_full_name()}",
                message=f"نوع مدرک: {doc_type}\nکاربر: {current_user.get_full_name()}\nلطفاً بررسی کنید.",
                link="/hr-manager/dashboard"
            )
        
        return jsonify({'success': True, 'message': 'مدرک با موفقیت آپلود شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
        return jsonify({'error': str(e)}), 500




# ==================== پنل مدیر منابع انسانی (HR Manager) ====================
@app.route('/hr-manager/dashboard')
@login_required
def hr_manager_dashboard():
    if current_user.role not in ['hr_manager', 'admin']:
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    today_date = jdatetime.datetime.now().strftime('%Y/%m/%d')
    return render_template('hr_manager/dashboard.html', today_date=today_date)


@app.route('/hr-manager/api/users', methods=['GET'])
@login_required
def hr_manager_api_users():
    if current_user.role not in ['hr_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    search = request.args.get('search', '')
    
    query = User.query
    if search:
        query = query.filter(
            db.or_(
                User.national_code.contains(search),
                User.first_name.contains(search),
                User.last_name.contains(search),
                User.personnel_code.contains(search)
            )
        )
    
    pagination = query.order_by(User.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'users': [{
            'id': u.id,
            'national_code': u.national_code,
            'first_name': u.first_name,
            'last_name': u.last_name,
            'full_name': u.get_full_name(),
            'phone': u.phone or '',
            'personnel_code': u.personnel_code or '',
            'role': u.role,
            'role_persian': u.get_role_persian(),
            'is_active': u.is_active,
            'is_approved': u.is_approved,
            'created_at': u.get_jalali_created_date(),
            'last_login': u.last_login.strftime('%Y/%m/%d %H:%M') if u.last_login else 'هرگز'
        } for u in pagination.items],
        'total': pagination.total,
        'page': page,
        'per_page': per_page,
        'pages': pagination.pages
    })


@app.route('/hr-manager/api/users/<int:user_id>/edit', methods=['PUT'])
@login_required
def hr_manager_api_user_edit(user_id):
    if current_user.role not in ['hr_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    user = User.query.get_or_404(user_id)
    data = request.get_json()
    
    user.first_name = data.get('first_name', user.first_name)
    user.last_name = data.get('last_name', user.last_name)
    user.phone = data.get('phone', user.phone)
    user.personnel_code = data.get('personnel_code', user.personnel_code)
    
    db.session.commit()
    invalidate_cache()
    
    return jsonify({'success': True, 'message': 'اطلاعات کاربر با موفقیت به‌روزرسانی شد'})


@app.route('/hr-manager/api/documents/<int:user_id>', methods=['GET'])
@login_required
def hr_manager_api_documents(user_id):
    if current_user.role not in ['hr_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    docs = UserDocument.query.filter_by(user_id=user_id).all()
    user = User.query.get(user_id)
    
    return jsonify({
        'user_name': user.get_full_name() if user else '',
        'user_national_code': user.national_code if user else '',
        'documents': [{
            'id': d.id,
            'doc_type': d.doc_type,
            'doc_filename': d.doc_filename,
            'doc_original_name': d.doc_original_name,
            'doc_size': d.doc_size,
            'status': d.status,
            'admin_note': d.admin_note,
            'created_at': d.created_at.strftime('%Y/%m/%d %H:%M')
        } for d in docs]
    })


@app.route('/hr-manager/api/documents/<int:doc_id>/review', methods=['POST'])
@login_required
def hr_manager_api_review_document(doc_id):
    if current_user.role not in ['hr_manager', 'admin']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    doc = UserDocument.query.get_or_404(doc_id)
    data = request.get_json()
    status = data.get('status')
    note = data.get('note', '')
    
    if status not in ['approved', 'rejected']:
        return jsonify({'error': 'وضعیت نامعتبر است'}), 400
    
    doc.status = status
    doc.admin_note = note
    doc.reviewed_by = current_user.id
    doc.reviewed_at = datetime.now()
    db.session.commit()
    
    send_workflow_notification(
        to_user_id=doc.user_id,
        title=f"📎 وضعیت مدرک {doc.doc_type}",
        message=f"مدرک شما توسط {current_user.get_full_name()} {('تایید' if status == 'approved' else 'رد')} شد.\nتوضیحات: {note or '-'}",
        link="/profile"
    )
    
    return jsonify({'success': True, 'message': 'وضعیت مدرک به‌روزرسانی شد'})


@app.route('/api/documents/my', methods=['GET'])
@login_required
def api_my_documents():
    """دریافت مدارک کاربر جاری"""
    docs = UserDocument.query.filter_by(user_id=current_user.id).order_by(UserDocument.created_at.desc()).all()
    return jsonify([{
        'id': d.id,
        'doc_type': d.doc_type,
        'doc_filename': d.doc_filename,
        'doc_original_name': d.doc_original_name,
        'doc_size': d.doc_size,
        'status': d.status,
        'admin_note': d.admin_note,
        'created_at': d.created_at.strftime('%Y/%m/%d %H:%M')
    } for d in docs])


@app.route('/api/documents/<int:doc_id>/delete', methods=['DELETE'])
@login_required
def api_delete_document(doc_id):
    """حذف مدرک توسط کاربر"""
    doc = UserDocument.query.get_or_404(doc_id)
    
    if doc.user_id != current_user.id and current_user.role not in ['admin', 'hr_manager']:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'documents')
    file_path = os.path.join(upload_dir, doc.doc_filename)
    if os.path.exists(file_path):
        os.remove(file_path)
    
    db.session.delete(doc)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'مدرک با موفقیت حذف شد'})

# ==================== آپلود عکس پروفایل ====================

# ==================== آپلود عکس پروفایل ====================

@app.route('/profile/upload-avatar', methods=['POST'])
@login_required
def upload_avatar():
    try:
        if 'avatar' not in request.files:
            return jsonify({'error': 'فایلی ارسال نشده'}), 400
        
        file = request.files['avatar']
        if file.filename == '':
            return jsonify({'error': 'فایلی انتخاب نشده'}), 400
        
        # بررسی پسوند
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if ext not in ['jpg', 'jpeg', 'png', 'gif']:
            return jsonify({'error': 'فقط فایل‌های تصویری مجاز هستند'}), 400
        
        # بررسی حجم (حداکثر 5MB)
        file.seek(0, 2)
        file_size_mb = file.tell() // (1024 * 1024)
        file.seek(0)
        if file_size_mb > 5:
            return jsonify({'error': 'حجم فایل نباید بیشتر از 5 مگابایت باشد'}), 400
        
        # ========== ذخیره با کد ملی ==========
        upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'avatars')
        os.makedirs(upload_dir, exist_ok=True)
        
        # نام فایل = کد ملی کاربر + پسوند
        filename = f"{current_user.national_code}.{ext}"
        file_path = os.path.join(upload_dir, filename)
        
        # ========== حذف عکس قبلی ==========
        # اگر عکس قبلی با نام دیگری وجود داشت، حذفش کن
        if current_user.profile_picture:
            old_path = os.path.join(upload_dir, current_user.profile_picture)
            if os.path.exists(old_path) and old_path != file_path:
                os.remove(old_path)
        
        # ذخیره فایل جدید (جایگزینی)
        file.save(file_path)
        
        # به‌روزرسانی در دیتابیس
        current_user.profile_picture = filename
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'url': f'/static/uploads/avatars/{filename}',
            'filename': filename
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error uploading avatar: {e}")
        return jsonify({'error': str(e)}), 500
        
def get_unit_supervisor(user_id):
    """دریافت سرپرست واحد یک کاربر"""
    # ابتدا پرسنل مرتبط با این کاربر را پیدا کن
    personnel = Personnel.query.filter_by(national_code=user_id).first()
    if not personnel:
        # اگر کاربر مستقیماً با personnel مرتبط نیست، از طریق national_code جستجو کن
        personnel = Personnel.query.filter_by(national_code=user_id).first()
    
    if not personnel:
        return None
    
    # پیدا کردن سرپرست واحد
    unit_supervisor = UnitSupervisor.query.filter_by(unit_id=personnel.unit_id).first()
    if unit_supervisor:
        return User.query.get(unit_supervisor.user_id)
    return None


def send_request_notification(to_user_id, request_type, request_id, message):
    """ارسال اعلان برای درخواست جدید"""
    if not to_user_id:
        return
    
    type_persian = {
        'overtime': 'اضافه کار',
        'deficiency': 'ثبت نقص',
        'daily_mission': 'ماموریت روزانه',
        'official_mission': 'ماموریت اداری',
        'arbaeen': 'سفر اربعین',
        'annual_leave': 'مرخصی روزانه',
        'hourly_leave': 'مرخصی ساعتی'
    }.get(request_type, 'درخواست')
    
    notif = Notification(
        user_id=to_user_id,
        notification_type='request',
        title=f"📋 درخواست جدید {type_persian}",
        message=message[:200],
        link=f"/requests/{request_id}",
        is_read=False
    )
    db.session.add(notif)
    db.session.commit()
    

@app.route('/admin/users/<int:user_id>/edit', methods=['PUT'])
@login_required
def admin_user_edit(user_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    try:
        user = User.query.get_or_404(user_id)
        data = request.get_json()
        
        print(f"📥 داده دریافتی برای ویرایش کاربر {user_id}:", data)
        
        # به‌روزرسانی اطلاعات پایه کاربر
        user.first_name = data.get('first_name', user.first_name)
        user.last_name = data.get('last_name', user.last_name)
        user.phone = data.get('phone', user.phone)
        user.role = data.get('role', user.role)
        user.personnel_code = data.get('personnel_code', user.personnel_code)
        
        new_password = data.get('password')
        if new_password and new_password.strip():
            user.set_password(new_password.strip())
        
        db.session.commit()
        
        # ========== ذخیره انتصاب ==========
        unit_id = data.get('assign_unit_id')
        start_date = data.get('assign_start_date')
        
        print(f"📌 داده‌های انتصاب - unit_id: {unit_id}, start_date: {start_date}")
        
        if unit_id and start_date:
            # پیدا کردن یا ایجاد پرسنل
            personnel = Personnel.query.filter_by(
                national_code=user.national_code, 
                is_deleted=False
            ).first()
            
            # پیدا کردن واحد و department_id
            unit_obj = Unit.query.get(unit_id)
            department_id = unit_obj.department_id if unit_obj else None
            
            if not personnel:
                # ایجاد پرسنل جدید
                personnel = Personnel(
                    national_code=user.national_code,
                    first_name=user.first_name,
                    last_name=user.last_name,
                    phone=user.phone or '',
                    department_id=department_id,
                    unit_id=unit_id,
                    period_id=None
                )
                db.session.add(personnel)
                db.session.commit()
                print(f"✅ پرسنل جدید ایجاد شد: ID {personnel.id}")
            
            # به‌روزرسانی اطلاعات پرسنل
            personnel.first_name = user.first_name
            personnel.last_name = user.last_name
            personnel.phone = user.phone or ''
            personnel.unit_id = unit_id
            personnel.department_id = department_id
            db.session.commit()
            
            # پایان دادن به انتصاب فعال قبلی
            old_assignment = PersonnelAssignment.query.filter_by(
                personnel_id=personnel.id, 
                is_active=True
            ).first()
            
            if old_assignment:
                old_assignment.is_active = False
                old_assignment.end_date = start_date
                db.session.commit()
                print(f"✅ انتصاب قبلی غیرفعال شد: ID {old_assignment.id}")
            
            # ایجاد انتصاب جدید
            new_assignment = PersonnelAssignment(
                personnel_id=personnel.id,
                unit_id=unit_id,
                period_id=None,
                start_date=start_date,
                assignment_type='initial',
                description=data.get('assign_description', 'انتصاب از طریق مدیریت کاربران'),
                created_by=current_user.id,
                is_active=True
            )
            db.session.add(new_assignment)
            db.session.commit()
            print(f"✅ انتصاب جدید ایجاد شد: ID {new_assignment.id}")
        else:
            print("⚠️ اطلاعات انتصاب کامل نیست")
        
        return jsonify({'success': True, 'message': 'اطلاعات کاربر با موفقیت به‌روزرسانی شد'})
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ خطا در ویرایش کاربر: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'خطا در سرور: {str(e)}'}), 500


# ==================== مسیرهای صفحات درخواست‌ها ====================

@app.route('/requests')
@login_required
def my_requests_page():
    """لیست درخواست‌های من"""
    if current_user.role not in ['subordinate', 'unit_supervisor']:
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    return render_template('user/my_requests.html')


@app.route('/requests/overtime/new')
@login_required
def overtime_request_page():
    """فرم ثبت درخواست اضافه کار ساعتی"""
    if current_user.role != 'subordinate':
        flash('فقط کاربران عادی می‌توانند درخواست ثبت کنند', 'error')
        return redirect(url_for('dashboard'))
    return render_template('requests/overtime.html')

@app.route('/requests/hourly-leave/new')
@login_required
def hourly_leave_request_page():
    if current_user.role != 'subordinate':
        flash('فقط کاربران عادی می‌توانند درخواست ثبت کنند', 'error')
        return redirect(url_for('dashboard'))
    return render_template('requests/hourly_leave.html')

@app.route('/requests/daily-mission/new')
@login_required
def daily_mission_request_page():
    if current_user.role != 'subordinate':
        flash('فقط کاربران عادی می‌توانند درخواست ثبت کنند', 'error')
        return redirect(url_for('dashboard'))
    return render_template('requests/daily_mission.html')
    
    
@app.route('/requests/official-mission/new')
@login_required
def official_mission_request_page():
    if current_user.role != 'subordinate':
        flash('فقط کاربران عادی می‌توانند درخواست ثبت کنند', 'error')
        return redirect(url_for('dashboard'))
    return render_template('requests/official_mission.html')

@app.route('/requests/arbaeen/new')
@login_required
def arbaeen_request_page():
    if current_user.role != 'subordinate':
        flash('فقط کاربران عادی می‌توانند درخواست ثبت کنند', 'error')
        return redirect(url_for('dashboard'))
    return render_template('requests/arbaeen.html')
    
    
    
@app.route('/unit-supervisor/requests')
@login_required
def unit_supervisor_requests_page():
    """لیست درخواست‌های پرسنل برای سرپرست واحد"""
    if current_user.role != 'unit_supervisor':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    return render_template('unit_supervisor/requests_list.html')
    
    
@app.route('/api/supervisor/requests')
@login_required
def api_supervisor_requests():
    """API دریافت لیست درخواست‌های پرسنل زیرمجموعه برای سرپرست واحد"""
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '')
    request_type = request.args.get('type', '')
    requester_id = request.args.get('requester_id', type=int)
    
    # دریافت واحدهای تحت سرپرستی
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    unit_ids = [u.id for u in supervised_units]
    
    # دریافت پرسنل این واحدها
    personnel_list = Personnel.query.filter(Personnel.unit_id.in_(unit_ids)).all()
    
    # دریافت کاربران مرتبط با این پرسنل
    user_ids = []
    for p in personnel_list:
        user = User.query.filter_by(national_code=p.national_code).first()
        if user:
            user_ids.append(user.id)
    
    query = Request.query.filter(Request.requester_id.in_(user_ids))
    
    if status:
        query = query.filter_by(status=status)
    if request_type:
        query = query.filter_by(request_type=request_type)
    if requester_id:
        query = query.filter_by(requester_id=requester_id)
    
    pagination = query.order_by(Request.request_date.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    # آمار
    stats = {
        'total': Request.query.filter(Request.requester_id.in_(user_ids)).count(),
        'pending': Request.query.filter(Request.requester_id.in_(user_ids), Request.status=='pending_unit').count(),
        'approved': Request.query.filter(Request.requester_id.in_(user_ids), Request.status=='approved').count(),
        'rejected': Request.query.filter(Request.requester_id.in_(user_ids), Request.status=='rejected').count(),
        'revision': Request.query.filter(Request.requester_id.in_(user_ids), Request.status=='revision').count()
    }
    
    # لیست درخواست‌دهندگان برای فیلتر
    requesters = db.session.query(User.id, User.first_name, User.last_name).filter(User.id.in_(user_ids)).all()
    requesters_list = [{'id': r.id, 'name': f"{r.first_name} {r.last_name}"} for r in requesters]
    
    return jsonify({
        'requests': [r.to_dict() for r in pagination.items],
        'total': pagination.total,
        'page': page,
        'per_page': per_page,
        'pages': pagination.pages,
        'stats': stats,
        'requesters': requesters_list
    })
    
@app.route('/requests/<int:request_id>/edit')
@login_required
def edit_request_page(request_id):
    """صفحه ویرایش درخواست (فقط برای وضعیت revision)"""
    if current_user.role != 'subordinate':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    req = Request.query.get_or_404(request_id)
    if req.requester_id != current_user.id:
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    if req.status != 'revision':
        flash('این درخواست قابل ویرایش نیست', 'error')
        return redirect(url_for('my_requests_page'))
    
    return render_template('requests/edit_request.html')
    
    
# مسیرهای صفحات درخواست‌ها
@app.route('/requests/deficiency/new')
@login_required
def deficiency_request_page():
    if current_user.role != 'subordinate':
        flash('فقط کاربران عادی می‌توانند درخواست ثبت کنند', 'error')
        return redirect(url_for('dashboard'))
    return render_template('requests/deficiency.html')

@app.route('/requests/annual-leave/new')
@login_required
def annual_leave_request_page():
    if current_user.role != 'subordinate':
        flash('فقط کاربران عادی می‌توانند درخواست ثبت کنند', 'error')
        return redirect(url_for('dashboard'))
    return render_template('requests/annual_leave.html')


@app.route('/api/requests')
@login_required
def api_get_requests():
    """API دریافت لیست درخواست‌ها"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '')
    request_type = request.args.get('type', '')
    
    if current_user.role == 'unit_supervisor':
        # سرپرست واحد: درخواست‌های پرسنل زیرمجموعه
        # دریافت واحدهای تحت سرپرستی
        supervised_units = db.session.query(Unit).join(
            UnitSupervisor, UnitSupervisor.unit_id == Unit.id
        ).filter(UnitSupervisor.user_id == current_user.id).all()
        unit_ids = [u.id for u in supervised_units]
        
        # دریافت پرسنل این واحدها
        personnel_list = Personnel.query.filter(Personnel.unit_id.in_(unit_ids)).all()
        personnel_ids = [p.id for p in personnel_list]
        
        # دریافت کاربران مرتبط با این پرسنل (با کد ملی)
        user_ids = []
        for p in personnel_list:
            user = User.query.filter_by(national_code=p.national_code).first()
            if user:
                user_ids.append(user.id)
        
        query = Request.query.filter(Request.requester_id.in_(user_ids))
        
    else:
        # کاربر عادی: فقط درخواست‌های خودش
        query = Request.query.filter_by(requester_id=current_user.id)
    
    if status:
        query = query.filter_by(status=status)
    if request_type:
        query = query.filter_by(request_type=request_type)
    
    pagination = query.order_by(Request.request_date.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'requests': [r.to_dict() for r in pagination.items],
        'total': pagination.total,
        'page': page,
        'per_page': per_page,
        'pages': pagination.pages
    })


@app.route('/api/requests/<int:request_id>')
@login_required
def api_get_request(request_id):
    """API دریافت جزئیات یک درخواست"""
    req = Request.query.get_or_404(request_id)
    
    # بررسی دسترسی
    if current_user.role != 'admin' and current_user.id != req.requester_id and current_user.id != req.unit_supervisor_id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    return jsonify(req.to_dict())


@app.route('/api/requests/<int:request_id>/approve', methods=['POST'])
@login_required
def api_approve_request(request_id):
    """تأیید درخواست توسط سرپرست واحد"""
    req = Request.query.get_or_404(request_id)
    
    # فقط سرپرست واحد می‌تواند تأیید کند
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # بررسی اینکه این درخواست متعلق به واحد تحت سرپرستی است
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    unit_ids = [u.id for u in supervised_units]
    
    personnel = Personnel.query.filter_by(national_code=req.requester.national_code).first()
    if not personnel or personnel.unit_id not in unit_ids:
        return jsonify({'error': 'شما به این درخواست دسترسی ندارید'}), 403
    
    if req.status != 'pending_unit':
        return jsonify({'error': 'این درخواست قبلاً بررسی شده است'}), 400
    
    req.status = 'approved'
    req.reviewed_at = datetime.now()
    db.session.commit()
    
    # ارسال اعلان به کاربر
    send_workflow_notification(
        to_user_id=req.requester_id,
        title=f"✅ درخواست {req.get_request_type_persian()} شما تأیید شد",
        message=f"درخواست شما توسط {current_user.get_full_name()} تأیید شد.",
        link=f"/requests/{req.id}"
    )
    
    return jsonify({'success': True, 'message': 'درخواست با موفقیت تأیید شد'})


@app.route('/api/requests/<int:request_id>/reject', methods=['POST'])
@login_required
def api_reject_request(request_id):
    """رد درخواست توسط سرپرست واحد"""
    req = Request.query.get_or_404(request_id)
    
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # بررسی دسترسی (همانند approve)
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    unit_ids = [u.id for u in supervised_units]
    
    personnel = Personnel.query.filter_by(national_code=req.requester.national_code).first()
    if not personnel or personnel.unit_id not in unit_ids:
        return jsonify({'error': 'شما به این درخواست دسترسی ندارید'}), 403
    
    if req.status != 'pending_unit':
        return jsonify({'error': 'این درخواست قبلاً بررسی شده است'}), 400
    
    data = request.get_json()
    reason = data.get('reason', '').strip()
    
    req.status = 'rejected'
    req.reject_reason = reason
    req.reviewed_at = datetime.now()
    db.session.commit()
    
    # ارسال اعلان به کاربر
    send_workflow_notification(
        to_user_id=req.requester_id,
        title=f"❌ درخواست {req.get_request_type_persian()} شما رد شد",
        message=f"درخواست شما توسط {current_user.get_full_name()} رد شد.\nدلیل: {reason or 'بدون توضیح'}",
        link=f"/requests/{req.id}"
    )
    
    return jsonify({'success': True, 'message': 'درخواست رد شد'})


@app.route('/api/requests/<int:request_id>/revision', methods=['POST'])
@login_required
def api_request_revision(request_id):
    """درخواست اصلاح توسط سرپرست واحد"""
    req = Request.query.get_or_404(request_id)
    
    if current_user.role != 'unit_supervisor':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    # بررسی دسترسی
    supervised_units = db.session.query(Unit).join(
        UnitSupervisor, UnitSupervisor.unit_id == Unit.id
    ).filter(UnitSupervisor.user_id == current_user.id).all()
    unit_ids = [u.id for u in supervised_units]
    
    personnel = Personnel.query.filter_by(national_code=req.requester.national_code).first()
    if not personnel or personnel.unit_id not in unit_ids:
        return jsonify({'error': 'شما به این درخواست دسترسی ندارید'}), 403
    
    if req.status != 'pending_unit':
        return jsonify({'error': 'این درخواست قبلاً بررسی شده است'}), 400
    
    data = request.get_json()
    revision_note = data.get('revision_note', '').strip()
    
    if not revision_note:
        return jsonify({'error': 'لطفاً توضیحات اصلاحی را وارد کنید'}), 400
    
    req.status = 'revision'
    req.revision_note = revision_note
    req.reviewed_at = datetime.now()
    db.session.commit()
    
    # ارسال اعلان به کاربر
    send_workflow_notification(
        to_user_id=req.requester_id,
        title=f"🔄 درخواست {req.get_request_type_persian()} شما نیاز به اصلاح دارد",
        message=f"لطفاً درخواست خود را بر اساس نظر {current_user.get_full_name()} اصلاح کنید.\nتوضیحات: {revision_note}",
        link=f"/requests/{req.id}/edit"
    )
    
    return jsonify({'success': True, 'message': 'درخواست برای اصلاح برگشت داده شد'})


@app.route('/api/requests/<int:request_id>/edit', methods=['PUT'])
@login_required
def api_edit_request(request_id):
    """ویرایش درخواست توسط کاربر (فقط در وضعیت revision)"""
    req = Request.query.get_or_404(request_id)
    
    # فقط خود کاربر می‌تواند ویرایش کند
    if current_user.id != req.requester_id:
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    if req.status != 'revision':
        return jsonify({'error': 'این درخواست قابل ویرایش نیست'}), 400
    
    data = request.get_json()
    extra_data = data.get('extra_data', {})
    
    req.set_extra_data(extra_data)
    req.status = 'pending_unit'  # دوباره به صف تایید برمی‌گردد
    req.revision_note = None  # پاک کردن یادداشت اصلاح قبلی
    db.session.commit()
    
    # ارسال اعلان دوباره به سرپرست
    supervisor = get_unit_supervisor(req.requester.national_code)
    if supervisor:
        send_request_notification(
            to_user_id=supervisor.id,
            request_type=req.request_type,
            request_id=req.id,
            message=f"{req.requester.get_full_name()} درخواست خود را اصلاح و مجدداً ارسال کرد."
        )
    
    return jsonify({'success': True, 'message': 'درخواست با موفقیت ویرایش و مجدداً ارسال شد'})
    
# ==================== ثبت درخواست اضافه کار ساعتی ====================

@app.route('/api/requests/overtime/create', methods=['POST'])
@login_required
def api_create_overtime_request():
    """ثبت درخواست اضافه کار ساعتی"""
    if current_user.role not in ['subordinate']:
        return jsonify({'error': 'فقط کاربران عادی می‌توانند درخواست ثبت کنند'}), 403
    
    data = request.get_json()
    
    # اعتبارسنجی فیلدها
    start_time = data.get('start_time', '').strip()
    end_time = data.get('end_time', '').strip()
    date = data.get('date', '').strip()
    subject = data.get('subject', '').strip()
    duration = data.get('duration', '').strip()
    
    if not start_time or not end_time or not date or not subject:
        return jsonify({'error': 'تمامی فیلدهای الزامی را پر کنید'}), 400
    
    # محاسبه مجدد مدت در بک‌اند (برای امنیت)
    try:
        start_h, start_m = map(int, start_time.split(':'))
        end_h, end_m = map(int, end_time.split(':'))
        start_total = start_h * 60 + start_m
        end_total = end_h * 60 + end_m
        duration_minutes = end_total - start_total
        if duration_minutes <= 0:
            return jsonify({'error': 'زمان پایان باید بعد از زمان شروع باشد'}), 400
        duration_hours = duration_minutes / 60
        duration_str = f"{int(duration_hours)} ساعت و {duration_minutes % 60} دقیقه"
    except:
        duration_str = duration
    
    # پیدا کردن سرپرست واحد
    supervisor = get_unit_supervisor(current_user.national_code)
    if not supervisor:
        return jsonify({'error': 'سرپرست واحدی برای شما تعریف نشده است'}), 400
    
    # ایجاد درخواست
    req = Request(
        request_type='overtime',
        requester_id=current_user.id,
        unit_supervisor_id=supervisor.id,
        status='pending_unit'
    )
    
    extra_data = {
        'start_time': start_time,
        'end_time': end_time,
        'date': date,
        'subject': subject,
        'duration': duration_str
    }
    req.set_extra_data(extra_data)
    
    db.session.add(req)
    db.session.commit()
    
    # ارسال اعلان به سرپرست
    send_request_notification(
        to_user_id=supervisor.id,
        request_type='overtime',
        request_id=req.id,
        message=f"{current_user.get_full_name()} درخواست اضافه کار برای تاریخ {date} ثبت کرد."
    )
    
    return jsonify({
        'success': True,
        'message': 'درخواست اضافه کار با موفقیت ثبت شد',
        'request_id': req.id
    })
    


# ==================== ثبت درخواست ثبت نواقص ====================

@app.route('/api/requests/deficiency/create', methods=['POST'])
@login_required
def api_create_deficiency_request():
    """ثبت درخواست ثبت نواقص (تأخیر/تعجیل)"""
    if current_user.role not in ['subordinate']:
        return jsonify({'error': 'فقط کاربران عادی می‌توانند درخواست ثبت کنند'}), 403
    
    data = request.get_json()
    
    # اعتبارسنجی فیلدها
    date = data.get('date', '').strip()
    time = data.get('time', '').strip()
    deficiency_type = data.get('type', '').strip()  # 'ورود' یا 'خروج'
    description = data.get('description', '').strip()
    
    if not date or not time or not deficiency_type or not description:
        return jsonify({'error': 'تمامی فیلدهای الزامی را پر کنید'}), 400
    
    if deficiency_type not in ['ورود', 'خروج']:
        return jsonify({'error': 'نوع نقص باید ورود یا خروج باشد'}), 400
    
    # پیدا کردن سرپرست واحد
    supervisor = get_unit_supervisor(current_user.national_code)
    if not supervisor:
        return jsonify({'error': 'سرپرست واحدی برای شما تعریف نشده است'}), 400
    
    # ایجاد درخواست
    req = Request(
        request_type='deficiency',
        requester_id=current_user.id,
        unit_supervisor_id=supervisor.id,
        status='pending_unit'
    )
    
    extra_data = {
        'date': date,
        'time': time,
        'type': deficiency_type,
        'description': description
    }
    req.set_extra_data(extra_data)
    
    db.session.add(req)
    db.session.commit()
    
    # ارسال اعلان به سرپرست
    send_request_notification(
        to_user_id=supervisor.id,
        request_type='deficiency',
        request_id=req.id,
        message=f"{current_user.get_full_name()} درخواست ثبت نقص ({deficiency_type}) برای تاریخ {date} ثبت کرد."
    )
    
    return jsonify({
        'success': True,
        'message': 'درخواست ثبت نقص با موفقیت ثبت شد',
        'request_id': req.id
    })
    

    
# ==================== ثبت درخواست مرخصی روزانه ====================

@app.route('/api/requests/annual-leave/create', methods=['POST'])
@login_required
def api_create_annual_leave_request():
    """ثبت درخواست مرخصی روزانه"""
    if current_user.role not in ['subordinate']:
        return jsonify({'error': 'فقط کاربران عادی می‌توانند درخواست ثبت کنند'}), 403
    
    data = request.get_json()
    
    # اعتبارسنجی فیلدها
    request_date = data.get('request_date', '').strip()
    start_date = data.get('start_date', '').strip()
    end_date = data.get('end_date', '').strip()
    leave_type = data.get('leave_type', '').strip()
    description = data.get('description', '').strip()
    
    if not start_date or not end_date or not leave_type:
        return jsonify({'error': 'فیلدهای تاریخ شروع، پایان و نوع مرخصی الزامی است'}), 400
    
    if leave_type not in ['استحقاقی', 'استعلاجی', 'تشویقی']:
        return jsonify({'error': 'نوع مرخصی نامعتبر است'}), 400
    
    # محاسبه مدت (تعداد روزها)
    try:
        # تبدیل تاریخ شمسی به میلادی برای محاسبه دقیق
        start_parts = start_date.split('/')
        end_parts = end_date.split('/')
        
        start_jalali = jdatetime.date(int(start_parts[0]), int(start_parts[1]), int(start_parts[2]))
        end_jalali = jdatetime.date(int(end_parts[0]), int(end_parts[1]), int(end_parts[2]))
        
        duration_days = (end_jalali - start_jalali).days + 1
        if duration_days <= 0:
            return jsonify({'error': 'تاریخ پایان باید بعد از تاریخ شروع باشد'}), 400
        
        duration_str = f"{duration_days} روز"
    except:
        duration_str = data.get('duration', '')
    
    # پیدا کردن سرپرست واحد
    supervisor = get_unit_supervisor(current_user.national_code)
    if not supervisor:
        return jsonify({'error': 'سرپرست واحدی برای شما تعریف نشده است'}), 400
    
    # ایجاد درخواست
    req = Request(
        request_type='annual_leave',
        requester_id=current_user.id,
        unit_supervisor_id=supervisor.id,
        status='pending_unit'
    )
    
    extra_data = {
        'request_date': request_date,
        'start_date': start_date,
        'end_date': end_date,
        'leave_type': leave_type,
        'duration': duration_str,
        'description': description
    }
    req.set_extra_data(extra_data)
    
    db.session.add(req)
    db.session.commit()
    
    # ارسال اعلان به سرپرست
    send_request_notification(
        to_user_id=supervisor.id,
        request_type='annual_leave',
        request_id=req.id,
        message=f"{current_user.get_full_name()} درخواست مرخصی {leave_type} به مدت {duration_str} ثبت کرد."
    )
    
    return jsonify({
        'success': True,
        'message': 'درخواست مرخصی با موفقیت ثبت شد',
        'request_id': req.id
    })
    
# ==================== ثبت درخواست مرخصی ساعتی ====================

@app.route('/api/requests/hourly-leave/create', methods=['POST'])
@login_required
def api_create_hourly_leave_request():
    """ثبت درخواست مرخصی ساعتی"""
    if current_user.role not in ['subordinate']:
        return jsonify({'error': 'فقط کاربران عادی می‌توانند درخواست ثبت کنند'}), 403
    
    data = request.get_json()
    
    # اعتبارسنجی فیلدها
    request_date = data.get('request_date', '').strip()
    start_time = data.get('start_time', '').strip()
    end_time = data.get('end_time', '').strip()
    leave_type = data.get('leave_type', '').strip()
    description = data.get('description', '').strip()
    
    if not start_time or not end_time or not leave_type:
        return jsonify({'error': 'فیلدهای زمان شروع، پایان و نوع مرخصی الزامی است'}), 400
    
    if leave_type not in ['استحقاقی', 'استعلاجی', 'تشویقی']:
        return jsonify({'error': 'نوع مرخصی نامعتبر است'}), 400
    
    # محاسبه مدت (ساعت و دقیقه)
    try:
        start_h, start_m = map(int, start_time.split(':'))
        end_h, end_m = map(int, end_time.split(':'))
        start_total = start_h * 60 + start_m
        end_total = end_h * 60 + end_m
        
        if end_total <= start_total:
            return jsonify({'error': 'زمان پایان باید بعد از زمان شروع باشد'}), 400
        
        duration_minutes = end_total - start_total
        hours = duration_minutes // 60
        minutes = duration_minutes % 60
        
        if hours > 0 and minutes > 0:
            duration_str = f"{hours} ساعت و {minutes} دقیقه"
        elif hours > 0:
            duration_str = f"{hours} ساعت"
        else:
            duration_str = f"{minutes} دقیقه"
    except:
        duration_str = data.get('duration', '')
    
    # پیدا کردن سرپرست واحد
    supervisor = get_unit_supervisor(current_user.national_code)
    if not supervisor:
        return jsonify({'error': 'سرپرست واحدی برای شما تعریف نشده است'}), 400
    
    # ایجاد درخواست
    req = Request(
        request_type='hourly_leave',
        requester_id=current_user.id,
        unit_supervisor_id=supervisor.id,
        status='pending_unit'
    )
    
    extra_data = {
        'request_date': request_date,
        'start_time': start_time,
        'end_time': end_time,
        'leave_type': leave_type,
        'duration': duration_str,
        'description': description
    }
    req.set_extra_data(extra_data)
    
    db.session.add(req)
    db.session.commit()
    
    # ارسال اعلان به سرپرست
    send_request_notification(
        to_user_id=supervisor.id,
        request_type='hourly_leave',
        request_id=req.id,
        message=f"{current_user.get_full_name()} درخواست مرخصی ساعتی {leave_type} به مدت {duration_str} ثبت کرد."
    )
    
    return jsonify({
        'success': True,
        'message': 'درخواست مرخصی ساعتی با موفقیت ثبت شد',
        'request_id': req.id
    })
    
# ==================== ثبت درخواست مأموریت روزانه ====================

@app.route('/api/requests/daily-mission/create', methods=['POST'])
@login_required
def api_create_daily_mission_request():
    """ثبت درخواست مأموریت روزانه"""
    if current_user.role not in ['subordinate']:
        return jsonify({'error': 'فقط کاربران عادی می‌توانند درخواست ثبت کنند'}), 403
    
    data = request.get_json()
    
    # اعتبارسنجی فیلدهای الزامی
    mission_type = data.get('mission_type', '').strip()  # 'خارج از شهر' یا 'داخل شهر'
    subject = data.get('subject', '').strip()
    vehicle = data.get('vehicle', '').strip()
    origin = data.get('origin', '').strip()
    destination = data.get('destination', '').strip()
    departure_date = data.get('departure_date', '').strip()
    return_date = data.get('return_date', '').strip()
    departure_time = data.get('departure_time', '').strip()
    return_time = data.get('return_time', '').strip()
    location = data.get('location', '').strip()
    advance_amount = data.get('advance_amount', '').strip()
    order_number = data.get('order_number', '').strip()
    travel_details = data.get('travel_details', '').strip()
    passengers = data.get('passengers', '').strip()
    description = data.get('description', '').strip()
    
    if not mission_type or not subject or not origin or not destination:
        return jsonify({'error': 'فیلدهای نوع مأموریت، موضوع، مبدأ و مقصد الزامی است'}), 400
    
    if not departure_date or not return_date:
        return jsonify({'error': 'تاریخ رفت و برگشت الزامی است'}), 400
    
    # محاسبه مدت مأموریت (تعداد روزها)
    try:
        start_parts = departure_date.split('/')
        end_parts = return_date.split('/')
        
        start_jalali = jdatetime.date(int(start_parts[0]), int(start_parts[1]), int(start_parts[2]))
        end_jalali = jdatetime.date(int(end_parts[0]), int(end_parts[1]), int(end_parts[2]))
        
        duration_days = (end_jalali - start_jalali).days + 1
        if duration_days <= 0:
            return jsonify({'error': 'تاریخ برگشت باید بعد از تاریخ رفت باشد'}), 400
        
        duration_str = f"{duration_days} روز"
    except:
        duration_str = data.get('duration', '')
    
    # پیدا کردن سرپرست واحد
    supervisor = get_unit_supervisor(current_user.national_code)
    if not supervisor:
        return jsonify({'error': 'سرپرست واحدی برای شما تعریف نشده است'}), 400
    
    # ایجاد درخواست
    req = Request(
        request_type='daily_mission',
        requester_id=current_user.id,
        unit_supervisor_id=supervisor.id,
        status='pending_unit'
    )
    
    extra_data = {
        'mission_type': mission_type,
        'subject': subject,
        'vehicle': vehicle,
        'origin': origin,
        'destination': destination,
        'departure_date': departure_date,
        'return_date': return_date,
        'departure_time': departure_time,
        'return_time': return_time,
        'location': location,
        'duration': duration_str,
        'advance_amount': advance_amount,
        'order_number': order_number,
        'request_date': datetime.now().strftime('%Y/%m/%d'),
        'travel_details': travel_details,
        'passengers': passengers,
        'description': description
    }
    req.set_extra_data(extra_data)
    
    db.session.add(req)
    db.session.commit()
    
    # ارسال اعلان به سرپرست
    send_request_notification(
        to_user_id=supervisor.id,
        request_type='daily_mission',
        request_id=req.id,
        message=f"{current_user.get_full_name()} درخواست مأموریت {mission_type} به {destination} ثبت کرد."
    )
    
    return jsonify({
        'success': True,
        'message': 'درخواست مأموریت با موفقیت ثبت شد',
        'request_id': req.id
    })
    
# ==================== ثبت درخواست مأموریت اداری ====================

@app.route('/api/requests/official-mission/create', methods=['POST'])
@login_required
def api_create_official_mission_request():
    """ثبت درخواست مأموریت اداری"""
    if current_user.role not in ['subordinate']:
        return jsonify({'error': 'فقط کاربران عادی می‌توانند درخواست ثبت کنند'}), 403
    
    data = request.get_json()
    
    # اعتبارسنجی فیلدها
    mission_type = data.get('mission_type', '').strip()  # اداری، فنی، بازرسی، مأموریت طرح بسیج، آموزشی
    date = data.get('date', '').strip()
    start_time = data.get('start_time', '').strip()
    end_time = data.get('end_time', '').strip()
    subject = data.get('subject', '').strip()
    description = data.get('description', '').strip()
    
    if not mission_type or not date or not start_time or not end_time or not subject:
        return jsonify({'error': 'تمامی فیلدهای الزامی را پر کنید'}), 400
    
    valid_types = ['اداری', 'فنی', 'بازرسی', 'ماموریت طرح بسیج', 'آموزشی']
    if mission_type not in valid_types:
        return jsonify({'error': 'نوع مأموریت نامعتبر است'}), 400
    
    # محاسبه مدت (ساعت و دقیقه)
    try:
        start_h, start_m = map(int, start_time.split(':'))
        end_h, end_m = map(int, end_time.split(':'))
        start_total = start_h * 60 + start_m
        end_total = end_h * 60 + end_m
        
        if end_total <= start_total:
            return jsonify({'error': 'زمان پایان باید بعد از زمان شروع باشد'}), 400
        
        duration_minutes = end_total - start_total
        hours = duration_minutes // 60
        minutes = duration_minutes % 60
        
        if hours > 0 and minutes > 0:
            duration_str = f"{hours} ساعت و {minutes} دقیقه"
        elif hours > 0:
            duration_str = f"{hours} ساعت"
        else:
            duration_str = f"{minutes} دقیقه"
    except:
        duration_str = data.get('duration', '')
    
    # پیدا کردن سرپرست واحد
    supervisor = get_unit_supervisor(current_user.national_code)
    if not supervisor:
        return jsonify({'error': 'سرپرست واحدی برای شما تعریف نشده است'}), 400
    
    # ایجاد درخواست
    req = Request(
        request_type='official_mission',
        requester_id=current_user.id,
        unit_supervisor_id=supervisor.id,
        status='pending_unit'
    )
    
    extra_data = {
        'mission_type': mission_type,
        'date': date,
        'start_time': start_time,
        'end_time': end_time,
        'duration': duration_str,
        'subject': subject,
        'description': description
    }
    req.set_extra_data(extra_data)
    
    db.session.add(req)
    db.session.commit()
    
    # ارسال اعلان به سرپرست
    send_request_notification(
        to_user_id=supervisor.id,
        request_type='official_mission',
        request_id=req.id,
        message=f"{current_user.get_full_name()} درخواست مأموریت {mission_type} برای تاریخ {date} ثبت کرد."
    )
    
    return jsonify({
        'success': True,
        'message': 'درخواست مأموریت اداری با موفقیت ثبت شد',
        'request_id': req.id
    })
    
    
# ==================== ثبت درخواست سفر اربعین ====================

@app.route('/api/requests/arbaeen/create', methods=['POST'])
@login_required
def api_create_arbaeen_request():
    """ثبت درخواست سفر اربعین"""
    if current_user.role not in ['subordinate']:
        return jsonify({'error': 'فقط کاربران عادی می‌توانند درخواست ثبت کنند'}), 403
    
    data = request.get_json()
    
    # اعتبارسنجی فیلدها
    departure_date = data.get('departure_date', '').strip()
    return_date = data.get('return_date', '').strip()
    exit_border = data.get('exit_border', '').strip()
    entry_border = data.get('entry_border', '').strip()
    
    if not departure_date or not return_date or not exit_border or not entry_border:
        return jsonify({'error': 'تمامی فیلدهای الزامی را پر کنید'}), 400
    
    # محاسبه مدت سفر
    try:
        start_parts = departure_date.split('/')
        end_parts = return_date.split('/')
        
        start_jalali = jdatetime.date(int(start_parts[0]), int(start_parts[1]), int(start_parts[2]))
        end_jalali = jdatetime.date(int(end_parts[0]), int(end_parts[1]), int(end_parts[2]))
        
        duration_days = (end_jalali - start_jalali).days + 1
        if duration_days <= 0:
            return jsonify({'error': 'تاریخ برگشت باید بعد از تاریخ رفت باشد'}), 400
        
        duration_str = f"{duration_days} روز"
    except:
        duration_str = data.get('duration', '')
    
    # پیدا کردن سرپرست واحد
    supervisor = get_unit_supervisor(current_user.national_code)
    if not supervisor:
        return jsonify({'error': 'سرپرست واحدی برای شما تعریف نشده است'}), 400
    
    # ایجاد درخواست
    req = Request(
        request_type='arbaeen',
        requester_id=current_user.id,
        unit_supervisor_id=supervisor.id,
        status='pending_unit'
    )
    
    extra_data = {
        'request_date': datetime.now().strftime('%Y/%m/%d'),
        'departure_date': departure_date,
        'return_date': return_date,
        'exit_border': exit_border,
        'entry_border': entry_border,
        'duration': duration_str
    }
    req.set_extra_data(extra_data)
    
    db.session.add(req)
    db.session.commit()
    
    # ارسال اعلان به سرپرست
    send_request_notification(
        to_user_id=supervisor.id,
        request_type='arbaeen',
        request_id=req.id,
        message=f"{current_user.get_full_name()} درخواست سفر اربعین از مرز {exit_border} ثبت کرد."
    )
    
    return jsonify({
        'success': True,
        'message': 'درخواست سفر اربعین با موفقیت ثبت شد',
        'request_id': req.id
    })
    
    
    



@app.route('/admin/api/personnel-assign', methods=['POST'])
@login_required
def admin_api_personnel_assign():
    """ثبت انتصاب/انتقال پرسنل به واحد"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    personnel_id = data.get('personnel_id')
    unit_id = data.get('unit_id')
    period_id = data.get('period_id')
    start_date = data.get('start_date', '').strip()
    assignment_type = data.get('assignment_type', 'initial')
    description = data.get('description', '')
    
    if not personnel_id or not unit_id or not start_date:
        return jsonify({'error': 'اطلاعات ناقص است'}), 400
    
    personnel = Personnel.query.get(personnel_id)
    if not personnel:
        return jsonify({'error': 'پرسنل یافت نشد'}), 404
    
    unit = Unit.query.get(unit_id)
    if not unit:
        return jsonify({'error': 'واحد یافت نشد'}), 404
    
    # پایان دادن به انتصاب فعال قبلی (در صورت وجود)
    old_assignment = PersonnelAssignment.query.filter_by(
        personnel_id=personnel_id, 
        is_active=True
    ).first()
    
    if old_assignment:
        old_assignment.is_active = False
        old_assignment.end_date = start_date
    
    # ایجاد انتصاب جدید
    new_assignment = PersonnelAssignment(
        personnel_id=personnel_id,
        unit_id=unit_id,
        period_id=period_id,
        start_date=start_date,
        assignment_type=assignment_type,
        description=description,
        created_by=current_user.id,
        is_active=True
    )
    db.session.add(new_assignment)
    
    # به‌روزرسانی فیلدهای unit_id و department_id در جدول personnel
    personnel.unit_id = unit_id
    personnel.department_id = unit.department_id
    personnel.period_id = period_id
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'انتصاب با موفقیت ثبت شد'})


@app.route('/admin/api/personnel-assignments/<int:personnel_id>')
@login_required
def admin_api_personnel_assignments(personnel_id):
    """دریافت تاریخچه انتصابات پرسنل"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    assignments = PersonnelAssignment.query.filter_by(
        personnel_id=personnel_id
    ).order_by(PersonnelAssignment.start_date.desc()).all()
    
    result = []
    for a in assignments:
        unit = Unit.query.get(a.unit_id)
        dept = Department.query.get(unit.department_id) if unit else None
        creator = User.query.get(a.created_by)
        
        result.append({
            'id': a.id,
            'personnel_id': a.personnel_id,
            'unit_id': a.unit_id,
            'unit_name': unit.name if unit else '-',
            'department_id': dept.id if dept else None,
            'department_name': dept.name if dept else '-',
            'period_id': a.period_id,
            'assignment_type': a.assignment_type,
            'assignment_type_persian': {
                'initial': 'انتصاب اولیه',
                'transfer': 'انتقال',
                'promotion': 'ارتقا'
            }.get(a.assignment_type, a.assignment_type),
            'start_date': a.start_date,
            'end_date': a.end_date,
            'is_active': a.is_active,
            'description': a.description,
            'created_by_name': creator.get_full_name() if creator else '-',
            'created_at': a.created_at.strftime('%Y/%m/%d %H:%M') if a.created_at else '-'
        })
    
    return jsonify(result)
    
# ========== APIهای مدیریت انتصابات ==========

@app.route('/admin/api/unassigned-personnel')
@login_required
def admin_api_unassigned_personnel():
    """دریافت لیست پرسنل بدون انتصاب فعال"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    search = request.args.get('search', '')
    
    # پرسنلی که هیچ انتصاب فعالی ندارند
    query = Personnel.query.filter_by(is_deleted=False)
    
    # پرسنلی که انتصاب فعال دارند
    assigned_ids = [a.personnel_id for a in PersonnelAssignment.query.filter_by(is_active=True).all()]
    query = query.filter(~Personnel.id.in_(assigned_ids))
    
    if search:
        query = query.filter(
            db.or_(
                Personnel.national_code.contains(search),
                Personnel.first_name.contains(search),
                Personnel.last_name.contains(search)
            )
        )
    
    personnel = query.all()
    result = [{
        'id': p.id,
        'national_code': p.national_code,
        'first_name': p.first_name or '',
        'last_name': p.last_name or '',
        'full_name': p.get_full_name(),
        'phone': p.phone or '',
        'position': p.position or ''
    } for p in personnel]
    
    return jsonify({'personnel': result})


@app.route('/admin/api/active-assignments')
@login_required
def admin_api_active_assignments():
    """دریافت لیست انتصابات فعال"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    dept_id = request.args.get('dept_id', type=int)
    unit_id = request.args.get('unit_id', type=int)
    
    query = PersonnelAssignment.query.filter_by(is_active=True)
    
    if unit_id:
        query = query.filter_by(unit_id=unit_id)
    elif dept_id:
        # پیدا کردن واحدهای آن اداره
        units = Unit.query.filter_by(department_id=dept_id).all()
        unit_ids = [u.id for u in units]
        query = query.filter(PersonnelAssignment.unit_id.in_(unit_ids))
    
    assignments = query.order_by(PersonnelAssignment.start_date.desc()).all()
    result = []
    
    for a in assignments:
        personnel = Personnel.query.get(a.personnel_id)
        unit = Unit.query.get(a.unit_id)
        dept = Department.query.get(unit.department_id) if unit else None
        
        if personnel and not personnel.is_deleted:
            result.append({
                'id': a.id,
                'personnel_id': a.personnel_id,
                'personnel_name': personnel.get_full_name(),
                'national_code': personnel.national_code,
                'unit_id': a.unit_id,
                'unit_name': unit.name if unit else '-',
                'department_name': dept.name if dept else '-',
                'start_date': a.start_date,
                'assignment_type': a.assignment_type,
                'is_active': a.is_active
            })
    
    return jsonify({'assignments': result})


# اضافه کن به بخش APIهای مدیریت انتصابات

@app.route('/admin/api/assignments-history')
@login_required
def admin_api_assignments_history():
    """دریافت تاریخچه انتصابات برای سال‌های داینامیک"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    assignments = PersonnelAssignment.query.order_by(PersonnelAssignment.start_date.desc()).limit(500).all()
    
    result = []
    for a in assignments:
        personnel = Personnel.query.get(a.personnel_id)
        unit = Unit.query.get(a.unit_id)
        dept = Department.query.get(unit.department_id) if unit else None
        creator = User.query.get(a.created_by)
        
        if personnel:
            result.append({
                'id': a.id,
                'personnel_id': a.personnel_id,
                'personnel_name': personnel.get_full_name(),
                'national_code': personnel.national_code,
                'unit_name': unit.name if unit else '-',
                'department_name': dept.name if dept else '-',
                'assignment_type': a.assignment_type,
                'assignment_type_persian': {
                    'initial': 'انتصاب اولیه',
                    'transfer': 'انتقال',
                    'promotion': 'ارتقا'
                }.get(a.assignment_type, a.assignment_type),
                'start_date': a.start_date,
                'end_date': a.end_date,
                'is_active': a.is_active,
                'description': a.description,
                'created_by_name': creator.get_full_name() if creator else '-',
                'created_at': a.created_at.strftime('%Y/%m/%d %H:%M')
            })
    
    return jsonify({'history': result})
    
    
# ========== APIهای پیشرفته مدیریت کاربران ==========

@app.route('/admin/api/users-advanced')
@login_required
def admin_api_users_advanced():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    search = request.args.get('search', '')
    role = request.args.get('role', '')
    dept_id = request.args.get('dept_id', type=int)
    unit_id = request.args.get('unit_id', type=int)
    period_id = request.args.get('period_id', type=int)
    status = request.args.get('status', '')
    online = request.args.get('online', '')
    unassigned_only = request.args.get('unassigned_only', 'false').lower() == 'true'  # اضافه کن
    
    query = User.query
    
    if search:
        query = query.filter(
            db.or_(
                User.national_code.contains(search),
                User.first_name.contains(search),
                User.last_name.contains(search),
                User.personnel_code.contains(search)
            )
        )
    
    if role:
        query = query.filter_by(role=role)
    
    if status == 'active':
        query = query.filter_by(is_active=True, is_approved=True)
    elif status == 'inactive':
        query = query.filter_by(is_active=False)
    elif status == 'pending':
        query = query.filter_by(is_approved=False, is_active=True)
    
    # ========== فیلتر برای نمایش فقط کاربران بدون انتصاب ==========
    if unassigned_only:
        # دریافت پرسنلی که انتصاب فعال دارند
        assigned_personnel_ids = [a.personnel_id for a in PersonnelAssignment.query.filter_by(is_active=True).all()]
        
        # پرسنلی که انتصاب فعال ندارند
        unassigned_personnel = Personnel.query.filter(
            ~Personnel.id.in_(assigned_personnel_ids) if assigned_personnel_ids else True,
            Personnel.is_deleted == False
        ).all()
        
        # کدهای ملی پرسنل بدون انتصاب
        unassigned_national_codes = [p.national_code for p in unassigned_personnel if p.national_code]
        
        if unassigned_national_codes:
            query = query.filter(User.national_code.in_(unassigned_national_codes))
        else:
            query = query.filter(False)
    
    # آمار
    active_count = User.query.filter_by(is_active=True, is_approved=True).count()
    pending_count = User.query.filter_by(is_approved=False, is_active=True).count()
    inactive_count = User.query.filter_by(is_active=False).count()
    
    # آنلاین (آخرین ورود در 60 دقیقه اخیر)
    from datetime import timedelta
    one_hour_ago = datetime.now() - timedelta(minutes=60)
    online_count = User.query.filter(User.last_login > one_hour_ago).count()
    
    # پرسنل بدون انتصاب
    assigned_ids = [a.personnel_id for a in PersonnelAssignment.query.filter_by(is_active=True).all()]
    unassigned_count = Personnel.query.filter(~Personnel.id.in_(assigned_ids), Personnel.is_deleted==False).count()
    
    pagination = query.order_by(User.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    users = []
    for u in pagination.items:
        # پیدا کردن اداره و واحد از طریق پرسنل
        personnel = Personnel.query.filter_by(national_code=u.national_code, is_deleted=False).first()
        department_name = personnel.department.name if personnel and personnel.department else '-'
        unit_name = personnel.unit.name if personnel and personnel.unit else '-'
        is_online = u.last_login and (datetime.now() - u.last_login).total_seconds() < 3600
        
        # بررسی وجود انتصاب فعال
        has_assignment = PersonnelAssignment.query.filter_by(personnel_id=personnel.id, is_active=True).first() is not None if personnel else False
        
        users.append({
            'id': u.id,
            'national_code': u.national_code,
            'first_name': u.first_name,
            'last_name': u.last_name,
            'full_name': u.get_full_name(),
            'phone': u.phone or '',
            'role': u.role,
            'role_persian': u.get_role_persian(),
            'is_active': u.is_active,
            'is_approved': u.is_approved,
            'department_name': department_name,
            'unit_name': unit_name,
            'has_assignment': has_assignment,
            'is_online': is_online,
            'created_at': u.get_jalali_created_date(),
            'last_login': u.last_login.strftime('%Y/%m/%d %H:%M') if u.last_login else None
        })
    
    return jsonify({
        'users': users,
        'total': pagination.total,
        'page': page,
        'per_page': per_page,
        'pages': pagination.pages,
        'active_count': active_count,
        'pending_count': pending_count,
        'inactive_count': inactive_count,
        'online_count': online_count,
        'unassigned_count': unassigned_count
    })


@app.route('/admin/api/user-full-detail/<int:user_id>')
@login_required
def admin_api_user_full_detail(user_id):
    """دریافت اطلاعات کامل یک کاربر برای نمایش در مودال"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    user = User.query.get_or_404(user_id)
    personnel = Personnel.query.filter_by(national_code=user.national_code, is_deleted=False).first()
    
    return jsonify({
        'id': user.id,
        'national_code': user.national_code,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'full_name': user.get_full_name(),
        'phone': user.phone or '',
        'role': user.role,
        'role_persian': user.get_role_persian(),
        'personnel_code': user.personnel_code or '',
        'profile_picture': user.profile_picture,
        'department_name': personnel.department.name if personnel and personnel.department else '-',
        'unit_name': personnel.unit.name if personnel and personnel.unit else '-',
        'position': personnel.position if personnel else '-',
        'created_at': user.get_jalali_created_date(),
        'last_login': user.last_login.strftime('%Y/%m/%d %H:%M') if user.last_login else 'هرگز'
    })





@app.route('/admin/api/user-requests/<int:user_id>')
@login_required
def admin_api_user_requests(user_id):
    """دریافت لیست درخواست‌های کاربر برای نمایش در مودال"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    year = request.args.get('year', 'all')
    user = User.query.get_or_404(user_id)
    
    query = Request.query.filter_by(requester_id=user_id).order_by(Request.request_date.desc())
    
    requests = []
    for req in query.all():
        if year != 'all':
            jalali_year = str(jdatetime.datetime.fromgregorian(datetime=req.request_date).year) if req.request_date else ''
            if jalali_year != year:
                continue
        requests.append(req.to_dict())
    
    return jsonify(requests)
    

# ========== دانلود قالب اکسل کاربران ==========
@app.route('/admin/users/download-template')
@login_required
def admin_users_download_template():
    """دانلود قالب اکسل برای آپلود کاربران"""
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import tempfile
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "قالب کاربران"
    
    # هدرها بر اساس فیلدهای پروفایل
    headers = ['کد ملی', 'نام', 'نام خانوادگی', 'نقش', 'کد پرسنلی', 'شماره تماس']
    
    # اضافه کردن فیلدهای داینامیک فعال
    dynamic_fields = DynamicField.query.filter_by(is_active=True).all()
    for field in dynamic_fields:
        if not field.is_key and field.title not in ['نام', 'نام خانوادگی', 'شماره تماس', 'سمت']:
            headers.append(field.title)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12, color='ffffff')
        cell.fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # تنظیم عرض ستون‌ها
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    # یک ردیف نمونه
    sample_row = ['1234567890', 'احمد', 'رضایی', 'subordinate', 'EMP001', '09123456789']
    for col, val in enumerate(sample_row, 1):
        ws.cell(row=2, column=col, value=val)
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    
    return send_file(temp_file.name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='قالب_کاربران_آوان.xlsx')


# ========== خروجی اکسل کاربران ==========
# جایگزین کن تابع admin_export_users_excel را با این نسخه:

@app.route('/admin/export-users-excel')
@login_required
def admin_export_users_excel():
    """خروجی اکسل لیست کاربران با قالب تنظیمات"""
    if current_user.role != 'admin':
        flash('دسترسی غیرمجاز', 'error')
        return redirect(url_for('dashboard'))
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile
    
    users = User.query.all()
    
    # دریافت قالب از دیتابیس
    template = ExcelTemplate.query.first()
    if not template:
        template = ExcelTemplate()
    
    # رنگ‌ها از قالب
    header_bg = template.header_bg_color.replace('#', '') if template and template.header_bg_color else '1e293b'
    header_text = template.header_text_color.replace('#', '') if template and template.header_text_color else 'ffffff'
    even_color = template.even_row_color.replace('#', '') if template and template.even_row_color else 'f8f9fa'
    odd_color = template.odd_row_color.replace('#', '') if template and template.odd_row_color else 'ffffff'
    font_name = template.font_name if template and template.font_name else 'Calibri'
    header_font_size = template.header_font_size if template and template.header_font_size else 12
    data_font_size = template.data_font_size if template and template.data_font_size else 11
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "لیست کاربران"
    
    # هدرها
    headers = ['ردیف', 'کد ملی', 'نام', 'نام خانوادگی', 'شماره تماس', 'نقش', 'کد پرسنلی', 'وضعیت', 'تاریخ عضویت']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name=font_name, size=header_font_size, bold=True, color=header_text)
        cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # داده‌ها با رنگ‌بندی ردیف‌های زوج و فرد
    for idx, user in enumerate(users, 1):
        row_num = idx + 1
        bg_color = even_color if (row_num % 2 == 0) else odd_color
        
        ws.cell(row=row_num, column=1, value=idx)
        ws.cell(row=row_num, column=2, value=user.national_code)
        ws.cell(row=row_num, column=3, value=user.first_name)
        ws.cell(row=row_num, column=4, value=user.last_name)
        ws.cell(row=row_num, column=5, value=user.phone or '')
        ws.cell(row=row_num, column=6, value=user.get_role_persian())
        ws.cell(row=row_num, column=7, value=user.personnel_code or '')
        ws.cell(row=row_num, column=8, value='فعال' if user.is_active and user.is_approved else ('در انتظار تایید' if not user.is_approved else 'غیرفعال'))
        ws.cell(row=row_num, column=9, value=user.get_jalali_created_date())
        
        # اعمال رنگ پس‌زمینه به همه سلول‌های ردیف
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            cell.font = Font(name=font_name, size=data_font_size)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # تنظیم عرض ستون‌ها
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    
    today = jdatetime.datetime.now().strftime('%Y%m%d')
    filename = f"لیست_کاربران_{today}.xlsx"
    
    return send_file(temp_file.name, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=filename)

    
# ========== APIهای کمکی برای مدیریت کاربران ==========

@app.route('/admin/api/personnel-by-code')
@login_required
def admin_api_personnel_by_code():
    """دریافت پرسنل بر اساس کد ملی"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    national_code = request.args.get('code', '')
    if not national_code:
        return jsonify({'error': 'کد ملی ارسال نشده است'}), 400
    
    personnel = Personnel.query.filter_by(national_code=national_code, is_deleted=False).first()
    if not personnel:
        return jsonify({'error': 'پرسنل یافت نشد'}), 404
    
    return jsonify({
        'id': personnel.id,
        'national_code': personnel.national_code,
        'full_name': personnel.get_full_name()
    })

# یک route موقت در app.py اضافه کنید
@app.route('/fix-department-id')
@login_required
def fix_department_id():
    if current_user.role != 'admin':
        return 'دسترسی غیرمجاز', 403
    
    import sqlite3
    db_path = os.path.join(os.path.dirname(__file__), 'instance', 'avan_system.db')
    
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # حذف constraint NOT NULL از department_id
    cursor.execute("PRAGMA foreign_keys=OFF")
    cursor.execute("CREATE TABLE personnel_temp AS SELECT * FROM personnel")
    cursor.execute("DROP TABLE personnel")
    cursor.execute('''
        CREATE TABLE personnel (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            national_code VARCHAR(10) NOT NULL,
            first_name VARCHAR(50),
            last_name VARCHAR(50),
            phone VARCHAR(20),
            position VARCHAR(100),
            hire_date VARCHAR(20),
            department_id INTEGER REFERENCES departments(id),
            unit_id INTEGER NOT NULL REFERENCES units(id),
            period_id INTEGER,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            is_deleted BOOLEAN DEFAULT 0
        )
    ''')
    cursor.execute("INSERT INTO personnel SELECT * FROM personnel_temp")
    cursor.execute("DROP TABLE personnel_temp")
    cursor.execute("PRAGMA foreign_keys=ON")
    
    conn.commit()
    conn.close()
    
    return "✅ فیلد department_id اصلاح شد. حالا می‌توانید از انتصاب استفاده کنید."
# ========== APIهای کمکی برای مدیریت کاربران ==========

@app.route('/admin/api/user-active-assignment/<int:user_id>')
@login_required
def admin_api_user_active_assignment(user_id):
    """دریافت انتصاب فعال کاربر"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    user = User.query.get_or_404(user_id)
    personnel = Personnel.query.filter_by(national_code=user.national_code, is_deleted=False).first()
    
    if not personnel:
        return jsonify({'assignment': None})
    
    assignment = PersonnelAssignment.query.filter_by(
        personnel_id=personnel.id, 
        is_active=True
    ).first()
    
    if not assignment:
        return jsonify({'assignment': None})
    
    unit = Unit.query.get(assignment.unit_id)
    return jsonify({
        'assignment': {
            'id': assignment.id,
            'unit_id': assignment.unit_id,
            'unit_name': unit.name if unit else '-',
            'department_id': unit.department_id if unit else None,
            'department_name': unit.department.name if unit and unit.department else '-',
            'start_date': assignment.start_date,
            'assignment_type': assignment.assignment_type,
            'is_active': assignment.is_active
        }
    })


@app.route('/admin/api/user-stats/<int:user_id>')
@login_required
def admin_api_user_stats(user_id):
    """دریافت آمار درخواست‌های کاربر بر اساس سال"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    year = request.args.get('year', 'all')
    user = User.query.get_or_404(user_id)
    
    # آمار پیش‌فرض
    stats = {
        'annual_leave_total': 0,
        'annual_leave_used': 0,
        'annual_leave_remaining': 0,
        'hourly_leave_total': 0,
        'hourly_leave_used': 0,
        'hourly_leave_remaining': 0,
        'overtime_total': 0,
        'mission': 0,
        'sick_leave': 0,
        'deficiency': 0,
        'arbaeen': 0,
        'daily_leave_limit': 25,
        'hourly_leave_limit': 120
    }
    
    # دریافت درخواست‌های تایید شده کاربر
    requests = Request.query.filter_by(requester_id=user_id, status='approved').all()
    
    for req in requests:
        extra = req.get_extra_data()
        jalali_year = ''
        if req.request_date:
            try:
                jalali_year = str(jdatetime.datetime.fromgregorian(datetime=req.request_date).year)
            except:
                pass
        
        if year != 'all' and jalali_year != year:
            continue
        
        if req.request_type == 'annual_leave':
            stats['annual_leave_total'] += 1
            stats['annual_leave_used'] += 1
            if extra.get('leave_type') == 'استعلاجی':
                stats['sick_leave'] += 1
        elif req.request_type == 'hourly_leave':
            stats['hourly_leave_total'] += 1
            stats['hourly_leave_used'] += 1
        elif req.request_type in ['daily_mission', 'official_mission']:
            stats['mission'] += 1
        elif req.request_type == 'overtime':
            stats['overtime_total'] += 1
        elif req.request_type == 'deficiency':
            stats['deficiency'] += 1
        elif req.request_type == 'arbaeen':
            stats['arbaeen'] += 1
    
    stats['annual_leave_remaining'] = stats['daily_leave_limit'] - stats['annual_leave_used']
    stats['hourly_leave_remaining'] = stats['hourly_leave_limit'] - stats['hourly_leave_used']
    
    return jsonify(stats)
    
@app.route('/fix-department-column')
@login_required
def fix_department_column():
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    import sqlite3
    db_path = os.path.join(os.path.dirname(__file__), 'instance', 'avan_system.db')
    
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # حذف constraint NOT NULL
        cursor.execute("PRAGMA foreign_keys=OFF")
        cursor.execute("CREATE TABLE personnel_new AS SELECT * FROM personnel")
        cursor.execute("DROP TABLE personnel")
        cursor.execute('''
            CREATE TABLE personnel (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                national_code VARCHAR(10) NOT NULL,
                first_name VARCHAR(50),
                last_name VARCHAR(50),
                phone VARCHAR(20),
                position VARCHAR(100),
                hire_date VARCHAR(20),
                department_id INTEGER REFERENCES departments(id),
                unit_id INTEGER NOT NULL REFERENCES units(id),
                period_id INTEGER,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                is_deleted BOOLEAN DEFAULT 0
            )
        ''')
        cursor.execute("INSERT INTO personnel SELECT * FROM personnel_new")
        cursor.execute("DROP TABLE personnel_new")
        cursor.execute("PRAGMA foreign_keys=ON")
        
        conn.commit()
        conn.close()
        
        return "✅ ستون department_id اصلاح شد. اکنون nullable=True است."
        
    except Exception as e:
        return f"❌ خطا: {str(e)}"


# ==================== API دریافت کاربران با فیلتر ====================
@app.route('/admin/api/users-with-filters')
@login_required
def admin_api_users_with_filters():
    """دریافت کاربران با فیلترهای اداره و واحد برای افزودن به پرسنل"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    dept_id = request.args.get('dept_id', type=int)
    unit_id = request.args.get('unit_id', type=int)
    search = request.args.get('search', '')
    
    query = User.query.filter_by(is_active=True, is_approved=True)
    
    # فیلتر بر اساس اداره از طریق پرسنل
    if dept_id:
        personnel_in_dept = Personnel.query.filter_by(department_id=dept_id, is_deleted=False).all()
        national_codes = [p.national_code for p in personnel_in_dept]
        if national_codes:
            query = query.filter(User.national_code.in_(national_codes))
    
    # فیلتر بر اساس واحد از طریق پرسنل
    if unit_id:
        personnel_in_unit = Personnel.query.filter_by(unit_id=unit_id, is_deleted=False).all()
        national_codes = [p.national_code for p in personnel_in_unit]
        if national_codes:
            query = query.filter(User.national_code.in_(national_codes))
    
    # جستجو
    if search:
        query = query.filter(
            db.or_(
                User.national_code.contains(search),
                User.first_name.contains(search),
                User.last_name.contains(search)
            )
        )
    
    users = query.order_by(User.last_name).all()
    
    result = []
    for u in users:
        personnel = Personnel.query.filter_by(national_code=u.national_code, is_deleted=False).first()
        department_name = personnel.department.name if personnel and personnel.department else '-'
        unit_name = personnel.unit.name if personnel and personnel.unit else '-'
        department_id = personnel.department_id if personnel else None
        unit_id = personnel.unit_id if personnel else None
        
        has_assignment = False
        if personnel:
            assignment = PersonnelAssignment.query.filter_by(
                personnel_id=personnel.id, 
                is_active=True
            ).first()
            has_assignment = bool(assignment)
        
        result.append({
            'id': u.id,
            'national_code': u.national_code,
            'first_name': u.first_name,
            'last_name': u.last_name,
            'full_name': u.get_full_name(),
            'phone': u.phone or '',
            'department_id': department_id,
            'department_name': department_name,
            'unit_id': unit_id,
            'unit_name': unit_name,
            'has_assignment': has_assignment
        })
    
    return jsonify(result)


# ==================== API دریافت پرسنل یک دوره (برای بررسی تکراری‌ها) ====================
@app.route('/admin/api/personnel-by-period')
@login_required
def admin_api_personnel_by_period():
    """دریافت کدهای ملی پرسنل یک دوره برای بررسی تکراری‌ها"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    period_id = request.args.get('period_id', type=int)
    if not period_id:
        return jsonify([])
    
    personnel = Personnel.query.filter_by(period_id=period_id, is_deleted=False).all()
    return jsonify([{'national_code': p.national_code} for p in personnel])


# ==================== API دریافت ادارات با واحدها (برای فیلتر مودال) ====================
@app.route('/admin/api/departments-with-units')
@login_required
def admin_api_departments_with_units():
    """دریافت ادارات با واحدهای آنها برای فیلتر مودال افزودن کاربران"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    departments = Department.query.filter_by(is_active=True).all()
    result = []
    for dept in departments:
        units = Unit.query.filter_by(department_id=dept.id, is_active=True).all()
        result.append({
            'id': dept.id,
            'name': dept.name,
            'units': [{'id': u.id, 'name': u.name} for u in units]
        })
    
    return jsonify(result)


# ==================== API افزودن گروهی کاربران به پرسنل ====================
@app.route('/admin/api/add-users-to-personnel', methods=['POST'])
@login_required
def admin_api_add_users_to_personnel():
    """افزودن گروهی کاربران به پرسنل با انتخاب دوره"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    user_ids = data.get('user_ids', [])
    period_id = data.get('period_id')
    
    if not user_ids:
        return jsonify({'error': 'هیچ کاربری انتخاب نشده است'}), 400
    
    if not period_id:
        return jsonify({'error': 'لطفاً دوره مورد نظر را انتخاب کنید'}), 400
    
    period = WorkPeriod.query.get(period_id)
    if not period:
        return jsonify({'error': 'دوره یافت نشد'}), 404
    
    users = User.query.filter(User.id.in_(user_ids)).all()
    
    results = {
        'success': [],
        'skipped': [],
        'errors': []
    }
    
    for user in users:
        try:
            # پیدا کردن پرسنل موجود برای این کاربر در دوره انتخابی
            existing_personnel = Personnel.query.filter_by(
                national_code=user.national_code,
                period_id=period_id,
                is_deleted=False
            ).first()
            
            if existing_personnel:
                results['skipped'].append({
                    'national_code': user.national_code,
                    'name': user.get_full_name(),
                    'reason': 'کد ملی قبلاً در این دوره ثبت شده است'
                })
                continue
            
            # پیدا کردن پرسنل عمومی کاربر (بدون دوره)
            existing_personnel_general = Personnel.query.filter_by(
                national_code=user.national_code,
                is_deleted=False
            ).first()
            
            if existing_personnel_general:
                # استفاده از اطلاعات موجود
                department_id = existing_personnel_general.department_id
                unit_id = existing_personnel_general.unit_id
            else:
                department_id = None
                unit_id = None
            
            # ایجاد پرسنل جدید در دوره انتخابی
            new_personnel = Personnel(
                national_code=user.national_code,
                first_name=user.first_name,
                last_name=user.last_name,
                phone=user.phone or '',
                department_id=department_id,
                unit_id=unit_id,
                period_id=period_id
            )
            db.session.add(new_personnel)
            db.session.flush()
            
            # کپی فیلدهای داینامیک از پرسنل قبلی (اگر وجود داشته باشد)
            if existing_personnel_general:
                old_values = PersonnelValue.query.filter_by(
                    personnel_id=existing_personnel_general.id
                ).all()
                for old_val in old_values:
                    if old_val.period_id is None:  # فقط مقادیر عمومی
                        new_val = PersonnelValue(
                            personnel_id=new_personnel.id,
                            field_id=old_val.field_id,
                            period_id=period_id,
                            value_text=old_val.value_text,
                            value_number=old_val.value_number,
                            value_date=old_val.value_date
                        )
                        db.session.add(new_val)
            
            # ایجاد وضعیت کارکرد پیش‌فرض
            work_status = PersonnelWorkStatus(
                personnel_id=new_personnel.id,
                period_id=period_id,
                status='draft'
            )
            db.session.add(work_status)
            
            results['success'].append({
                'national_code': user.national_code,
                'name': user.get_full_name()
            })
            
        except Exception as e:
            db.session.rollback()
            results['errors'].append({
                'national_code': user.national_code,
                'name': user.get_full_name(),
                'error': str(e)
            })
    
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'خطا در ذخیره‌سازی: {str(e)}'}), 500
    
    return jsonify({
        'success': True,
        'results': results,
        'total': len(user_ids),
        'added': len(results['success']),
        'skipped': len(results['skipped']),
        'errors': len(results['errors'])
    })

@app.context_processor
def utility_processor():
    from datetime import datetime
    return {'now': datetime.now}
    
@app.route('/admin/api/sync-avatars', methods=['POST'])
@login_required
def admin_api_sync_avatars():
    """همگام‌سازی عکس‌های پروفایل با کد ملی کاربران"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    import os
    avatars_dir = os.path.join(app.root_path, 'static', 'uploads', 'avatars')
    updated = 0
    errors = []
    
    if not os.path.exists(avatars_dir):
        return jsonify({'error': 'پوشه آواتارها وجود ندارد'}), 404
    
    for filename in os.listdir(avatars_dir):
        # بررسی پسوندهای مجاز
        ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        if ext not in ['jpg', 'jpeg', 'png', 'gif']:
            continue
            
        # استخراج کد ملی از نام فایل (بدون پسوند)
        name_without_ext = os.path.splitext(filename)[0]
        
        # بررسی اینکه کد ملی 10 رقمی باشد
        if name_without_ext.isdigit() and len(name_without_ext) == 10:
            user = User.query.filter_by(national_code=name_without_ext).first()
            if user:
                # اگر کاربر عکس نداشت یا عکسش با فایل موجود مطابقت نداشت
                if user.profile_picture != filename:
                    # حذف عکس قبلی اگر وجود داشت
                    if user.profile_picture:
                        old_path = os.path.join(avatars_dir, user.profile_picture)
                        if os.path.exists(old_path) and old_path != os.path.join(avatars_dir, filename):
                            # عکس قبلی رو حذف نکن، چون ممکنه کاربر دیگری ازش استفاده کنه
                            pass
                    user.profile_picture = filename
                    updated += 1
            else:
                errors.append(f'کاربر با کد ملی {name_without_ext} یافت نشد')
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'updated': updated,
        'errors': errors,
        'message': f'✅ {updated} عکس با موفقیت همگام‌سازی شد' + (f' | ❌ {len(errors)} خطا' if errors else '')
    })
    
    
    
@app.route('/admin/api/upload-user-avatar', methods=['POST'])
@login_required
def admin_api_upload_user_avatar():
    """آپلود عکس پروفایل برای کاربر توسط ادمین"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    if 'avatar' not in request.files:
        return jsonify({'error': 'فایلی ارسال نشده'}), 400
    
    file = request.files['avatar']
    if file.filename == '':
        return jsonify({'error': 'فایلی انتخاب نشده'}), 400
    
    user_id = request.form.get('user_id')
    if not user_id:
        return jsonify({'error': 'شناسه کاربر ارسال نشده'}), 400
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'کاربر یافت نشد'}), 404
    
    # بررسی پسوند
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    if ext not in ['jpg', 'jpeg', 'png', 'gif']:
        return jsonify({'error': 'فقط فایل‌های تصویری مجاز هستند'}), 400
    
    # بررسی حجم (حداکثر 5MB)
    file.seek(0, 2)
    file_size_mb = file.tell() // (1024 * 1024)
    file.seek(0)
    if file_size_mb > 100:
        return jsonify({'error': 'حجم فایل نباید بیشتر از 100 مگابایت باشد'}), 400
    
    # ========== ذخیره با کد ملی ==========
    upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'avatars')
    os.makedirs(upload_dir, exist_ok=True)
    
    # ========== حذف همه عکس‌های قبلی این کاربر ==========
    # 1. حذف عکس با نام قبلی (اگر در دیتابیس ثبت شده)
    if user.profile_picture:
        old_path = os.path.join(upload_dir, user.profile_picture)
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except:
                pass
    
    # 2. حذف همه عکس‌های با کد ملی کاربر (با پسوندهای مختلف)
    for old_ext in ['jpg', 'jpeg', 'png', 'gif']:
        old_file = os.path.join(upload_dir, f"{user.national_code}.{old_ext}")
        if os.path.exists(old_file):
            try:
                os.remove(old_file)
            except:
                pass
    
    # نام فایل = کد ملی کاربر + پسوند
    filename = f"{user.national_code}.{ext}"
    file_path = os.path.join(upload_dir, filename)
    
    # ذخیره فایل جدید
    file.save(file_path)
    
    # به‌روزرسانی در دیتابیس (فقط نام فایل را ذخیره کن)
    user.profile_picture = filename
    db.session.commit()
    
    # ثبت لاگ
    try:
        log = ActivityLog(
            user_id=current_user.id,
            user_name=current_user.get_full_name(),
            message=f"عکس پروفایل کاربر {user.get_full_name()} (کد ملی: {user.national_code}) تغییر یافت",
            badge="تغییر عکس کاربر",
            log_type="success"
        )
        db.session.add(log)
        db.session.commit()
    except:
        pass
    
    return jsonify({
        'success': True,
        'url': f'/static/uploads/avatars/{filename}?t={datetime.now().timestamp()}',
        'filename': filename,
        'user_first_name': user.first_name
    })
    
    
    
# ==================== مدل اعلان‌های سراسری (بنر) ====================
class GlobalAnnouncement(db.Model):
    __tablename__ = 'global_announcements'
    
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=True)
    message = db.Column(db.Text, nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    priority = db.Column(db.Integer, default=0)
    
    # ✅ سایز فیلدها را افزایش دهید
    bg_color = db.Column(db.String(100), default='rgba(59, 130, 246, 0.12)')
    border_color = db.Column(db.String(100), default='rgba(59, 130, 246, 0.3)')
    text_color = db.Column(db.String(20), default='#1e293b')
    icon = db.Column(db.String(50), default='📢')
    
    start_date = db.Column(db.String(20), nullable=True)
    end_date = db.Column(db.String(20), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    created_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    
    creator = db.relationship('User', foreign_keys=[created_by])
    def to_dict(self):
        return {
            'id': self.id,
            'title': self.title,
            'message': self.message,
            'is_active': self.is_active,
            'priority': self.priority,
            'bg_color': self.bg_color,
            'border_color': self.border_color,
            'text_color': self.text_color,
            'icon': self.icon,
            'start_date': self.start_date,
            'end_date': self.end_date,
            'created_at': self.created_at.strftime('%Y/%m/%d %H:%M') if self.created_at else '',
            'created_by': self.creator.get_full_name() if self.creator else 'سیستم'
        }
        
        
# ==================== APIهای اعلان‌های سراسری ====================

@app.route('/api/announcements/active')
def api_get_active_announcements():
    """دریافت اعلان‌های فعال برای نمایش در هدر (عمومی)"""
    now = datetime.now()
    today = jdatetime.datetime.now().strftime('%Y/%m/%d')
    
    query = GlobalAnnouncement.query.filter_by(is_active=True)
    
    # فیلتر بر اساس تاریخ
    announcements = query.order_by(GlobalAnnouncement.priority.desc()).all()
    
    result = []
    for ann in announcements:
        # بررسی تاریخ (اگر تاریخ شروع و پایان تنظیم شده باشد)
        if ann.start_date and ann.start_date > today:
            continue
        if ann.end_date and ann.end_date < today:
            continue
        result.append(ann.to_dict())
    
    return jsonify(result)


@app.route('/admin/api/announcements')
@login_required
def admin_api_get_announcements():
    """دریافت همه اعلان‌ها برای پنل ادمین"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    announcements = GlobalAnnouncement.query.order_by(
        GlobalAnnouncement.priority.desc(),
        GlobalAnnouncement.created_at.desc()
    ).all()
    
    return jsonify([ann.to_dict() for ann in announcements])


@app.route('/admin/api/announcements/create', methods=['POST'])
@login_required
def admin_api_create_announcement():
    """ایجاد اعلان جدید"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    data = request.get_json()
    
    # اعتبارسنجی
    if not data.get('message'):
        return jsonify({'error': 'متن پیام الزامی است'}), 400
    
    announcement = GlobalAnnouncement(
        title=data.get('title', '').strip(),
        message=data.get('message', '').strip(),
        is_active=data.get('is_active', True),
        priority=int(data.get('priority', 0)),
        bg_color=data.get('bg_color', 'rgba(59, 130, 246, 0.12)'),
        border_color=data.get('border_color', 'rgba(59, 130, 246, 0.3)'),
        text_color=data.get('text_color', '#1e293b'),
        icon=data.get('icon', '📢'),
        start_date=data.get('start_date', '').strip(),
        end_date=data.get('end_date', '').strip(),
        created_by=current_user.id
    )
    
    db.session.add(announcement)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'اعلان با موفقیت ایجاد شد', 'id': announcement.id})


@app.route('/admin/api/announcements/<int:ann_id>/edit', methods=['PUT'])
@login_required
def admin_api_edit_announcement(ann_id):
    """ویرایش اعلان"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    ann = GlobalAnnouncement.query.get_or_404(ann_id)
    data = request.get_json()
    
    ann.title = data.get('title', ann.title)
    ann.message = data.get('message', ann.message)
    ann.is_active = data.get('is_active', ann.is_active)
    ann.priority = int(data.get('priority', ann.priority))
    ann.bg_color = data.get('bg_color', ann.bg_color)
    ann.border_color = data.get('border_color', ann.border_color)
    ann.text_color = data.get('text_color', ann.text_color)
    ann.icon = data.get('icon', ann.icon)
    ann.start_date = data.get('start_date', ann.start_date)
    ann.end_date = data.get('end_date', ann.end_date)
    ann.updated_at = datetime.now()
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'اعلان با موفقیت ویرایش شد'})


@app.route('/admin/api/announcements/<int:ann_id>/delete', methods=['DELETE'])
@login_required
def admin_api_delete_announcement(ann_id):
    """حذف اعلان"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    ann = GlobalAnnouncement.query.get_or_404(ann_id)
    db.session.delete(ann)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'اعلان با موفقیت حذف شد'})


@app.route('/admin/api/announcements/<int:ann_id>/toggle', methods=['POST'])
@login_required
def admin_api_toggle_announcement(ann_id):
    """فعال/غیرفعال کردن اعلان"""
    if current_user.role != 'admin':
        return jsonify({'error': 'دسترسی غیرمجاز'}), 403
    
    ann = GlobalAnnouncement.query.get_or_404(ann_id)
    ann.is_active = not ann.is_active
    db.session.commit()
    
    return jsonify({
        'success': True, 
        'is_active': ann.is_active,
        'message': 'اعلان ' + ('فعال' if ann.is_active else 'غیرفعال') + ' شد'
    })
    
    
@app.route('/favicon.ico')
def favicon():
    return '', 204
# ==================== راه‌اندازی ====================
if __name__ == '__main__':
    import sys
    import os
    
    os.environ['PYTHONTHREADDEBUG'] = '0'
    
    with app.app_context():
        db.create_all()
        init_default_settings()
        init_default_admin()
        
        print("\n" + "="*60)
        print("🚀 سامانه آوان با موفقیت راه‌اندازی شد!")
        base_url = Setting.get('base_url', '10.86.109.219')
        port = Setting.get('port', '5000')
        print(f"📍 آدرس: http://{base_url}:{port}")
        print("👤 ادمین: 1234567890 / 1234")
        print("="*60 + "\n")
    
    try:
        app.run(host='0.0.0.0', port=5000, debug=True, use_reloader=False)
    except KeyboardInterrupt:
        print("\n🛑 سامانه متوقف شد")
        sys.exit(0)